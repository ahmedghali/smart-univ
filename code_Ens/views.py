from asyncio.log import logger
from datetime import datetime, timedelta
from pyexpat.errors import messages
from django.http import Http404, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from apps.academique.affectation.models import *
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
import pandas as pd
from apps.academique.departement.models import NivSpeDep, Specialite
from apps.academique.affectation.models import Ens_Dep, Classe
from apps.noyau.commun.models import AnneeUniversitaire
from apps.noyau.authentification.decorators import enseignant_access_required
from apps.noyau.authentification.forms import *
from apps.noyau.commun.models import Niveau, Reforme
from .models import *
from .forms import *
from django.db.models import Sum
from decimal import Decimal
from decimal import Decimal
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db import transaction, models
from django.urls import reverse
from django.views.decorators.http import require_http_methods
from django import template
from django.db.models import Count, Q, Prefetch 
from collections import defaultdict
import math



#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@require_http_methods(["GET"])
def get_student_totals_ajax(request, dep_id, etudiant_id, classe_id):
    """
    Vue AJAX pour obtenir les totaux actuels d'un étudiant
    """
    try:
        etudiant = get_object_or_404(Etudiant, id=etudiant_id)
        classe = get_object_or_404(Classe, id=classe_id)
        
        total_seance = calculate_total_sup_seance_for_etudiant(etudiant, classe)
        # total_examen = calculate_total_sup_examen_for_etudiant(etudiant, classe)
        
        return JsonResponse({
            'success': True,
            'total_seance': float(total_seance),
            # 'total_examen': float(total_examen),
            'total_combined': float(total_seance)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)




#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
def auto_save_points_ajax(request, dep_id, sea_id):
    """
    Sauvegarde automatique des points via AJAX
    """
    try:
        abs_etu_seance_id = request.POST.get('abs_id')
        points_seance = request.POST.get('points_seance', '0.0')
        # points_examen = request.POST.get('points_examen', '0.0')
        
        abs_etu_seance = get_object_or_404(Abs_Etu_Seance, id=abs_etu_seance_id)
        
        # Mettre à jour les points
        abs_etu_seance.points_sup_seance = Decimal(str(points_seance))
        # abs_etu_seance.points_sup_examen = Decimal(str(points_examen))
        abs_etu_seance.save()
        
        # Recalculer les totaux
        recalculate_totals_for_etudiant(
            abs_etu_seance.etudiant, 
            abs_etu_seance.seance.classe
        )
        
        # Obtenir les nouveaux totaux
        new_total_seance = calculate_total_sup_seance_for_etudiant(
            abs_etu_seance.etudiant, 
            abs_etu_seance.seance.classe
        )
        # new_total_examen = calculate_total_sup_examen_for_etudiant(
        #     abs_etu_seance.etudiant, 
        #     abs_etu_seance.seance.classe
        # )
        
        return JsonResponse({
            'success': True,
            'new_total_seance': float(new_total_seance),
            # 'new_total_examen': float(new_total_examen),
            'message': 'تم الحفظ تلقائياً'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)





#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
def debug_totals_for_classe(classe_id):
    """
    Fonction de debug pour vérifier les totaux d'une classe
    """
    try:
        classe = Classe.objects.get(id=classe_id)
        abs_etu_classes = Abs_Etu_Classe.objects.filter(classe=classe)
        
        for abs_etu_classe in abs_etu_classes:
            # Calculer depuis les séances
            total_seance_calc = calculate_total_sup_seance_for_etudiant(
                abs_etu_classe.etudiant, classe
            )
            
    except Exception as e:
        pass




#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
def fix_all_totals_command():
    """
    Fonction pour corriger tous les totaux dans toutes les classes
    À utiliser en cas de problème de synchronisation
    """
    try:
        classes = Classe.objects.all()
        total_fixed = 0
        
        for classe in classes:
            fixed_count = recalculate_all_totals_for_classe_complete(classe.id)
            total_fixed += fixed_count
            # print(f"✅ Classe {classe}: {fixed_count} étudiants corrigés")
        
        # print(f"\n🎯 TOTAL: {total_fixed} enregistrements corrigés")
        return total_fixed
        
    except Exception as e:
        # print(f"❌ Erreur correction globale: {e}")
        return 0




#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def dashboard_Ens(request, dep_id, enseignant, departement):       
    try:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    except:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant)

    context = {
        'title': 'لوحة التحكم',
        'my_Ens': enseignant,
        'my_Dep': departement,  # Passer le département au template
        'my_Fac': departement.faculte,  # Passer le département au template
        'real_Dep': real_Dep,
    }
    return render(request, 'enseignant/dashboard_Ens.html', context)




#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def dashboard_Ens(request, dep_id, enseignant, departement):  
    """
    Dashboard principal pour l'enseignant - Version simplifiée et robuste
    """
    try:
        # Données de base
        my_Ens = enseignant
        my_Dep = departement
        my_Fac = my_Dep.faculte
        today = timezone.now().date()
        
        # Mapping des jours pour aujourd'hui
        day_mapping = {
            'Monday': 'Lundi',
            'Tuesday': 'Mardi', 
            'Wednesday': 'Mercredi',
            'Thursday': 'Jeudi',
            'Friday': 'Vendredi',
            'Saturday': 'Samedi',
            'Sunday': 'Dimanche'
        }
        today_french = day_mapping.get(today.strftime('%A'), 'Lundi')
        
        # Gestion du semestre
        semestre_selected = request.GET.get('semestre', '1')
        try:
            semestre_obj = Semestre.objects.get(numero=semestre_selected)
        except Semestre.DoesNotExist:
            semestre_obj = Semestre.objects.filter(numero='1').first()
            if not semestre_obj:
                # Créer un semestre par défaut si aucun n'existe
                semestre_obj = Semestre.objects.create(numero='1', nom='السداسي الأول')
            semestre_selected = str(semestre_obj.numero)
        
        all_semestres = Semestre.objects.all().order_by('numero')
        
        # Classes de l'enseignant
        try:
            mes_classes = Classe.objects.filter(
                enseignant__enseignant=my_Ens,
                semestre=semestre_obj
            ).select_related('matiere', 'niv_spe_dep_sg', 'semestre').order_by('jour', 'temps', 'matiere__nom_fr')
        except Exception as e:
            # print(f"Erreur récupération classes: {e}")
            mes_classes = Classe.objects.none()
        
        # ========== STATISTIQUES DE BASE ==========
        
        # Comptage sécurisé des étudiants
        total_etudiants = 0
        try:
            niv_spe_dep_sgs = mes_classes.values_list('niv_spe_dep_sg', flat=True).distinct()
            for niv_spe_dep_sg_id in niv_spe_dep_sgs:
                count = Etudiant.objects.filter(niv_spe_dep_sg_id=niv_spe_dep_sg_id).count()
                total_etudiants += count
        except Exception as e:
            # print(f"Erreur comptage étudiants: {e}")
            total_etudiants = 0
        
        # Statistiques principales
        stats_enseignant = {
            'total_classes': mes_classes.count(),
            'total_matieres': mes_classes.values('matiere').distinct().count(),
            'total_etudiants': total_etudiants,
            'niveaux_count': mes_classes.values('niv_spe_dep_sg__niv_spe_dep').distinct().count(),
            'specialites_count': mes_classes.values('niv_spe_dep_sg__niv_spe_dep__specialite').distinct().count(),
            'volume_horaire': round(mes_classes.count() * 1.5, 1),
        }
        
        # Taux de réalisation
        classes_avec_seances = mes_classes.filter(seance_created=True).count()
        if mes_classes.count() > 0:
            taux_realisation = round((classes_avec_seances / mes_classes.count()) * 100, 1)
        else:
            taux_realisation = 0
        stats_enseignant['taux_realisation'] = taux_realisation
        
        # Répartition par type
        stats_repartition = {
            'cours': mes_classes.filter(type='Cours').count(),
            'td': mes_classes.filter(type='TD').count(),
            'tp': mes_classes.filter(type='TP').count(),
            'ss': mes_classes.filter(type='Sortie Scientifique').count(),
        }
        
        # ========== EMPLOI DU TEMPS AUJOURD'HUI ==========
        
        try:
            classes_today = mes_classes.filter(jour=today_french).order_by('temps')
            
            # Mise à jour simple des taux d'avancement
            for classe in classes_today:
                if classe.seance_created and not hasattr(classe, 'taux_avancement'):
                    try:
                        all_seances = Seance.objects.filter(classe=classe).count()
                        seances_faites = Seance.objects.filter(classe=classe, fait=True).count()
                        if all_seances > 0:
                            classe.taux_avancement = round((seances_faites / all_seances) * 100)
                        else:
                            classe.taux_avancement = 0
                    except:
                        classe.taux_avancement = 0
        except Exception as e:
            # print(f"Erreur classes aujourd'hui: {e}")
            classes_today = []
        
        # ========== STATISTIQUES D'ASSIDUITÉ SIMPLIFIÉES ==========
        
        stats_assiduite = {
            'taux_presence_general': 85.0,  # Valeur par défaut
            'total_absences': 0,
            'absences_justifiees': 0,
            'etudiants_risque': 0,
        }
        
        try:
            # Calcul simple des absences
            absences_total = 0
            absences_justifiees_total = 0
            
            for classe in mes_classes:
                if classe.abs_liste_Etu:
                    absences_classe = Abs_Etu_Classe.objects.filter(classe=classe)
                    absences_total += sum([abs.nbr_absence or 0 for abs in absences_classe])
                    absences_justifiees_total += sum([abs.nbr_absence_justifiee or 0 for abs in absences_classe])
            
            stats_assiduite.update({
                'total_absences': absences_total,
                'absences_justifiees': absences_justifiees_total,
                'etudiants_risque': max(0, absences_total - absences_justifiees_total) // 3,  # Estimation
            })
            
            # Recalcul du taux de présence si on a des données
            if absences_total > 0:
                total_seances_estimees = mes_classes.count() * 10  # Estimation
                taux_presence = max(0, ((total_seances_estimees - absences_total) / total_seances_estimees) * 100)
                stats_assiduite['taux_presence_general'] = round(taux_presence, 1)
                
        except Exception as e:
            pass
            # print(f"Erreur statistiques assiduité: {e}")
        
        # ========== ACTIVITÉS RÉCENTES SIMPLIFIÉES ==========
        
        activites_recentes = []
        try:
            # Ajouter quelques activités basiques
            for i, classe in enumerate(mes_classes[:3]):
                activites_recentes.append({
                    'type': 'classe',
                    'titre': f'{classe.matiere.nom_fr}',
                    'description': f'{classe.type} - {classe.niv_spe_dep_sg}',
                    'date': today - timedelta(days=i),
                })
        except Exception as e:
            pass
            # print(f"Erreur activités récentes: {e}")
        
        # ========== NOTIFICATIONS SIMPLES ==========
        
        notifications = []
        
        # Classes sans séances
        classes_sans_seances = mes_classes.filter(seance_created=False).count()
        if classes_sans_seances > 0:
            notifications.append({
                'type': 'warning',
                'icon': 'exclamation-triangle',
                'titre': 'تنبيه:',
                'message': f'لديك {classes_sans_seances} حصة بدون دروس منشأة'
            })
        
        # Classes d'aujourd'hui
        if classes_today:
            notifications.append({
                'type': 'info',
                'icon': 'calendar-day',
                'titre': 'تذكير:',
                'message': f'لديك {len(classes_today)} حصة مجدولة اليوم'
            })
        
        # Message de bienvenue si tout va bien
        if not notifications:
            notifications.append({
                'type': 'success',
                'icon': 'check-circle',
                'titre': 'ممتاز!',
                'message': 'جميع حصصك منظمة بشكل جيد'
            })
        
        # ========== CONTEXTE FINAL ==========
        
        context = {
            'title': 'لوحة التحكم - الأستاذ',
            'my_Ens': my_Ens,
            'my_Dep': my_Dep,
            'my_Fac': my_Fac,
            'today': today,
            'all_semestres': all_semestres,
            'semestre_selected': semestre_selected,
            'semestre_obj': semestre_obj,
            'mes_classes': mes_classes,
            'classes_today': classes_today,
            'stats_enseignant': stats_enseignant,
            'stats_repartition': stats_repartition,
            'stats_assiduite': stats_assiduite,
            'activites_recentes': activites_recentes,
            'notifications': notifications,
        }
        
        return render(request, 'enseignant/dashboard_Ens.html', context)
        
    except Exception as e:
        # print(f"Erreur générale dashboard: {e}")
        messages.error(request, f'حدث خطأ في تحميل لوحة التحكم: {str(e)}')
        
        # Context minimal de fallback
        context = {
            'title': 'لوحة التحكم - الأستاذ',
            'my_Ens': enseignant,
            'my_Dep': departement,
            'my_Fac': departement.faculte,
            'today': timezone.now().date(),
            'all_semestres': Semestre.objects.all().order_by('numero'),
            'semestre_selected': '1',
            'mes_classes': [],
            'classes_today': [],
            'stats_enseignant': {
                'total_classes': 0,
                'total_matieres': 0,
                'total_etudiants': 0,
                'niveaux_count': 0,
                'specialites_count': 0,
                'volume_horaire': 0,
                'taux_realisation': 0,
            },
            'stats_repartition': {'cours': 0, 'td': 0, 'tp': 0, 'ss': 0},
            'stats_assiduite': {
                'taux_presence_general': 0,
                'total_absences': 0,
                'absences_justifiees': 0,
                'etudiants_risque': 0,
            },
            'activites_recentes': [],
            'notifications': [{
                'type': 'danger',
                'icon': 'exclamation-circle',
                'titre': 'خطأ:',
                'message': 'حدث خطأ في تحميل البيانات'
            }],
        }
        return render(request, 'enseignant/dashboard_Ens.html', context)




#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def dashboard_ajax_stats(request, enseignant, departement): 
    """
    Vue AJAX pour mettre à jour les statistiques en temps réel
    """
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            semestre_selected = request.GET.get('semestre', '1')
            
            # Recalculer les statistiques rapidement
            semestre_obj = Semestre.objects.get(numero=semestre_selected)
            mes_classes = Classe.objects.filter(
                enseignant__enseignant=enseignant,
                semestre=semestre_obj
            )
            
            stats = {
                'total_classes': mes_classes.count(),
                'classes_avec_seances': mes_classes.filter(seance_created=True).count(),
                'taux_realisation': 0,
            }
            
            if mes_classes.exists():
                stats['taux_realisation'] = round(
                    (stats['classes_avec_seances'] / stats['total_classes']) * 100, 1
                )
            
            return JsonResponse({'success': True, 'stats': stats})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Requête non autorisée'})



#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def profile_Ens(request, dep_id, enseignant, departement): 
    try:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    except:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant)   

    # Aff de l'Ens dans des différente Fac et Dep
    try:
        ALL_Dep = Ens_Dep.objects.filter(
            enseignant = enseignant,
            ).exclude(departement=real_Dep.departement)
    except:
        ALL_Dep = None

    context = {
        'title': 'صفحة التعريف',
        'my_Ens': enseignant,
        'my_Dep': departement,  # Passer le département au template
        'my_Fac': departement.faculte,  # Passer le département au template
        'real_Dep': real_Dep,
        'ALL_Dep': ALL_Dep,
    }
    return render(request, 'enseignant/profile_Ens.html', context)




#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def profileUpdate_Ens(request, dep_id, enseignant, departement): 
    try:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    except:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant)  

    if request.method == 'POST':
        User_form = UserUpdateForm(request.POST, instance=request.user)
        Ens_form = profileUpdate_Ens_Form(request.POST, instance=enseignant)
        
        # Débogage détaillé
        print("=" * 50)
        print("POST Data:", request.POST)
        print("=" * 50)
        print("User_form is_valid:", User_form.is_valid())
        print("User_form errors:", User_form.errors)
        print("=" * 50)
        print("Ens_form is_valid:", Ens_form.is_valid())
        print("Ens_form errors:", Ens_form.errors)
        print("=" * 50)
        
        if User_form.is_valid() and Ens_form.is_valid():
            User_form.save()
            Ens_form.save()
            messages.success(request, 'تم تحديث المعلومات بنجاح')
            return redirect('ense:profile_Ens', dep_id=departement.id)
        else:
            messages.error(request, 'يرجى تصحيح الأخطاء في النموذج')
    else:
        User_form = UserUpdateForm(instance=request.user)
        Ens_form = profileUpdate_Ens_Form(instance=enseignant)

    context = {
        'title': 'تعديل المعلومات',
        'User_form': User_form,
        'Ens_form': Ens_form,
        'my_Ens': enseignant,
        'my_Dep': departement,
        'my_Fac': departement.faculte,
        'real_Dep': real_Dep,
    }
    return render(request, 'enseignant/profileUpdate_Ens.html', context)




#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def change_password_Ens(request, dep_id, enseignant, departement): 
    try:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    except:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant) 

    if request.method == 'POST':
        old_password = request.POST['old_password']
        new_password = request.POST['new_password']
        confirm_password = request.POST['confirm_password']
        user = request.user
        # Vérifier si l'ancien mot de passe est correct
        if not user.check_password(old_password):
            messages.warning(request, "كلمة المرور القديمة خاطئة.")
            return redirect(reverse('ense:change_password_Ens', kwargs={'dep_id': dep_id}))
        # Vérifier si les deux nouveaux mots de passe correspondent
        if new_password != confirm_password:
            messages.warning(request, "كلمة المرور غير متطابقة.")
            return redirect(reverse('ense:change_password_Ens', kwargs={'dep_id': dep_id}))
        # Mettre à jour le mot de passe
        user.set_password(new_password)
        user.save()
        # Maintenir l'utilisateur connecté après le changement de mot de passe
        update_session_auth_hash(request, user)
        messages.success(request, "تم تغيير كلمة المرور بنجاح.")
        # return redirect('timeTable_Ens')  # Rediriger vers le profil ou autre page après le succès
        # return render(request, 'app_enseignant/dashboard_Ens.html', context)
        return redirect(reverse('ense:dashboard_Ens', kwargs={'dep_id': dep_id}))

    context = {
        'title': 'تغيير كلمة المرور',
        'my_Ens': enseignant,
        'my_Dep': departement,  # Passer le département au template
        'my_Fac': departement.faculte,  # Passer le département au template
        'real_Dep': real_Dep,
    }
    return render(request, 'enseignant/change_password_Ens.html',context)





#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def timeTable_Ens(request, dep_id, enseignant, departement): 
    try:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    except:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant)  
    
    # Récupérer le paramètre semestre (par défaut S1)
    semestre_selected = request.GET.get('semestre', '1')
    
    # Récupérer les objets semestre
    try:
        semestre_obj = Semestre.objects.get(numero=semestre_selected)
    except:
        semestre_obj = Semestre.objects.filter(numero='1').first()
        semestre_selected = '1'
    
    my_Dep = Departement.objects.get(id=departement.id)
    
    # Variables communes
    ttEns_0 = 0
    all_Classe_Ens = []
    nbr_Cours = 0
    nbr_TP = 0
    nbr_TD = 0
    nbr_SS = 0
    reload = False

    all_Ens_Dep = Ens_Dep.objects.filter(departement=departement.id).order_by('enseignant__nom_ar')
    
    # Filtrer par semestre sélectionné
    base_filter = {
        'enseignant__departement': my_Dep.id,
        'enseignant__enseignant': enseignant.id,
        'semestre': semestre_obj
    }
    
    # Récupérer les classes par jour
    jours = ['Samedi', 'Dimanche', 'Lundi', 'Mardi', 'Mercredi', 'Jeudi']
    for jour in jours:
        classes_jour = Classe.objects.filter(
            **base_filter,
            jour=jour
        ).order_by('temps')
        all_Classe_Ens.extend(classes_jour)

    # print("Base filter:", base_filter)
    # for jour in jours:
    #     print(f"Recherche pour jour: {jour}")
    #     classes_jour = Classe.objects.filter(**base_filter, jour=jour)
    #     print(f"  -> {classes_jour.count()} classes trouvées")
    
    # print('*'*50)
    # print(all_Classe_Ens)
    # print('*'*50)
    
    # Récupérer les classes par horaire
    all_Classe_Time = []
    horaires = [
        '09:30-08:00',
        '11:10-09:40', 
        '12:50-11:20',
        '14:40-13:10',
        '16:20-14:50',
        '18:00-16:30'
    ]
    
    for horaire in horaires:
        classes_horaire = Classe.objects.filter(
            **base_filter,
            temps=horaire
        )
        all_Classe_Time.extend(classes_horaire)

    sixClasses = horaires
    
    # Compter les types de classes et calculer les taux d'avancement
    for idx1 in all_Classe_Ens:
        if "Cours" == idx1.type:
            nbr_Cours += 1
        elif "TP" == idx1.type:
            nbr_TP += 1
        elif "TD" == idx1.type:
            nbr_TD += 1
        elif "Sortie Scientifique" == idx1.type:
            nbr_SS += 1
        
        if idx1.seance_created == True:
            all_Seances = Seance.objects.filter(classe=idx1.id)
            nbr_Sea_fait = Seance.objects.filter(classe=idx1.id, fait=True)
            
            if nbr_Sea_fait.exists() and all_Seances.exists():
                taux_avancement = ((nbr_Sea_fait.count()) / all_Seances.count()) * 100
                taux_avancement = round(taux_avancement)
                idx1.taux_avancement = taux_avancement
                idx1.save()

    all_classes = nbr_Cours + nbr_TP + nbr_TD + nbr_SS    
    
    sevenDays = ['Samedi', 'Dimanche', 'Lundi', 'Mardi', 'Mercredi', 'Jeudi']
    
    # Récupérer les statistiques depuis Ens_Dep selon le semestre
    if semestre_selected == '1':
        stats_classes = {
            'total': real_Dep.nbrClas_in_Dep_S1,
            'cours': real_Dep.nbrClas_Cours_in_Dep_S1,
            'td': real_Dep.nbrClas_TD_in_Dep_S1,
            'tp': real_Dep.nbrClas_TP_in_Dep_S1,
            'ss': real_Dep.nbrClas_SS_in_Dep_S1,
            'jours': real_Dep.nbrJour_in_Dep_S1,
            'vol_horaire': real_Dep.volHor_in_Dep_S1
        }
    else:
        stats_classes = {
            'total': real_Dep.nbrClas_in_Dep_S2,
            'cours': real_Dep.nbrClas_Cours_in_Dep_S2,
            'td': real_Dep.nbrClas_TD_in_Dep_S2,
            'tp': real_Dep.nbrClas_TP_in_Dep_S2,
            'ss': real_Dep.nbrClas_SS_in_Dep_S2,
            'jours': real_Dep.nbrJour_in_Dep_S2,
            'vol_horaire': real_Dep.volHor_in_Dep_S2
        }
    
    # Récupérer tous les semestres disponibles
    all_semestres = Semestre.objects.all().order_by('numero')

    context = {
        'title': 'إستعمال الزمن',
        'all_Classe_Ens': all_Classe_Ens,
        'all_Ens_Dep': all_Ens_Dep,
        'reload': reload,
        'my_Dep': departement,
        'my_Fac': departement.faculte,
        'my_Ens': enseignant,
        'ttEns_0': ttEns_0,
        'all_Classe_Time': all_Classe_Time,
        'nbr_Cours': nbr_Cours,
        'nbr_TP': nbr_TP,
        'sevenDays': sevenDays,
        'sixClasses': sixClasses,
        'nbr_TD': nbr_TD,
        'nbr_SS': nbr_SS,
        'all_classes': all_classes,
        'semestre_selected': semestre_selected,
        'semestre_obj': semestre_obj,
        'all_semestres': all_semestres,
        'stats_classes': stats_classes,
    }
    return render(request, 'enseignant/timeTable_Ens.html', context)











@enseignant_access_required
def update_classe_moodle(request, dep_id, classe_id, enseignant, departement):
    """
    Vue pour mettre à jour le lien Moodle et l'observation d'une classe
    """
    if request.method == 'POST':
        try:
            classe = Classe.objects.get(
                id=classe_id,
                enseignant__enseignant=enseignant,
                enseignant__departement=departement
            )
            
            # Mettre à jour les champs
            lien_moodle = request.POST.get('lien_moodle', '').strip()
            observation = request.POST.get('observation', '').strip()
            
            classe.lien_moodle = lien_moodle if lien_moodle else None
            classe.observation = observation
            classe.save()
            
            return JsonResponse({
                'success': True,
                'message': 'تم حفظ التعديلات بنجاح'
            })
            
        except Classe.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'الحصة غير موجودة'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'error': 'طريقة غير مسموحة'
    }, status=405)
















#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def fichePedagogique(request, dep_id, enseignant, departement): 
    """
    Vue pour la fiche pédagogique globale (deux semestres)
    """
    try:
        # Récupérer le département
        my_Dep = get_object_or_404(Departement, id=dep_id)
        try:
            my_Ens = Ens_Dep.objects.get(enseignant=enseignant, departement=departement)
        except (Enseignant.DoesNotExist, Ens_Dep.DoesNotExist):
            pass
        
        # Filtre de base pour les classes
        base_filter_S1 = {
            'enseignant__enseignant': my_Ens.enseignant,
            'enseignant__departement': my_Dep,
            'matiere__semestre__numero': 1
        }
        
        base_filter_S2 = {
            'enseignant__enseignant': my_Ens.enseignant,
            'enseignant__departement': my_Dep,
            'matiere__semestre__numero': 2
        }
        
        # Liste des jours dans l'ordre souhaité
        jours = ['Samedi', 'Dimanche', 'Lundi', 'Mardi', 'Mercredi', 'Jeudi']
        
        # Récupérer les classes pour le semestre 1 par jour
        all_Classe_Ens_S1 = []
        for jour in jours:
            classes_jour = Classe.objects.filter(
                **base_filter_S1,
                jour=jour
            ).order_by('temps')
            all_Classe_Ens_S1.extend(classes_jour)
        
        # Récupérer les classes pour le semestre 2 par jour
        all_Classe_Ens_S2 = []
        for jour in jours:
            classes_jour = Classe.objects.filter(
                **base_filter_S2,
                jour=jour
            ).order_by('temps')
            all_Classe_Ens_S2.extend(classes_jour)
        
        # Calculer les statistiques S1
        all_classes_S1 = len(all_Classe_Ens_S1)
        vol_hor_cours_S1 = sum([2.25 for x in all_Classe_Ens_S1 if x.type == 'Cours'])
        vol_hor_TD_S1 = sum([1.5 for x in all_Classe_Ens_S1 if x.type == 'TD'])
        vol_hor_TP_S1 = sum([1.5 for x in all_Classe_Ens_S1 if x.type == 'TP'])
        vol_hor_Total_S1 = vol_hor_cours_S1 + vol_hor_TD_S1 + vol_hor_TP_S1
        
        # Calculer les statistiques S2
        all_classes_S2 = len(all_Classe_Ens_S2)
        vol_hor_cours_S2 = sum([2.25 for x in all_Classe_Ens_S2 if x.type == 'Cours'])
        vol_hor_TD_S2 = sum([1.5 for x in all_Classe_Ens_S2 if x.type == 'TD'])
        vol_hor_TP_S2 = sum([1.5 for x in all_Classe_Ens_S2 if x.type == 'TP'])
        vol_hor_Total_S2 = vol_hor_cours_S2 + vol_hor_TD_S2 + vol_hor_TP_S2
        
        # Total général
        vol_hor_Total_General = vol_hor_Total_S1 + vol_hor_Total_S2
        
        # Informations de l'université
        univ = "جامعة قاصدي مرباح ورقلة"
        my_Fac = my_Dep.faculte.nom_ar if my_Dep.faculte else "كلية غير محددة"
        
        context = {
            'my_Dep': my_Dep,
            'my_Ens': my_Ens,
            'all_Classe_Ens_S1': all_Classe_Ens_S1,
            'all_Classe_Ens_S2': all_Classe_Ens_S2,
            'all_classes_S1': all_classes_S1,
            'all_classes_S2': all_classes_S2,
            'vol_hor_cours_S1': vol_hor_cours_S1,
            'vol_hor_TD_S1': vol_hor_TD_S1,
            'vol_hor_TP_S1': vol_hor_TP_S1,
            'vol_hor_Total_S1': vol_hor_Total_S1,
            'vol_hor_cours_S2': vol_hor_cours_S2,
            'vol_hor_TD_S2': vol_hor_TD_S2,
            'vol_hor_TP_S2': vol_hor_TP_S2,
            'vol_hor_Total_S2': vol_hor_Total_S2,
            'vol_hor_Total_General': vol_hor_Total_General,
            'univ': univ,
            'my_Fac': my_Fac,
        }
        
        return render(request, 'enseignant/fichePedagogique.html', context)
        
    except Exception as e:
        # Redirection simple sans message d'erreur
        return redirect('ense:timeTable_Ens', dep_id=dep_id)




#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def new_Sceance_Ens(request, dep_id, clas_id, enseignant, departement): 
    try:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    except:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant)  
    # -------------------------------
    my_Dep = Departement.objects.get(id = departement.id)
    
    myClasse = Classe.objects.get(id=clas_id)
    
    # Si la méthode est POST, traiter la création manuelle d'une séance
    if request.method == 'POST':
        date_seance = request.POST.get('date_seance')
        temps_seance = request.POST.get('temps_seance')
        intitule = request.POST.get('intitule', '')
        type_audience = request.POST.get('type_audience', 'groupe_complet')
        sous_groupe_unique_id = request.POST.get('sous_groupe_unique')
        sous_groupes_multiples_ids = request.POST.getlist('sous_groupes_multiples')
        
        try:
            # Vérifier si la séance existe déjà
            existing_seance = Seance.objects.filter(
                classe=myClasse,
                date=date_seance,
                temps=temps_seance
            ).first()
            
            if existing_seance:
                messages.error(request, f'توجد حصة مسجلة مسبقاً في نفس التاريخ والوقت')
                return redirect(reverse('ense:new_Sceance_Ens', kwargs={'dep_id': dep_id, 'clas_id': clas_id}))
            
            # Créer la nouvelle séance
            new_seance = Seance(
                classe=myClasse,
                date=date_seance,
                temps=temps_seance,
                intitule=intitule,
                type_audience=type_audience
            )
            
            # Gérer les sous-groupes selon le type d'audience
            if type_audience == 'sous_groupe' and sous_groupe_unique_id:
                sous_groupe = SousGroupe.objects.get(id=sous_groupe_unique_id)
                new_seance.sous_groupe_unique = sous_groupe
            
            new_seance.save()
            
            # Gérer les sous-groupes multiples après la sauvegarde
            if type_audience == 'multi_sous_groupes' and sous_groupes_multiples_ids:
                sous_groupes = SousGroupe.objects.filter(id__in=sous_groupes_multiples_ids)
                new_seance.sous_groupes_multiples.set(sous_groupes)
            
            messages.success(request, f'تم إنشاء الحصة بنجاح')
            return redirect(reverse('ense:list_Sea_Ens', kwargs={'dep_id': dep_id, 'clas_id': clas_id}))
            
        except Exception as e:
            messages.error(request, f'خطأ في إنشاء الحصة: {str(e)}')
            return redirect(reverse('ense:new_Sceance_Ens', kwargs={'dep_id': dep_id, 'clas_id': clas_id}))
    
    # Si la méthode est GET, afficher le formulaire ou créer automatiquement
    action = request.GET.get('action', 'form')
    
    if action == 'auto_create':
        # Logique de création automatique (améliorée)
        dateStart = myClasse.semestre.date_debut
        dateEnd = myClasse.semestre.date_fin
        
        dic_Days = {'Samedi':'Saturday', 'Dimanche':'Sunday', 'Lundi':'Monday', 'Mardi': 'Tuesday', 'Mercredi':'Wednesday', 'Jeudi':'Thursday', 'Vendredi':'Friday'}
        
        seances_created = 0
        seances_skipped = 0
        
        current_date = dateStart
        while current_date <= dateEnd:
            for key, val in dic_Days.items():
                dateStartName = pd.to_datetime(str(current_date).split(" - ")).day_name()
                
                if dateStartName == val:
                    dateStartName_fr = key
                    
                    if (myClasse.jour == dateStartName_fr) and (current_date <= dateEnd):
                        # Vérifier si une séance existe déjà
                        existing_seance = Seance.objects.filter(
                            classe=myClasse,
                            date=current_date,
                            temps=myClasse.temps
                        ).first()
                        
                        if not existing_seance:
                            new_Seance = Seance(
                                classe=myClasse,
                                date=current_date,
                                temps=myClasse.temps,
                                type_audience='groupe_complet'  # Par défaut
                            )
                            new_Seance.save()
                            seances_created += 1
                        else:
                            seances_skipped += 1
                        
                        current_date = current_date + timedelta(days=1)
                        break
            else:
                current_date = current_date + timedelta(days=1)
        
        myClasse.seance_created = True
        myClasse.save()
        
        if seances_created > 0:
            messages.success(request, f'تم إنشاء {seances_created} حصة جديدة')
        if seances_skipped > 0:
            messages.info(request, f'تم تجاهل {seances_skipped} حصة موجودة مسبقاً')
        
        return redirect(reverse('ense:list_Sea_Ens', kwargs={'dep_id': dep_id, 'clas_id': clas_id}))
    
    # Préparer les données pour le formulaire
    sous_groupes_disponibles = SousGroupe.objects.filter(
        groupe_principal=myClasse.niv_spe_dep_sg,
        actif=True
    ).order_by('ordre_affichage', 'nom')
    
    # Calculer les statistiques
    all_Classe_Ens = Classe.objects.filter(
        enseignant__departement=departement.id,
        enseignant=enseignant.id
    )
    
    nbr_Cours = all_Classe_Ens.filter(type="Cours").count()
    nbr_TP = all_Classe_Ens.filter(type="TP").count()
    nbr_TD = all_Classe_Ens.filter(type="TD").count()
    nbr_SS = all_Classe_Ens.filter(type="Sortie Scientifique").count()
    all_classes = nbr_Cours + nbr_TP + nbr_TD + nbr_SS
    
    context = {
        'title': 'إنشاء حصة جديدة',
        'myClasse': myClasse,
        'sous_groupes_disponibles': sous_groupes_disponibles,
        'my_Dep': departement,
        'my_Fac': departement.faculte,
        'my_Ens': enseignant,
        'nbr_Cours': nbr_Cours,
        'nbr_TP': nbr_TP,
        'nbr_TD': nbr_TD,
        'nbr_SS': nbr_SS,
        'all_classes': all_classes,
    }
    
    return render(request, 'enseignant/new_Seance_Ens.html', context)





#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def list_Sea_Ens(request, dep_id, clas_id, enseignant, departement): 
    try:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    except:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant)  
    # -------------------------------
    my_Dep = Departement.objects.get(id = departement.id)

    taux_avancement = 0
    nbr_Sea = 0
    nbr_Sea_fait = 0
    nbr_Sea_annuler = 0
    nbr_Sea_reste = 0
    nbr_Sea_remplacer = 0

    all_Seances = Seance.objects.filter(classe = clas_id).order_by('date') # plus Ens get JSON
    nbr_Sea_fait = Seance.objects.filter(classe = clas_id, fait = True) # plus Ens get JSON
    nbr_Sea_remplacer = Seance.objects.filter(classe = clas_id, remplacer = True) # plus Ens get JSON
    nbr_Sea_annuler = Seance.objects.filter(classe = clas_id, annuler = True) # plus Ens get JSON
    nbr_Sea_reste = all_Seances.count() - nbr_Sea_fait.count()

    if nbr_Sea_fait.count() != 0 and all_Seances.count() != 0 :
        taux_avancement = (nbr_Sea_fait.count() / all_Seances.count()) * 100
        taux_avancement = round(taux_avancement)
        
    sea_name = Classe.objects.get(id = clas_id)

    context = {
        'title': 'جميع الدروس',
        'all_Seances': all_Seances,
        'nbr_Sea_fait': nbr_Sea_fait,
        'sea_name': sea_name,
        'nbr_Sea_annuler': nbr_Sea_annuler,
        'nbr_Sea_reste': nbr_Sea_reste,
        'taux_avancement': taux_avancement,
        'my_Dep': departement,  # Passer le département au template
        'my_Fac': departement.faculte,  # Passer le département au template
        'my_Ens': enseignant,
    }
    return render(request, 'enseignant/list_Sea_Ens.html', context)







#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def update_seance(request, dep_id, sea_id, enseignant, departement): 
    try:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    except:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant)  
    
    my_Dep = Departement.objects.get(id=departement.id)
    my_Sea = Seance.objects.get(id=sea_id)
    
    # Récupérer les sous-groupes disponibles pour cette classe
    sous_groupes_disponibles = SousGroupe.objects.filter(
        groupe_principal=my_Sea.classe.niv_spe_dep_sg,
        actif=True
    ).order_by('ordre_affichage')
    
    sea_form = Seance_Form(instance=my_Sea)
    my_Abs_Etu_Clas = Abs_Etu_Classe.objects.filter(classe=my_Sea.classe) 

    if request.method == 'POST':
        sea_form = Seance_Form(request.POST, instance=my_Sea)
        if sea_form.is_valid():
            # Accéder au champ 'remplacer' après validation
            if sea_form.cleaned_data.get('remplacer', False):
                sea_form.instance.fait = True
            
            # ✅ NOUVEAU - Traitement simplifié du groupe/sous-groupe
            audience_selection = request.POST.get('audience_selection', 'groupe_complet')

            
            
            if audience_selection == 'groupe_complet':
                # Groupe complet
                my_Sea.type_audience = 'groupe_complet'
                my_Sea.sous_groupe_unique = None
                my_Sea.sous_groupes_multiples.clear()
                
            elif audience_selection.startswith('sous_groupe_'):
                # Sous-groupe spécifique
                sous_groupe_id = audience_selection.replace('sous_groupe_', '')
                try:
                    sous_groupe = SousGroupe.objects.get(
                        id=sous_groupe_id,
                        groupe_principal=my_Sea.classe.niv_spe_dep_sg,
                        actif=True
                    )
                    my_Sea.type_audience = 'sous_groupe'
                    my_Sea.sous_groupe_unique = sous_groupe
                    my_Sea.sous_groupes_multiples.clear()
                except SousGroupe.DoesNotExist:
                    messages.error(request, 'المجموعة الفرعية المختارة غير موجودة')
                    my_Sea.type_audience = 'groupe_complet'
                    my_Sea.sous_groupe_unique = None
            
            # Sauvegarder la séance avec les modifications
            sea_form.save()
            my_Sea.save()

            this_Seance = Seance.objects.get(id=sea_id)

            

            # Si le type de la Classe est different de Cours --> Reste TD/TP/SS
            if this_Seance.classe.type != "Cours":

                if ((this_Seance.fait == True) or (this_Seance.remplacer == True)) and (this_Seance.annuler == False) and (this_Seance.list_abs_etudiant_generee == False):
                    this_Seance.list_abs_etudiant_generee = True

                    if this_Seance.remplacer == True:
                        this_Seance.fait = True
                    this_Seance.save()

                    # Ajoutez ce code dans votre vue update_seance, juste avant l'appel à get_etudiants_concernes()

                    # print("="*80)
                    # print("DIAGNOSTIC COMPLET DE LA STRUCTURE DES DONNÉES")
                    # print("="*80)

                    # # 1. Informations sur la séance
                    # print(f"1. SÉANCE:")
                    # print(f"   ID: {this_Seance.id}")
                    # print(f"   Type audience: {this_Seance.type_audience}")
                    # print(f"   Classe: {this_Seance.classe}")

                    # # 2. Informations sur la classe
                    # print(f"\n2. CLASSE:")
                    # classe = this_Seance.classe
                    # print(f"   ID: {classe.id}")
                    # print(f"   Type: {classe.type}")
                    # print(f"   Groupe (niv_spe_dep_sg): {classe.niv_spe_dep_sg}")

                    # # 3. Diagnostiquer le groupe
                    # print(f"\n3. GROUPE (niv_spe_dep_sg):")
                    # groupe = classe.niv_spe_dep_sg
                    # print(f"   ID: {groupe.id if hasattr(groupe, 'id') else 'N/A'}")
                    # print(f"   Type: {type(groupe)}")

                    # # 4. Lister TOUS les attributs et méthodes du groupe
                    # print(f"\n4. ATTRIBUTS DU GROUPE:")
                    # attributs_groupe = []
                    # for attr in dir(groupe):
                    #     if not attr.startswith('_'):
                    #         try:
                    #             value = getattr(groupe, attr)
                    #             if hasattr(value, 'all'):  # C'est probablement une relation
                    #                 try:
                    #                     count = value.all().count()
                    #                     attributs_groupe.append(f"   ✅ {attr} (Manager): {count} objets")
                    #                 except:
                    #                     attributs_groupe.append(f"   ❓ {attr} (Manager): erreur count")
                    #             elif callable(value):
                    #                 attributs_groupe.append(f"   📞 {attr} (Méthode)")
                    #             else:
                    #                 attributs_groupe.append(f"   📄 {attr}: {value}")
                    #         except Exception as e:
                    #             attributs_groupe.append(f"   ❌ {attr}: erreur ({e})")

                    # for attr in attributs_groupe[:20]:  # Limiter l'affichage
                    #     print(attr)

                    # # 5. Diagnostiquer la classe
                    # print(f"\n5. ATTRIBUTS DE LA CLASSE:")
                    # attributs_classe = []
                    # for attr in dir(classe):
                    #     if not attr.startswith('_') and ('etudiant' in attr.lower() or 'student' in attr.lower()):
                    #         try:
                    #             value = getattr(classe, attr)
                    #             if hasattr(value, 'all'):
                    #                 try:
                    #                     count = value.all().count()
                    #                     attributs_classe.append(f"   ✅ {attr} (Manager): {count} objets")
                    #                 except:
                    #                     attributs_classe.append(f"   ❓ {attr} (Manager): erreur count")
                    #             else:
                    #                 attributs_classe.append(f"   📄 {attr}: {value}")
                    #         except Exception as e:
                    #             attributs_classe.append(f"   ❌ {attr}: erreur ({e})")

                    # for attr in attributs_classe:
                    #     print(attr)

                    # # 6. Chercher tous les modèles qui pourraient contenir des étudiants
                    # print(f"\n6. RECHERCHE DE MODÈLES ÉTUDIANTS:")
                    # from django.apps import apps

                    # modeles_etudiants = []
                    # for model in apps.get_models():
                    #     model_name = model.__name__.lower()
                    #     if any(word in model_name for word in ['etudiant', 'student', 'eleve', 'apprenant']):
                    #         try:
                    #             count = model.objects.filter(actif=True).count() if hasattr(model, 'objects') else 0
                    #             modeles_etudiants.append(f"   📚 {model.__name__}: {count} actifs")
                    #         except:
                    #             try:
                    #                 count = model.objects.all().count() if hasattr(model, 'objects') else 0
                    #                 modeles_etudiants.append(f"   📚 {model.__name__}: {count} total")
                    #             except:
                    #                 modeles_etudiants.append(f"   📚 {model.__name__}: erreur")

                    # for modele in modeles_etudiants:
                    #     print(modele)

                    # # 7. Chercher des modèles d'inscription/affectation
                    # print(f"\n7. MODÈLES D'INSCRIPTION/AFFECTATION:")
                    # modeles_inscription = []
                    # for model in apps.get_models():
                    #     model_name = model.__name__.lower()
                    #     if any(word in model_name for word in ['inscription', 'affectation', 'classe', 'groupe', 'enrollment']):
                    #         try:
                    #             count = model.objects.all().count() if hasattr(model, 'objects') else 0
                    #             modeles_inscription.append(f"   📝 {model.__name__}: {count} enregistrements")
                    #         except:
                    #             modeles_inscription.append(f"   📝 {model.__name__}: erreur")

                    # for modele in modeles_inscription[:10]:  # Limiter l'affichage
                    #     print(modele)

                    # # 8. Tester des requêtes spécifiques si on trouve des modèles pertinents
                    # print(f"\n8. TESTS SPÉCIFIQUES:")
                    # try:
                    #     # Essayer de trouver un modèle Etudiant
                    #     from django.apps import apps
                    #     etudiant_model = None
                        
                    #     for model in apps.get_models():
                    #         if model.__name__.lower() in ['etudiant', 'student']:
                    #             etudiant_model = model
                    #             break
                        
                    #     if etudiant_model:
                    #         print(f"   ✅ Modèle étudiant trouvé: {etudiant_model.__name__}")
                            
                    #         # Lister les champs de ce modèle
                    #         print(f"   📋 Champs du modèle étudiant:")
                    #         for field in etudiant_model._meta.get_fields():
                    #             print(f"      - {field.name} ({type(field).__name__})")
                            
                    #         # Compter les étudiants
                    #         total = etudiant_model.objects.all().count()
                    #         print(f"   📊 Total étudiants: {total}")
                            
                    #         if hasattr(etudiant_model, 'objects'):
                    #             try:
                    #                 actifs = etudiant_model.objects.filter(actif=True).count()
                    #                 print(f"   📊 Étudiants actifs: {actifs}")
                    #             except:
                    #                 print(f"   📊 Pas de champ 'actif'")
                    #     else:
                    #         print(f"   ❌ Aucun modèle étudiant trouvé")

                    # except Exception as e:
                    #     print(f"   ❌ Erreur dans les tests: {e}")

                    # print("="*80)

                    # print('/'*50)
                    # print(this_Seance)
                    # print('/'*50)

                    # ✅ NOUVEAU - Utiliser la nouvelle méthode pour récupérer les étudiants selon le type d'audience
                    etudiants_concernes = this_Seance.get_etudiants_concernes()
                    # print('*'*50)
                    # print(etudiants_concernes)
                    # print('*'*50)
                    
                    # Créer les absences pour les étudiants concernés
                    for etudiant in etudiants_concernes:
                        # Déterminer le sous-groupe concerné
                        sous_groupe_concerne = None
                        if this_Seance.type_audience == 'sous_groupe':
                            sous_groupe_concerne = this_Seance.sous_groupe_unique
                        elif this_Seance.type_audience == 'multi_sous_groupes':
                            # Trouver le sous-groupe de cet étudiant parmi ceux sélectionnés
                            for sg in this_Seance.sous_groupes_multiples.all():
                                if EtudiantSousGroupe.objects.filter(etudiant=etudiant, sous_groupe=sg, actif=True).exists():
                                    sous_groupe_concerne = sg
                                    break
                        
                        new_Abs_Etu = Abs_Etu_Seance(
                            seance=this_Seance, 
                            etudiant=etudiant,
                            sous_groupe_concerne=sous_groupe_concerne,
                            type_audience_lors_creation=this_Seance.type_audience
                        )
                        new_Abs_Etu.save()
                
                    # Mettre à jour Abs_Etu_Classe pour tous les étudiants concernés
                    for etudiant in etudiants_concernes:
                        rep = Abs_Etu_Classe.objects.filter(classe=my_Sea.classe.id, etudiant=etudiant)
                        if not rep.exists():
                            new_Abs_Etu_Clas = Abs_Etu_Classe(classe=this_Seance.classe, etudiant=etudiant)
                            new_Abs_Etu_Clas.save()
                        my_Classe = Classe.objects.get(id=this_Seance.classe.id)
                        my_Classe.abs_liste_Etu = True
                        my_Classe.save()
                            
                elif ((this_Seance.fait == False) and (this_Seance.remplacer == False) and 
                    (this_Seance.annuler == False) and (this_Seance.list_abs_etudiant_generee == True)) or \
                    ((this_Seance.fait == False) and (this_Seance.remplacer == False) and 
                    (this_Seance.annuler == True) and (this_Seance.list_abs_etudiant_generee == True)):
                    
                    this_Seance.list_abs_etudiant_generee = False
                    this_Seance.nbr_absence = 0
                    this_Seance.nbr_absence_justifiee = 0
                    this_Seance.save()

                    all_Abs_Etu = Abs_Etu_Seance.objects.filter(seance=this_Seance)
                    for idx in all_Abs_Etu:
                        idx.delete()

                    all_Sea_Cette_Classe = Seance.objects.filter(classe=my_Sea.classe)
                    
                    total = 0 
                    for idxx in all_Sea_Cette_Classe:
                        all_Abs_Etu = Abs_Etu_Seance.objects.filter(seance=idxx)
                        total = total + all_Abs_Etu.count()
                    
                    if total > 0:
                        my_Abs_Etu_Clas = Abs_Etu_Classe.objects.filter(classe=my_Sea.classe)

                        for ascc in all_Sea_Cette_Classe:
                            my_all_Abs_Etu = Abs_Etu_Seance.objects.filter(seance=ascc)
                            for x in my_all_Abs_Etu:
                                for x_i in my_Abs_Etu_Clas:
                                    if x.etudiant == x_i.etudiant:
                                        x_i.nbr_absence = 0
                                        x_i.nbr_absence_justifiee = 0
                                        x_i.save()

                        for ascc in all_Sea_Cette_Classe:
                            my_all_Abs_Etu = Abs_Etu_Seance.objects.filter(seance=ascc)
                            for x in my_all_Abs_Etu:
                                for x_i in my_Abs_Etu_Clas:
                                    if x.etudiant == x_i.etudiant:
                                        if (x.present == False) and (x.justifiee == False):
                                            x_i.nbr_absence += 1
                                        elif (x.present == False) and (x.justifiee == True):
                                            x_i.nbr_absence_justifiee += 1
                                        x_i.save()

                    else: # c'est les derniers enregistrement
                        my_Abs_Etu_Clas = Abs_Etu_Classe.objects.filter(classe=my_Sea.classe)
                        for x_i in my_Abs_Etu_Clas:
                            x_i.delete()
                        
                        my_Classe = Classe.objects.get(id=this_Seance.classe.id)
                        my_Classe.abs_liste_Etu = False
                        my_Classe.save()
                        
                messages.success(request, 'تم تعديل الدرس بنجاح')
                clas_id = my_Sea.classe.id
                return redirect(reverse('ense:list_Sea_Ens', kwargs={'dep_id': dep_id, 'clas_id': clas_id}))
            
            else:
                # Le type == Cours
                messages.success(request, 'تم تعديل الدرس بنجاح')
                clas_id = my_Sea.classe.id
                return redirect(reverse('ense:list_Sea_Ens', kwargs={'dep_id': dep_id, 'clas_id': clas_id}))
        
    else:
        sea_form = Seance_Form(instance=my_Sea)
    
    context = {
        'title': 'تعديل الدرس',
        'my_Sea': my_Sea,
        'sea_form': sea_form,
        'sous_groupes_disponibles': sous_groupes_disponibles,  # ✅ NOUVEAU
        'my_Dep': departement,
        'my_Fac': departement.faculte,
        'my_Ens': enseignant,
    }
    return render(request, 'enseignant/update_seance.html', context)





#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def list_Abs_Etu(request, dep_id, sea_id, enseignant, departement):
    try:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    except:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant)  
    
    my_Dep = Departement.objects.get(id=departement.id)
    my_Sea = Seance.objects.get(id=sea_id)
    
    titre_00 = my_Sea.classe.niv_spe_dep_sg
    titre_01 = my_Sea.classe.matiere.nom_fr
    titre_02 = str(my_Sea.date) + " / " + str(my_Sea.temps)
    
    all_Abs_Etu = Abs_Etu_Seance.objects.filter(seance=sea_id).order_by('etudiant__nom_fr', 'etudiant__prenom_fr')
    
    # Calculer les compteurs initiaux
    present_count = Abs_Etu_Seance.objects.filter(seance=sea_id, present=True).count()
    justified_count = Abs_Etu_Seance.objects.filter(seance=sea_id, justifiee=True).count()

    if request.method == 'POST':
        if 'btn_valide' in request.POST:
            # Sauvegarder les modifications - les signaux feront le reste automatiquement
            for x in all_Abs_Etu:
                abs_id = str(x.id)
                present_key = f"present_{abs_id}"
                justifiee_key = f"justifiee_{abs_id}"
                participation_key = f"participation_{abs_id}"
                points_seance_key = f"points_sup_seance_{abs_id}"
                # points_examen_key = f"points_sup_examen_{abs_id}"
                obs_key = f"observation_{abs_id}"

                x.present = present_key in request.POST and request.POST[present_key] == '1'
                x.justifiee = justifiee_key in request.POST and request.POST[justifiee_key] == '1'
                x.participation = participation_key in request.POST and request.POST[participation_key] == '1'
                
                # Conversion sécurisée en Decimal
                try:
                    x.points_sup_seance = Decimal(str(request.POST.get(points_seance_key, '0.0')))
                except:
                    x.points_sup_seance = Decimal('0.0')
                    
                # try:
                #     x.points_sup_examen = Decimal(str(request.POST.get(points_examen_key, '0.0')))
                # except:
                #     x.points_sup_examen = Decimal('0.0')
                    
                x.obs = request.POST.get(obs_key, '').strip()

                x.save()  # ✅ Les signaux vont automatiquement mettre à jour Abs_Etu_Classe

            # Mettre à jour les compteurs dans Seance
            nbr_Abs = Abs_Etu_Seance.objects.filter(seance=sea_id, present=False, justifiee=False).count()
            nbr_Abs_Just = Abs_Etu_Seance.objects.filter(seance=sea_id, present=False, justifiee=True).count()
            my_Sea.nbr_absence = nbr_Abs
            my_Sea.nbr_absence_justifiee = nbr_Abs_Just
            my_Sea.save()

            messages.success(request, 'تم تسجيل غيابات الطلاب وتحديث النقاط تلقائياً')
            clas_id = my_Sea.classe.id
            return redirect(reverse('ense:list_Sea_Ens', kwargs={'dep_id': dep_id, 'clas_id': clas_id}))

    context = {
        'title': 'غيابات الطلبة',
        'all_Abs_Etu': all_Abs_Etu,
        'titre_00': titre_00,
        'titre_01': titre_01,
        'titre_02': titre_02,
        'my_Dep': departement,
        'my_Fac': departement.faculte,
        'my_Ens': enseignant,
        'present_count': present_count,
        'justified_count': justified_count,
    }
    return render(request, 'enseignant/list_Abs_Etu.html', context)





#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
def recalculate_student_totals(etudiant, classe):
    """
    Recalcule les totaux pour un étudiant - Version simplifiée
    """
    try:       
        # Récupérer ou créer l'enregistrement Abs_Etu_Classe
        abs_etu_classe, created = Abs_Etu_Classe.objects.get_or_create(
            etudiant=etudiant,
            classe=classe,
            defaults={
                'nbr_absence': 0,
                'nbr_absence_justifiee': 0,
                'nbr_seances_totales': 0,
                'total_sup_seance': Decimal('0.0'),
                # 'total_sup_examen': Decimal('0.0'),
                'note_presence': Decimal('0.0'),
                'note_participe_HW': Decimal('0.0'),
                'note_controle_1': Decimal('0.0'),
                'note_controle_2': Decimal('0.0'),
                'note_finale': Decimal('0.0'),
                'obs': ''
            }
        )
        
        # Calculer les totaux depuis toutes les séances
        seances = Abs_Etu_Seance.objects.filter(
            etudiant=etudiant,
            seance__classe=classe
        )
        
        total_seance = Decimal('0.0')        
        for seance in seances:
            if seance.points_sup_seance:
                total_seance += Decimal(str(seance.points_sup_seance))
        
        # Appliquer les limites
        total_seance = min(total_seance, Decimal('3.00'))
        
        # Mettre à jour si différent
        if abs_etu_classe.total_sup_seance != total_seance :# or 
            abs_etu_classe.total_sup_seance = total_seance
            # abs_etu_classe.total_sup_examen = total_examen
            abs_etu_classe.save()
        
    except Exception as e:
        pass
        



#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
def update_absence_counts_simple(classe, my_Abs_Etu_Clas):
    """
    Version simplifiée de la mise à jour des compteurs d'absences
    """
    try:        
        all_Sea_Cette_Classe = Seance.objects.filter(classe=classe).order_by('date')
        
        # Réinitialiser les compteurs
        for abs_classe in my_Abs_Etu_Clas:
            abs_classe.nbr_absence = 0
            abs_classe.nbr_absence_justifiee = 0
            abs_classe.save(update_fields=['nbr_absence', 'nbr_absence_justifiee'])

        # Recalculer les compteurs
        for seance in all_Sea_Cette_Classe:
            my_all_Abs_Etu = Abs_Etu_Seance.objects.filter(seance=seance)
            for x in my_all_Abs_Etu:
                for x_i in my_Abs_Etu_Clas:
                    if x.etudiant == x_i.etudiant:
                        if x.present == False and x.justifiee == False:
                            x_i.nbr_absence += 1
                        elif x.present == False and x.justifiee == True:
                            x_i.nbr_absence_justifiee += 1
                        x_i.save(update_fields=['nbr_absence', 'nbr_absence_justifiee'])
                
    except Exception as e:
        pass





#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
def test_quick_calculation(classe_id, etudiant_id):
    """
    Test rapide pour vérifier le calcul
    """
    try:
        from .models import Classe, Etudiant
        classe = Classe.objects.get(id=classe_id)
        etudiant = Etudiant.objects.get(id=etudiant_id)
            
        # Lister les séances avec points
        seances = Abs_Etu_Seance.objects.filter(
            etudiant=etudiant,
            seance__classe=classe
        )
        
        total_s = 0
        total_e = 0
        for s in seances:
            if s.points_sup_seance > 0 :
                total_s += float(s.points_sup_seance or 0)
        
        # Forcer le recalcul
        recalculate_student_totals(etudiant, classe)
        
    except Exception as e:
        pass
        





#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
def recalculate_totals_for_etudiant(etudiant, classe):
    """
    Recalcule les totaux total_sup_seance et total_sup_examen 
    pour un étudiant dans une classe donnée
    """
    try:
        # Récupérer ou créer l'enregistrement Abs_Etu_Classe
        abs_etu_classe, created = Abs_Etu_Classe.objects.get_or_create(
            etudiant=etudiant,
            classe=classe,
            defaults={
                'nbr_absence': 0,
                'nbr_absence_justifiee': 0,
                'nbr_seances_totales': 0,
                'total_sup_seance': Decimal('0.0'),
                # 'total_sup_examen': Decimal('0.0'),
                'note_presence': Decimal('0.0'),
                'note_participe_HW': Decimal('0.0'),
                'note_controle_1': Decimal('0.0'),
                'note_controle_2': Decimal('0.0'),
                'note_finale': Decimal('0.0'),
                'obs': ''
            }
        )
        

        # Calculer les nouveaux totaux depuis toutes les séances
        total_seance = calculate_total_sup_seance_for_etudiant(etudiant, classe)
        # total_examen = calculate_total_sup_examen_for_etudiant(etudiant, classe)
        
        # Mettre à jour seulement si les valeurs ont changé
        if abs_etu_classe.total_sup_seance != total_seance: # or 
            
            old_total_seance = abs_etu_classe.total_sup_seance            
            abs_etu_classe.total_sup_seance = total_seance
            abs_etu_classe.save()  # La note finale sera recalculée automatiquement

    except Exception as e:
        pass






#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
# ✅ FONCTIONS DE CALCUL DES TOTAUX
def calculate_total_sup_seance_for_etudiant(etudiant, classe):
    """
    Calcule la somme de tous les points_sup_seance pour un étudiant dans une classe
    """
    try:
        total = Abs_Etu_Seance.objects.filter(
            etudiant=etudiant,
            seance__classe=classe
        ).aggregate(
            total=models.Sum('points_sup_seance')
        )['total'] or Decimal('0.0')
        
        # Limiter au maximum autorisé (3.0)
        result = min(Decimal('3.00'), Decimal(str(total)))
        return result
        
    except Exception as e:
        return Decimal('0.0')



#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
# # ✅ FONCTION POUR METTRE À JOUR LES COMPTEURS D'ABSENCES
def update_absence_counts_for_classe(classe):
    """
    Met à jour les compteurs d'absences pour tous les étudiants d'une classe
    """
    try:
        all_Sea_Cette_Classe = Seance.objects.filter(classe=classe).order_by('date')
        my_Abs_Etu_Clas = Abs_Etu_Classe.objects.filter(classe=classe)
        
        # Réinitialiser les compteurs
        for abs_classe in my_Abs_Etu_Clas:
            abs_classe.nbr_absence = 0
            abs_classe.nbr_absence_justifiee = 0
            abs_classe.save(update_fields=['nbr_absence', 'nbr_absence_justifiee'])
        
        # Recalculer les compteurs
        for seance in all_Sea_Cette_Classe:
            abs_seances = Abs_Etu_Seance.objects.filter(seance=seance)
            
            for abs_seance in abs_seances:
                for abs_classe in my_Abs_Etu_Clas:
                    if abs_seance.etudiant == abs_classe.etudiant:
                        if not abs_seance.present and not abs_seance.justifiee:
                            abs_classe.nbr_absence += 1
                        elif not abs_seance.present and abs_seance.justifiee:
                            abs_classe.nbr_absence_justifiee += 1
                        abs_classe.save(update_fields=['nbr_absence', 'nbr_absence_justifiee'])
        
    except Exception as e:
        pass





#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
# ✅ FONCTION DE DEBUG POUR VÉRIFIER LES TOTAUX
def debug_totals_for_student(etudiant, classe):
    """
    Fonction de debug pour vérifier les totaux d'un étudiant
    """
    try:        
        # Lister toutes les séances avec points
        seances = Abs_Etu_Seance.objects.filter(
            etudiant=etudiant,
            seance__classe=classe
        ).select_related('seance')
        
        total_seance_calc = Decimal('0.0')
        # total_examen_calc = Decimal('0.0')
        
        # print("📋 Détail des séances:")
        for seance in seances:
            points_s = seance.points_sup_seance or Decimal('0.0')
            total_seance_calc += points_s
        try:
            abs_etu_classe = Abs_Etu_Classe.objects.get(etudiant=etudiant, classe=classe)                
        except Abs_Etu_Classe.DoesNotExist:
            pass
            
    except Exception as e:
        pass




#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
# ✅ FONCTION UTILITAIRE POUR RECALCULER TOUS LES TOTAUX D'UNE CLASSE
def recalculate_all_totals_for_classe_complete(classe_id):
    """
    Recalcule tous les totaux pour une classe complète
    """
    try:
        classe = Classe.objects.get(id=classe_id)
        
        # Récupérer tous les étudiants qui ont des séances dans cette classe
        etudiants_avec_seances = Abs_Etu_Seance.objects.filter(
            seance__classe=classe
        ).values_list('etudiant', flat=True).distinct()
        
        updated_count = 0
        for etudiant_id in etudiants_avec_seances:
            etudiant = Etudiant.objects.get(id=etudiant_id)
            recalculate_totals_for_etudiant(etudiant, classe)
            updated_count += 1
        return updated_count
        
    except Exception as e:
        return 0




#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def list_Abs_Etu_Classe(request, dep_id, clas_id, enseignant, departement):
    try:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    except:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant)  
    # -------------------------------
    # my_Dep = Departement.objects.get(id=departement.id)
    
    my_Dep = Departement.objects.get(id=dep_id)
    my_Clas = Classe.objects.get(id=clas_id)
    all_sea_fait = Seance.objects.filter(classe=my_Clas, fait=True).order_by('date')
    all_Abs_Classe = Abs_Etu_Classe.objects.filter(classe=my_Clas).order_by('etudiant__nom_fr')
    nbr_Abs_Classe = Abs_Etu_Classe.objects.filter(classe=my_Clas)

    # Précalculer les absences par séance pour chaque étudiant
    all_Abs_Classe_with_absences = []
    for abs_classe in all_Abs_Classe:
        absences_per_seance = []
        for sea in all_sea_fait:
            absence = Abs_Etu_Seance.objects.filter(seance=sea, etudiant=abs_classe.etudiant).first()
            absences_per_seance.append(absence if absence else None)
        abs_classe.absences_per_seance = absences_per_seance
        all_Abs_Classe_with_absences.append(abs_classe)

    titre_00 = my_Clas.niv_spe_dep_sg
    titre_01 = my_Clas.matiere.nom_fr
    # enseignant = request.user  # Récupérer l'enseignant à partir de l'utilisateur connecté
    departement = my_Dep  # Utiliser le département récupéré

    context = {
        'titre_00': titre_00,
        'titre_01': titre_01,
        'all_sea_fait': all_sea_fait,
        'all_Abs_Classe_with_absences': all_Abs_Classe_with_absences,
        'nbr_Abs_Classe': nbr_Abs_Classe,
        'my_Dep': my_Dep,
        'my_Fac': my_Dep.faculte,
        'my_Ens': enseignant,
    }
    return render(request, 'enseignant/list_Abs_Etu_Classe.html', context)






from django.db.models import Q, Count, Avg, Min

@enseignant_access_required
def list_enseignants_ens(request, dep_id, enseignant, departement, semestre_num=1):
    """
    Vue pour afficher la liste des enseignants du département par semestre
    Version simplifiée pour les enseignants
    """
    
    # Récupérer l'année universitaire courante
    # from structure.models import AnneeUniversitaire
    annee_courante = AnneeUniversitaire.get_courante()
    
    if not annee_courante:
        messages.error(request, 'لا توجد سنة جامعية محددة كسنة حالية')
        return redirect('ense:dashboard_Ens', dep_id=dep_id)

    # Valider le numéro de semestre
    if semestre_num not in [1, 2]:
        messages.error(request, 'رقم السداسي غير صحيح')
        return redirect('ense:list_enseignants_ens', dep_id=dep_id, semestre_num=1)

    # Récupérer tous les enseignants du département (année courante + semestre)
    # from departement.models import Ens_Dep
    # from structure.models import Classe
    
    # Filtre avec semestre
    filter_kwargs = {
        'departement': departement,
        'annee_univ': annee_courante,
        'est_inscrit': True,
        f'semestre_{semestre_num}': True
    }
    
    all_Ens_Dep = Ens_Dep.objects.filter(**filter_kwargs).select_related(
        'enseignant__grade',
        'enseignant__user__poste_principal',
        'enseignant'
    )
    
    # Calculer les taux pour le semestre sélectionné
    for ens_dep in all_Ens_Dep:
        classes = Classe.objects.filter(
            enseignant__enseignant=ens_dep.enseignant,
            enseignant__departement=departement,
            semestre=semestre_num
        )
        
        if classes.exists():
            stats = classes.aggregate(
                taux_min=Min('taux_avancement'),
                taux_moy=Avg('taux_avancement')
            )
            ens_dep.taux_min_display = int(stats.get('taux_min', 0) or 0)
            ens_dep.taux_moy_display = round(stats.get('taux_moy', 0) or 0, 1)
        else:
            ens_dep.taux_min_display = 0
            ens_dep.taux_moy_display = 0.0
    
    # Tri personnalisé: statut puis nom alphabétique
    # Ordre des statuts: Permanent, Permanent & Vacataire, Vacataire, Associe, Doctorant
    statut_order = {
        'Permanent': 1,
        'Permanent & Vacataire': 2,
        'Vacataire': 3,
        'Associe': 4,
        'Doctorant': 5
    }
    
    all_Ens_Dep_sorted = sorted(
        all_Ens_Dep,
        key=lambda x: (statut_order.get(x.statut, 99), x.enseignant.nom_ar, x.enseignant.prenom_ar)
    )
    
    # Statistiques
    total_enseignants = len(all_Ens_Dep_sorted)
    count_per = len([x for x in all_Ens_Dep_sorted if x.statut == 'Permanent'])
    count_temp = len([x for x in all_Ens_Dep_sorted if x.statut != 'Permanent'])
    count_scholar = len([x for x in all_Ens_Dep_sorted if x.enseignant.googlescholar])
    count_email = len([x for x in all_Ens_Dep_sorted if x.enseignant.email_prof])
    
    grade_stats = all_Ens_Dep.values('enseignant__grade__nom_ar').annotate(
        count=Count('id')
    ).order_by('-count')
    
    context = {
        'title': 'قائمة الأساتذة',
        'my_Dep': departement,
        'my_Fac': departement.faculte,
        'my_Ens': enseignant,
        'annee_courante': annee_courante,
        'semestre_num': semestre_num,
        'all_Ens_Dep': all_Ens_Dep_sorted,
        'total_enseignants': total_enseignants,
        'count_per': count_per,
        'count_temp': count_temp,
        'count_scholar': count_scholar,
        'count_email': count_email,
        'grade_stats': grade_stats,
    }
    
    return render(request, 'enseignant/list_enseignants_ens.html', context)


















#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def list_Etudiant_Ens(request, dep_id, enseignant, departement):
    """
    Version corrigée qui ignore les sections conteneurs
    """
    
    if request.method == 'POST':
        niv_spe_dep_sg_id = request.POST.get('niv_spe_dep_sg')
        # print(f"🔍 DEBUG: ID reçu = '{niv_spe_dep_sg_id}'")
        
        if niv_spe_dep_sg_id:
            
            if niv_spe_dep_sg_id.startswith('tous_'):
                # print("🔍 DEBUG: Cas TOUS LES ÉTUDIANTS")
                
                # CORRECTION: Extraire niveau_id et reforme_id de l'ID
                if '_' in niv_spe_dep_sg_id:
                    parts = niv_spe_dep_sg_id.replace('tous_', '').split('_')
                    niveau_id = parts[0]
                    reforme_id = parts[1] if len(parts) > 1 else None
                else:
                    niveau_id = niv_spe_dep_sg_id.replace('tous_', '')
                    reforme_id = None
                
                # print(f"🔍 Extrait: niveau_id={niveau_id}, reforme_id={reforme_id}")
                
                # Filtres STRICTS
                groupes_filters = {
                    'niv_spe_dep__departement': departement,
                    'niv_spe_dep__niveau_id': niveau_id
                }
                if reforme_id:
                    groupes_filters['niv_spe_dep__specialite__reforme_id'] = reforme_id
                
                # Récupérer tous les NivSpeDep_SG correspondants
                groupes_niveau = NivSpeDep_SG.objects.filter(**groupes_filters).exclude(
                    section__isnull=False,
                    groupe__isnull=True
                ).values_list('id', flat=True)
                
                # print(f"🔍 Groupes niveau trouvés: {len(list(groupes_niveau))}")
                
                etudiants = Etudiant.objects.filter(
                    niv_spe_dep_sg_id__in=groupes_niveau
                ).distinct().order_by('nom_ar').values(
                    'id', 'nom_ar', 'prenom_ar', 'nom_fr', 'prenom_fr', 'matricule', 'inscrit_univ'
                )
                
            elif niv_spe_dep_sg_id.startswith('section_'):
                # print("🔍 DEBUG: Cas SECTION COMPLÈTE")
                
                # CORRECTION: Extraire section_name et reforme_id de l'ID  
                if '_' in niv_spe_dep_sg_id:
                    parts = niv_spe_dep_sg_id.replace('section_', '').split('_')
                    section_name = parts[0]
                    reforme_id = parts[1] if len(parts) > 1 else None
                else:
                    section_name = niv_spe_dep_sg_id.replace('section_', '')
                    reforme_id = None
                
                # print(f"🔍 Extrait: section_name={section_name}, reforme_id={reforme_id}")
                
                # Filtres STRICTS pour section
                section_filters = {
                    'niv_spe_dep__departement': departement,
                    'section__nom_ar': section_name,
                    'section__isnull': False,
                    'groupe__isnull': False
                }
                if reforme_id:
                    section_filters['niv_spe_dep__specialite__reforme_id'] = reforme_id
                
                groupes_section = NivSpeDep_SG.objects.filter(**section_filters)
                
                # print(f"🔍 Groupes section trouvés: {groupes_section.count()}")
                
                groupes_ids = list(groupes_section.values_list('id', flat=True))
                
                etudiants = Etudiant.objects.filter(
                    niv_spe_dep_sg_id__in=groupes_ids
                ).distinct().order_by('nom_ar').values(
                    'id', 'nom_ar', 'prenom_ar', 'nom_fr', 'prenom_fr', 'matricule', 'inscrit_univ'
                )
                
            else:
                # print("🔍 DEBUG: Cas GROUPE INDIVIDUEL")
                etudiants = Etudiant.objects.filter(
                    niv_spe_dep_sg_id=niv_spe_dep_sg_id
                ).order_by('nom_ar').values(
                    'id', 'nom_ar', 'prenom_ar', 'nom_fr', 'prenom_fr', 'matricule', 'inscrit_univ'
                )
            
            result_data = list(etudiants)
            # print(f"🔍 DEBUG: Données finales: {len(result_data)} étudiants")
            return JsonResponse({'data': result_data})
            
        return JsonResponse({'data': []})


    context = {
    'title': 'قائمة الطلبة',
    'my_Dep': departement,
    'my_Fac': departement.faculte,
    'my_Ens': enseignant,
    
    # CORRECTION: Utiliser les bonnes URLs pour les étudiants
    'refs_url': reverse('ense:get_reformes_json_ens', kwargs={'dep_id': dep_id}),
    'nivs_url': reverse('ense:get_niveaux_json_ens', kwargs={'dep_id': dep_id, 'reforme_id': 0}),
    'niv_spe_dep_sg_url': reverse('ense:get_niv_spe_dep_sg_json_ens', kwargs={'dep_id': dep_id, 'niveau_id': 0}),
    }
    return render(request, 'enseignant/list_etudiants_ens.html', context)




#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
# Vues AJAX adaptées du département
@enseignant_access_required
def get_reformes_json_ens(request, dep_id, enseignant, departement):
    """Récupérer les réformes - Adaptée du département"""
    # Même logique que get_reformes_json du département
    reformes = Reforme.objects.filter(Specialites_reforme__departement=departement).values('id', 'nom_ar', 'nom_fr').distinct()
    return JsonResponse({'data': list(reformes)})


#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required  
def get_niveaux_json_ens(request, dep_id, reforme_id, enseignant, departement):
    """Récupérer les niveaux - Adaptée du département"""
    # Même logique que get_niveaux_json du département
    niveaux = NivSpeDep.objects.filter(
        departement=departement,
        specialite__reforme_id=reforme_id
    ).values('niveau__id', 'niveau__nom_ar', 'niveau__nom_fr').distinct()
    return JsonResponse({'data': list(niveaux)})


#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def get_niv_spe_dep_sg_json_ens(request, dep_id, niveau_id, enseignant, departement):
    """Récupérer les groupes - VERSION DEBUG PROGRESSIVE"""
    try:
        # print(f"🔍 get_niv_spe_dep_sg_json_ens: département = {departement.nom_ar}, niveau_id = {niveau_id}")
        
        # ÉTAPE 1: Récupérer reforme_id depuis la requête
        reforme_id = request.GET.get('reforme_id')
        # print(f"🔍 reforme_id reçu: {reforme_id}")
        
        # ÉTAPE 2: Vérifier les données de base
        total_nivspedep = NivSpeDep.objects.filter(
            departement=departement,
            niveau_id=niveau_id
        ).count()
        # print(f"🔍 NivSpeDep trouvés pour ce département/niveau: {total_nivspedep}")
        
        if reforme_id:
            nivspedep_avec_reforme = NivSpeDep.objects.filter(
                departement=departement,
                niveau_id=niveau_id,
                specialite__reforme_id=reforme_id
            ).count()
            # print(f"🔍 NivSpeDep avec reforme_id {reforme_id}: {nivspedep_avec_reforme}")
        
        # ÉTAPE 3: Construire les filtres
        groupes_filters = {
            'niv_spe_dep__departement': departement,
            'niv_spe_dep__niveau_id': niveau_id
        }
        
        if reforme_id:
            groupes_filters['niv_spe_dep__specialite__reforme_id'] = reforme_id
        
        # print(f"🔍 Filtres utilisés: {groupes_filters}")
        
        # ÉTAPE 4: Récupérer TOUS les NivSpeDep_SG (même vides pour debug)
        groupes = NivSpeDep_SG.objects.filter(**groupes_filters).select_related(
            'niv_spe_dep__niveau',
            'niv_spe_dep__specialite',
            'niv_spe_dep__specialite__reforme',
            'section',
            'groupe'
        ).order_by('section__nom_ar', 'groupe__nom_ar')
        
        data = []
        total_etudiants = 0
        
        for groupe in groupes:
            total_etudiants += groupe.nbr_etudiants_SG
            
            # Construire le nom complet
            niveau_spe = f"{groupe.niv_spe_dep.niveau.nom_ar} - {groupe.niv_spe_dep.specialite.nom_ar}"
            
            if groupe.section and groupe.groupe:
                full_name = f"{niveau_spe} - Section: {groupe.section.nom_ar} - Groupe: {groupe.groupe.nom_ar} ({groupe.nbr_etudiants_SG} étudiants)"
            elif groupe.section:
                full_name = f"{niveau_spe} - Section: {groupe.section.nom_ar} ({groupe.nbr_etudiants_SG} étudiants)"
            elif groupe.groupe:
                full_name = f"{niveau_spe} - Groupe: {groupe.groupe.nom_ar} ({groupe.nbr_etudiants_SG} étudiants)"
            else:
                full_name = f"{niveau_spe} - Groupe {groupe.id} ({groupe.nbr_etudiants_SG} étudiants)"
            
            data.append({
                'id': groupe.id,
                'full_name': full_name
            })
        
        total_etudiants = math.ceil(total_etudiants/2)

        # ÉTAPE 7: Ajouter "TOUS LES ÉTUDIANTS" si des groupes existent
        if groupes.exists():
            niveau_spe = f"{groupes.first().niv_spe_dep.niveau.nom_ar} - {groupes.first().niv_spe_dep.specialite.nom_ar}"
            tous_id = f"tous_{niveau_id}"
            if reforme_id:
                tous_id += f"_{reforme_id}"
                
            data.insert(0, {
                'id': tous_id,
                'full_name': f"{niveau_spe} - TOUS LES ÉTUDIANTS ({total_etudiants} étudiants)"
            })
        
        
        return JsonResponse({'data': data})
        
    except Exception as e:
        # print(f"❌ Erreur get_niv_spe_dep_sg: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)})





#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def list_Mat_Niv_Ens(request, dep_id, enseignant, departement):
    """
    Liste des matières/modules adaptée pour l'enseignant
    """
    
    context = {
        'title': 'قائمة المواد',
        'my_Dep': departement,
        'my_Fac': departement.faculte,
        'my_Ens': enseignant,
        # URLs adaptées pour l'enseignant
        'refs_url': reverse('ense:refs_json_ens', kwargs={'dep_id': dep_id}),
        'nivs_url': reverse('ense:nivs_json_ens', kwargs={'dep_id': dep_id, 'reforme_id': 0}),
        'spts_url': reverse('ense:spts_json_ens', kwargs={'dep_id': dep_id, 'niveau_id': 0}),
        'semestres_url': reverse('ense:semestres_json_ens', kwargs={'dep_id': dep_id}),
        'matieres_url': reverse('ense:matieres_json_ens', kwargs={'dep_id': dep_id}),
    }
    return render(request, 'enseignant/list_Mat_Niv_Ens.html', context)





#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def refs_json_ens(request, dep_id, enseignant, departement):
    """Réformes réelles du département - VERSION PROGRESSIVE"""
    try:       
        nivspedep_dept = NivSpeDep.objects.filter(departement=departement).count()       
        # Version simple d'abord - toutes les réformes
        if nivspedep_dept == 0:
            reformes = Reforme.objects.all().values('id', 'nom_ar', 'nom_fr')[:10]
        else:
            # Filtrage par département via NivSpeDep
            reformes = Reforme.objects.filter(
                Specialites_reforme__nivspedep__departement=departement
            ).values('id', 'nom_ar', 'nom_fr').distinct()
        
        result_data = list(reformes)            
        return JsonResponse({'data': result_data})
        
    except Exception as e:
        # print(f"❌ Erreur refs: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)})




#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required  
def nivs_json_ens(request, dep_id, reforme_id, enseignant, departement):
    """Niveaux réels - VERSION PROGRESSIVE"""
    try:
        total_niveaux = Niveau.objects.count()        
        # Version simple - utiliser NivSpeDep
        niveaux = NivSpeDep.objects.filter(
            departement=departement,
            specialite__reforme_id=reforme_id
        ).values('niveau__id', 'niveau__nom_ar', 'niveau__nom_fr').distinct()
        
        result_data = list(niveaux)        
        return JsonResponse({'data': result_data})
        
    except Exception as e:
        # print(f"❌ Erreur nivs: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)})




#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def spts_json_ens(request, dep_id, niveau_id, enseignant, departement):
    """Spécialités réelles - VERSION PROGRESSIVE"""
    try:
        # Filtrer par département ET niveau
        specialites = NivSpeDep.objects.filter(
            departement=departement,
            niveau_id=niveau_id
        ).values('specialite__id', 'specialite__nom_ar', 'specialite__nom_fr').distinct()
        
        result_data = list(specialites)            
        return JsonResponse({'data': result_data})
        
    except Exception as e:
        # print(f"❌ Erreur spts: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)})




#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def semestres_json_ens(request, dep_id, enseignant, departement):
    """Semestres réels"""
    try:
        # print(f"🔍 semestres_json_ens appelée")
        
        semestres = Semestre.objects.all().values('id', 'nom_ar', 'numero').order_by('numero')
        result_data = list(semestres)
        # print(f"🔍 Semestres trouvés: {len(result_data)}")
        
        return JsonResponse({'data': result_data})
        
    except Exception as e:
        # print(f"❌ Erreur semestres: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)})


@enseignant_access_required
def matieres_json_ens(request, dep_id, enseignant, departement):
    """Matières réelles - VERSION PROGRESSIVE"""
    try:        
        if request.method == 'POST':
            reforme_id = request.POST.get('reforme')
            niveau_id = request.POST.get('niveau') 
            specialite_id = request.POST.get('specialite')
            semestre_id = request.POST.get('semestre')
            
            # Test 1: Vérifier si Matiere existe
            try:
                total_matieres = Matiere.objects.count()
                # print(f"🔍 Total matières dans la base: {total_matieres}")
                
                if total_matieres == 0:
                    return JsonResponse({'data': [], 'message': 'Aucune matière dans la base'})
                    
            except Exception as e:
                return JsonResponse({'error': f'Table Matiere inaccessible: {e}'})
            
            # Test 2: Matières du département
            try:
                matieres_dept = Matiere.objects.filter(niv_spe_dep__departement=departement).count()
                # print(f"🔍 Matières du département: {matieres_dept}")
                
                if matieres_dept == 0:
                    # print("⚠️ Aucune matière pour ce département")
                    return JsonResponse({'data': [], 'message': 'Aucune matière pour ce département'})
                    
            except Exception as e:
                # print(f"❌ Erreur filtrage département: {e}")
                return JsonResponse({'error': f'Erreur filtrage département: {e}'})
            
            # Construire les filtres progressivement
            filters = {'niv_spe_dep__departement': departement}
            
            if reforme_id:
                filters['niv_spe_dep__specialite__reforme_id'] = reforme_id
            if niveau_id:
                filters['niv_spe_dep__niveau_id'] = niveau_id
            if specialite_id:
                filters['niv_spe_dep__specialite_id'] = specialite_id
            if semestre_id:
                filters['semestre_id'] = semestre_id
            
            # print(f"🔍 Filtres appliqués: {filters}")
            
            # Récupérer les matières avec les BONS noms de champs
            matieres = Matiere.objects.filter(**filters).values(
                'id', 'nom_ar', 'nom_fr', 'code', 'coeff', 'credit'  # 'coeff' au lieu de 'coefficient'
            ).order_by('code')
            
            result_data = list(matieres)
            
            return JsonResponse({'data': result_data})
        
        # print("🔍 Method non POST")
        return JsonResponse({'data': []})
        
    except Exception as e:
        # print(f"❌ Erreur matieres: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)})
    


#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def list_Specialite_Ens(request, dep_id, enseignant, departement):
    """
    Liste des spécialités adaptée pour l'enseignant
    Identique à la logique du département
    """
    # Récupérer toutes les spécialités du département
    all_Specialite_Ens = Specialite.objects.filter(
        departement=departement.id
    ).select_related('reforme', 'identification').order_by('reforme')
    
    context = {
        'title': 'قائمة التخصصات',
        'my_Fac': departement.faculte,
        'my_Dep': departement,
        'my_Ens': enseignant,
        'all_Specialite_Ens': all_Specialite_Ens,
    }
    return render(request, 'enseignant/list_Specialite_Ens.html', context)




#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def profile_Etu_Ens(request, dep_id, etudiant_id, enseignant, departement):
    etudiant = get_object_or_404(Etudiant, id=etudiant_id)

    context = {
        'title': f'ملف الطالب - {etudiant.nom_ar}',
        'my_Fac': departement.faculte,
        'my_Dep': departement,
        'my_Ens': enseignant,
        'my_Etu': etudiant,
    }
    return render(request, 'enseignant/profile_Etu_Ens.html', context)



#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def list_NivSpeDep_Ens(request, dep_id, enseignant, departement):
    """Liste des niveaux/spécialités pour l'enseignant"""
    
    all_NivSpeDep = NivSpeDep.objects.filter(
        departement=departement
    ).order_by('specialite__reforme', 'niveau', 'specialite__identification')

    # print(f"🔍 DEBUG: Nombre NivSpeDep trouvés: {all_NivSpeDep.count()}")

    # Calcul pour NivSpeDep (nbr_matieres et nbr_etudiants)
    for x in all_NivSpeDep:
        x.nbr_matieres_s1 = Matiere.objects.filter(niv_spe_dep=x, semestre__numero=1).count()
        x.nbr_matieres_s2 = Matiere.objects.filter(niv_spe_dep=x, semestre__numero=2).count()
        x.nbr_etudiants = Etudiant.objects.filter(niv_spe_dep_sg__niv_spe_dep=x).count()
        x.save()

    # ✅ CALCUL INTELLIGENT pour NivSpeDep_SG selon le type d'affectation
    all_niv_spe_dep_sg = NivSpeDep_SG.objects.filter(niv_spe_dep__departement=departement)

    for niv_spe_dep_sg in all_niv_spe_dep_sg:
        if niv_spe_dep_sg.type_affectation == 'par_groupe':
            # 🎯 AFFECTATION PAR GROUPE
            niv_spe_dep_sg.nbr_etudiants_SG = Etudiant.objects.filter(
                niv_spe_dep_sg=niv_spe_dep_sg
            ).count()
            
        elif niv_spe_dep_sg.type_affectation == 'par_section':
            # 🎯 AFFECTATION PAR SECTION: Version optimisée avec une seule requête
            niv_spe_dep_sg.nbr_etudiants_SG = Etudiant.objects.filter(
                niv_spe_dep_sg__niv_spe_dep=niv_spe_dep_sg.niv_spe_dep,
                niv_spe_dep_sg__section=niv_spe_dep_sg.section,
                niv_spe_dep_sg__type_affectation='par_groupe'
            ).count()
            
        elif niv_spe_dep_sg.type_affectation == 'tous_etudiants':
            # 🎯 AFFECTATION TOUS ÉTUDIANTS
            niv_spe_dep_sg.nbr_etudiants_SG = Etudiant.objects.filter(
                niv_spe_dep_sg__niv_spe_dep=niv_spe_dep_sg.niv_spe_dep
            ).count()
        
        niv_spe_dep_sg.save()

    # Construction des listes pour l'affichage
    all_NivSpeDep_S1 = []
    all_NivSpeDep_S2 = []
    
    for x in all_NivSpeDep:
        # Pour S1
        s1_matieres = Matiere.objects.filter(niv_spe_dep=x, semestre__numero=1)
        if s1_matieres.exists():
            # Créer un objet factice pour l'affichage avec les infos nécessaires
            fake_obj_s1 = type('obj', (object,), {
                'id': x.id,
                'niv_spe_dep': x,
                'semestre': type('semestre', (object,), {'numero': 1})()
            })()
            all_NivSpeDep_S1.append(fake_obj_s1)
        
        # Pour S2
        s2_matieres = Matiere.objects.filter(niv_spe_dep=x, semestre__numero=2)
        if s2_matieres.exists():
            # Créer un objet factice pour l'affichage avec les infos nécessaires
            fake_obj_s2 = type('obj', (object,), {
                'id': x.id,
                'niv_spe_dep': x,
                'semestre': type('semestre', (object,), {'numero': 2})()
            })()
            all_NivSpeDep_S2.append(fake_obj_s2)

    context = {
        'title': 'قائمة المستويات',
        'my_Fac': departement.faculte,
        'my_Dep': departement,
        'my_Ens': enseignant,
        'all_NivSpeDep': all_NivSpeDep,
        'all_NivSpeDep_S1': all_NivSpeDep_S1,
        'all_NivSpeDep_S2': all_NivSpeDep_S2,
    }
    return render(request, 'enseignant/list_NivSpeDep_Ens.html', context)




#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def timeTable_Niv_Ens(request, dep_id, niv_spe_dep_id, semestre_num, enseignant, departement):
    """Emploi du temps d'un NivSpeDep - Version optimisée"""
    try:
        niv_spe_dep = NivSpeDep.objects.get(id=niv_spe_dep_id)
    except NivSpeDep.DoesNotExist:
        return render(request, 'enseignant/timeTable_Niv_Ens.html', {
            'error': f'المستوى/التخصص برقم {niv_spe_dep_id} غير موجود',
            'all_classes': 0,
            'my_Dep': departement,
            'my_Fac': departement.faculte,
            'my_Ens': enseignant,
        })

    try:
        semestre = Semestre.objects.get(numero=semestre_num)
    except Semestre.DoesNotExist:
        return render(request, 'enseignant/timeTable_Niv_Ens.html', {
            'error': f'السداسي رقم {semestre_num} غير موجود',
            'all_classes': 0,
            'my_Dep': departement,
            'my_Fac': departement.faculte,
            'my_Ens': enseignant,
        })

    # ✅ Précharger TOUTES les relations importantes
    all_Classe_Niv = Classe.objects.filter(
        enseignant__departement=departement,
        niv_spe_dep_sg__niv_spe_dep=niv_spe_dep,
        semestre__numero=semestre_num
    ).select_related(
        'enseignant__enseignant',
        'matiere',
        'niv_spe_dep_sg__niv_spe_dep__niveau',
        'niv_spe_dep_sg__niv_spe_dep__specialite__reforme',
        'niv_spe_dep_sg__niv_spe_dep__specialite__identification',
        'niv_spe_dep_sg__niv_spe_dep__specialite__parcours',
        'niv_spe_dep_sg__section',
        'niv_spe_dep_sg__groupe',
    ).order_by('jour', 'temps')

    # Même queryset pour all_Classe_Time
    all_Classe_Time = all_Classe_Niv

    # ✅ Statistiques optimisées (1 seule requête par type)
    nbr_Cours = all_Classe_Niv.filter(type="Cours").count()
    nbr_TP = all_Classe_Niv.filter(type="TP").count()
    nbr_TD = all_Classe_Niv.filter(type="TD").count()
    nbr_SS = all_Classe_Niv.filter(type="Sortie Scientifique").count()

    # ✅ Calcul du taux d'avancement OPTIMISÉ
    # Récupérer toutes les classes avec seance_created = True
    classes_avec_seances = [c for c in all_Classe_Niv if hasattr(c, 'seance_created') and c.seance_created]
    
    if classes_avec_seances:
        # Récupérer tous les IDs des classes concernées
        classe_ids = [c.id for c in classes_avec_seances]
        
        # ✅ 2 requêtes au lieu de 100+ !
        from django.db.models import Count, Case, When, IntegerField
        
        # Compter toutes les séances et celles faites par classe
        seances_stats = Seance.objects.filter(
            classe_id__in=classe_ids
        ).values('classe_id').annotate(
            total=Count('id'),
            faites=Count(Case(When(fait=True, then=1), output_field=IntegerField()))
        )
        
        # Créer un dictionnaire pour accès rapide
        stats_dict = {s['classe_id']: s for s in seances_stats}
        
        # Mettre à jour les classes
        classes_to_update = []
        for classe in classes_avec_seances:
            if classe.id in stats_dict:
                stats = stats_dict[classe.id]
                if stats['total'] > 0:
                    taux = (stats['faites'] / stats['total']) * 100
                    classe.taux_avancement = round(taux)
                    classes_to_update.append(classe)
        
        # Sauvegarde en bulk
        if classes_to_update:
            Classe.objects.bulk_update(classes_to_update, ['taux_avancement'])

    all_classes = nbr_Cours + nbr_TP + nbr_TD + nbr_SS

    sixClassesTimes = [
        '09:30-08:00', '11:10-09:40', '12:50-11:20',
        '14:40-13:10', '16:20-14:50', '18:00-16:30',
    ]
    sevenDays = ['Samedi', 'Dimanche', 'Lundi', 'Mardi', 'Mercredi', 'Jeudi']

    context = {
        'title': 'إستعمال الزمن للمستوى',
        'all_Classe_Niv': all_Classe_Niv,
        'my_Dep': departement,
        'my_Fac': departement.faculte,
        'my_Ens': enseignant,
        'all_Classe_Time': all_Classe_Time,
        'nbr_Cours': nbr_Cours,
        'nbr_TP': nbr_TP,
        'nbr_TD': nbr_TD,
        'nbr_SS': nbr_SS,
        'all_classes': all_classes,
        'sevenDays': sevenDays,
        'sixClassesTimes': sixClassesTimes,
        'niv_spe_dep': niv_spe_dep,
        'niv_spe_dep_id': niv_spe_dep_id,
        'semestre_num': semestre_num,
    }

    return render(request, 'enseignant/timeTable_Niv_Ens.html', context)



#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def fichPeda_Ens_Semestre(request, dep_id, enseignant, departement):
    """
    Vue pour la fiche pédagogique semestrielle
    """
    try:
        # Récupérer le département
        my_Dep = get_object_or_404(Departement, id=dep_id)
        
        # Récupérer l'enseignant connecté
        try:
            my_Ens = Ens_Dep.objects.get(enseignant=enseignant, departement=departement)
        except (Enseignant.DoesNotExist, Ens_Dep.DoesNotExist):
            pass
        
        # Récupérer le semestre sélectionné
        semestre_selected = request.GET.get('semestre', '1')
        
        # Filtre de base pour les classes
        base_filter = {
            'enseignant__enseignant': my_Ens.enseignant,
            'enseignant__departement': my_Dep,
            'matiere__semestre__numero': semestre_selected
        }
        
        # Récupérer les classes par jour dans l'ordre souhaité
        jours = ['Samedi', 'Dimanche', 'Lundi', 'Mardi', 'Mercredi', 'Jeudi']
        all_Classe_Ens = []
        
        for jour in jours:
            classes_jour = Classe.objects.filter(
                **base_filter,
                jour=jour
            ).order_by('temps')
            all_Classe_Ens.extend(classes_jour)
        
        # Calculer les statistiques
        all_classes = len(all_Classe_Ens)
        
        # Calculer les volumes horaires
        vol_hor_cours = sum([2.25 for x in all_Classe_Ens if x.type == 'Cours'])
        vol_hor_TD = sum([1.5 for x in all_Classe_Ens if x.type == 'TD'])
        vol_hor_TP = sum([1.5 for x in all_Classe_Ens if x.type == 'TP'])
        vol_hor_Total = vol_hor_cours + vol_hor_TD + vol_hor_TP
        
        # Informations de l'université
        univ = "جامعة قاصدي مرباح ورقلة"
        my_Fac = my_Dep.faculte.nom_ar if my_Dep.faculte else "كلية غير محددة"
        
        context = {
            'my_Dep': my_Dep,
            'my_Ens': my_Ens,
            'semestre_selected': semestre_selected,
            'all_Classe_Ens': all_Classe_Ens,
            'all_classes': all_classes,
            'vol_hor_cours': vol_hor_cours,
            'vol_hor_TD': vol_hor_TD,
            'vol_hor_TP': vol_hor_TP,
            'vol_hor_Total': vol_hor_Total,
            'univ': univ,
            'my_Fac': my_Fac,
        }
        
        return render(request, 'enseignant/fichPeda_Ens_Semestre.html', context)
        
    except Exception as e:
        messages.error(request, f"خطأ في تحميل البطاقة البيداغوجية: {str(e)}")
        return redirect('ense:timeTable_Ens', dep_id=dep_id)
    



#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def list_Amphi_Ens(request, dep_id, enseignant, departement):
# def list_Amphi_Ens(request, dep_id):
    """
    Vue pour afficher la liste des amphithéâtres du département pour un enseignant
    """
    try:
        # Récupération de l'enseignant connecté
        enseignant = request.user.enseignant_profile
        
        # Récupération du département spécifique
        departement = get_object_or_404(Departement, id=dep_id)
        
        # Vérification que l'enseignant appartient bien à ce département
        try:
            real_Dep = Ens_Dep.objects.get(
                enseignant=enseignant, 
                departement=departement,
                statut='Permanent'
            )
        except Ens_Dep.DoesNotExist:
            # Si pas de relation permanente, chercher une relation temporaire
            try:
                real_Dep = Ens_Dep.objects.get(
                    enseignant=enseignant, 
                    departement=departement
                )
            except Ens_Dep.DoesNotExist:
                messages.error(request, 'غير مصرح لك بالوصول إلى هذا القسم')
                return redirect('ense:dashboard_Ens', dep_id=enseignant.departements.first().id)
        
        # Récupération des amphithéâtres par semestre
        all_Amphi_Dep_S1 = Amphi_Dep.objects.filter(
            departement=departement.id, 
            semestre__numero=1
        ).select_related('amphi', 'semestre').order_by('amphi__numero')
        
        all_Amphi_Dep_S2 = Amphi_Dep.objects.filter(
            departement=departement.id, 
            semestre__numero=2
        ).select_related('amphi', 'semestre').order_by('amphi__numero')
        
        context = {
            'title': 'قائمة المدرجات',
            'my_Fac': departement.faculte,
            'my_Dep': departement,
            'my_Ens': enseignant,
            'real_Dep': real_Dep,
            'all_Amphi_Dep_S1': all_Amphi_Dep_S1,
            'all_Amphi_Dep_S2': all_Amphi_Dep_S2,
        }
        
        return render(request, 'enseignant/list_Amphi_Ens.html', context)
        
    except Exception as e:
        messages.error(request, f'خطأ في تحميل قائمة المدرجات: {str(e)}')
        return redirect('ense:dashboard_Ens', dep_id=dep_id)




#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def timeTable_Amphi_Ens(request, dep_id, amphi_dep_id, semestre_numero, enseignant, departement): 
# def timeTable_Amphi_Ens(request, dep_id, amphi_dep_id, semestre_numero):
    """
    Vue pour afficher l'emploi du temps d'un amphithéâtre spécifique pour un enseignant
    """
    try:
        # Récupération de l'enseignant connecté
        enseignant = request.user.enseignant_profile
        
        # Récupération du département
        departement = get_object_or_404(Departement, id=dep_id)
        
        # Vérification des permissions
        try:
            real_Dep = Ens_Dep.objects.get(
                enseignant=enseignant, 
                departement=departement
            )
        except Ens_Dep.DoesNotExist:
            messages.error(request, 'غير مصرح لك بالوصول إلى هذا القسم')
            return redirect('ense:list_Amphi_Ens', dep_id=dep_id)
        
        # Récupération de l'amphithéâtre-département
        amphi_dep = get_object_or_404(
            Amphi_Dep, 
            id=amphi_dep_id,
            departement=departement,
            semestre__numero=semestre_numero
        )

        # Initialisation des variables
        nbr_Cours = 0
        nbr_TP = 0
        nbr_TD = 0
        nbr_SS = 0

        # Définition des créneaux horaires
        sixClasses = [
            '09:30-08:00',
            '11:10-09:40',
            '12:50-11:20',
            '14:40-13:10',
            '16:20-14:50',
            '18:00-16:30',
        ]

        # Définition des jours
        sevenDays = ['Samedi', 'Dimanche', 'Lundi', 'Mardi', 'Mercredi', 'Jeudi']

        # ContentType pour Amphi_Dep
        lieu_content_type = ContentType.objects.get_for_model(Amphi_Dep)

        # Récupération de toutes les classes pour cet amphithéâtre
        all_Classe_Amp = []

        # Récupération par jour
        for jour in sevenDays:
            classes_jour = Classe.objects.filter(
                enseignant__departement=departement.id,
                content_type=lieu_content_type,
                semestre__numero=semestre_numero,
                object_id=amphi_dep_id,
                jour=jour
            ).select_related(
                'enseignant__enseignant',
                'matiere',
                'niv_spe_dep_sg__niv_spe_dep__niveau',
                'niv_spe_dep_sg__niv_spe_dep__specialite',
                'niv_spe_dep_sg__niv_spe_dep__departement'
            ).order_by('temps')
            
            all_Classe_Amp.extend(classes_jour)

        # Récupération par créneau horaire
        all_Classe_Time = []
        for creneau in sixClasses:
            classes_temps = Classe.objects.filter(
                enseignant__departement=departement.id,
                content_type=lieu_content_type,
                semestre__numero=semestre_numero,
                object_id=amphi_dep_id,
                temps=creneau
            ).select_related(
                'enseignant__enseignant',
                'matiere',
                'niv_spe_dep_sg__niv_spe_dep__niveau',
                'niv_spe_dep_sg__niv_spe_dep__specialite',
                'niv_spe_dep_sg__niv_spe_dep__departement'
            )
            
            all_Classe_Time.extend(classes_temps)

        # Calcul des statistiques et mise à jour du taux d'avancement
        for classe in all_Classe_Amp:
            # Comptage par type
            if classe.type == "Cours":
                nbr_Cours += 1
            elif classe.type == "TP":
                nbr_TP += 1
            elif classe.type == "TD":
                nbr_TD += 1
            elif classe.type == "Sortie Scientifique":
                nbr_SS += 1
            
            # Calcul du taux d'avancement si les séances sont créées
            if hasattr(classe, 'seance_created') and classe.seance_created:
                try:
                    all_seances = Seance.objects.filter(classe=classe.id)
                    seances_faites = all_seances.filter(fait=True)
                    
                    if all_seances.exists():
                        taux_avancement = (seances_faites.count() / all_seances.count()) * 100
                        classe.taux_avancement = round(taux_avancement)
                        classe.save()
                except Exception:
                    classe.taux_avancement = 0

        all_classes = nbr_Cours + nbr_TP + nbr_TD + nbr_SS
        
        context = {
            'title': f'استعمال الزمن - {amphi_dep.amphi.nom_ar}',
            'my_Fac': departement.faculte,
            'my_Dep': departement,
            'my_Ens': enseignant,
            'real_Dep': real_Dep,
            'amphi_dep': amphi_dep,
            'semestre_numero': semestre_numero,
            'all_Classe_Amp': all_Classe_Amp,
            'all_Classe_Time': all_Classe_Time,
            'sevenDays': sevenDays,
            'sixClasses': sixClasses,
            'nbr_Cours': nbr_Cours,
            'nbr_TP': nbr_TP,
            'nbr_TD': nbr_TD,
            'nbr_SS': nbr_SS,
            'all_classes': all_classes,
        }
        
        return render(request, 'enseignant/timeTable_Amphi_Ens.html', context)
        
    except Exception as e:
        messages.error(request, f'خطأ في تحميل استعمال الزمن: {str(e)}')
        return redirect('ense:list_Amphi_Ens', dep_id=dep_id)    
    
    
#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
register = template.Library()

@register.simple_tag
def set_flag(value, as_var):
    """Tag pour définir une variable dans le template"""
    return value

@register.filter
def slice_reverse(value, arg):
    """Filtre pour inverser un slice"""
    if isinstance(value, str):
        try:
            if arg.startswith('-') and ':' in arg:
                # Pour "-11:" retourne les 11 derniers caractères
                parts = arg.split(':')
                start = int(parts[0])
                return value[start:]
            elif ':' in arg:
                # Pour "0:15" retourne les 15 premiers caractères
                parts = arg.split(':')
                start = int(parts[0]) if parts[0] else None
                end = int(parts[1]) if parts[1] else None
                return value[start:end]
        except (ValueError, IndexError):
            pass
    return value

@register.filter  
def format_time_range(time_str):
    """Formate les créneaux horaires"""
    if isinstance(time_str, str) and '-' in time_str:
        parts = time_str.split('-')
        if len(parts) == 2:
            start_time = parts[1]  # Début (ex: 08:00)
            end_time = parts[0]    # Fin (ex: 09:30)
            return f"{start_time} - {end_time}"
    return time_str

@register.filter
def get_time_period(time_str, period_type):
    """Extrait une partie spécifique du créneau horaire"""
    if isinstance(time_str, str) and '-' in time_str:
        parts = time_str.split('-')
        if len(parts) == 2:
            if period_type == 'start':
                return parts[1]  # Heure de début
            elif period_type == 'end':
                return parts[0]  # Heure de fin
    return time_str

@register.inclusion_tag('enseignant/tags/class_item.html')
def render_class_item(class_obj):
    """Tag d'inclusion pour rendre un élément de cours"""
    return {
        'class': class_obj,
        'type_color': {
            'Cours': 'success',
            'TD': 'warning', 
            'TP': 'info',
            'Sortie Scientifique': 'danger'
        }.get(class_obj.type, 'secondary')
    }

@register.filter
def default_if_none(value, default):
    """Retourne une valeur par défaut si la valeur est None"""
    return default if value is None else value








#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def list_Salle_Ens(request, dep_id, enseignant, departement): 
# def list_Salle_Ens(request, dep_id):
    """
    Vue pour afficher la liste des salles du département pour un enseignant
    """
    try:
        # Récupération de l'enseignant connecté
        enseignant = request.user.enseignant_profile
        
        # Récupération du département spécifique
        departement = get_object_or_404(Departement, id=dep_id)
        
        # Vérification que l'enseignant appartient bien à ce département
        try:
            real_Dep = Ens_Dep.objects.get(
                enseignant=enseignant, 
                departement=departement,
                statut='Permanent'
            )
        except Ens_Dep.DoesNotExist:
            # Si pas de relation permanente, chercher une relation temporaire
            try:
                real_Dep = Ens_Dep.objects.get(
                    enseignant=enseignant, 
                    departement=departement
                )
            except Ens_Dep.DoesNotExist:
                messages.error(request, 'غير مصرح لك بالوصول إلى هذا القسم')
                return redirect('ense:dashboard_Ens', dep_id=enseignant.departements.first().id)
        
        # Récupération des salles par semestre
        all_Salle_Dep_S1 = Salle_Dep.objects.filter(
            departement=departement.id, 
            semestre__numero=1
        ).select_related('salle', 'semestre').order_by('salle__numero')
        
        all_Salle_Dep_S2 = Salle_Dep.objects.filter(
            departement=departement.id, 
            semestre__numero=2
        ).select_related('salle', 'semestre').order_by('salle__numero')
        
        # Calcul du total
        all_Salle_Dep = all_Salle_Dep_S1.count() + all_Salle_Dep_S2.count()
        
        context = {
            'title': 'قائمة القاعات',
            'my_Fac': departement.faculte,
            'my_Dep': departement,
            'my_Ens': enseignant,
            'real_Dep': real_Dep,
            'all_Salle_Dep_S1': all_Salle_Dep_S1,
            'all_Salle_Dep_S2': all_Salle_Dep_S2,
            'all_Salle_Dep': all_Salle_Dep,
        }
        
        return render(request, 'enseignant/list_Salle_Ens.html', context)
        
    except Exception as e:
        messages.error(request, f'خطأ في تحميل قائمة القاعات: {str(e)}')
        return redirect('ense:dashboard_Ens', dep_id=dep_id)





#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def timeTable_Salle_Ens(request, dep_id, salle_dep_id, semestre_numero, enseignant, departement): 
# def timeTable_Salle_Ens(request, dep_id, salle_dep_id, semestre_numero):
    """
    Vue pour afficher l'emploi du temps d'une salle spécifique pour un enseignant
    """
    try:
        # Récupération de l'enseignant connecté
        enseignant = request.user.enseignant_profile
        
        # Récupération du département
        departement = get_object_or_404(Departement, id=dep_id)
        
        # Vérification des permissions
        try:
            real_Dep = Ens_Dep.objects.get(
                enseignant=enseignant, 
                departement=departement
            )
        except Ens_Dep.DoesNotExist:
            messages.error(request, 'غير مصرح لك بالوصول إلى هذا القسم')
            return redirect('ense:list_Salle_Ens', dep_id=dep_id)
        
        # Récupération de la salle-département
        salle_dep = get_object_or_404(
            Salle_Dep, 
            id=salle_dep_id,
            departement=departement,
            semestre__numero=semestre_numero
        )

        # Initialisation des variables
        nbr_Cours = 0
        nbr_TP = 0
        nbr_TD = 0
        nbr_SS = 0

        # Définition des créneaux horaires
        sixClasses = [
            '09:30-08:00',
            '11:10-09:40',
            '12:50-11:20',
            '14:40-13:10',
            '16:20-14:50',
            '18:00-16:30',
        ]

        # Définition des jours
        sevenDays = ['Samedi', 'Dimanche', 'Lundi', 'Mardi', 'Mercredi', 'Jeudi']

        # ContentType pour Salle_Dep
        lieu_content_type = ContentType.objects.get_for_model(Salle_Dep)

        # Récupération de toutes les classes pour cette salle
        all_Classe_Salle = []

        # Récupération par jour
        for jour in sevenDays:
            classes_jour = Classe.objects.filter(
                enseignant__departement=departement.id,
                content_type=lieu_content_type,
                semestre__numero=semestre_numero,
                object_id=salle_dep_id,
                jour=jour
            ).select_related(
                'enseignant__enseignant',
                'matiere',
                'niv_spe_dep_sg__niv_spe_dep__niveau',
                'niv_spe_dep_sg__niv_spe_dep__specialite',
                'niv_spe_dep_sg__niv_spe_dep__departement'
            ).order_by('temps')
            
            all_Classe_Salle.extend(classes_jour)

        # Récupération par créneau horaire
        all_Classe_Time = []
        for creneau in sixClasses:
            classes_temps = Classe.objects.filter(
                enseignant__departement=departement.id,
                content_type=lieu_content_type,
                semestre__numero=semestre_numero,
                object_id=salle_dep_id,
                temps=creneau
            ).select_related(
                'enseignant__enseignant',
                'matiere',
                'niv_spe_dep_sg__niv_spe_dep__niveau',
                'niv_spe_dep_sg__niv_spe_dep__specialite',
                'niv_spe_dep_sg__niv_spe_dep__departement'
            )
            
            all_Classe_Time.extend(classes_temps)

        # Calcul des statistiques et mise à jour du taux d'avancement
        for classe in all_Classe_Salle:
            # Comptage par type
            if classe.type == "Cours":
                nbr_Cours += 1
            elif classe.type == "TP":
                nbr_TP += 1
            elif classe.type == "TD":
                nbr_TD += 1
            elif classe.type == "Sortie Scientifique":
                nbr_SS += 1
            
            # Calcul du taux d'avancement si les séances sont créées
            if hasattr(classe, 'seance_created') and classe.seance_created:
                try:
                    all_seances = Seance.objects.filter(classe=classe.id)
                    seances_faites = all_seances.filter(fait=True)
                    
                    if all_seances.exists():
                        taux_avancement = (seances_faites.count() / all_seances.count()) * 100
                        classe.taux_avancement = round(taux_avancement)
                        classe.save()
                except Exception:
                    classe.taux_avancement = 0

        all_classes = nbr_Cours + nbr_TP + nbr_TD + nbr_SS
        
        context = {
            'title': f'استعمال الزمن - {salle_dep.salle.nom_ar}',
            'my_Fac': departement.faculte,
            'my_Dep': departement,
            'my_Ens': enseignant,
            'real_Dep': real_Dep,
            'salle_dep': salle_dep,
            'semestre_numero': semestre_numero,
            'all_Classe_Salle': all_Classe_Salle,
            'all_Classe_Time': all_Classe_Time,
            'sevenDays': sevenDays,
            'sixClasses': sixClasses,
            'nbr_Cours': nbr_Cours,
            'nbr_TP': nbr_TP,
            'nbr_TD': nbr_TD,
            'nbr_SS': nbr_SS,
            'all_classes': all_classes,
        }
        
        return render(request, 'enseignant/timeTable_Salle_Ens.html', context)
        
    except Exception as e:
        messages.error(request, f'خطأ في تحميل استعمال الزمن: {str(e)}')
        return redirect('ense:list_Salle_Ens', dep_id=dep_id)
    





#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def list_Salle_Ens(request, dep_id, enseignant, departement): 
# def list_Salle_Ens(request, dep_id):
    """
    Vue pour afficher la liste des salles du département pour un enseignant
    """
    try:
        # Récupération de l'enseignant connecté
        enseignant = request.user.enseignant_profile
        
        # Récupération du département spécifique
        departement = get_object_or_404(Departement, id=dep_id)
        
        # Vérification que l'enseignant appartient bien à ce département
        try:
            real_Dep = Ens_Dep.objects.get(
                enseignant=enseignant, 
                departement=departement,
                statut='Permanent'
            )
        except Ens_Dep.DoesNotExist:
            # Si pas de relation permanente, chercher une relation temporaire
            try:
                real_Dep = Ens_Dep.objects.get(
                    enseignant=enseignant, 
                    departement=departement
                )
            except Ens_Dep.DoesNotExist:
                messages.error(request, 'غير مصرح لك بالوصول إلى هذا القسم')
                return redirect('ense:dashboard_Ens', dep_id=enseignant.departements.first().id)
        
        # Récupération des salles par semestre
        all_Salle_Dep_S1 = Salle_Dep.objects.filter(
            departement=departement.id, 
            semestre__numero=1
        ).select_related('salle', 'semestre').order_by('salle__numero')
        
        all_Salle_Dep_S2 = Salle_Dep.objects.filter(
            departement=departement.id, 
            semestre__numero=2
        ).select_related('salle', 'semestre').order_by('salle__numero')
        
        # Calcul du total
        all_Salle_Dep = all_Salle_Dep_S1.count() + all_Salle_Dep_S2.count()
        
        context = {
            'title': 'قائمة القاعات',
            'my_Fac': departement.faculte,
            'my_Dep': departement,
            'my_Ens': enseignant,
            'real_Dep': real_Dep,
            'all_Salle_Dep_S1': all_Salle_Dep_S1,
            'all_Salle_Dep_S2': all_Salle_Dep_S2,
            'all_Salle_Dep': all_Salle_Dep,
        }
        
        return render(request, 'enseignant/list_Salle_Ens.html', context)
        
    except Exception as e:
        messages.error(request, f'خطأ في تحميل قائمة القاعات: {str(e)}')
        return redirect('ense:dashboard_Ens', dep_id=dep_id)






#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def timeTable_Salle_Ens(request, dep_id, salle_dep_id, semestre_numero, enseignant, departement):  
# def timeTable_Salle_Ens(request, dep_id, salle_dep_id, semestre_numero):
    """
    Vue pour afficher l'emploi du temps d'une salle spécifique pour un enseignant
    """
    try:
        # Récupération de l'enseignant connecté
        enseignant = request.user.enseignant_profile
        
        # Récupération du département
        departement = get_object_or_404(Departement, id=dep_id)
        
        # Vérification des permissions
        try:
            real_Dep = Ens_Dep.objects.get(
                enseignant=enseignant, 
                departement=departement
            )
        except Ens_Dep.DoesNotExist:
            messages.error(request, 'غير مصرح لك بالوصول إلى هذا القسم')
            return redirect('ense:list_Salle_Ens', dep_id=dep_id)
        
        # Récupération de la salle-département
        salle_dep = get_object_or_404(
            Salle_Dep, 
            id=salle_dep_id,
            departement=departement,
            semestre__numero=semestre_numero
        )

        # Initialisation des variables
        nbr_Cours = 0
        nbr_TP = 0
        nbr_TD = 0
        nbr_SS = 0

        # Définition des créneaux horaires
        sixClasses = [
            '09:30-08:00',
            '11:10-09:40',
            '12:50-11:20',
            '14:40-13:10',
            '16:20-14:50',
            '18:00-16:30',
        ]

        # Définition des jours
        sevenDays = ['Samedi', 'Dimanche', 'Lundi', 'Mardi', 'Mercredi', 'Jeudi']

        # ContentType pour Salle_Dep
        lieu_content_type = ContentType.objects.get_for_model(Salle_Dep)

        # Récupération de toutes les classes pour cette salle
        all_Classe_Salle = []

        # Récupération par jour
        for jour in sevenDays:
            classes_jour = Classe.objects.filter(
                enseignant__departement=departement.id,
                content_type=lieu_content_type,
                semestre__numero=semestre_numero,
                object_id=salle_dep_id,
                jour=jour
            ).select_related(
                'enseignant__enseignant',
                'matiere',
                'niv_spe_dep_sg__niv_spe_dep__niveau',
                'niv_spe_dep_sg__niv_spe_dep__specialite',
                'niv_spe_dep_sg__niv_spe_dep__departement'
            ).order_by('temps')
            
            all_Classe_Salle.extend(classes_jour)

        # Récupération par créneau horaire
        all_Classe_Time = []
        for creneau in sixClasses:
            classes_temps = Classe.objects.filter(
                enseignant__departement=departement.id,
                content_type=lieu_content_type,
                semestre__numero=semestre_numero,
                object_id=salle_dep_id,
                temps=creneau
            ).select_related(
                'enseignant__enseignant',
                'matiere',
                'niv_spe_dep_sg__niv_spe_dep__niveau',
                'niv_spe_dep_sg__niv_spe_dep__specialite',
                'niv_spe_dep_sg__niv_spe_dep__departement'
            )
            
            all_Classe_Time.extend(classes_temps)

        # Calcul des statistiques et mise à jour du taux d'avancement
        for classe in all_Classe_Salle:
            # Comptage par type
            if classe.type == "Cours":
                nbr_Cours += 1
            elif classe.type == "TP":
                nbr_TP += 1
            elif classe.type == "TD":
                nbr_TD += 1
            elif classe.type == "Sortie Scientifique":
                nbr_SS += 1
            
            # Calcul du taux d'avancement si les séances sont créées
            if hasattr(classe, 'seance_created') and classe.seance_created:
                try:
                    all_seances = Seance.objects.filter(classe=classe.id)
                    seances_faites = all_seances.filter(fait=True)
                    
                    if all_seances.exists():
                        taux_avancement = (seances_faites.count() / all_seances.count()) * 100
                        classe.taux_avancement = round(taux_avancement)
                        classe.save()
                except Exception:
                    classe.taux_avancement = 0

        all_classes = nbr_Cours + nbr_TP + nbr_TD + nbr_SS
        
        context = {
            'title': f'استعمال الزمن - {salle_dep.salle.nom_ar}',
            'my_Fac': departement.faculte,
            'my_Dep': departement,
            'my_Ens': enseignant,
            'real_Dep': real_Dep,
            'salle_dep': salle_dep,
            'semestre_numero': semestre_numero,
            'all_Classe_Salle': all_Classe_Salle,
            'all_Classe_Time': all_Classe_Time,
            'sevenDays': sevenDays,
            'sixClasses': sixClasses,
            'nbr_Cours': nbr_Cours,
            'nbr_TP': nbr_TP,
            'nbr_TD': nbr_TD,
            'nbr_SS': nbr_SS,
            'all_classes': all_classes,
        }
        
        return render(request, 'enseignant/timeTable_Salle_Ens.html', context)
        
    except Exception as e:
        messages.error(request, f'خطأ في تحميل استعمال الزمن: {str(e)}')
        return redirect('ense:list_Salle_Ens', dep_id=dep_id)





#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def list_Labo_Ens(request, dep_id, enseignant, departement):  
# def list_Labo_Ens(request, dep_id):
    """
    Vue pour afficher la liste des laboratoires du département pour un enseignant
    """
    try:
        # Récupération de l'enseignant connecté
        enseignant = request.user.enseignant_profile
        
        # Récupération du département spécifique
        departement = get_object_or_404(Departement, id=dep_id)
        
        # Vérification que l'enseignant appartient bien à ce département
        try:
            real_Dep = Ens_Dep.objects.get(
                enseignant=enseignant, 
                departement=departement,
                statut='Permanent'
            )
        except Ens_Dep.DoesNotExist:
            # Si pas de relation permanente, chercher une relation temporaire
            try:
                real_Dep = Ens_Dep.objects.get(
                    enseignant=enseignant, 
                    departement=departement
                )
            except Ens_Dep.DoesNotExist:
                messages.error(request, 'غير مصرح لك بالوصول إلى هذا القسم')
                return redirect('ense:dashboard_Ens', dep_id=enseignant.departements.first().id)
        
        # Récupération des laboratoires par semestre
        all_Labo_Dep_S1 = Laboratoire_Dep.objects.filter(
            departement=departement.id, 
            semestre__numero=1
        ).select_related('laboratoire', 'semestre').order_by('laboratoire__numero')
        
        all_Labo_Dep_S2 = Laboratoire_Dep.objects.filter(
            departement=departement.id, 
            semestre__numero=2
        ).select_related('laboratoire', 'semestre').order_by('laboratoire__numero')
        
        context = {
            'title': 'قائمة المخابر',
            'my_Fac': departement.faculte,
            'my_Dep': departement,
            'my_Ens': enseignant,
            'real_Dep': real_Dep,
            'all_Labo_Dep_S1': all_Labo_Dep_S1,
            'all_Labo_Dep_S2': all_Labo_Dep_S2,
        }
        
        return render(request, 'enseignant/list_Labo_Ens.html', context)
        
    except Exception as e:
        messages.error(request, f'خطأ في تحميل قائمة المخابر: {str(e)}')
        return redirect('ense:dashboard_Ens', dep_id=dep_id)






#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def timeTable_Labo_Ens(request, dep_id, labo_dep_id, semestre_numero, enseignant, departement):  
# def timeTable_Labo_Ens(request, dep_id, labo_dep_id, semestre_numero):
    """
    Vue pour afficher l'emploi du temps d'un laboratoire spécifique pour un enseignant
    """
    try:
        # Récupération de l'enseignant connecté
        enseignant = request.user.enseignant_profile
        
        # Récupération du département
        departement = get_object_or_404(Departement, id=dep_id)
        
        # Vérification des permissions
        try:
            real_Dep = Ens_Dep.objects.get(
                enseignant=enseignant, 
                departement=departement
            )
        except Ens_Dep.DoesNotExist:
            messages.error(request, 'غير مصرح لك بالوصول إلى هذا القسم')
            return redirect('ense:list_Labo_Ens', dep_id=dep_id)
        
        # Récupération du laboratoire-département
        labo_dep = get_object_or_404(
            Laboratoire_Dep, 
            id=labo_dep_id,
            departement=departement,
            semestre__numero=semestre_numero
        )

        # Initialisation des variables
        nbr_Cours = 0
        nbr_TP = 0
        nbr_TD = 0
        nbr_SS = 0

        # Définition des créneaux horaires
        sixClasses = [
            '09:30-08:00',
            '11:10-09:40',
            '12:50-11:20',
            '14:40-13:10',
            '16:20-14:50',
            '18:00-16:30',
        ]

        # Définition des jours
        sevenDays = ['Samedi', 'Dimanche', 'Lundi', 'Mardi', 'Mercredi', 'Jeudi']

        # ContentType pour Laboratoire_Dep
        lieu_content_type = ContentType.objects.get_for_model(Laboratoire_Dep)

        # Récupération de toutes les classes pour ce laboratoire
        all_Classe_Labo = []

        # Récupération par jour
        for jour in sevenDays:
            classes_jour = Classe.objects.filter(
                enseignant__departement=departement.id,
                content_type=lieu_content_type,
                semestre__numero=semestre_numero,
                object_id=labo_dep_id,
                jour=jour
            ).select_related(
                'enseignant__enseignant',
                'matiere',
                'niv_spe_dep_sg__niv_spe_dep__niveau',
                'niv_spe_dep_sg__niv_spe_dep__specialite',
                'niv_spe_dep_sg__niv_spe_dep__departement'
            ).order_by('temps')
            
            all_Classe_Labo.extend(classes_jour)

        # Récupération par créneau horaire
        all_Classe_Time = []
        for creneau in sixClasses:
            classes_temps = Classe.objects.filter(
                enseignant__departement=departement.id,
                content_type=lieu_content_type,
                semestre__numero=semestre_numero,
                object_id=labo_dep_id,
                temps=creneau
            ).select_related(
                'enseignant__enseignant',
                'matiere',
                'niv_spe_dep_sg__niv_spe_dep__niveau',
                'niv_spe_dep_sg__niv_spe_dep__specialite',
                'niv_spe_dep_sg__niv_spe_dep__departement'
            )
            
            all_Classe_Time.extend(classes_temps)

        # Calcul des statistiques et mise à jour du taux d'avancement
        for classe in all_Classe_Labo:
            # Comptage par type
            if classe.type == "Cours":
                nbr_Cours += 1
            elif classe.type == "TP":
                nbr_TP += 1
            elif classe.type == "TD":
                nbr_TD += 1
            elif classe.type == "Sortie Scientifique":
                nbr_SS += 1
            
            # Calcul du taux d'avancement si les séances sont créées
            if hasattr(classe, 'seance_created') and classe.seance_created:
                try:
                    all_seances = Seance.objects.filter(classe=classe.id)
                    seances_faites = all_seances.filter(fait=True)
                    
                    if all_seances.exists():
                        taux_avancement = (seances_faites.count() / all_seances.count()) * 100
                        classe.taux_avancement = round(taux_avancement)
                        classe.save()
                except Exception:
                    classe.taux_avancement = 0

        all_classes = nbr_Cours + nbr_TP + nbr_TD + nbr_SS
        
        context = {
            'title': f'استعمال الزمن - {labo_dep.laboratoire.nom_ar}',
            'my_Fac': departement.faculte,
            'my_Dep': departement,
            'my_Ens': enseignant,
            'real_Dep': real_Dep,
            'labo_dep': labo_dep,
            'semestre_numero': semestre_numero,
            'all_Classe_Labo': all_Classe_Labo,
            'all_Classe_Time': all_Classe_Time,
            'sevenDays': sevenDays,
            'sixClasses': sixClasses,
            'nbr_Cours': nbr_Cours,
            'nbr_TP': nbr_TP,
            'nbr_TD': nbr_TD,
            'nbr_SS': nbr_SS,
            'all_classes': all_classes,
        }
        
        return render(request, 'enseignant/timeTable_Labo_Ens.html', context)
        
    except Exception as e:
        messages.error(request, f'خطأ في تحميل استعمال الزمن: {str(e)}')
        return redirect('ense:list_Labo_Ens', dep_id=dep_id)
    





#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@login_required
def mes_sections(request):
    """Page qui affiche les sections et groupes que j'enseigne"""
    
    try:
        enseignant = request.user.enseignant_profile
    except AttributeError:
        messages.error(request, 'Profil enseignant non trouvé')
        return redirect('admin:index')
    
    # Récupérer l'année universitaire courante
    annee_courante = AnneeUniversitaire.get_courante()
    if not annee_courante:
        messages.error(request, 'لا توجد سنة جامعية محددة كسنة حالية')
        return redirect('admin:index')
    
    # Récupérer mes classes/sections avec optimisation
    mes_classes = Classe.objects.filter(
        enseignant=enseignant,
        annee_univ=annee_courante
    ).select_related(
        'section__niveau',
        'section__specialite',
        'matiere'
    ).order_by(
        'semestre',
        'section__niveau__ordre', 
        'section__nom',
        'matiere__nom'
    )
    
    # Grouper par section
    sections_dict = {}
    for classe in mes_classes:
        section_id = classe.section.id
        
        if section_id not in sections_dict:
            sections_dict[section_id] = {
                'section': classe.section,
                'matieres': [],
                'semestres': set()
            }
        
        sections_dict[section_id]['matieres'].append({
            'nom': classe.matiere.nom,
            'semestre': classe.semestre,
            'type': classe.type_classe.nom if classe.type_classe else 'Cours'
        })
        sections_dict[section_id]['semestres'].add(classe.semestre)
    
    # Convertir en liste
    mes_sections = []
    for data in sections_dict.values():
        data['semestres'] = sorted(list(data['semestres']))
        mes_sections.append(data)
    
    # Trier par niveau puis nom
    mes_sections.sort(key=lambda x: (
        x['section'].niveau.ordre if x['section'].niveau else 0,
        x['section'].nom
    ))
    
    context = {
        'title': 'الأقسام والمجموعات التي أدرسها',
        'mes_sections': mes_sections,
        'annee_courante': annee_courante,
        'total_sections': len(mes_sections)
    }
    
    return render(request, 'enseignant/mes_sections.html', context)








#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def niveaux_enseigner(request, dep_id, enseignant, departement):
    """
    Vue pour afficher la liste des niveaux enseignés par un enseignant
    selon ses affectations (sections et groupes)
    """
    try:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    except:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant)
    
    # Récupérer le paramètre semestre (par défaut S1)
    semestre_selected = request.GET.get('semestre', '1')
    
    # Récupérer les objets semestre
    try:
        semestre_obj = Semestre.objects.get(numero=semestre_selected)
    except:
        semestre_obj = Semestre.objects.filter(numero='1').first()
        semestre_selected = '1'
    
    my_Dep = Departement.objects.get(id=departement.id)
    
    def traduire_jour_en_arabe(jour_francais):
        """Traduire les jours en arabe"""
        traduction_jours = {
            'Samedi': 'السبت',
            'Dimanche': 'الأحد', 
            'Lundi': 'الإثنين',
            'Mardi': 'الثلاثاء',
            'Mercredi': 'الأربعاء',
            'Jeudi': 'الخميس',
            'Vendredi': 'الجمعة'
        }
        return traduction_jours.get(jour_francais, jour_francais)

    def compter_etudiants_section_ou_groupe(niv_spe_dep_sg, type_affectation, section_numero=None):
        """Compter les étudiants selon le type (section ou groupe)"""
        try:
            if type_affectation == 'par_section' and section_numero is not None:
                # ✅ POUR UNE SECTION : Compter TOUS les groupes de cette section par NUMERO
                total_etudiants = 0
                
                # Trouver tous les NivSpeDep_SG de cette section par numero
                groupes_de_la_section = NivSpeDep_SG.objects.filter(
                    niv_spe_dep=niv_spe_dep_sg.niv_spe_dep,
                    section__numero=section_numero,  # ✅ UTILISER LE NUMERO
                    type_affectation='par_groupe'
                )
                
                # Additionner les étudiants de chaque groupe
                for groupe in groupes_de_la_section:
                    nb_etudiants_groupe = Etudiant.objects.filter(niv_spe_dep_sg=groupe).count()
                    total_etudiants += nb_etudiants_groupe
                
                return total_etudiants
                
            elif type_affectation == 'par_groupe':
                # ✅ POUR UN GROUPE : Compter directement
                return Etudiant.objects.filter(niv_spe_dep_sg=niv_spe_dep_sg).count()
            
            else:
                return 0
                
        except Exception as e:
            print(f"Erreur comptage étudiants: {e}")
            return 0
        
    # Récupérer toutes les classes de l'enseignant pour le semestre sélectionné
    base_filter = {
        'enseignant__departement': my_Dep.id,
        'enseignant__enseignant': enseignant.id,
        'semestre': semestre_obj
    }
    
    all_classes = Classe.objects.filter(**base_filter).select_related(
        'niv_spe_dep_sg',
        'niv_spe_dep_sg__niv_spe_dep__niveau',
        'niv_spe_dep_sg__niv_spe_dep__specialite', 
        'niv_spe_dep_sg__niv_spe_dep__departement',
        'niv_spe_dep_sg__section',  # ✅ AJOUTER
        'niv_spe_dep_sg__groupe',   # ✅ AJOUTER
        'matiere'
    )
    
    # Organiser les données par niveau, spécialité et type (section/groupe)
    niveaux_data = defaultdict(lambda: {
        'niveau': None,
        'specialite': None,
        'departement': None,
        'sections': defaultdict(lambda: {
            'matiere_count': 0,
            'matieres': [],
            'types_cours': set(),
            'total_heures': 0,
            'section_numero': None  # ✅ AJOUTER pour le tri
        }),
        'groupes': defaultdict(lambda: {
            'matiere_count': 0,
            'matieres': [],
            'types_cours': set(),
            'total_heures': 0,
            'groupe_numero': None  # ✅ AJOUTER pour le tri
        })
    })
    
    # Traiter chaque classe
    for classe in all_classes:
        niv_spe_dep_sg = classe.niv_spe_dep_sg
        
        # Clé unique pour identifier le niveau-spécialité-département
        key = f"{niv_spe_dep_sg.niv_spe_dep.niveau.nom_fr}_{niv_spe_dep_sg.niv_spe_dep.specialite.nom_fr}_{niv_spe_dep_sg.niv_spe_dep.departement.nom_fr}"
        
        # Remplir les informations générales
        niveaux_data[key]['niveau'] = niv_spe_dep_sg.niv_spe_dep.niveau
        niveaux_data[key]['specialite'] = niv_spe_dep_sg.niv_spe_dep.specialite
        niveaux_data[key]['departement'] = niv_spe_dep_sg.niv_spe_dep.departement
        
        # Calculer le nombre de séances créées et faites
        nbr_seances_total = 0
        nbr_seances_faites = 0
        
        if classe.seance_created:
            seances = Seance.objects.filter(classe=classe)
            nbr_seances_total = seances.count()
            nbr_seances_faites = seances.filter(fait=True).count()

        sous_groupes_count = 0
        if niv_spe_dep_sg.type_affectation == 'par_groupe':
            sous_groupes_count = SousGroupe.objects.filter(
                groupe_principal=niv_spe_dep_sg,
                actif=True
            ).count()
        
        # ✅ DÉTERMINER LE TYPE ET COMPTER LES ÉTUDIANTS
        if niv_spe_dep_sg.type_affectation == 'par_groupe':
            # C'est un groupe
            groupe_nom = niv_spe_dep_sg.groupe.nom_ar
            container = niveaux_data[key]['groupes'][groupe_nom]
            container['groupe_numero'] = niv_spe_dep_sg.groupe.numero  # ✅ STOCKER LE NUMERO
            nb_etudiants = compter_etudiants_section_ou_groupe(niv_spe_dep_sg, 'par_groupe')
        else:
            # C'est une section
            section_nom = niv_spe_dep_sg.section.nom_ar if niv_spe_dep_sg.section else f"Section {niv_spe_dep_sg.numero_section}"
            container = niveaux_data[key]['sections'][section_nom]
            section_numero = niv_spe_dep_sg.section.numero if niv_spe_dep_sg.section else niv_spe_dep_sg.numero_section
            container['section_numero'] = section_numero  # ✅ STOCKER LE NUMERO
            nb_etudiants = compter_etudiants_section_ou_groupe(niv_spe_dep_sg, 'par_section', section_numero)

        # Ajouter les informations de la matière
        matiere_info = {
            'matiere': classe.matiere,
            'type_cours': classe.type,
            'jour': traduire_jour_en_arabe(classe.jour),
            'temps': classe.temps,
            'taux_avancement': classe.taux_avancement,
            'seances_total': nbr_seances_total,
            'seances_faites': nbr_seances_faites,
            'lieu': classe.lieu if classe.lieu else None,
            'classe_id': classe.id,
            'sous_groupes_count': sous_groupes_count,
            'niv_spe_dep_sg_id': niv_spe_dep_sg.id,
            'nb_etudiants': nb_etudiants
        }
        
        container['matieres'].append(matiere_info)
        container['types_cours'].add(classe.type)
        container['matiere_count'] += 1
        container['total_heures'] += 1.5

    # Convertir en structure plus simple pour le template
    niveaux_final = []
    for key, data in niveaux_data.items():
        # ✅ TRIER LES SECTIONS PAR NUMERO
        sections_sorted = dict(sorted(
            data['sections'].items(), 
            key=lambda x: x[1]['section_numero'] if x[1]['section_numero'] is not None else 999
        ))
        
        # ✅ TRIER LES GROUPES PAR NUMERO
        groupes_sorted = dict(sorted(
            data['groupes'].items(), 
            key=lambda x: x[1]['groupe_numero'] if x[1]['groupe_numero'] is not None else 999
        ))
        
        niveau_info = {
            'niveau': data['niveau'],
            'specialite': data['specialite'],
            'departement': data['departement'],
            'sections': sections_sorted,
            'groupes': groupes_sorted,
            'total_sections': len(data['sections']),
            'total_groupes': len(data['groupes']),
            'total_matieres': sum(s['matiere_count'] for s in data['sections'].values()) + 
                             sum(g['matiere_count'] for g in data['groupes'].values())
        }
        niveaux_final.append(niveau_info)
    
    # Trier par niveau et spécialité
    niveaux_final.sort(key=lambda x: (x['niveau'].nom_fr, x['specialite'].nom_fr))
    
    # Statistiques générales
    stats = {
        'total_niveaux': len(niveaux_final),
        'total_sections': sum(n['total_sections'] for n in niveaux_final),
        'total_groupes': sum(n['total_groupes'] for n in niveaux_final),
        'total_matieres': sum(n['total_matieres'] for n in niveaux_final),
        'total_classes': all_classes.count()
    }
    
    # Répartition par type de cours
    type_repartition = all_classes.values('type').annotate(
        count=Count('id')
    ).order_by('type')
    
    # Récupérer tous les semestres disponibles
    all_semestres = Semestre.objects.all().order_by('numero')
    
    context = {
        'title': 'المستويات المدرسة',
        'niveaux_enseigner': niveaux_final,
        'stats': stats,
        'type_repartition': type_repartition,
        'semestre_selected': semestre_selected,
        'semestre_obj': semestre_obj,
        'all_semestres': all_semestres,
        'my_Dep': departement,
        'my_Fac': departement.faculte,
        'my_Ens': enseignant,
        'real_Dep': real_Dep,
    }
    
    return render(request, 'enseignant/niveaux_enseigner.html', context)








#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def page_nombre_sous_groupes(request, dep_id, classe_id, enseignant, departement):
    """Page pour choisir le nombre de sous-groupes"""
    classe = get_object_or_404(Classe, id=classe_id)
    
    if request.method == 'POST':
        nombre = int(request.POST.get('nombre', 2))
        
        # Créer les sous-groupes avec votre SousGroupeManager
        SousGroupeManager.creer_sous_groupes_automatiques(
            groupe_principal=classe.niv_spe_dep_sg,
            nombre_sous_groupes=nombre,
            enseignant=enseignant
        )
        
        # Rediriger vers la page d'affectation des étudiants
        return redirect('ense:affecter_etudiants_sous_groupes', dep_id=dep_id, classe_id=classe_id)
    
    context = {
        'classe': classe,
        'my_Dep': departement,
        'my_Ens': enseignant,
    }
    return render(request, 'enseignant/choisir_nombre_sous_groupes.html', context)





#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def affecter_etudiants_sous_groupes(request, dep_id, classe_id, enseignant, departement):
    """Page pour affecter les étudiants aux sous-groupes"""
    classe = get_object_or_404(Classe, id=classe_id)
    
    # Récupérer les sous-groupes
    sous_groupes = SousGroupe.objects.filter(
        groupe_principal=classe.niv_spe_dep_sg,
        actif=True
    ).order_by('ordre_affichage')
    
    # ✅ CORRECTION - Supprimer le filtre 'actif=True'
    etudiants = Etudiant.objects.filter(
        niv_spe_dep_sg=classe.niv_spe_dep_sg
    ).order_by('nom_ar', 'prenom_ar')
    
    if request.method == 'POST':
        # Traitement des affectations
        for etudiant in etudiants:
            sous_groupe_id = request.POST.get(f'etudiant_{etudiant.id}')
            if sous_groupe_id:
                sous_groupe = get_object_or_404(SousGroupe, id=sous_groupe_id)
                
                # Supprimer ancienne affectation
                EtudiantSousGroupe.objects.filter(etudiant=etudiant).delete()
                
                # Créer nouvelle affectation
                EtudiantSousGroupe.objects.create(
                    etudiant=etudiant,
                    sous_groupe=sous_groupe,
                    affecte_par=enseignant
                )
        
        messages.success(request, "تم توزيع الطلاب على المجموعات الفرعية بنجاح!")
        return redirect('enseignant:niveaux_enseigner', dep_id=dep_id)
    
    context = {
        'classe': classe,
        'sous_groupes': sous_groupes,
        'etudiants': etudiants,
        'my_Dep': departement,
        'my_Ens': enseignant,
    }
    return render(request, 'enseignant/affecter_etudiants_sous_groupes.html', context)





#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def liste_sous_groupes(request, dep_id, niv_spe_dep_sg_id, enseignant, departement):
    """Afficher la liste des sous-groupes et leurs étudiants"""
    niv_spe_dep_sg = get_object_or_404(NivSpeDep_SG, id=niv_spe_dep_sg_id)
    
    # Récupérer les sous-groupes avec leurs étudiants
    sous_groupes = SousGroupe.objects.filter(
        groupe_principal=niv_spe_dep_sg,
        actif=True
    ).order_by('ordre_affichage')
    
    # ✅ NOUVEAU - Récupérer les étudiants NON AFFECTÉS
    tous_etudiants = Etudiant.objects.filter(
        niv_spe_dep_sg=niv_spe_dep_sg
    ).order_by('nom_ar', 'prenom_ar')
    
    etudiants_affectes_ids = EtudiantSousGroupe.objects.filter(
        sous_groupe__in=sous_groupes,
        actif=True
    ).values_list('etudiant_id', flat=True)
    
    etudiants_non_affectes = tous_etudiants.exclude(id__in=etudiants_affectes_ids)
    
    # Pour chaque sous-groupe, récupérer ses étudiants
    sous_groupes_avec_etudiants = []
    for sous_groupe in sous_groupes:
        etudiants = Etudiant.objects.filter(
            affectations_sous_groupes__sous_groupe=sous_groupe,
            affectations_sous_groupes__actif=True
        ).order_by('nom_ar', 'prenom_ar')
        
        sous_groupes_avec_etudiants.append({
            'sous_groupe': sous_groupe,
            'etudiants': etudiants
        })
    
    context = {
        'niv_spe_dep_sg': niv_spe_dep_sg,
        'sous_groupes_avec_etudiants': sous_groupes_avec_etudiants,
        'etudiants_non_affectes': etudiants_non_affectes,  # ✅ NOUVEAU
        'my_Dep': departement,
        'my_Ens': enseignant,
    }
    return render(request, 'enseignant/liste_sous_groupes.html', context)





#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def affecter_direct_sous_groupes(request, dep_id, niv_spe_dep_sg_id, enseignant, departement):
    """Traiter l'affectation directe des étudiants aux sous-groupes - VERSION CORRIGÉE"""
    
    if request.method == 'POST':
        niv_spe_dep_sg = get_object_or_404(NivSpeDep_SG, id=niv_spe_dep_sg_id)
        
        # Récupérer tous les sous-groupes
        sous_groupes = SousGroupe.objects.filter(
            groupe_principal=niv_spe_dep_sg,
            actif=True
        )
        
        # Récupérer TOUS les étudiants du groupe (affectés ET non affectés)
        tous_etudiants = Etudiant.objects.filter(
            niv_spe_dep_sg=niv_spe_dep_sg
        )
        
        # Traiter les affectations pour TOUS les étudiants
        affectations_reussies = 0
        modifications_reussies = 0
        
        for etudiant in tous_etudiants:
            sous_groupe_id = request.POST.get(f'etudiant_{etudiant.id}')
            
            if sous_groupe_id:
                try:
                    sous_groupe = get_object_or_404(SousGroupe, id=sous_groupe_id)
                    
                    # Vérifier que le sous-groupe appartient au même groupe principal
                    if sous_groupe.groupe_principal == niv_spe_dep_sg:
                        # Vérifier s'il existe déjà une affectation
                        affectation_existante = EtudiantSousGroupe.objects.filter(
                            etudiant=etudiant,
                            sous_groupe__groupe_principal=niv_spe_dep_sg,
                            actif=True
                        ).first()
                        
                        if affectation_existante:
                            # Si l'étudiant change de sous-groupe
                            if affectation_existante.sous_groupe.id != int(sous_groupe_id):
                                # Supprimer l'ancienne affectation
                                affectation_existante.delete()
                                
                                # Créer la nouvelle affectation
                                EtudiantSousGroupe.objects.create(
                                    etudiant=etudiant,
                                    sous_groupe=sous_groupe,
                                    affecte_par=enseignant
                                )
                                modifications_reussies += 1
                        else:
                            # Nouvelle affectation pour un étudiant non affecté
                            EtudiantSousGroupe.objects.create(
                                etudiant=etudiant,
                                sous_groupe=sous_groupe,
                                affecte_par=enseignant
                            )
                            affectations_reussies += 1
                            
                except Exception as e:
                    print(f"Erreur affectation étudiant {etudiant.id}: {e}")
        
        # Messages de succès
        if affectations_reussies > 0 and modifications_reussies > 0:
            messages.success(request, f"تم تعيين {affectations_reussies} طالب جديد وتعديل {modifications_reussies} تعيين موجود بنجاح!")
        elif affectations_reussies > 0:
            messages.success(request, f"تم تعيين {affectations_reussies} طالب جديد بنجاح!")
        elif modifications_reussies > 0:
            messages.success(request, f"تم تعديل {modifications_reussies} تعيين بنجاح!")
        else:
            messages.warning(request, "لم يتم إجراء أي تغييرات. تأكد من اختيار المجموعات الفرعية.")
        
        # Rediriger vers la même page pour voir les résultats
        return redirect('ense:liste_sous_groupes', dep_id=dep_id, niv_spe_dep_sg_id=niv_spe_dep_sg_id)
    
    # Si pas POST, rediriger vers la liste
    return redirect('ense:liste_sous_groupes', dep_id=dep_id, niv_spe_dep_sg_id=niv_spe_dep_sg_id)




#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def delete_seance(request, dep_id, sea_id, enseignant, departement):
    """Supprimer une séance et toutes ses données associées"""
    try:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    except:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant)
    
    my_Dep = Departement.objects.get(id=departement.id)
    
    try:
        # Récupérer la séance à supprimer
        seance_to_delete = get_object_or_404(Seance, id=sea_id)
        classe_id = seance_to_delete.classe.id
        classe_obj = seance_to_delete.classe
        
        # Vérifier que l'enseignant a le droit de supprimer cette séance
        if seance_to_delete.classe.enseignant.enseignant != enseignant:
            messages.error(request, "ليس لديك صلاحية لحذف هذه الحصة")
            return redirect('ense:list_Sea_Ens', dep_id=dep_id, clas_id=classe_id)
        
        if request.method == 'POST':
            # Confirmation de suppression reçue
            
            # 📝 ÉTAPE 1: Supprimer toutes les absences liées à cette séance
            absences_seance = Abs_Etu_Seance.objects.filter(seance=seance_to_delete)
            absences_count = absences_seance.count()
            
            # Récupérer les étudiants concernés avant suppression
            etudiants_concernes = set(abs.etudiant for abs in absences_seance)
            
            # Supprimer les absences de la séance
            absences_seance.delete()
            
            # 📝 ÉTAPE 2: Supprimer la séance elle-même
            seance_date = seance_to_delete.date
            seance_intitule = seance_to_delete.intitule or "Séance sans titre"
            seance_to_delete.delete()
            
            # 📝 ÉTAPE 3: Mettre à jour les compteurs Abs_Etu_Classe
            # Récupérer toutes les séances restantes de cette classe
            seances_restantes = Seance.objects.filter(classe=classe_obj)
            
            if seances_restantes.exists():
                # Il reste des séances, recalculer les compteurs
                for etudiant in etudiants_concernes:
                    try:
                        abs_etu_classe = Abs_Etu_Classe.objects.get(
                            classe=classe_obj, 
                            etudiant=etudiant
                        )
                        
                        # Remettre les compteurs à zéro
                        abs_etu_classe.nbr_absence = 0
                        abs_etu_classe.nbr_absence_justifiee = 0
                        
                        # Recalculer à partir des séances restantes
                        for seance_restante in seances_restantes:
                            absences_etudiant = Abs_Etu_Seance.objects.filter(
                                seance=seance_restante,
                                etudiant=etudiant
                            )
                            
                            for absence in absences_etudiant:
                                if not absence.present and not absence.justifiee:
                                    abs_etu_classe.nbr_absence += 1
                                elif not absence.present and absence.justifiee:
                                    abs_etu_classe.nbr_absence_justifiee += 1
                        
                        abs_etu_classe.save()
                        
                    except Abs_Etu_Classe.DoesNotExist:
                        # L'enregistrement n'existe pas, pas de problème
                        continue
            else:
                # Plus de séances dans cette classe, supprimer tous les Abs_Etu_Classe
                Abs_Etu_Classe.objects.filter(
                    classe=classe_obj,
                    etudiant__in=etudiants_concernes
                ).delete()
                
                # Marquer la classe comme n'ayant plus de liste d'absences
                classe_obj.abs_liste_Etu = False
                classe_obj.save()
            
            # 📝 ÉTAPE 4: Vérifier s'il faut mettre à jour le statut seance_created de la classe
            if not seances_restantes.exists():
                # Plus aucune séance, marquer seance_created à False
                classe_obj.seance_created = False
                classe_obj.save()
            
            # Message de succès
            messages.success(
                request, 
                f'تم حذف الحصة بتاريخ {seance_date} ({seance_intitule}) بنجاح، '
                f'وتم حذف {absences_count} سجل غياب مرتبط بها وتحديث إحصائيات الغيابات'
            )
            
            return redirect('ense:list_Sea_Ens', dep_id=dep_id, clas_id=classe_id)
        
        # GET request - afficher la page de confirmation
        context = {
            'title': 'تأكيد حذف الحصة',
            'seance': seance_to_delete,
            'my_Dep': departement,
            'my_Fac': departement.faculte,
            'my_Ens': enseignant,
            'classe_id': classe_id,
        }
        return render(request, 'enseignant/confirm_delete_seance.html', context)
        
    except Exception as e:
        messages.error(request, f'حدث خطأ أثناء حذف الحصة: {str(e)}')
        return redirect('ense:list_Sea_Ens', dep_id=dep_id, clas_id=classe_id)
    



#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def init_Notes_Etu_Classe(request, dep_id, clas_id, enseignant, departement):
    """
    Initialise les notes pour tous les étudiants d'une classe
    """
    try:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    except:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant)
    
    my_Dep = Departement.objects.get(id=dep_id)
    my_Clas = get_object_or_404(Classe, id=clas_id)
    
    # Vérifier si les notes sont déjà initialisées
    if hasattr(my_Clas, 'notes_liste_Etu') and my_Clas.notes_liste_Etu:
        messages.info(request, 'النقاط موجودة مسبقاً لهذا الفصل')
        return redirect('ense:list_Notes_Etu_Classe', dep_id=dep_id, clas_id=clas_id)
    
    try:
        with transaction.atomic():
            # ✅ CORRECTION : Récupérer les étudiants sans le filtre 'actif'
            etudiants = None
                        
            # Méthode 3 : Via d'autres relations possibles
            if etudiants is None:
                try:
                    # Adaptez selon votre structure de modèle
                    etudiants = Etudiant.objects.filter(
                        niv_spe_dep_sg_id=my_Clas.niv_spe_dep_sg.id
                    )
                except:
                    etudiants = None
            
            if not etudiants.exists():
                messages.error(request, 'لا يوجد طلاب في هذا المستوى')
                return redirect('ense:timeTable_Ens', dep_id=dep_id)
            
            # Créer les enregistrements de notes pour chaque étudiant
            notes_created = 0
            notes_updated = 0
            
            for etudiant in etudiants:
                # Calculer le nombre de séances
                total_seances = 0
                try:
                    total_seances = my_Clas.seances_classe.filter(annuler=False).count()
                except:
                    total_seances = 0
                
                note_obj, created = Abs_Etu_Classe.objects.get_or_create(
                    classe=my_Clas,
                    etudiant=etudiant,
                    defaults={
                        'nbr_absence': 0,
                        'nbr_absence_justifiee': 0,
                        'nbr_seances_totales': total_seances,
                        'note_presence': 0.0,
                        'note_participe_HW': 0.0,
                        'note_controle_1': 0.0,
                        'note_controle_2': 0.0,
                        'total_sup_seance': 0.0,
                        # 'total_sup_examen': 0.0,
                        'note_finale': 0.0,
                        'obs': ''
                    }
                )
                
                if created:
                    notes_created += 1
                else:
                    # Mettre à jour le nombre de séances si l'enregistrement existe déjà
                    note_obj.nbr_seances_totales = total_seances
                    note_obj.save()
                    notes_updated += 1
            
            # Marquer la classe comme ayant ses notes initialisées
            if hasattr(my_Clas, 'notes_liste_Etu'):
                my_Clas.notes_liste_Etu = True
                my_Clas.save()
            
            if notes_created > 0:
                messages.success(request, f'تم إنشاء النقاط لـ {notes_created} طالب بنجاح')
            if notes_updated > 0:
                messages.info(request, f'تم تحديث النقاط لـ {notes_updated} طالب')
            
            if notes_created == 0 and notes_updated == 0:
                messages.warning(request, 'لم يتم إنشاء أو تحديث أي نقاط')
                
    except Exception as e:
        messages.error(request, f'خطأ في إنشاء النقاط: {str(e)}')
        print(f"DEBUG ERROR: {str(e)}")  # Pour debugging
        return redirect('ense:timeTable_Ens', dep_id=dep_id)
    
    # Redirection vers la liste des notes
    return redirect('ense:list_Notes_Etu_Classe', dep_id=dep_id, clas_id=clas_id)






#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def list_Notes_Etu_Classe(request, dep_id, clas_id, enseignant, departement):
    """
    Affiche et permet de modifier les notes des étudiants d'une classe
    Les totaux supplémentaires sont calculés automatiquement depuis les séances
    """
    try:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    except:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant)
    
    my_Dep = Departement.objects.get(id=dep_id)
    my_Clas = get_object_or_404(Classe, id=clas_id)
    
    # Traitement du formulaire de mise à jour
    if request.method == 'POST':
        try:
            with transaction.atomic():
                updated_count = 0
                
                for key, value in request.POST.items():
                    # ✅ Exclure les champs total_sup_ car ils sont calculés automatiquement
                    if key.startswith('note_') or key.startswith('obs_'):
                        parts = key.split('_')
                        etudiant_id = parts[-1]
                        
                        try:
                            note_obj = Abs_Etu_Classe.objects.get(
                                classe=my_Clas,
                                etudiant_id=etudiant_id
                            )
                            
                            if key.startswith('note_presence_'):
                                note_obj.note_presence = Decimal(str(value or 0))
                            elif key.startswith('note_participe_HW_'):
                                note_obj.note_participe_HW = Decimal(str(value or 0))
                            elif key.startswith('note_controle_1'):
                                note_obj.note_controle_1 = Decimal(str(value or 0))
                            elif key.startswith('note_controle_2'):
                                note_obj.note_controle_2 = Decimal(str(value or 0))
                            elif key.startswith('obs_'):
                                note_obj.obs = value
                            
                            # ✅ Calculer automatiquement les totaux supplémentaires
                            note_obj.total_sup_seance = calculate_total_sup_seance_from_seances(note_obj.etudiant, my_Clas)
                            # note_obj.total_sup_examen = calculate_total_sup_examen_from_seances(note_obj.etudiant, my_Clas)
                            
                            # Sauvegarder (le calcul de la note finale se fait dans save())
                            note_obj.save()
                            updated_count += 1
                            
                        except (Abs_Etu_Classe.DoesNotExist, ValueError):
                            continue
                
                messages.success(request, f'تم تحديث النقاط بنجاح')
                
        except Exception as e:
            messages.error(request, f'خطأ في تحديث النقاط: {str(e)}')
    
    # ✅ Récalculer tous les totaux automatiquement avant affichage
    # ✅ CORRECTION : Recalculer TOUTES les notes avant affichage
    all_Notes_Classe = Abs_Etu_Classe.objects.filter(
        classe=my_Clas
    ).select_related('etudiant').order_by('etudiant__nom_ar', 'etudiant__prenom_ar')

    # ÉTAPE 1: Mettre à jour les notes de présence pour tous les étudiants
    for note in all_Notes_Classe:
        # Recalculer le nombre total de séances et la note de présence
        note.calculate_note_presence()
        
        # Calculer les totaux depuis les séances individuelles
        try:
            new_total_seance = calculate_total_sup_seance_from_seances(note.etudiant, my_Clas)
            note.total_sup_seance = new_total_seance
        except:
            pass  # Si la fonction n'existe pas
        
        # Sauvegarder avec recalcul de la note finale
        note.save()

    
    # Calculer les statistiques de la classe
    try:
        statistiques = Abs_Etu_Classe.get_statistiques_classe(my_Clas)
    except:
        statistiques = None
    
    titre_00 = my_Clas.niv_spe_dep_sg
    titre_01 = my_Clas.matiere.nom_fr
    
    
    context = {
        'title': 'نقاط الطلاب',
        'titre_00': titre_00,
        'titre_01': titre_01,
        'my_Clas': my_Clas,
        'all_Notes_Classe': all_Notes_Classe,
        'statistiques': statistiques,
        'my_Dep': my_Dep,
        'my_Fac': my_Dep.faculte,
        'my_Ens': enseignant,
    }
    
    return render(request, 'enseignant/list_Notes_Etu_Classe.html', context)





#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
# ✅ FONCTIONS DE CALCUL AUTOMATIQUE DEPUIS LES SÉANCES
def calculate_total_sup_seance_from_seances(etudiant, classe):
    """
    Calcule le total des points_sup_seance depuis toutes les séances de cet étudiant
    """
    try:
        # Récupérer toutes les absences/présences de cet étudiant pour cette classe
        absences_seances = Abs_Etu_Seance.objects.filter(
            etudiant=etudiant,
            seance__classe=classe
        )
        
        total = Decimal('0.0')
        for abs_seance in absences_seances:
            # Additionner tous les points_sup_seance
            if abs_seance.points_sup_seance:
                total += Decimal(str(abs_seance.points_sup_seance))
        
        # Limiter au maximum autorisé (3.0)
        return min(Decimal('3.00'), total)
        
    except Exception as e:
        print(f"Erreur calcul total_sup_seance: {e}")
        return Decimal('0.0')





#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
# ✅ FONCTION UTILITAIRE POUR RECALCULER TOUS LES TOTAUX D'UNE CLASSE
def recalculate_all_totals_for_classe(classe_id):
    """
    Recalcule tous les totaux supplémentaires pour une classe donnée
    À appeler quand les séances sont modifiées
    """
    try:
        classe = Classe.objects.get(id=classe_id)
        notes = Abs_Etu_Classe.objects.filter(classe=classe)
        
        updated_count = 0
        for note in notes:
            new_total_seance = calculate_total_sup_seance_from_seances(note.etudiant, classe)
            # new_total_examen = calculate_total_sup_examen_from_seances(note.etudiant, classe)
            
            if note.total_sup_seance != new_total_seance or note.total_sup_examen:# != new_total_examen:
                note.total_sup_seance = new_total_seance
                # note.total_sup_examen = new_total_examen
                note.save()
                updated_count += 1
        
        return updated_count
        
    except Exception as e:
        print(f"Erreur recalcul totaux classe: {e}")
        return 0
    



#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def valider_notes_classe(request, dep_id, clas_id, enseignant, departement):
    """
    Valide toutes les notes d'une classe
    """
    if request.method == 'POST':
        my_Clas = get_object_or_404(Classe, id=clas_id)
        
        try:
            with transaction.atomic():
                # Valider toutes les notes de la classe
                notes_updated = Abs_Etu_Classe.objects.filter(
                    classe=my_Clas
                ).update(
                    validee_par_enseignant=True,
                    date_validation=timezone.now()
                )
                
                messages.success(request, f'تم التصديق على {notes_updated} نقطة بنجاح')
                
        except Exception as e:
            messages.error(request, f'خطأ في التصديق: {str(e)}')
    else:
        # Si ce n'est pas une requête POST, rediriger
        messages.warning(request, 'طريقة الوصول غير صحيحة')
    
    return redirect('ense:list_Notes_Etu_Classe', dep_id=dep_id, clas_id=clas_id)







#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@enseignant_access_required
def export_notes_classe(request, dep_id, clas_id, enseignant, departement):
    """
    Exporte les notes d'une classe en format Excel
    """
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    
    my_Clas = get_object_or_404(Classe, id=clas_id)
    # ✅ TRI PAR NOM_FR au lieu de nom_ar pour correspondre à l'affichage de la table
    all_Notes_Classe = Abs_Etu_Classe.objects.filter(
        classe=my_Clas
    ).select_related('etudiant').order_by('etudiant__nom_fr')

    # Créer un workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Notes_{my_Clas.matiere.nom_fr}"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ✅ STYLES COLORÉS pour les colonnes de notes
    fill_primary = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")  # حضور
    fill_info = PatternFill(start_color="0dcaf0", end_color="0dcaf0", fill_type="solid")     # واجبات
    fill_warning = PatternFill(start_color="ffc107", end_color="ffc107", fill_type="solid")  # امتحان 1
    fill_danger = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")   # امتحان 2
    font_dark = Font(bold=True, color="000000")  # Pour le fond jaune

    # ✅ EN-TÊTES MODIFIÉS : nouvel ordre des colonnes + ajout de "عدد المشاركات" et "%حضور"
    headers = [
        '#', 'Nom', 'Prénom', 'اللقب', 'الاسم',
        'نقطة الحضور (0-5)', 'نقطة المشاركة (0-5)', 'نقطة الامتحان 1 (0-5)', 'نقطة الامتحان 2 (0-5)',
        'نقاط إضافية', '%حضور', 'عدد المشاركات', 'النقطة النهائية (0-20)', 'التقييم', 'ملاحظات'
    ]

    # ✅ COLORATION DES EN-TÊTES
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Appliquer les couleurs spécifiques
        if col == 6:  # نقطة الحضور
            cell.fill = fill_primary
            cell.font = header_font
        elif col == 7:  # نقطة المشاركة
            cell.fill = fill_info
            cell.font = header_font
        elif col == 8:  # نقطة الامتحان 1
            cell.fill = fill_warning
            cell.font = font_dark
        elif col == 9:  # نقطة الامتحان 2
            cell.fill = fill_danger
            cell.font = header_font
        else:
            cell.fill = header_fill
            cell.font = header_font

    # ✅ DONNÉES MODIFIÉES : nouvel ordre + ajout de taux_presence_pourcentage et nombre_participations
    for row, note in enumerate(all_Notes_Classe, 2):
        data = [
            row - 1,  # #
            note.etudiant.nom_fr or '',  # Nom
            note.etudiant.prenom_fr or '',  # Prénom
            note.etudiant.nom_ar or '',  # اللقب
            note.etudiant.prenom_ar or '',  # الاسم
            float(note.note_presence),  # نقطة الحضور
            float(note.note_participe_HW),  # نقطة المشاركة
            float(note.note_controle_1),  # نقطة الامتحان 1
            float(note.note_controle_2),  # نقطة الامتحان 2
            float(note.total_sup_seance),  # نقاط إضافية
            note.taux_presence_pourcentage,  # %حضور (NOUVEAU)
            note.nombre_participations,  # عدد المشاركات (NOUVEAU)
            f"{float(note.note_finale):.2f}",  # ✅ النقطة النهائية avec format ##.##
            note.mention,  # التقييم
            note.obs or ''  # ملاحظات
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Préparer la réponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="notes_{my_Clas.matiere.nom_fr}.xlsx"'
    
    wb.save(response)
    return response




#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
# ✅ FONCTION CORRIGÉE AVEC GESTION DECIMAL
def calculate_total_from_seances_decimal(etudiant, classe, type_points):
    """
    Calcule le total des points depuis les séances individuelles - Version Decimal
    """
    try:
        # Vérifier si le modèle Abs_Etu_Seance existe
        from apps.academique.affectation.models import Abs_Etu_Seance
        
        # Récupérer toutes les absences/séances de cet étudiant pour cette classe
        absences_seances = Abs_Etu_Seance.objects.filter(
            etudiant=etudiant,
            seance__classe=classe
        )
        
        total = Decimal('0.0')
        if type_points == 'seance':
            # Sommer tous les points_sup_seance
            for abs_seance in absences_seances:
                points = abs_seance.points_sup_seance or Decimal('0.0')
                total += Decimal(str(points))
        return total
        
    except Exception as e:
        # print(f"Erreur calcul total séances: {e}")
        return Decimal('0.0')






#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
# ✅ FONCTION POUR CALCULER LES TOTAUX DEPUIS LES SÉANCES INDIVIDUELLES
def calculate_total_from_seances(etudiant, classe, type_points):
    """
    Calcule le total des points depuis les séances individuelles
    etudiant: L'étudiant concerné
    classe: La classe concernée
    type_points: 'seance' pour points_sup_seance ou 'examen' pour points_sup_examen
    """
    try:
        # Récupérer toutes les absences/séances de cet étudiant pour cette classe
        absences_seances = Abs_Etu_Seance.objects.filter(
            etudiant=etudiant,
            seance__classe=classe
        )
        
        total = 0.0
        if type_points == 'seance':
            # Sommer tous les points_sup_seance
            total = sum(float(abs_seance.points_sup_seance or 0) for abs_seance in absences_seances)
        return round(total, 2)
        
    except Exception as e:
        # print(f"Erreur calcul total séances: {e}")
        return 0.0

        


#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
# ✅ NOUVELLES FONCTIONS DE CALCUL AUTOMATIQUE
def calculate_total_sup_seance(note_obj):
    """
    Calcule automatiquement total_sup_seance basé sur différents critères
    """
    total = 0.0
    
    # Critère 1 : Bonus pour excellent taux de présence
    if note_obj.taux_presence_pourcentage >= 95:
        total += 1.0
    elif note_obj.taux_presence_pourcentage >= 90:
        total += 0.5
    
    # Critère 2 : Bonus pour participation élevée
    if note_obj.note_participe_HW >= 4.5:
        total += 1.0
    elif note_obj.note_participe_HW >= 4.0:
        total += 0.5
    
    # Critère 3 : Bonus combiné (présence + participation)
    if note_obj.taux_presence_pourcentage >= 90 and note_obj.note_participe_HW >= 4.0:
        total += 0.5
    
    # Limiter au maximum autorisé (3.0)
    return min(3.0, total)



#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
# ✅ FONCTION UTILITAIRE POUR RECALCULER TOUTES LES NOTES D'UNE CLASSE
def recalculate_all_bonus_for_classe(classe_id):
    """
    Recalcule tous les bonus pour une classe donnée
    """
    notes = Abs_Etu_Classe.objects.filter(classe_id=classe_id)
    
    for note in notes:
        note.total_sup_seance = calculate_total_sup_seance(note)
        # note.total_sup_examen = calculate_total_sup_examen(note)
        note.save(update_fields=['total_sup_seance',])
    
    return notes.count()



#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
def calculate_total_sup_seance_from_seances(etudiant, classe):
    """
    Calcule le total des points_sup_seance depuis toutes les séances de cet étudiant
    Version corrigée pour views.py
    """
    try:
        # ✅ CORRECTION : Import direct depuis apps.academique.affectation.models
        from apps.academique.affectation.models import Abs_Etu_Seance
        
        # Récupérer toutes les absences/présences de cet étudiant pour cette classe
        absences_seances = Abs_Etu_Seance.objects.filter(
            etudiant=etudiant,
            seance__classe=classe
        )
        
        total = Decimal('0.0')
        for abs_seance in absences_seances:
            # Additionner tous les points_sup_seance
            if abs_seance.points_sup_seance:
                total += Decimal(str(abs_seance.points_sup_seance))
        
        # Limiter au maximum autorisé (3.0)
        return min(Decimal('3.00'), total)
        
    except Exception as e:
        # print(f"Erreur calcul total_sup_seance: {e}")
        return Decimal('0.0')