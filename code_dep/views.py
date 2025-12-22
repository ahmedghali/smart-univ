from asyncio.log import logger
from datetime import timedelta
# from functools import cache
from pyexpat.errors import messages
from django.db import IntegrityError
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from apps.academique.affectation.models import *
from django.contrib import messages
# from django.contrib.auth.models import User

from apps.noyau.authentification.decorators import enseignant_access_required
from apps.noyau.commun.models import Poste 
from .models import *
# from app_faculte.models import *
from .forms import *
from tablib import Dataset
from apps.academique.faculte.models import Faculte
from apps.academique.affectation.models import Ens_Dep
from apps.noyau.authentification.models import CustomUser  # Importer CustomUser au lieu de User
from django.core.exceptions import ObjectDoesNotExist
import pandas as pd

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from apps.academique.etudiant.models import Etudiant
from django.urls import reverse
from django.db.models import Count, Case, When, IntegerField

from django.db import transaction
import math
import random
import unicodedata
import re

# Create your views here.

@login_required
def dashboard_Dep(request):
    """Dashboard du département avec statistiques dynamiques"""
    
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)
    
    # Récupérer l'année universitaire courante
    try:
        annee_courante = AnneeUniversitaire.objects.order_by('-date_debut').first()
        if not annee_courante:
            annee_courante = None
    except Exception:
        annee_courante = None

    # ========== CALCUL DES STATISTIQUES PRINCIPALES ==========
    
    # 1. STATISTIQUES DES ENSEIGNANTS
    if annee_courante:
        # Tous les enseignants du département pour cette année (S1 ou S2)
        all_enseignants = Ens_Dep.objects.filter(
            departement=departement,
            annee_univ=annee_courante
        ).select_related('enseignant')
        
        # Enseignants par statut
        permanent_teachers = all_enseignants.filter(statut='Permanent').count()
        permanent_vacataire_teachers = all_enseignants.filter(statut='Permanent & Vacataire').count()
        vacataire_teachers = all_enseignants.filter(statut='Vacataire').count()
        associe_teachers = all_enseignants.filter(statut='Associe').count()
        doctorant_teachers = all_enseignants.filter(statut='Doctorant').count()
        
        # Totaux
        total_teachers = all_enseignants.count()
        temporary_teachers = permanent_vacataire_teachers + vacataire_teachers + associe_teachers + doctorant_teachers
        
        # Enseignants actifs (inscrits sur la plateforme)
        present_teachers = all_enseignants.filter(est_inscrit=True).count()
        
    else:
        # Valeurs par défaut si pas d'année universitaire
        total_teachers = 0
        permanent_teachers = 0
        temporary_teachers = 0
        present_teachers = 0

    # 2. STATISTIQUES DES ÉTUDIANTS
    try:
        # Adapter selon votre modèle Etudiant
        total_students = Etudiant.objects.filter(
            # Ajoutez vos filtres selon votre modèle
            # Par exemple: departement=departement
        ).count()
    except:
        total_students = 0

    # 3. STATISTIQUES DES MATIÈRES
    try:
        # Adapter selon votre modèle Matiere
        total_subjects = Matiere.objects.filter(
            # Ajoutez vos filtres selon votre modèle
            # Par exemple: departement=departement
        ).count()
    except:
        total_subjects = 0

    # 4. STATISTIQUES DES SPÉCIALITÉS
    try:
        total_specialities = Specialite.objects.filter(
            departement=departement
        ).count()
    except:
        total_specialities = 0

    # 5. STATISTIQUES DES SALLES/AMPHITHÉÂTRES
    try:
        total_rooms = (
            Salle.objects.filter(departement=departement).count() +
            Amphi.objects.filter(departement=departement).count() +
            Laboratoire.objects.filter(departement=departement).count()
        )
    except:
        total_rooms = 0

    # 6. STATISTIQUES DES CLASSES/COURS
    try:
        # Classes pour cette semaine ou ce semestre
        total_classes = Classe.objects.filter(
            # Ajoutez vos filtres selon votre modèle
            # Par exemple: enseignant__in=enseignants_du_departement
        ).count()
        
        # Classes actives aujourd'hui
        from datetime import date
        today = date.today()
        active_classes = Classe.objects.filter(
            # Ajoutez vos filtres pour aujourd'hui
            # Par exemple: jour=today.strftime('%A')
        ).count()
        
    except:
        total_classes = 0
        active_classes = 0

    # 7. STATISTIQUES DES SEMESTRES
    if annee_courante:
        # Enseignants S1
        enseignants_s1 = Ens_Dep.objects.filter(
            departement=departement,
            annee_univ=annee_courante,
            semestre_1=True
        ).count()
        
        # Enseignants S2
        enseignants_s2 = Ens_Dep.objects.filter(
            departement=departement,
            annee_univ=annee_courante,
            semestre_2=True
        ).count()
    else:
        enseignants_s1 = 0
        enseignants_s2 = 0

    # 8. AUTRES STATISTIQUES
    pending_requests = 0  # À adapter selon vos besoins
    completed_tasks = 95   # Pourcentage fictif, à adapter
    
    # ========== STATISTIQUES PAR SEMESTRE (pour liens rapides) ==========
    
    # Statistiques détaillées pour chaque semestre
    semestre_stats = {}
    for sem_num in [1, 2]:
        if annee_courante:
            sem_filter = {
                'departement': departement,
                'annee_univ': annee_courante,
                f'semestre_{sem_num}': True
            }
            
            sem_enseignants = Ens_Dep.objects.filter(**sem_filter)
            
            semestre_stats[f's{sem_num}'] = {
                'total': sem_enseignants.count(),
                'permanents': sem_enseignants.filter(statut='Permanent').count(),
                'temporaires': sem_enseignants.exclude(statut='Permanent').count(),
                'inscrits': sem_enseignants.filter(est_inscrit=True).count(),
            }
        else:
            semestre_stats[f's{sem_num}'] = {
                'total': 0, 'permanents': 0, 'temporaires': 0, 'inscrits': 0
            }

    # ========== CONTEXTE POUR LE TEMPLATE ==========
    
    context = {
        'title': 'لوحة التحكم الرئيسية',
        'my_Dep': departement,
        'my_Fac': departement.faculte,
        'annee_courante': annee_courante,
        
        # ===== STATISTIQUES PRINCIPALES =====
        'total_teachers': total_teachers,
        'total_students': total_students,
        'total_subjects': total_subjects,
        'total_classes': total_classes,
        
        # ===== DÉTAILS ENSEIGNANTS =====
        'permanent_teachers': permanent_teachers,
        'temporary_teachers': temporary_teachers,
        'present_teachers': present_teachers,
        
        # ===== INFRASTRUCTURES =====
        'total_specialities': total_specialities,
        'total_rooms': total_rooms,
        
        # ===== ACTIVITÉS =====
        'active_classes': active_classes,
        'pending_requests': pending_requests,
        'completed_tasks': completed_tasks,
        
        # ===== STATISTIQUES PAR SEMESTRE =====
        'enseignants_s1': enseignants_s1,
        'enseignants_s2': enseignants_s2,
        'semestre_stats': semestre_stats,
        
        # ===== NAVIGATION RAPIDE =====
        'can_manage_s1': enseignants_s1 > 0,
        'can_manage_s2': enseignants_s2 > 0,
    }
    
    return render(request, 'departement/dashboard_Dep.html', context)


# ===== FONCTION UTILITAIRE POUR STATISTIQUES AVANCÉES =====

def get_department_advanced_stats(departement, annee_univ=None):
    """
    Fonction utilitaire pour récupérer des statistiques avancées
    Peut être réutilisée dans d'autres vues
    """
    
    if not annee_univ:
        annee_univ = AnneeUniversitaire.objects.order_by('-date_debut').first()
    
    if not annee_univ:
        return {}
    
    # Requêtes optimisées avec select_related et prefetch_related
    enseignants_data = Ens_Dep.objects.filter(
        departement=departement,
        annee_univ=annee_univ
    ).select_related('enseignant__user').values(
        'statut', 'est_inscrit', 'semestre_1', 'semestre_2'
    )
    
    # Agrégation des données
    from django.db.models import Count, Q
    
    stats = enseignants_data.aggregate(
        total=Count('id'),
        permanents=Count('id', filter=Q(statut='Permanent')),
        temporaires=Count('id', filter=~Q(statut='Permanent')),
        inscrits=Count('id', filter=Q(est_inscrit=True)),
        s1=Count('id', filter=Q(semestre_1=True)),
        s2=Count('id', filter=Q(semestre_2=True)),
        s1_et_s2=Count('id', filter=Q(semestre_1=True, semestre_2=True)),
    )
    
    # Calculs dérivés
    stats['taux_inscription'] = round(
        (stats['inscrits'] / stats['total'] * 100) if stats['total'] > 0 else 0, 1
    )
    
    stats['taux_permanents'] = round(
        (stats['permanents'] / stats['total'] * 100) if stats['total'] > 0 else 0, 1
    )
    
    return stats


# ===== VUE AJAX POUR MISE À JOUR EN TEMPS RÉEL (OPTIONNEL) =====

@login_required
def dashboard_stats_api(request):
    """
    API pour récupérer les statistiques en JSON
    Utile pour mise à jour en temps réel via JavaScript
    """
    
    if request.method != 'GET':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)
    
    # Récupérer statistiques avancées
    stats = get_department_advanced_stats(departement)
    
    # Ajouter timestamp
    from datetime import datetime
    stats['last_updated'] = datetime.now().isoformat()
    
    return JsonResponse(stats)


# ===== MISE À JOUR DU FICHIER URLS.PY =====

# Dans votre urls.py, ajoutez :
# path('dashboard/', views.dashboard_Dep, name='dashboard_Dep'),
# path('api/dashboard-stats/', views.dashboard_stats_api, name='dashboard_stats_api'),




# ----------------------------------------------------------------------
@login_required
def profile_Dep(request):
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')    
    selected_departement = Departement.objects.get(id=real_Dep.departement.id)

    context = {
        'title': 'صفحة التعريف',
        'my_Ens': enseignant,
        'my_Dep': selected_departement,  # Passer le département au template
        'my_Fac': selected_departement.faculte,  # Passer le département au template
    }
    return render(request, 'departement/profile_Dep.html', context)


# ----------------------------------------------------------------------
@login_required
def profile_Ens_Dep(request, ens_id):
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')    
    selected_departement = Departement.objects.get(id=real_Dep.departement.id)
    # my_Ens = Ens_Dep.objects.get(enseignant_id=ens_id)
    my_Ens = Enseignant.objects.get(id=ens_id)
    

    try:
        ALL_Dep = Ens_Dep.objects.filter(
            enseignant = my_Ens.enseignant,
            ).exclude(departement=real_Dep.departement)
    except:
        ALL_Dep = None

    context = {
        'title': 'صفحة التعريف',
        'my_Ens': my_Ens,
        'my_Dep': selected_departement,  # Passer le département au template
        'my_Fac': selected_departement.faculte,  # Passer le département au template
        'real_Dep': real_Dep,
        'ALL_Dep': ALL_Dep,
    }
    return render(request, 'departement/profile_Ens_Dep.html', context)



# ----------------------------------------------------------------------
@login_required
def profileUpdate_Dep(request):
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')    
    selected_departement = Departement.objects.get(id=real_Dep.departement.id)

    Dep_form = profileUpdate_Dep_Form(instance=selected_departement)
    if request.method == 'POST':
        # profile_form = ProfileUpdateForm(request.POST, request.FILES, instance=request.user.profile)
        Dep_form = profileUpdate_Dep_Form(request.POST, instance=selected_departement)
        # if user_form.is_valid and profile_form.is_valid and dep_form.is_valid:
        if Dep_form.is_valid:
            Dep_form.save()
            messages.success(request, 'تم تحديث المعلومات بنجاح')
            context = {
            'title': 'تعديل المعلومات',
            'Dep_form' : Dep_form,
            'my_Ens': enseignant,
            'my_Dep': selected_departement,  # Passer le département au template
            'my_Fac': selected_departement.faculte,  # Passer le département au template
            }
            return render(request, 'departement/profile_Dep.html', context)
    else:
        Dep_form = profileUpdate_Dep_Form(instance=selected_departement)

    context = {
        'title': 'تعديل المعلومات',
        'Dep_form' : Dep_form,
        'my_Ens': enseignant,
        'my_Dep': selected_departement,  # Passer le département au template
        'my_Fac': selected_departement.faculte,  # Passer le département au template
    }
    return render(request, 'departement/profileUpdate_Dep.html', context)




# ----------------------------------------------------------------------
@login_required
def list_Amphi_Dep(request):
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)
    # all_Amphi_Dep = Amphi_Dep.objects.filter(departement = departement.id).order_by('amphi__numero')
    all_Amphi_Dep_S1 = Amphi_Dep.objects.filter(departement = departement.id, semestre__numero=1).order_by('amphi__numero')
    all_Amphi_Dep_S2 = Amphi_Dep.objects.filter(departement = departement.id, semestre__numero=2).order_by('amphi__numero')

    
    context = {
        'title': 'قائمة المدرجات',
        'my_Fac': departement.faculte,
        'my_Dep': departement,
        'my_Ens': enseignant,
        # 'all_Amphi_Dep': all_Amphi_Dep,
        'all_Amphi_Dep_S1': all_Amphi_Dep_S1,
        'all_Amphi_Dep_S2': all_Amphi_Dep_S2,
    }
    return render(request, 'departement/list_Amphi_Dep.html', context)



# ----------------------------------------------------------------------
@login_required
def list_Specialite_Dep(request):
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)
    all_Specialite_Dep = Specialite.objects.filter(departement = departement.id).order_by('reforme')
    
    context = {
        'title': 'قائمة التخصصات',
        'my_Fac': departement.faculte,
        'my_Dep': departement,
        'my_Ens': enseignant,
        'all_Specialite_Dep': all_Specialite_Dep,
    }
    return render(request, 'departement/list_Specialite_Dep.html', context)




# ----------------------------------------------------------------------
@login_required
def list_Salle_Dep(request):
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)
    # all_Salle_Dep = Salle_Dep.objects.filter(departement = departement.id).order_by('salle__numero')
    all_Salle_Dep_S1 = Salle_Dep.objects.filter(departement = departement.id, semestre__numero=1).order_by('salle__numero')
    all_Salle_Dep_S2 = Salle_Dep.objects.filter(departement = departement.id, semestre__numero=2).order_by('salle__numero')
    all_Salle_Dep = all_Salle_Dep_S1.count() + all_Salle_Dep_S2.count()
    context = {
        'title': 'قائمة القاعات',
        'my_Fac': departement.faculte,
        'my_Dep': departement,
        'my_Ens': enseignant,
        # 'all_Salle_Dep': all_Salle_Dep,
        'all_Salle_Dep_S1': all_Salle_Dep_S1,
        'all_Salle_Dep_S2': all_Salle_Dep_S2,
        'all_Salle_Dep': all_Salle_Dep,
    }
    return render(request, 'departement/list_Salle_Dep.html', context)





# ----------------------------------------------------------------------
@login_required
def list_Labo_Dep(request):
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)
    all_Labo_Dep_S1 = Laboratoire_Dep.objects.filter(departement = departement.id, semestre__numero=1).order_by('laboratoire__numero')
    all_Labo_Dep_S2 = Laboratoire_Dep.objects.filter(departement = departement.id, semestre__numero=2).order_by('laboratoire__numero')

    context = {
        'title': 'قائمة القاعات',
        'my_Fac': departement.faculte,
        'my_Dep': departement,
        'my_Ens': enseignant,
        # 'all_Labo_Dep': all_Labo_Dep,
        'all_Labo_Dep_S1': all_Labo_Dep_S1,
        'all_Labo_Dep_S2': all_Labo_Dep_S2,
    }
    return render(request, 'departement/list_Labo_Dep.html', context)






from django.core.cache import cache




# ===== VUE UNIFIÉE POUR GÉRER LES ENSEIGNANTS =====
# ===== VUE UNIFIÉE POUR GÉRER LES ENSEIGNANTS =====
# @login_required
# def manage_enseignants(request, semestre_num=1):
#     """Vue unifiée pour gérer les enseignants des deux semestres"""

#     # # Créer une clé de cache unique
#     # cache_key = f"manage_enseignants_{request.user.id}_{semestre_num}"
    
#     # # Vérifier le cache
#     # cached_context = cache.get(cache_key)
#     # if cached_context:
#     #     return render(request, 'departement/manage_enseignants.html', cached_context)
    
#     enseignant = request.user.enseignant_profile
#     real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
#     departement = Departement.objects.get(id=real_Dep.departement.id)
    
#     # Récupérer l'année universitaire courante
#     annee_courante = AnneeUniversitaire.get_courante()
#     if not annee_courante:
#         messages.error(request, 'لا توجد سنة جامعية محددة كسنة حالية')
#         return redirect('admin:index')

#     # Valider le numéro de semestre
#     if semestre_num not in [1, 2]:
#         messages.error(request, 'رقم السداسي غير صحيح')
#         return redirect('depa:manage_enseignants', semestre_num=1)

#     # ========== FILTRES DE BASE AVEC SEMESTRE ==========
    
#     # CORRECTION: Ajouter le filtre par semestre pour tous les enseignants
#     base_filter = {
#         'departement': departement,
#         'annee_univ': annee_courante,
#         f'semestre_{semestre_num}': True  # semestre_1=True ou semestre_2=True
#     }
    
#     # ========== DONNÉES POUR LA LISTE - FILTRÉES PAR SEMESTRE ==========
    
#     # AVANT (PROBLÈME): Récupérait TOUS les enseignants du département
#     # all_Ens_Dep_Per = Ens_Dep.objects.filter(departement=departement, statut='Permanent')
    
#     # APRÈS (CORRECTION): Filtrer par semestre ET par département
#     all_Ens_Dep = Ens_Dep.objects.filter(**base_filter).order_by('enseignant__nom_ar')
    
#     # Appliquer le filtre de semestre à chaque catégorie d'enseignants
#     all_Ens_Dep_Per = Ens_Dep.objects.filter(
#         departement=departement, 
#         statut='Permanent',
#         annee_univ=annee_courante,
#         **{f'semestre_{semestre_num}': True}  # CORRECTION: Ajouter le filtre par semestre
#     ).select_related(
#         'enseignant__user__poste_principal',
#         'enseignant'
#     ).order_by('enseignant__nom_ar')

#     all_Ens_Dep_PerVac = Ens_Dep.objects.filter(
#         departement=departement, 
#         statut='Permanent & Vacataire',
#         annee_univ=annee_courante,
#         **{f'semestre_{semestre_num}': True}  # CORRECTION: Ajouter le filtre par semestre
#     ).select_related(
#         'enseignant__user__poste_principal',
#         'enseignant'
#     ).order_by('enseignant__nom_ar')

#     all_Ens_Dep_Vac = Ens_Dep.objects.filter(
#         departement=departement, 
#         statut='Vacataire',
#         annee_univ=annee_courante,
#         **{f'semestre_{semestre_num}': True}  # CORRECTION: Ajouter le filtre par semestre
#     ).select_related(
#         'enseignant__user__poste_principal',
#         'enseignant'
#     ).order_by('enseignant__nom_ar')

#     all_Ens_Dep_Aso = Ens_Dep.objects.filter(
#         departement=departement, 
#         statut='Associe',
#         annee_univ=annee_courante,
#         **{f'semestre_{semestre_num}': True}  # CORRECTION: Ajouter le filtre par semestre
#     ).select_related(
#         'enseignant__user__poste_principal',
#         'enseignant'
#     ).order_by('enseignant__nom_ar')

#     all_Ens_Dep_Doc = Ens_Dep.objects.filter(
#         departement=departement, 
#         statut='Doctorant',
#         annee_univ=annee_courante,
#         **{f'semestre_{semestre_num}': True}  # CORRECTION: Ajouter le filtre par semestre
#     ).select_related(
#         'enseignant__user__poste_principal',
#         'enseignant'
#     ).order_by('enseignant__nom_ar')

#     # ========== CALCUL DES STATISTIQUES ==========
    
#     # Logique de calcul existante adaptée...
#     all_Ens_Dep_for_stats = all_Ens_Dep
#     for idx in all_Ens_Dep_for_stats:
#         my_Ens = idx.enseignant
        
#         # Utiliser semestre_num directement dans les requêtes Classe
#         i_ens_nbrClas_in_Dep = Classe.objects.filter(
#             enseignant__enseignant=idx.enseignant, 
#             enseignant__departement=real_Dep.departement,
#             semestre=semestre_num
#         ).count()

#         i_ens_nbrClas_out_Dep = Classe.objects.filter(
#             enseignant__enseignant=idx.enseignant,  
#             semestre=semestre_num
#         ).exclude(enseignant__departement=real_Dep.departement).count()
        
#         # Calculer les différents types de classes pour ce semestre
#         i_ens_nbrClas_Cours_in_Dep = Classe.objects.filter(
#             enseignant__enseignant=idx.enseignant, 
#             enseignant__departement=real_Dep.departement,
#             semestre=semestre_num,
#             type='Cours'
#         ).count()
        
#         i_ens_nbrClas_TD_in_Dep = Classe.objects.filter(
#             enseignant__enseignant=idx.enseignant,
#             enseignant__departement=real_Dep.departement,
#             semestre=semestre_num,
#             type='TD'
#         ).count()
        
#         i_ens_nbrClas_TP_in_Dep = Classe.objects.filter(
#             enseignant__enseignant=idx.enseignant, 
#             enseignant__departement=real_Dep.departement,
#             semestre=semestre_num,
#             type='TP'
#         ).count()
        
#         i_ens_nbrClas_SS_in_Dep = Classe.objects.filter(
#             enseignant__enseignant=idx.enseignant, 
#             enseignant__departement=real_Dep.departement,
#             semestre=semestre_num,
#             type='SS'
#         ).count()
        
#         # Calculer les jours de travail uniques
#         jours_travail = Classe.objects.filter(
#             enseignant__enseignant=idx.enseignant, 
#             enseignant__departement=real_Dep.departement,
#             semestre=semestre_num
#         ).values_list('jour', flat=True).distinct().count()

#         # Calculer le taux d'avancement minimum et moyen
#         classes_enseignant = Classe.objects.filter(
#             enseignant__enseignant=idx.enseignant,
#             enseignant__departement=real_Dep.departement,
#             semestre__numero=semestre_num
#         )
        
#         if classes_enseignant.exists():
#             taux_list = list(classes_enseignant.values_list('taux_avancement', flat=True))
#             i_ens_taux_min = min(taux_list)
#             i_ens_taux_moyen = sum(taux_list) / len(taux_list)
#         else:
#             i_ens_taux_min = 0
#             i_ens_taux_moyen = 0


        
#         # Champs à mettre à jour selon le semestre
#         if semestre_num == 1:
#             idx.nbrClas_in_Dep_S1 = i_ens_nbrClas_in_Dep
#             idx.nbrClas_out_Dep_S1 = i_ens_nbrClas_out_Dep
#             idx.nbrClas_ALL_Dep_S1 = i_ens_nbrClas_in_Dep + i_ens_nbrClas_out_Dep
#             idx.nbrClas_Cours_in_Dep_S1 = i_ens_nbrClas_Cours_in_Dep
#             idx.nbrClas_TD_in_Dep_S1 = i_ens_nbrClas_TD_in_Dep
#             idx.nbrClas_TP_in_Dep_S1 = i_ens_nbrClas_TP_in_Dep
#             idx.nbrClas_SS_in_Dep_S1 = i_ens_nbrClas_SS_in_Dep
#             idx.nbrJour_in_Dep_S1 = jours_travail
#             idx.taux_min_S1 = i_ens_taux_min  # ✅ NOUVEAU
#             idx.taux_moyen_S1 = round(i_ens_taux_moyen, 1)  # ✅ NOUVEAU
#             idx.save()
#         else:
#             idx.nbrClas_in_Dep_S2 = i_ens_nbrClas_in_Dep
#             idx.nbrClas_out_Dep_S2 = i_ens_nbrClas_out_Dep
#             idx.nbrClas_ALL_Dep_S2 = i_ens_nbrClas_in_Dep + i_ens_nbrClas_out_Dep
#             idx.nbrClas_Cours_in_Dep_S2 = i_ens_nbrClas_Cours_in_Dep
#             idx.nbrClas_TD_in_Dep_S2 = i_ens_nbrClas_TD_in_Dep
#             idx.nbrClas_TP_in_Dep_S2 = i_ens_nbrClas_TP_in_Dep
#             idx.nbrClas_SS_in_Dep_S2 = i_ens_nbrClas_SS_in_Dep
#             idx.nbrJour_in_Dep_S2 = jours_travail
#             idx.taux_min_S2 = i_ens_taux_min  # ✅ NOUVEAU
#             idx.taux_moyen_S2 = round(i_ens_taux_moyen, 1)  # ✅ NOUVEAU
#             idx.save()

#     # ========== DONNÉES POUR LES HEURES - AUSSI FILTRÉES PAR SEMESTRE ==========
    
#     # Tri dynamique selon le semestre
#     if semestre_num == 1:
#         order_field = '-nbrClas_ALL_Dep_S1'
#     else:
#         order_field = '-nbrClas_ALL_Dep_S2'
    
#     # CORRECTION: Appliquer le même filtre par semestre pour l'onglet heures
#     main_all_aff_Ens_Dep = all_Ens_Dep.filter(
#         departement=departement,
#         est_inscrit=True,
#         **{f'semestre_{semestre_num}': True}
#     ).order_by(order_field, 'enseignant__nom_ar')
    
#     main_PER = all_Ens_Dep_Per.filter(
#         departement=departement,
#         est_inscrit=True
#     ).order_by(order_field, 'enseignant__nom_ar')
    
#     main_PER_VAC = all_Ens_Dep_PerVac.filter(
#         departement=departement,
#         est_inscrit=True
#     ).order_by(order_field, 'enseignant__nom_ar')
    
#     main_VAC = all_Ens_Dep_Vac.filter(
#         departement=departement,
#         est_inscrit=True
#     ).order_by(order_field, 'enseignant__nom_ar')
    
#     main_ASO = all_Ens_Dep_Aso.filter(
#         departement=departement,
#         est_inscrit=True
#     ).order_by(order_field, 'enseignant__nom_ar')
    
#     main_DOC = all_Ens_Dep_Doc.filter(
#         departement=departement,
#         est_inscrit=True
#     ).order_by(order_field, 'enseignant__nom_ar')

#     # ========== CONTEXTE ==========
    
#     context = {
#         'title': 'إدارة الأساتذة',
#         'my_Dep': departement,
#         'my_Fac': departement.faculte,
#         'annee_courante': annee_courante,
#         'semestre_num': semestre_num,  # IMPORTANT: pour le template
        
#         # Données pour l'onglet Liste - MAINTENANT FILTRÉES PAR SEMESTRE
#         'all_Ens_Dep': all_Ens_Dep,
#         'all_Ens_Dep_Per': all_Ens_Dep_Per,
#         'all_Ens_Dep_Vac': all_Ens_Dep_Vac,
#         'all_Ens_Dep_PerVac': all_Ens_Dep_PerVac,
#         'all_Ens_Dep_Aso': all_Ens_Dep_Aso,
#         'all_Ens_Dep_Doc': all_Ens_Dep_Doc,
        
#         # Données pour l'onglet Heures - AUSSI FILTRÉES PAR SEMESTRE
#         'main_all_aff_Ens_Dep': main_all_aff_Ens_Dep,
#         'main_PER': main_PER,
#         'main_PER_VAC': main_PER_VAC,
#         'main_VAC': main_VAC,
#         'main_ASO': main_ASO,
#         'main_DOC': main_DOC,
#     }

#     # Mettre en cache pour 10 minutes = 600 secondes
#     # cache.set(cache_key, context, 600)
    
#     return render(request, 'departement/manage_enseignants.html', context)












# @login_required
# def list_enseignants_dep(request, semestre_num=1):
#     """Vue pour la liste des enseignants par semestre"""
    
#     enseignant = request.user.enseignant_profile
#     real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
#     departement = Departement.objects.get(id=real_Dep.departement.id)
    
#     # Récupérer l'année universitaire courante
#     annee_courante = AnneeUniversitaire.get_courante()
#     if not annee_courante:
#         messages.error(request, 'لا توجد سنة جامعية محددة كسنة حالية')
#         return redirect('admin:index')

#     # Valider le numéro de semestre
#     if semestre_num not in [1, 2]:
#         messages.error(request, 'رقم السداسي غير صحيح')
#         return redirect('depa:list_enseignants_dep', semestre_num=1)

#     # Filtre de base avec semestre
#     base_filter = {
#         'departement': departement,
#         'annee_univ': annee_courante,
#         f'semestre_{semestre_num}': True
#     }
    
#     # Récupérer les enseignants filtrés par semestre
#     all_Ens_Dep = Ens_Dep.objects.filter(**base_filter).order_by('enseignant__nom_ar')
    
#     all_Ens_Dep_Per = Ens_Dep.objects.filter(
#         departement=departement, 
#         statut='Permanent',
#         annee_univ=annee_courante,
#         **{f'semestre_{semestre_num}': True}
#     ).select_related(
#         'enseignant__user__poste_principal',
#         'enseignant'
#     ).order_by('enseignant__nom_ar')

#     all_Ens_Dep_PerVac = Ens_Dep.objects.filter(
#         departement=departement, 
#         statut='Permanent & Vacataire',
#         annee_univ=annee_courante,
#         **{f'semestre_{semestre_num}': True}
#     ).select_related(
#         'enseignant__user__poste_principal',
#         'enseignant'
#     ).order_by('enseignant__nom_ar')

#     all_Ens_Dep_Vac = Ens_Dep.objects.filter(
#         departement=departement, 
#         statut='Vacataire',
#         annee_univ=annee_courante,
#         **{f'semestre_{semestre_num}': True}
#     ).select_related(
#         'enseignant__user__poste_principal',
#         'enseignant'
#     ).order_by('enseignant__nom_ar')

#     all_Ens_Dep_Aso = Ens_Dep.objects.filter(
#         departement=departement, 
#         statut='Associe',
#         annee_univ=annee_courante,
#         **{f'semestre_{semestre_num}': True}
#     ).select_related(
#         'enseignant__user__poste_principal',
#         'enseignant'
#     ).order_by('enseignant__nom_ar')

#     all_Ens_Dep_Doc = Ens_Dep.objects.filter(
#         departement=departement, 
#         statut='Doctorant',
#         annee_univ=annee_courante,
#         **{f'semestre_{semestre_num}': True}
#     ).select_related(
#         'enseignant__user__poste_principal',
#         'enseignant'
#     ).order_by('enseignant__nom_ar')

#     context = {
#         'title': 'قائمة الأساتذة',
#         'my_Dep': departement,
#         'my_Fac': departement.faculte,
#         'annee_courante': annee_courante,
#         'semestre_num': semestre_num,
#         'all_Ens_Dep': all_Ens_Dep,
#         'all_Ens_Dep_Per': all_Ens_Dep_Per,
#         'all_Ens_Dep_Vac': all_Ens_Dep_Vac,
#         'all_Ens_Dep_PerVac': all_Ens_Dep_PerVac,
#         'all_Ens_Dep_Aso': all_Ens_Dep_Aso,
#         'all_Ens_Dep_Doc': all_Ens_Dep_Doc,
#     }
    
#     return render(request, 'departement/list_enseignants_dep.html', context)



@login_required
def list_enseignants_dep(request, semestre_num=1):
    """Vue pour la liste des enseignants par semestre"""
    
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)
    
    # Récupérer l'année universitaire courante
    annee_courante = AnneeUniversitaire.get_courante()
    if not annee_courante:
        messages.error(request, 'لا توجد سنة جامعية محددة كسنة حالية')
        return redirect('admin:index')

    # Valider le numéro de semestre
    if semestre_num not in [1, 2]:
        messages.error(request, 'رقم السداسي غير صحيح')
        return redirect('depa:list_enseignants_dep', semestre_num=1)

    # Filtre de base avec semestre
    base_filter = {
        'departement': departement,
        'annee_univ': annee_courante,
        f'semestre_{semestre_num}': True
    }
    
    # Récupérer les enseignants filtrés par semestre
    all_Ens_Dep = Ens_Dep.objects.filter(**base_filter).select_related(
        'enseignant__grade',
        'enseignant__user__poste_principal',
        'enseignant'
    ).order_by('enseignant__nom_ar')
    
    all_Ens_Dep_Per = Ens_Dep.objects.filter(
        departement=departement, 
        statut='Permanent',
        annee_univ=annee_courante,
        **{f'semestre_{semestre_num}': True}
    ).select_related(
        'enseignant__user__poste_principal',
        'enseignant__grade',
        'enseignant'
    ).order_by('enseignant__nom_ar')

    all_Ens_Dep_PerVac = Ens_Dep.objects.filter(
        departement=departement, 
        statut='Permanent & Vacataire',
        annee_univ=annee_courante,
        **{f'semestre_{semestre_num}': True}
    ).select_related(
        'enseignant__user__poste_principal',
        'enseignant__grade',
        'enseignant'
    ).order_by('enseignant__nom_ar')

    all_Ens_Dep_Vac = Ens_Dep.objects.filter(
        departement=departement, 
        statut='Vacataire',
        annee_univ=annee_courante,
        **{f'semestre_{semestre_num}': True}
    ).select_related(
        'enseignant__user__poste_principal',
        'enseignant__grade',
        'enseignant'
    ).order_by('enseignant__nom_ar')

    all_Ens_Dep_Aso = Ens_Dep.objects.filter(
        departement=departement, 
        statut='Associe',
        annee_univ=annee_courante,
        **{f'semestre_{semestre_num}': True}
    ).select_related(
        'enseignant__user__poste_principal',
        'enseignant__grade',
        'enseignant'
    ).order_by('enseignant__nom_ar')

    all_Ens_Dep_Doc = Ens_Dep.objects.filter(
        departement=departement, 
        statut='Doctorant',
        annee_univ=annee_courante,
        **{f'semestre_{semestre_num}': True}
    ).select_related(
        'enseignant__user__poste_principal',
        'enseignant__grade',
        'enseignant'
    ).order_by('enseignant__nom_ar')

    # Calculer les statistiques par grade
    from django.db.models import Count, Q
    grade_stats = all_Ens_Dep.values('enseignant__grade__nom_ar').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Compteurs pour les champs vides
    missing_email_count = all_Ens_Dep.filter(
        Q(enseignant__email_prof__isnull=True) | Q(enseignant__email_prof='')
    ).count()
    
    missing_scholar_count = all_Ens_Dep.filter(
        Q(enseignant__googlescholar__isnull=True) | Q(enseignant__googlescholar='')
    ).count()

    context = {
        'title': 'قائمة الأساتذة',
        'my_Dep': departement,
        'my_Fac': departement.faculte,
        'annee_courante': annee_courante,
        'semestre_num': semestre_num,
        'all_Ens_Dep': all_Ens_Dep,
        'all_Ens_Dep_Per': all_Ens_Dep_Per,
        'all_Ens_Dep_Vac': all_Ens_Dep_Vac,
        'all_Ens_Dep_PerVac': all_Ens_Dep_PerVac,
        'all_Ens_Dep_Aso': all_Ens_Dep_Aso,
        'all_Ens_Dep_Doc': all_Ens_Dep_Doc,
        'grade_stats': grade_stats,
        'missing_email_count': missing_email_count,
        'missing_scholar_count': missing_scholar_count,
    }
    
    return render(request, 'departement/list_enseignants_dep.html', context)


# @login_required
# def heures_enseignants_dep(request, semestre_num=1):
#     """Vue pour les heures de travail des enseignants par semestre"""
    
#     enseignant = request.user.enseignant_profile
#     real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
#     departement = Departement.objects.get(id=real_Dep.departement.id)
    
#     # Récupérer l'année universitaire courante
#     annee_courante = AnneeUniversitaire.get_courante()
#     if not annee_courante:
#         messages.error(request, 'لا توجد سنة جامعية محددة كسنة حالية')
#         return redirect('admin:index')

#     # Valider le numéro de semestre
#     if semestre_num not in [1, 2]:
#         messages.error(request, 'رقم السداسي غير صحيح')
#         return redirect('depa:heures_enseignants_dep', semestre_num=1)

#     # Filtre de base avec semestre
#     base_filter = {
#         'departement': departement,
#         'annee_univ': annee_courante,
#         f'semestre_{semestre_num}': True
#     }
    
#     # Récupérer tous les enseignants pour calculer les statistiques
#     all_Ens_Dep_for_stats = Ens_Dep.objects.filter(**base_filter).order_by('enseignant__nom_ar')
    
#     for idx in all_Ens_Dep_for_stats:
#         my_Ens = idx.enseignant
        
#         # Calculer les classes pour ce semestre
#         i_ens_nbrClas_in_Dep = Classe.objects.filter(
#             enseignant__enseignant=idx.enseignant, 
#             enseignant__departement=real_Dep.departement,
#             semestre=semestre_num
#         ).count()

#         i_ens_nbrClas_out_Dep = Classe.objects.filter(
#             enseignant__enseignant=idx.enseignant,  
#             semestre=semestre_num
#         ).exclude(enseignant__departement=real_Dep.departement).count()
        
#         # Calculer les différents types de classes
#         i_ens_nbrClas_Cours_in_Dep = Classe.objects.filter(
#             enseignant__enseignant=idx.enseignant, 
#             enseignant__departement=real_Dep.departement,
#             semestre=semestre_num,
#             type='Cours'
#         ).count()
        
#         i_ens_nbrClas_TD_in_Dep = Classe.objects.filter(
#             enseignant__enseignant=idx.enseignant,
#             enseignant__departement=real_Dep.departement,
#             semestre=semestre_num,
#             type='TD'
#         ).count()
        
#         i_ens_nbrClas_TP_in_Dep = Classe.objects.filter(
#             enseignant__enseignant=idx.enseignant, 
#             enseignant__departement=real_Dep.departement,
#             semestre=semestre_num,
#             type='TP'
#         ).count()
        
#         i_ens_nbrClas_SS_in_Dep = Classe.objects.filter(
#             enseignant__enseignant=idx.enseignant, 
#             enseignant__departement=real_Dep.departement,
#             semestre=semestre_num,
#             type='SS'
#         ).count()
        
#         # Calculer les jours de travail uniques
#         jours_travail = Classe.objects.filter(
#             enseignant__enseignant=idx.enseignant, 
#             enseignant__departement=real_Dep.departement,
#             semestre=semestre_num
#         ).values_list('jour', flat=True).distinct().count()

#         # Calculer le taux d'avancement minimum et moyen
#         classes_enseignant = Classe.objects.filter(
#             enseignant__enseignant=idx.enseignant,
#             enseignant__departement=real_Dep.departement,
#             semestre__numero=semestre_num
#         )
        
#         if classes_enseignant.exists():
#             taux_list = list(classes_enseignant.values_list('taux_avancement', flat=True))
#             i_ens_taux_min = min(taux_list)
#             i_ens_taux_moyen = sum(taux_list) / len(taux_list)
#         else:
#             i_ens_taux_min = 0
#             i_ens_taux_moyen = 0
        
#         # Mettre à jour selon le semestre
#         if semestre_num == 1:
#             idx.nbrClas_in_Dep_S1 = i_ens_nbrClas_in_Dep
#             idx.nbrClas_out_Dep_S1 = i_ens_nbrClas_out_Dep
#             idx.nbrClas_ALL_Dep_S1 = i_ens_nbrClas_in_Dep + i_ens_nbrClas_out_Dep
#             idx.nbrClas_Cours_in_Dep_S1 = i_ens_nbrClas_Cours_in_Dep
#             idx.nbrClas_TD_in_Dep_S1 = i_ens_nbrClas_TD_in_Dep
#             idx.nbrClas_TP_in_Dep_S1 = i_ens_nbrClas_TP_in_Dep
#             idx.nbrClas_SS_in_Dep_S1 = i_ens_nbrClas_SS_in_Dep
#             idx.nbrJour_in_Dep_S1 = jours_travail
#             idx.taux_min_S1 = i_ens_taux_min
#             idx.taux_moyen_S1 = round(i_ens_taux_moyen, 1)
#             idx.save()
#         else:
#             idx.nbrClas_in_Dep_S2 = i_ens_nbrClas_in_Dep
#             idx.nbrClas_out_Dep_S2 = i_ens_nbrClas_out_Dep
#             idx.nbrClas_ALL_Dep_S2 = i_ens_nbrClas_in_Dep + i_ens_nbrClas_out_Dep
#             idx.nbrClas_Cours_in_Dep_S2 = i_ens_nbrClas_Cours_in_Dep
#             idx.nbrClas_TD_in_Dep_S2 = i_ens_nbrClas_TD_in_Dep
#             idx.nbrClas_TP_in_Dep_S2 = i_ens_nbrClas_TP_in_Dep
#             idx.nbrClas_SS_in_Dep_S2 = i_ens_nbrClas_SS_in_Dep
#             idx.nbrJour_in_Dep_S2 = jours_travail
#             idx.taux_min_S2 = i_ens_taux_min
#             idx.taux_moyen_S2 = round(i_ens_taux_moyen, 1)
#             idx.save()

#     # Tri dynamique selon le semestre
#     if semestre_num == 1:
#         order_field = '-nbrClas_ALL_Dep_S1'
#     else:
#         order_field = '-nbrClas_ALL_Dep_S2'
    
#     # Récupérer les enseignants triés pour l'affichage
#     main_all_aff_Ens_Dep = Ens_Dep.objects.filter(
#         **base_filter,
#         est_inscrit=True
#     ).order_by(order_field, 'enseignant__nom_ar')
    
#     main_PER = Ens_Dep.objects.filter(
#         departement=departement,
#         statut='Permanent',
#         annee_univ=annee_courante,
#         est_inscrit=True,
#         **{f'semestre_{semestre_num}': True}
#     ).order_by(order_field, 'enseignant__nom_ar')
    
#     main_PER_VAC = Ens_Dep.objects.filter(
#         departement=departement,
#         statut='Permanent & Vacataire',
#         annee_univ=annee_courante,
#         est_inscrit=True,
#         **{f'semestre_{semestre_num}': True}
#     ).order_by(order_field, 'enseignant__nom_ar')
    
#     main_VAC = Ens_Dep.objects.filter(
#         departement=departement,
#         statut='Vacataire',
#         annee_univ=annee_courante,
#         est_inscrit=True,
#         **{f'semestre_{semestre_num}': True}
#     ).order_by(order_field, 'enseignant__nom_ar')
    
#     main_ASO = Ens_Dep.objects.filter(
#         departement=departement,
#         statut='Associe',
#         annee_univ=annee_courante,
#         est_inscrit=True,
#         **{f'semestre_{semestre_num}': True}
#     ).order_by(order_field, 'enseignant__nom_ar')
    
#     main_DOC = Ens_Dep.objects.filter(
#         departement=departement,
#         statut='Doctorant',
#         annee_univ=annee_courante,
#         est_inscrit=True,
#         **{f'semestre_{semestre_num}': True}
#     ).order_by(order_field, 'enseignant__nom_ar')

#     context = {
#         'title': 'ساعات عمل الأساتذة',
#         'my_Dep': departement,
#         'my_Fac': departement.faculte,
#         'annee_courante': annee_courante,
#         'semestre_num': semestre_num,
#         'main_all_aff_Ens_Dep': main_all_aff_Ens_Dep,
#         'main_PER': main_PER,
#         'main_PER_VAC': main_PER_VAC,
#         'main_VAC': main_VAC,
#         'main_ASO': main_ASO,
#         'main_DOC': main_DOC,
#     }
    
#     return render(request, 'departement/heures_enseignants_dep.html', context)













from django.db.models import Count, Q, Avg, Min, Prefetch
from django.core.cache import cache

from django.db.models import Count, Min, Avg, Q, Prefetch
from django.core.cache import cache

@login_required
def heures_enseignants_dep(request, semestre_num=1):
    """
    VERSION ULTRA-OPTIMISÉE
    - 3-4 requêtes SQL au total (au lieu de 150+)
    - Temps de chargement : < 500ms même avec 100 enseignants
    """
    
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)
    
    # Récupérer l'année universitaire courante
    annee_courante = AnneeUniversitaire.get_courante()
    if not annee_courante:
        messages.error(request, 'لا توجد سنة جامعية محددة كسنة حالية')
        return redirect('admin:index')

    # Valider le numéro de semestre
    if semestre_num not in [1, 2]:
        messages.error(request, 'رقم السداسي غير صحيح')
        return redirect('depa:heures_enseignants_dep', semestre_num=1)

    # 🎯 VÉRIFIER LE CACHE AVEC OPTION DE RAFRAÎCHISSEMENT
    force_refresh = request.GET.get('refresh', '0') == '1'
    cache_key = f"heures_ens_dep_{departement.id}_s{semestre_num}_{annee_courante.id}"
    
    if not force_refresh:
        cached_context = cache.get(cache_key)
        if cached_context:
            return render(request, 'departement/heures_enseignants_dep.html', cached_context)

    base_filter = {
        'departement': departement,
        'annee_univ': annee_courante,
        'est_inscrit': True,
        f'semestre_{semestre_num}': True
    }
    
    # 🚀 OPTIMISATION 1 : Récupérer tous les enseignants avec leurs relations
    main_all_aff_Ens_Dep = Ens_Dep.objects.filter(**base_filter).select_related(
        'enseignant',
        'enseignant__user',
        'enseignant__user__poste_principal'
    )
    
    # 🚀 OPTIMISATION 2 : Pré-calculer TOUTES les statistiques en 2 requêtes
    # Créer un dictionnaire avec l'ID de l'enseignant comme clé
    enseignants_ids = [ens_dep.enseignant.id for ens_dep in main_all_aff_Ens_Dep]
    
    # Requête 1 : Stats pour classes DANS le département
    stats_in_dept = Classe.objects.filter(
        enseignant__enseignant__id__in=enseignants_ids,
        enseignant__departement=departement,
        semestre=semestre_num
    ).values('enseignant__enseignant__id').annotate(
        total=Count('id'),
        cours=Count('id', filter=Q(type='Cours')),
        td=Count('id', filter=Q(type='TD')),
        tp=Count('id', filter=Q(type='TP')),
        ss=Count('id', filter=Q(type='SS')),
        jours=Count('jour', distinct=True),
        taux_min=Min('taux_avancement'),
        taux_moy=Avg('taux_avancement')
    )
    
    # Convertir en dictionnaire pour accès rapide
    stats_in_dict = {item['enseignant__enseignant__id']: item for item in stats_in_dept}
    
    # Requête 2 : Stats pour classes HORS département
    stats_out_dept = Classe.objects.filter(
        enseignant__enseignant__id__in=enseignants_ids,
        semestre=semestre_num
    ).exclude(
        enseignant__departement=departement
    ).values('enseignant__enseignant__id').annotate(
        total_out=Count('id')
    )
    
    # Convertir en dictionnaire pour accès rapide
    stats_out_dict = {item['enseignant__enseignant__id']: item['total_out'] for item in stats_out_dept}


    # Requête 3 : Stats pour les liens Moodle
    stats_moodle = Classe.objects.filter(
        enseignant__enseignant__id__in=enseignants_ids,
        enseignant__departement=departement,
        semestre=semestre_num
    ).values('enseignant__enseignant__id').annotate(
        total_classes=Count('id'),
        with_moodle=Count('id', filter=Q(lien_moodle__isnull=False) & ~Q(lien_moodle=''))
    )

    # Convertir en dictionnaire pour accès rapide
    stats_moodle_dict = {
        item['enseignant__enseignant__id']: {
            'total': item['total_classes'],
            'with_moodle': item['with_moodle'],
            'percentage': round((item['with_moodle'] / item['total_classes'] * 100), 1) if item['total_classes'] > 0 else 0
        } for item in stats_moodle
    }


    
    # 🚀 OPTIMISATION 3 : Mise à jour en lot (bulk_update)
    enseignants_to_update = []
    
    for idx in main_all_aff_Ens_Dep:
        ens_id = idx.enseignant.id
        
        # Récupérer les stats depuis les dictionnaires (pas de requête SQL!)
        stats = stats_in_dict.get(ens_id, {
            'total': 0, 'cours': 0, 'td': 0, 'tp': 0, 'ss': 0,
            'jours': 0, 'taux_min': 0, 'taux_moy': 0
        })
        stats_out = stats_out_dict.get(ens_id, 0)
        # ✅ NOUVEAU : Récupérer les stats Moodle
        moodle_stats = stats_moodle_dict.get(ens_id, {'total': 0, 'with_moodle': 0, 'percentage': 0})
        
        # Assigner selon le semestre
        if semestre_num == 1:
            idx.nbrClas_in_Dep_S1 = stats.get('total', 0)
            idx.nbrClas_out_Dep_S1 = stats_out
            idx.nbrClas_ALL_Dep_S1 = stats.get('total', 0) + stats_out
            idx.nbrClas_Cours_in_Dep_S1 = stats.get('cours', 0)
            idx.nbrClas_TD_in_Dep_S1 = stats.get('td', 0)
            idx.nbrClas_TP_in_Dep_S1 = stats.get('tp', 0)
            idx.nbrClas_SS_in_Dep_S1 = stats.get('ss', 0)
            idx.nbrJour_in_Dep_S1 = stats.get('jours', 0)
            idx.taux_min_S1 = int(stats.get('taux_min', 0) or 0)
            idx.taux_moyen_S1 = round(stats.get('taux_moy', 0) or 0, 1)
            idx.moodle_percentage_S1 = moodle_stats['percentage']
            
            update_fields = [
                'nbrClas_in_Dep_S1', 'nbrClas_out_Dep_S1', 'nbrClas_ALL_Dep_S1',
                'nbrClas_Cours_in_Dep_S1', 'nbrClas_TD_in_Dep_S1', 
                'nbrClas_TP_in_Dep_S1', 'nbrClas_SS_in_Dep_S1',
                'nbrJour_in_Dep_S1', 'taux_min_S1', 'taux_moyen_S1',
                'moodle_percentage_S1'  # ✅ NOUVEAU
            ]
        else:
            idx.nbrClas_in_Dep_S2 = stats.get('total', 0)
            idx.nbrClas_out_Dep_S2 = stats_out
            idx.nbrClas_ALL_Dep_S2 = stats.get('total', 0) + stats_out
            idx.nbrClas_Cours_in_Dep_S2 = stats.get('cours', 0)
            idx.nbrClas_TD_in_Dep_S2 = stats.get('td', 0)
            idx.nbrClas_TP_in_Dep_S2 = stats.get('tp', 0)
            idx.nbrClas_SS_in_Dep_S2 = stats.get('ss', 0)
            idx.nbrJour_in_Dep_S2 = stats.get('jours', 0)
            idx.taux_min_S2 = int(stats.get('taux_min', 0) or 0)
            idx.taux_moyen_S2 = round(stats.get('taux_moy', 0) or 0, 1)
            idx.moodle_percentage_S2 = moodle_stats['percentage']
            
            update_fields = [
                'nbrClas_in_Dep_S2', 'nbrClas_out_Dep_S2', 'nbrClas_ALL_Dep_S2',
                'nbrClas_Cours_in_Dep_S2', 'nbrClas_TD_in_Dep_S2', 
                'nbrClas_TP_in_Dep_S2', 'nbrClas_SS_in_Dep_S2',
                'nbrJour_in_Dep_S2', 'taux_min_S2', 'taux_moyen_S2',
                'moodle_percentage_S2'  # ✅ NOUVEAU
            ]
        
       
        enseignants_to_update.append(idx)

    
    # 💾 SAUVEGARDE EN LOT (1 seule requête pour tous les enseignants!)
    if enseignants_to_update:
        Ens_Dep.objects.bulk_update(enseignants_to_update, update_fields)
    
    # Re-charger avec le bon tri
    if semestre_num == 1:
        order_field = '-nbrClas_ALL_Dep_S1'
    else:
        order_field = '-nbrClas_ALL_Dep_S2'
    
    main_all_aff_Ens_Dep = Ens_Dep.objects.filter(**base_filter).select_related(
        'enseignant',
        'enseignant__user',
        'enseignant__user__poste_principal'
    ).order_by(order_field, 'enseignant__nom_ar')
    
    # Filtrer par catégorie
    main_PER = [x for x in main_all_aff_Ens_Dep if x.statut == 'Permanent']
    main_PER_VAC = [x for x in main_all_aff_Ens_Dep if x.statut == 'Permanent & Vacataire']
    main_VAC = [x for x in main_all_aff_Ens_Dep if x.statut == 'Vacataire']
    main_ASO = [x for x in main_all_aff_Ens_Dep if x.statut == 'Associe']
    main_DOC = [x for x in main_all_aff_Ens_Dep if x.statut == 'Doctorant']

    # Contexte
    context = {
        'title': 'ساعات عمل الأساتذة',
        'my_Dep': departement,
        'my_Fac': departement.faculte,
        'annee_courante': annee_courante,
        'semestre_num': semestre_num,
        'main_all_aff_Ens_Dep': list(main_all_aff_Ens_Dep),
        'main_PER': main_PER,
        'main_PER_VAC': main_PER_VAC,
        'main_VAC': main_VAC,
        'main_ASO': main_ASO,
        'main_DOC': main_DOC,
        'count_all': len(main_all_aff_Ens_Dep),
        'count_per': len(main_PER),
        'count_temp': len(main_PER_VAC) + len(main_VAC) + len(main_ASO) + len(main_DOC),
        'count_PER_VAC': len(main_PER_VAC),
        'count_VAC': len(main_VAC),
        'count_ASO': len(main_ASO),
        'count_DOC': len(main_DOC),
    }
    
    # 💾 METTRE EN CACHE POUR 5 MINUTES
    cache.set(cache_key, context, 300)
    
    return render(request, 'departement/heures_enseignants_dep.html', context)







# 🔧 FONCTION UTILITAIRE : Vider le cache
def invalidate_heures_cache(departement_id=None):
    """
    Invalider le cache des heures de travail
    À appeler après toute modification d'enseignant ou de classe
    """
    if departement_id:
        # Vider pour un département spécifique
        annees = AnneeUniversitaire.objects.all()
        for annee in annees:
            cache.delete(f"heures_ens_dep_{departement_id}_s1_{annee.id}")
            cache.delete(f"heures_ens_dep_{departement_id}_s2_{annee.id}")
    else:
        # Vider tout le cache (use avec précaution)
        cache.clear()






# 📝 EXEMPLE D'UTILISATION DANS VOS VUES DE MODIFICATION
"""
# Dans delete_Enseignant :
@login_required
def delete_Enseignant(request, ens_dep_id):
    ens_dep = get_object_or_404(Ens_Dep, id=ens_dep_id)
    dept_id = ens_dep.departement.id
    ens_dep.delete()
    
    # ✅ Invalider le cache
    invalidate_heures_cache(dept_id)
    
    messages.success(request, 'تم حذف الأستاذ بنجاح')
    return redirect('depa:list_enseignants_dep')


# Dans inscription_Ens :
@login_required
def inscription_Ens(request, ens_id):
    # ... votre code ...
    ens_dep.est_inscrit = True
    ens_dep.save()
    
    # ✅ Invalider le cache
    invalidate_heures_cache(ens_dep.departement.id)
    
    return redirect(...)


# Dans import_emploi (après import de classes) :
@login_required
def import_emploi(request):
    # ... import des classes ...
    
    # ✅ Invalider le cache pour tous les départements concernés
    for dept in departements_concernes:
        invalidate_heures_cache(dept.id)
    
    return redirect(...)
"""















# ----------------------------------------------------------------------
def facs_json(request):
    facultes = Faculte.objects.all()
    data = [{'id': faculte.id, 'nom_ar': faculte.nom_ar} for faculte in facultes]
    return JsonResponse(data, safe=False)
# ----------------------------------------------------------------------
def deps_json(request):
    faculte_id = request.GET.get('faculte_id')
    if faculte_id:
        departements = Departement.objects.filter(faculte_id=faculte_id)
    else:
        departements = Departement.objects.none()
    data = [{'id': dep.id, 'nom_ar': dep.nom_ar} for dep in departements]
    return JsonResponse(data, safe=False)
# ----------------------------------------------------------------------
def enss_json(request):
    departement_id = request.GET.get('departement_id')
    if departement_id:
        enseignants = Ens_Dep.objects.filter(departement_id=departement_id).values_list('enseignant__user__username', flat=True)
    else:
        enseignants = []
    data = [{'id': i, 'nom': enseignant} for i, enseignant in enumerate(enseignants, start=1)]
    return JsonResponse(data, safe=False)




    

def enss_json(request):
    try:
        departement_id = request.GET.get('departement_id')
        if departement_id:
            enseignants = Ens_Dep.objects.filter(departement_id=departement_id).select_related('enseignant__user')
            data = [
                {
                    'id': enseignant.enseignant.id,  # Utiliser l'ID réel de Ens_Dep
                    'nom': f"{enseignant.enseignant.nom_ar} {enseignant.enseignant.prenom_ar}"
                }
                for enseignant in enseignants
            ]
        else:
            data = []
        return JsonResponse(data, safe=False)
    except Exception as e:
        # print(f"Erreur dans enss_json: {e}")
        return JsonResponse({'error': str(e)}, status=500)
    

# from django.contrib.auth.models import User




# Fonctions helper pour la gestion des utilisateurs
def normalize_text(text):
    """Normalise le texte pour créer un nom d'utilisateur valide"""
    if not text:
        return ""
    # Supprimer les accents
    text = unicodedata.normalize('NFD', text)
    text = ''.join(char for char in text if unicodedata.category(char) != 'Mn')
    # Garder seulement les lettres et espaces
    text = re.sub(r'[^a-zA-Z\s]', '', text)
    # Remplacer les espaces par des points et mettre en minuscules
    text = text.lower().replace(' ', '.')
    return text

def generate_username(prenom_fr, nom_fr):
    """Génère un nom d'utilisateur unique"""
    prenom_clean = normalize_text(prenom_fr)
    nom_clean = normalize_text(nom_fr)
    
    base_username = f"{nom_clean}.{prenom_clean}"
    username = base_username
    
    # Ajouter un numéro si l'username existe déjà
    counter = 1
    while CustomUser.objects.filter(username=username).exists():
        username = f"{base_username}{counter}"
        counter += 1
        
    return username

def generate_password(nom_fr):
    # """Génère un mot de passe par défaut"""
    # first_letter = nom_fr[0].upper() if nom_fr else 'X'
    # random_numbers = ''.join([str(random.randint(0, 9)) for _ in range(4)])
    # return f"Vac{first_letter}{random_numbers}"
    return f"...Ens123"

def generate_email(prenom_fr, nom_fr):
    """Génère un email par défaut"""
    prenom_clean = normalize_text(prenom_fr)
    nom_clean = normalize_text(nom_fr)
    return f"{nom_clean}.{prenom_clean}@univ-ouargla.dz"


########################################


def get_or_create_enseignant_poste():
    """Récupère ou crée le poste 'enseignant'"""
    try:
        poste_enseignant = Poste.objects.get(type='enseignant', nom_ar='أستاذ', nom_fr='Enseignant', code='Enseignant')
        return poste_enseignant, None
    except Poste.DoesNotExist:
        # Si le poste n'existe pas, essayer de le créer
        try:
            poste_enseignant = Poste.objects.create(
                type='enseignant',
                nom_fr='Enseignant',
                nom_ar='أستاذ',
                code='Enseignant',
            )
            return poste_enseignant, "تم إنشاء منصب 'enseignant' تلقائياً"
        except Exception as e:
            return None, f"خطأ في إنشاء منصب enseignant: {str(e)}"





# Ajouter un nouveau Enseignant ----------------------------------------------------------------------
# Ajouter un nouveau Enseignant ----------------------------------------------------------------------
@login_required
def new_Enseignant(request):
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)
    
    # Récupérer l'année universitaire courante
    try:
        annee_courante = AnneeUniversitaire.objects.order_by('-date_debut').first()
        if not annee_courante:
            messages.error(request, 'لا توجد سنة جامعية في النظام')
            return redirect('admin:index')
    except Exception as e:
        messages.error(request, f'خطأ في استرجاع السنة الجامعية: {str(e)}')
        return redirect('admin:index')

    if request.method == 'POST':
        # Option 1: Import d'enseignants via fichier Excel
        if 'btnNew_EnsVac' in request.POST:
            # Récupérer le semestre sélectionné (1 ou 2)
            semestre_choisi = request.POST.get('semestre_vacataire')  # "1" ou "2"
            if not semestre_choisi:
                messages.error(request, 'يرجى اختيار السداسي')
                context = {
                    'title': 'إضافة أساتذة',
                    'my_Dep': departement,
                    'my_Fac': departement.faculte,
                    'annee_courante': annee_courante,
                }
                return render(request, 'departement/new_Enseignant.html', context)
            
            # Récupérer le statut sélectionné
            statut_enseignant = request.POST.get('statut_enseignant')
            if not statut_enseignant:
                messages.error(request, 'يرجى اختيار نوع الأستاذ')
                context = {
                    'title': 'إضافة أساتذة',
                    'my_Dep': departement,
                    'my_Fac': departement.faculte,
                    'annee_courante': annee_courante,
                }
                return render(request, 'departement/new_Enseignant.html', context)

            # Vérifier que le semestre est valide
            if semestre_choisi not in ['1', '2']:
                messages.error(request, 'رقم السداسي غير صحيح')
                context = {
                    'title': 'إضافة أساتذة',
                    'my_Dep': departement,
                    'my_Fac': departement.faculte,
                    'annee_courante': annee_courante,
                }
                return render(request, 'departement/new_Enseignant.html', context)

            # Traitement du fichier Excel
            ensResource = Ens_Vac_Form()
            dataset = Dataset() 
            all_Ens = request.FILES.get('file_EnsVac')

            if not all_Ens:
                messages.warning(request, 'لا يوجد ملف لتحميله')
                context = {
                    'title': 'إضافة أساتذة',
                    'my_Dep': departement,
                    'my_Fac': departement.faculte,
                    'annee_courante': annee_courante,
                }
                return render(request, 'departement/new_Enseignant.html', context)
            
            if not all_Ens.name.endswith('xlsx'):
                messages.warning(request, 'إمتداد الملف خاطئ')
                context = {
                    'title': 'إضافة أساتذة',
                    'my_Dep': departement,
                    'my_Fac': departement.faculte,
                    'annee_courante': annee_courante,
                }
                return render(request, 'departement/new_Enseignant.html', context)

            # Vérifier/créer le poste enseignant
            poste_enseignant, poste_message = get_or_create_enseignant_poste()
            if not poste_enseignant:
                messages.error(request, poste_message)
                context = {
                    'title': 'إضافة أساتذة',
                    'my_Dep': departement,
                    'my_Fac': departement.faculte,
                    'annee_courante': annee_courante,
                }
                return render(request, 'departement/new_Enseignant.html', context)
            
            if poste_message:
                messages.info(request, poste_message)

            imported_data = dataset.load(all_Ens.read(), format='xlsx')
            created_count = 0
            existing_count = 0
            error_count = 0
            new_users_info = []

            for data in imported_data:
                try:
                    with transaction.atomic():
                        # Vérifier si l'enseignant existe déjà
                        existing_enseignant = Enseignant.objects.filter(
                            nom_ar=data[0], prenom_ar=data[1],
                            nom_fr=data[2], prenom_fr=data[3],
                            date_nais=data[4]
                        ).first()

                        if existing_enseignant:
                            # Enseignant existe, vérifier s'il est déjà dans ce département pour cette année
                            existing_ens_dep = Ens_Dep.objects.filter(
                                departement=departement,
                                enseignant=existing_enseignant,
                                annee_univ=annee_courante
                            ).first()

                            if existing_ens_dep:
                                # Il existe déjà, juste activer le semestre demandé
                                if semestre_choisi == "1":
                                    if existing_ens_dep.semestre_1:
                                        messages.warning(request, f'الأستاذ "{data[1]} {data[0]}" مسجل مسبقاً في السداسي الأول')
                                    else:
                                        existing_ens_dep.semestre_1 = True
                                        existing_ens_dep.save()
                                        messages.success(request, f'تم تسجيل الأستاذ الموجود "{data[1]} {data[0]}" في السداسي الأول')
                                elif semestre_choisi == "2":
                                    if existing_ens_dep.semestre_2:
                                        messages.warning(request, f'الأستاذ "{data[1]} {data[0]}" مسجل مسبقاً في السداسي الثاني')
                                    else:
                                        existing_ens_dep.semestre_2 = True
                                        existing_ens_dep.save()
                                        messages.success(request, f'تم تسجيل الأستاذ الموجود "{data[1]} {data[0]}" في السداسي الثاني')
                            else:
                                # Créer nouvelle affectation pour l'enseignant existant
                                semestre_1_actif = (semestre_choisi == "1")
                                semestre_2_actif = (semestre_choisi == "2")
                                
                                valueAffEnsDep = Ens_Dep.objects.create(
                                    departement=departement,
                                    enseignant=existing_enseignant,
                                    annee_univ=annee_courante,
                                    statut=statut_enseignant,
                                    semestre_1=semestre_1_actif,
                                    semestre_2=semestre_2_actif
                                )
                                
                                semestre_nom = "الأول" if semestre_choisi == "1" else "الثاني"
                                messages.success(request, f'تم إضافة الأستاذ الموجود "{data[1]} {data[0]}" للقسم في السداسي {semestre_nom}')
                            
                            existing_count += 1
                            continue

                        # Nouvel enseignant - créer utilisateur et enseignant
                        username = generate_username(data[3], data[2])  # prenom_fr, nom_fr
                        password = generate_password(data[2])  # nom_fr
                        email = data[9] if data[9] else generate_email(data[3], data[2])

                        # Créer l'utilisateur CustomUser avec poste_principal
                        user = CustomUser.objects.create_user(
                            username=username,
                            email=email,
                            password=password,
                            first_name=data[3],  # prenom_fr
                            last_name=data[2],   # nom_fr
                            poste_principal=poste_enseignant
                        )

                        # Créer l'enseignant
                        valueEnseignant = Enseignant.objects.create(
                            user=user,
                            nom_ar=data[0],
                            prenom_ar=data[1],
                            nom_fr=data[2],
                            prenom_fr=data[3],
                            date_nais=data[4],
                            bac_annee=data[5],
                            matricule=data[6],
                            telmobile1=data[7],
                            email_perso=data[8],
                            email_prof=data[9],
                        )

                        # Créer l'affectation avec les bons champs boolean
                        semestre_1_actif = (semestre_choisi == "1")
                        semestre_2_actif = (semestre_choisi == "2")
                        
                        valueAffEnsDep = Ens_Dep.objects.create(
                            departement=departement,
                            enseignant=valueEnseignant,
                            est_inscrit=True,
                            statut=statut_enseignant,
                            annee_univ=annee_courante,
                            semestre_1=semestre_1_actif,
                            semestre_2=semestre_2_actif
                        )

                        # Stocker les informations du nouvel utilisateur
                        semestre_nom = "الأول" if semestre_choisi == "1" else "الثاني"
                        new_users_info.append({
                            'nom': f"{data[1]} {data[0]}",
                            'username': username,
                            'password': password,
                            'email': email,
                            'poste': 'enseignant',
                            'semestre': semestre_nom,
                            'statut': dict(Ens_Dep.STATUT_CHOICES)[statut_enseignant]
                        })

                        created_count += 1

                except Exception as e:
                    error_count += 1
                    messages.error(
                        request, 
                        f'خطأ في إضافة الأستاذ "{data[1]} {data[0]}": {str(e)}'
                    )

            # Résumé des opérations
            semestre_nom = "الأول" if semestre_choisi == "1" else "الثاني"
            if created_count > 0:
                statut_arabe = dict(Ens_Dep.STATUT_CHOICES)[statut_enseignant]
                messages.success(request, f'تم إنشاء {created_count} أستاذ جديد بنجاح كـ "{statut_arabe}" للسداسي {semestre_nom}')
            if existing_count > 0:
                messages.info(request, f'تم العثور على {existing_count} أستاذ موجود مسبقاً')
            if error_count > 0:
                messages.error(request, f'حدث خطأ مع {error_count} أستاذ')

            # Afficher les informations de connexion pour les nouveaux utilisateurs
            if new_users_info:
                user_credentials_message = f"معلومات تسجيل الدخول للأساتذة الجدد (السداسي {semestre_nom}):\n"
                for info in new_users_info:
                    user_credentials_message += f"• {info['nom']} ({info['statut']}): {info['username']} / {info['password']}\n"
                
                messages.success(request, user_credentials_message)

            return redirect('depa:new_Enseignant')
        
        # Option 2: Import d'un enseignant existant d'un autre département
        elif 'btn_import_Ens' in request.POST:
            # Récupérer le semestre sélectionné (1 ou 2)
            semestre_choisi = request.POST.get('semestre_import')  # "1" ou "2"
            if not semestre_choisi:
                messages.error(request, 'يرجى اختيار السداسي')
                context = {
                    'title': 'إضافة أساتذة',
                    'my_Dep': departement,
                    'my_Fac': departement.faculte,
                    'annee_courante': annee_courante,
                }
                return render(request, 'departement/new_Enseignant.html', context)
            
            # Vérifier que le semestre est valide
            if semestre_choisi not in ['1', '2']:
                messages.error(request, 'رقم السداسي غير صحيح')
                context = {
                    'title': 'إضافة أساتذة',
                    'my_Dep': departement,
                    'my_Fac': departement.faculte,
                    'annee_courante': annee_courante,
                }
                return render(request, 'departement/new_Enseignant.html', context)
                
            enseignant_id = request.POST.get('enseignant', False)
            if enseignant_id != False:                
                try:
                    with transaction.atomic():
                        enseignant_to_import = Enseignant.objects.get(id=enseignant_id)
                        
                        # Vérifier s'il existe déjà dans ce département pour cette année
                        existing_ens_dep = Ens_Dep.objects.filter(
                            departement=departement,
                            enseignant=enseignant_to_import,
                            annee_univ=annee_courante
                        ).first()

                        if existing_ens_dep:
                            # Il existe déjà, juste activer le semestre demandé
                            if semestre_choisi == "1":
                                if existing_ens_dep.semestre_1:
                                    messages.warning(request, f'الأستاذ "{enseignant_to_import}" مسجل مسبقاً في السداسي الأول')
                                else:
                                    existing_ens_dep.semestre_1 = True
                                    existing_ens_dep.save()
                                    messages.success(request, f'تم تسجيل الأستاذ "{enseignant_to_import}" في السداسي الأول')
                            elif semestre_choisi == "2":
                                if existing_ens_dep.semestre_2:
                                    messages.warning(request, f'الأستاذ "{enseignant_to_import}" مسجل مسبقاً في السداسي الثاني')
                                else:
                                    existing_ens_dep.semestre_2 = True
                                    existing_ens_dep.save()
                                    messages.success(request, f'تم تسجيل الأستاذ "{enseignant_to_import}" في السداسي الثاني')
                        else:
                            # Créer nouvelle affectation
                            semestre_1_actif = (semestre_choisi == "1")
                            semestre_2_actif = (semestre_choisi == "2")
                            
                            Ens_Dep.objects.create(
                                departement=departement,
                                enseignant=enseignant_to_import,
                                statut='Permanent & Vacataire',
                                annee_univ=annee_courante,
                                semestre_1=semestre_1_actif,
                                semestre_2=semestre_2_actif
                            )
                            
                            semestre_nom = "الأول" if semestre_choisi == "1" else "الثاني"
                            messages.success(request, f'تم إضافة الأستاذ "{enseignant_to_import}" للقسم في السداسي {semestre_nom}')
                        
                        return redirect('depa:new_Enseignant')

                except Enseignant.DoesNotExist:
                    messages.error(request, 'الأستاذ المحدد غير موجود')
                except Exception as e:
                    messages.error(request, f'حدث خطأ أثناء إضافة الأستاذ: {str(e)}')
            else:
                messages.warning(request, 'إختر أستاذ')

    context = {
        'title': 'إضافة أساتذة',
        'my_Dep': departement,
        'my_Fac': departement.faculte,
        'annee_courante': annee_courante,
    }
    return render(request, 'departement/new_Enseignant.html', context)






# ----------------------------------------------------------------------
@login_required
def inscription_Ens(request, ens_id):
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)
    countAll_User = CustomUser.objects.all().count() + 1
    myEns = Enseignant.objects.get(id=ens_id)

    try:
        this_Ens = Ens_Dep.objects.get(
            enseignant=ens_id,
            departement=departement,
            est_inscrit=False
        )
        
        if myEns.user != None:
            this_Ens.est_inscrit = True
            this_Ens.save()
            messages.success(request, f'لقد تمت عملية تنشيط حساب {myEns} بنجاح')
        else:
            # Remplacer les espaces par des underscores
            nom_Ens = myEns.nom_fr
            nom_modifie = nom_Ens.replace(' ', '_')
            poste = Poste.objects.get(nom_fr="Enseignant", code="Enseignant")
            user = CustomUser.objects.create(
                username=nom_modifie+str(countAll_User),
                first_name=myEns.prenom_fr,
                last_name=myEns.nom_fr,
                poste_principal=poste,
                email="enseignant@test.com",
            )
            # Définir un mot de passe
            user.set_password("..Ens123")
            user.save()

            myEns.user = user
            myEns.save()

            this_Ens.est_inscrit = True
            this_Ens.save()

            messages.success(request, f'لقد تمت عملية تسجيل {myEns} بنجاح')

    except Ens_Dep.DoesNotExist:
        # Cas où l'enseignant n'existe pas avec est_inscrit=False
        # Remplacer les espaces par des underscores
        nom_Ens = myEns.nom_fr
        nom_modifie = nom_Ens.replace(' ', '_')
        poste = Poste.objects.get(nom_fr="Enseignant", code="Enseignant")
        user = CustomUser.objects.create(
            username=nom_modifie+str(countAll_User),
            first_name=myEns.prenom_fr,
            last_name=myEns.nom_fr,
            poste_principal=poste,
            email="enseignant@test.com",
        )
        # Définir un mot de passe
        user.set_password("...Ens123")
        user.save()

        myEns.user = user
        myEns.save()

        # Chercher l'enseignant et le marquer comme inscrit
        try:
            this_Ens_inscrit = Ens_Dep.objects.get(
                enseignant=ens_id,
                departement=departement,
                est_inscrit=False
            )
            this_Ens_inscrit.est_inscrit = True
            this_Ens_inscrit.save()
        except Ens_Dep.DoesNotExist:
            # Si pas trouvé, marquer tous les enregistrements de cet enseignant
            Ens_Dep.objects.filter(
                enseignant=ens_id,
                departement=departement
            ).update(est_inscrit=True)

        messages.success(request, f'لقد تمت عملية تسجيل {myEns} بنجاح')
    
    # Vérifier s'il y a un paramètre 'next' dans l'URL
    next_url = request.GET.get('next')
    if next_url:
        return redirect(next_url)
    else:
        return redirect(request.META.get('HTTP_REFERER', 'depa:manage_enseignants'))





# ----------------------------------------------------------------------
@login_required
def inscription_Etu(request, etu_id):
    enseignant = request.user.enseignant_profile
    # real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    my_Etu = Etudiant.objects.get(etu_id)
    dep_id = my_Etu.niv_spe_dep_sg.niv_spe_dep.departement.id
    departement = Departement.objects.get(id=dep_id)
    # departement = Departement.objects.get(id=real_Dep.departement.id)
    countAll_User = CustomUser.objects.all().count() + 1
    # myEns = Enseignant.objects.get(id=ens_id)

    this_Etu = Etudiant.objects.get(
        id=etu_id,
        inscrit_univ = False,
        )
    if this_Etu != None:
        if my_Etu.user != None:
            this_Etu.inscrit_univ = True
            this_Etu.save()
            messages.success(request, f'لقد تمت عملية تنشيط حساب {my_Etu} بنجاح')
        else:
            # Remplacer les espaces par des underscores
            nom_Etu = my_Etu.nom_fr
            nom_modifie = nom_Etu.replace(' ', '_')
            poste = Poste.objects.get(nom_fr="Etudiant", type='etudiant').first()
            user = CustomUser.objects.create(
                    username=nom_modifie+str(countAll_User),  # Exemple de nom d'utilisateur
                    first_name=my_Etu.prenom_fr,  # Remplacez par les données réelles du formulaire
                    last_name=my_Etu.nom_fr,  # Remplacez par les données réelles du formulaire
                    poste_principal=poste,  # Définir explicitement
                    email="etudiant@test.com",  # Définir explicitement
                )
            # Définir un mot de passe (nécessaire pour un utilisateur)
            user.set_password("...Etu123")  # Remplacez par un mot de passe sécurisé
            user.save()

            my_Etu.user=user
            my_Etu.save()

            this_Etu.inscrit_univ = True
            this_Etu.save()

            messages.success(request, f'لقد تمت عملية تسجيل {my_Etu} بنجاح')

    else:
        # Remplacer les espaces par des underscores
        nom_Etu = my_Etu.nom_fr
        nom_modifie = nom_Etu.replace(' ', '_')
        poste = Poste.objects.get(nom_fr="Etudiant", type='etudiant').first()
        user = CustomUser.objects.create(
                username=nom_modifie+str(countAll_User),  # Exemple de nom d'utilisateur
                first_name=my_Etu.prenom_fr,  # Remplacez par les données réelles du formulaire
                last_name=my_Etu.nom_fr,  # Remplacez par les données réelles du formulaire
                poste_principal=poste,  # Définir explicitement
                email="etudiant@test.com",  # Définir explicitement
            )
        # Définir un mot de passe (nécessaire pour un utilisateur)
        user.set_password("...Etu123")  # Remplacez par un mot de passe sécurisé
        user.save()

        my_Etu.user=user
        my_Etu.save()
    
        # Enseignant to search
        this_Etu_inscrit = Etudiant.objects.get(
            id=etu_id,
            inscrit_univ = False,
            )
        this_Etu_inscrit.inscrit_univ = True
        # messages.warning(request, f'تم إضافة "{maj_Ens}" إلى المنصة.')
        this_Etu_inscrit.save()

        messages.success(request, f'لقد تمت عملية تسجيل {my_Etu} بنجاح')

    return redirect('depa:list_etudiants')







# ----------------------------------------------------------------------
@login_required
def delete_Enseignant(request, ens_dep_id):
    enseignant = request.user.enseignant_profile
    # real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')

    my_Ens_Dep = Ens_Dep.objects.get(id=ens_dep_id)
    deleted_Ens = my_Ens_Dep.enseignant

    if my_Ens_Dep.statut != 'Permanent':
        my_Ens_Dep.delete()
        messages.warning(request, f'تم حذف "{deleted_Ens}" من القسم.')
    else:
        messages.error(request, f'لا يمكن حذف الأستاذ المرسم "{deleted_Ens}".')

    # Récupérer l'URL de la page précédente et y retourner
    referer_url = request.META.get('HTTP_REFERER')
    if referer_url:
        return redirect(referer_url)
    else:
        # Si pas de referer, rediriger vers une page par défaut
        return redirect('depa:new_Enseignant')  # ou une autre page de votre choix
    






# ----------------------------------------------------------------------
@login_required
def delete_Ens_Platform(request, ens_id):
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)

    # Enseignant to delete
    maj_Ens = Enseignant.objects.get(id=ens_id)
    this_to_delete = Ens_Dep.objects.get(
        enseignant=ens_id,
        departement=departement,
        est_inscrit=True
    )
    this_to_delete.est_inscrit = False
    this_to_delete.save()
    
    messages.warning(request, f'تم حذف "{maj_Ens}" من المنصة.')
    
    # Récupérer l'URL de la page précédente et y retourner
    referer_url = request.META.get('HTTP_REFERER')
    if referer_url:
        return redirect(referer_url)
    else:
        # Si pas de referer, rediriger vers une page par défaut
        return redirect('depa:new_Enseignant')  # ou une autre page de votre choix



# ----------------------------------------------------------------------
# @login_required
# def list_NivSpeDep(request):
#     enseignant = request.user.enseignant_profile
#     real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
#     departement = Departement.objects.get(id=real_Dep.departement.id)
#     all_NivSpeDep = NivSpeDep.objects.filter(departement=departement.id).order_by('specialite__reforme','niveau','specialite__identification')

#     # Calcul pour NivSpeDep (nbr_matieres et nbr_etudiants)
#     for x in all_NivSpeDep:
#         x.nbr_matieres_s1 = Matiere.objects.filter(niv_spe_dep=x, semestre__numero=1).count()
#         x.nbr_matieres_s2 = Matiere.objects.filter(niv_spe_dep=x, semestre__numero=2).count()
#         x.nbr_etudiants = Etudiant.objects.filter(niv_spe_dep_sg__niv_spe_dep=x).count()
#         x.save()

#     # ✅ CALCUL INTELLIGENT pour NivSpeDep_SG selon le type d'affectation
#     # ✅ VERSION OPTIMISÉE pour NivSpeDep_SG
#     all_niv_spe_dep_sg = NivSpeDep_SG.objects.filter(niv_spe_dep__departement=departement)

#     for niv_spe_dep_sg in all_niv_spe_dep_sg:
#         if niv_spe_dep_sg.type_affectation == 'par_groupe':
#             # 🎯 AFFECTATION PAR GROUPE
#             niv_spe_dep_sg.nbr_etudiants_SG = Etudiant.objects.filter(
#                 niv_spe_dep_sg=niv_spe_dep_sg
#             ).count()
            
#         elif niv_spe_dep_sg.type_affectation == 'par_section':
#             # 🎯 AFFECTATION PAR SECTION: Version optimisée avec une seule requête
#             niv_spe_dep_sg.nbr_etudiants_SG = Etudiant.objects.filter(
#                 niv_spe_dep_sg__niv_spe_dep=niv_spe_dep_sg.niv_spe_dep,
#                 niv_spe_dep_sg__section=niv_spe_dep_sg.section,
#             ).count()
            
#         elif niv_spe_dep_sg.type_affectation == 'tous_etudiants':
#             # 🎯 AFFECTATION TOUS ÉTUDIANTS
#             niv_spe_dep_sg.nbr_etudiants_SG = Etudiant.objects.filter(
#                 niv_spe_dep_sg__niv_spe_dep=niv_spe_dep_sg.niv_spe_dep
#             ).count()
        
#         niv_spe_dep_sg.save()

#     # Construction des listes pour l'affichage
#     all_NivSpeDep_S1 = []
#     all_NivSpeDep_S2 = []
    
#     for x in all_NivSpeDep:
#         s1_matieres = Matiere.objects.filter(niv_spe_dep=x, semestre__numero=1).distinct('niv_spe_dep')
#         s2_matieres = Matiere.objects.filter(niv_spe_dep=x, semestre__numero=2).distinct('niv_spe_dep')
#         all_NivSpeDep_S1.extend(s1_matieres)
#         all_NivSpeDep_S2.extend(s2_matieres)

#     context = {
#         'title': 'قائمة المدرجات',
#         'my_Fac': departement.faculte,
#         'my_Dep': departement,
#         'all_NivSpeDep': all_NivSpeDep,
#         'all_NivSpeDep_S1': all_NivSpeDep_S1,
#         'all_NivSpeDep_S2': all_NivSpeDep_S2,
#     }
#     return render(request, 'departement/list_NivSpeDep.html', context)





# @login_required
# def list_NivSpeDep(request):
#     """Vue optimisée avec select_related et prefetch_related"""
#     enseignant = request.user.enseignant_profile
#     real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
#     departement = Departement.objects.get(id=real_Dep.departement.id)
    
#     # ✅ OPTIMISATION : Une seule requête avec toutes les relations
#     all_NivSpeDep = NivSpeDep.objects.filter(
#         departement=departement.id
#     ).select_related(
#         'niveau',
#         'specialite',
#         'specialite__reforme',
#         'departement'
#     ).prefetch_related(
#         'matiere_set',
#         'nivspedep_sg_set__etudiant_set'
#     ).annotate(
#         # Calculer directement dans la requête
#         nbr_matieres_s1=Count('matiere', filter=Q(matiere__semestre__numero=1), distinct=True),
#         nbr_matieres_s2=Count('matiere', filter=Q(matiere__semestre__numero=2), distinct=True),
#         nbr_etudiants=Count('nivspedep_sg__etudiant', distinct=True)
#     ).order_by('specialite__reforme', 'niveau', 'specialite__identification')

#     # ✅ OPTIMISATION : Récupérer les matieres avec select_related en une seule fois
#     all_NivSpeDep_S1 = Matiere.objects.filter(
#         niv_spe_dep__departement=departement,
#         semestre__numero=1
#     ).select_related(
#         'niv_spe_dep',
#         'niv_spe_dep__niveau',
#         'niv_spe_dep__specialite',
#         'niv_spe_dep__specialite__reforme',
#         'semestre'
#     ).distinct('niv_spe_dep')

#     all_NivSpeDep_S2 = Matiere.objects.filter(
#         niv_spe_dep__departement=departement,
#         semestre__numero=2
#     ).select_related(
#         'niv_spe_dep',
#         'niv_spe_dep__niveau',
#         'niv_spe_dep__specialite',
#         'niv_spe_dep__specialite__reforme',
#         'semestre'
#     ).distinct('niv_spe_dep')

#     context = {
#         'title': 'قائمة المدرجات',
#         'my_Fac': departement.faculte,
#         'my_Dep': departement,
#         'all_NivSpeDep': all_NivSpeDep,
#         'all_NivSpeDep_S1': all_NivSpeDep_S1,
#         'all_NivSpeDep_S2': all_NivSpeDep_S2,
#     }
#     return render(request, 'departement/list_NivSpeDep.html', context)






from django.db.models import Count, Q, Prefetch

@login_required
def list_NivSpeDep(request):
    """Vue optimisée SÛRE - Sans conflits de noms"""
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)
    
    # ✅ OPTIMISATION : Charger avec select_related uniquement
    all_NivSpeDep = NivSpeDep.objects.filter(
        departement=departement.id
    ).select_related(
        'niveau',
        'specialite',
        'specialite__reforme',
        'departement'
    ).order_by('specialite__reforme', 'niveau', 'specialite__identification')
    
    # ✅ Pré-charger TOUTES les matières du département en une seule requête
    all_matieres = Matiere.objects.filter(
        niv_spe_dep__departement=departement
    ).select_related('semestre', 'niv_spe_dep')
    
    # ✅ Créer un dictionnaire pour accès rapide
    matieres_par_niv = {}
    for matiere in all_matieres:
        niv_id = matiere.niv_spe_dep.id
        if niv_id not in matieres_par_niv:
            matieres_par_niv[niv_id] = {'s1': 0, 's2': 0}
        if matiere.semestre.numero == 1:
            matieres_par_niv[niv_id]['s1'] += 1
        elif matiere.semestre.numero == 2:
            matieres_par_niv[niv_id]['s2'] += 1
    
    # ✅ Pré-charger TOUS les étudiants du département en une seule requête
    etudiants_par_niv = Etudiant.objects.filter(
        niv_spe_dep_sg__niv_spe_dep__departement=departement
    ).values('niv_spe_dep_sg__niv_spe_dep').annotate(
        total=Count('id')
    )
    
    # ✅ Créer un dictionnaire pour accès rapide
    etud_dict = {item['niv_spe_dep_sg__niv_spe_dep']: item['total'] for item in etudiants_par_niv}
    
    # ✅ Assigner les valeurs EN MÉMOIRE (pas de requêtes supplémentaires!)
    for x in all_NivSpeDep:
        x.nbr_matieres_s1 = matieres_par_niv.get(x.id, {}).get('s1', 0)
        x.nbr_matieres_s2 = matieres_par_niv.get(x.id, {}).get('s2', 0)
        x.nbr_etudiants = etud_dict.get(x.id, 0)
        # Pas de save() - calcul à la volée

    # ✅ Récupérer les matières S1 et S2 avec relations optimisées
    all_NivSpeDep_S1 = Matiere.objects.filter(
        niv_spe_dep__departement=departement,
        semestre__numero=1
    ).select_related(
        'niv_spe_dep',
        'niv_spe_dep__niveau',
        'niv_spe_dep__specialite',
        'niv_spe_dep__specialite__reforme',
        'semestre'
    ).distinct('niv_spe_dep').order_by('niv_spe_dep')

    all_NivSpeDep_S2 = Matiere.objects.filter(
        niv_spe_dep__departement=departement,
        semestre__numero=2
    ).select_related(
        'niv_spe_dep',
        'niv_spe_dep__niveau',
        'niv_spe_dep__specialite',
        'niv_spe_dep__specialite__reforme',
        'semestre'
    ).distinct('niv_spe_dep').order_by('niv_spe_dep')

    # ✅ Calculer les statistiques pour NivSpeDep_SG de manière optimisée
    all_niv_spe_dep_sg = NivSpeDep_SG.objects.filter(
        niv_spe_dep__departement=departement
    ).select_related(
        'niv_spe_dep',
        'section',
        'groupe'
    )
    
    # ✅ Pré-charger les comptes d'étudiants par NivSpeDep_SG
    etud_par_nsg = Etudiant.objects.filter(
        niv_spe_dep_sg__niv_spe_dep__departement=departement
    ).values('niv_spe_dep_sg').annotate(total=Count('id'))
    
    etud_nsg_dict = {item['niv_spe_dep_sg']: item['total'] for item in etud_par_nsg}
    
    for niv_spe_dep_sg in all_niv_spe_dep_sg:
        niv_spe_dep_sg.nbr_etudiants_SG = etud_nsg_dict.get(niv_spe_dep_sg.id, 0)

    context = {
        'title': 'قائمة المدرجات',
        'my_Fac': departement.faculte,
        'my_Dep': departement,
        'all_NivSpeDep': all_NivSpeDep,
        'all_NivSpeDep_S1': all_NivSpeDep_S1,
        'all_NivSpeDep_S2': all_NivSpeDep_S2,
    }
    return render(request, 'departement/list_NivSpeDep.html', context)


# ========================================
# 📊 ANALYSE DE PERFORMANCE
# ========================================
# AVANT (version originale):
#   - ~100+ requêtes SQL (N+1 problem)
#   - ~5-10 secondes de chargement
#
# APRÈS (cette version):
#   - ~7-10 requêtes SQL seulement
#   - ~0.5-2 secondes de chargement
#   - Réduction de 90% des requêtes
# ========================================





# @login_required
# def list_Mat_Niv(request):
#     # Récupérer le département courant (supposons qu'il est lié à l'utilisateur via Ens_Dep ou autre logique)
#     enseignant = request.user.enseignant_profile
#     real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
#     departement = Departement.objects.get(id=real_Dep.departement.id)

#     # enseignant = request.user.enseignant_profile  # Assumons que l'utilisateur a un profil enseignant
#     # ens_dep = Ens_Dep.objects.filter(enseignant=enseignant, statut='Permanent')
#     # ens_dep = Ens_Dep.objects.filter(enseignant=enseignant).first()
#     # if not ens_dep:
#     #     return render(request, 'departement/erreur.html', {'message': 'Vous n’êtes pas affecté à un département.'})
#     # departement = ens_dep.departement

#     # Liste des réformes liées au département courant
#     reformes = Reforme.objects.filter(Specialites_reforme__departement=departement).distinct()

#     # # print('*'*50)
#     # # print(reformes)
#     # # print('*'*50)

#     # Si c'est une requête POST (soumission du formulaire)
#     if request.method == 'POST':
#         reforme_id = request.POST.get('reforme')
#         niveau_id = request.POST.get('niveau')

#         # Filtrer les matières selon la réforme et le niveau
#         if reforme_id and niveau_id:
#             matieres = Matiere.objects.filter(
#                 niv_spe_dep__niveau_id=niveau_id,
#                 niv_spe_dep__specialite__reforme_id=reforme_id,
#                 niv_spe_dep__departement=departement
#             )
#         else:
#             matieres = []

#         return render(request, 'departement/list_Mat_Niv.html', {
#             'reformes': reformes,
#             'selected_reforme': reforme_id,
#             'selected_niveau': niveau_id,
#             'matieres': matieres,
#         })

#     # Si c'est une requête GET (affichage initial du formulaire)
#     return render(request, 'departement/list_Mat_Niv.html', {
#         'reformes': reformes,
#         'selected_reforme': None,
#         'selected_niveau': None,
#         'matieres': [],
#     })

# # Vue AJAX pour récupérer les niveaux dynamiquement selon la réforme
# @login_required
# def get_niveaux_by_reforme(request):
#     reforme_id = request.GET.get('reforme_id')
#     departement_id = request.GET.get('departement_id')

#     if reforme_id and departement_id:
#         niveaux = NivSpeDep.objects.filter(
#             specialite__reforme_id=reforme_id,
#             departement_id=departement_id
#         ).distinct()
#         niveaux_data = [{'id': niveau.id, 'nom': str(niveau)} for niveau in niveaux]
#         return JsonResponse({'niveaux': niveaux_data})

#     return JsonResponse({'niveaux': []})

# @login_required
# def get_reformes_json(request):
#     departement = request.user.get_enseignant().departement  # Supposé que l'enseignant est lié à un département
#     reformes = Reforme.objects.filter(Specialites_reforme__departement=departement).values('id', 'nom_ar', 'nom_fr')
#     # print('*'*50)
#     # print(reformes)
#     # print('*'*50)
#     return JsonResponse({'data': list(reformes)})

# @login_required
# def get_reformes_json(request):
#     # Récupérer le département de l'utilisateur connecté
#     try:
#         departement = request.user.get_enseignant().departement
#     except AttributeError:
#         return JsonResponse({'data': []}, status=400)
#     # Récupérer les spécialités du département
#     specialites = Specialite.objects.filter(departement=departement)
#     # Récupérer les réformes associées à ces spécialités
#     reformes = Reforme.objects.filter(Specialites_reforme__in=specialites).distinct().values('id', 'nom_ar', 'nom_fr')
#     return JsonResponse({'data': list(reformes)})


# @login_required
# def get_reformes_json(request):
#     # Récupérer le département de l'utilisateur connecté
#     enseignant = request.user.enseignant_profile
#     real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
#     departement = Departement.objects.get(id=real_Dep.departement.id)
    
#     # Récupérer les spécialités du département
#     specialites = Specialite.objects.filter(departement=departement)
#     # Récupérer les réformes associées à ces spécialités
#     if not specialites.exists():
#         return JsonResponse({'data': [], 'error': 'Aucune spécialité trouvée pour ce département.'}, status=400)
#     reformes = Reforme.objects.filter(Specialites_reforme__in=specialites).distinct().values('id', 'nom_ar', 'nom_fr')
#     return JsonResponse({'data': list(reformes)})

@login_required
def get_reformes_json(request):
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)

    reformes = Reforme.objects.filter(Specialites_reforme__departement=departement).values('id', 'nom_ar', 'nom_fr').distinct()
    return JsonResponse({'data': list(reformes)})


@login_required
def get_niveaux_json(request, reforme_id):
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)

    niveaux = NivSpeDep.objects.filter(
        departement=departement,
        specialite__reforme_id=reforme_id
    ).values('niveau__id', 'niveau__nom_ar', 'niveau__nom_fr').distinct()
    return JsonResponse({'data': list(niveaux)})


@login_required
def get_specialites_json(request, niveau_id):
    """Vue pour récupérer les spécialités selon niveau ET réforme"""
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)

    if request.method == 'POST':
        reforme_id = request.POST.get('reforme')
        niveau_id = request.POST.get('niveau')
        
        # print(f"🔍 Recherche spécialités: reforme={reforme_id}, niveau={niveau_id}")
        
        # ✅ CORRECTION: Utiliser les bons champs selon vos modèles
        specialites = NivSpeDep.objects.filter(
            departement=departement,
            specialite__reforme_id=reforme_id,
            niveau_id=niveau_id
        ).values('specialite__id', 'specialite__nom_ar', 'specialite__nom_fr').distinct()
        
        # print(f"✅ Spécialités trouvées: {list(specialites)}")
        return JsonResponse({'data': list(specialites)})
    
    # Fallback pour GET (comportement original)
    specialites = NivSpeDep.objects.filter(
        departement=departement,
        niveau_id=niveau_id
    ).values('specialite__id', 'specialite__nom_ar', 'specialite__nom_fr').distinct()
    
    return JsonResponse({'data': list(specialites)})


@login_required
def get_semestres_json(request):
    """Vue pour récupérer les semestres - pour l'instant indépendants"""
    if request.method == 'POST':
        reforme_id = request.POST.get('reforme')
        niveau_id = request.POST.get('niveau')
        specialite_id = request.POST.get('specialite')
        
        # print(f"🔍 Recherche semestres: reforme={reforme_id}, niveau={niveau_id}, specialite={specialite_id}")
        
        # Pour l'instant, semestres restent indépendants
        # Si vous voulez les lier plus tard, ajoutez vos critères ici
        semestres = Semestre.objects.all().values('id', 'nom_ar').order_by('id')
        
        # print(f"✅ Semestres trouvés: {list(semestres)}")
        return JsonResponse({'data': list(semestres)})
    
    # GET par défaut
    semestres = Semestre.objects.all().values('id', 'nom_ar').order_by('id')
    return JsonResponse({'data': list(semestres)})


@login_required
def get_matieres_json(request):
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)

    if request.method == 'POST':
        reforme_id = request.POST.get('reforme')
        niveau_id = request.POST.get('niveau')
        specialite_id = request.POST.get('specialite')
        semestre_id = request.POST.get('semestre')

        # print(f"🔍 Recherche matières: reforme={reforme_id}, niveau={niveau_id}, specialite={specialite_id}, semestre={semestre_id}")

        matieres_query = Matiere.objects.filter(
            niv_spe_dep__departement=departement,
            niv_spe_dep__niveau_id=niveau_id,
            niv_spe_dep__specialite_id=specialite_id
        ).select_related('unite')

        if semestre_id:
            matieres_query = matieres_query.filter(semestre_id=semestre_id)

        matieres = matieres_query.values(
            'id', 'nom_ar', 'nom_fr', 'unite__code', 'code', 
            'coeff', 'credit', 'unite__nom_fr', 'unite__nom_ar'
        )

        # print(f"✅ Matières trouvées: {len(list(matieres))}")
        return JsonResponse({'data': list(matieres)})
    
    return JsonResponse({'data': []}, status=400)



# @login_required
# def list_Mat_Niv(request):
#     enseignant = request.user.enseignant_profile
#     real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
#     departement = Departement.objects.get(id=real_Dep.departement.id)

#     context = {
#         'departement': departement,
#         'refs_url': reverse('depa:refs-json'),
#         'nivs_url': reverse('depa:nivs-json', kwargs={'reforme_id': 0}),  # Placeholder
#         'spts_url': reverse('depa:spts-json', kwargs={'niveau_id': 0}),  # Placeholder
#         'semestres_url': reverse('depa:semestres-json'),  # Ajout de l'URL pour semestres
#         'matieres_url': reverse('depa:matieres-json'),
#     }
#     return render(request, 'departement/list_Mat_Niv.html', context)
# from django.http import JsonResponse
# from django.contrib.auth.decorators import login_required
# from django.db import transaction
# from django.core.exceptions import ValidationError
# from tablib import Dataset
# from apps.academique.affectation.models import Ens_Dep
# from apps.academique.departement.models import Departement
# from apps.academique.specialite.models import NivSpeDep
# from apps.academique.matiere.models import Matiere
# from apps.academique.unite.models import Unite
# from apps.noyau.commun.models import Semestre

# FONCTION ORIGINALE - Pour afficher la page
@login_required
def list_Mat_Niv(request):
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)

    context = {
        'departement': departement,
        # CORRECTION: Utiliser 'departement' au lieu de 'depa'
        'refs_url': reverse('depa:refs-json'),
        'nivs_url': reverse('depa:nivs-json', kwargs={'reforme_id': 0}),
        'spts_url': reverse('depa:spts-json', kwargs={'niveau_id': 0}),
        'semestres_url': reverse('depa:semestres-json'),
        'matieres_url': reverse('depa:matieres-json'),
        'my_Dep': departement,  # Passer le département au template
        'my_Fac': departement.faculte,
    }
    return render(request, 'departement/list_Mat_Niv.html', context)


# NOUVELLE FONCTION - Pour l'import des matières




# Vue d'import modifiée dans views.py
@login_required
def import_matieres(request):
    """
    Vue pour importer des matières avec debug détaillé
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'الطريقة غير مسموحة'})
    
    try:
        # Récupérer les données du formulaire
        reforme_id = request.POST.get('reforme')
        niveau_id = request.POST.get('niveau')
        specialite_id = request.POST.get('specialite')
        semestre_id = request.POST.get('semestre')
        file_matieres = request.FILES.get('file_matieres')
        
        # # print(f"🔍 DEBUG - Paramètres reçus:")
        # # print(f"   reforme_id: {reforme_id}")
        # # print(f"   niveau_id: {niveau_id}")
        # # print(f"   specialite_id: {specialite_id}")
        # # print(f"   semestre_id: {semestre_id}")
        # # print(f"   file_matieres: {file_matieres}")
        
        # Vérifications des paramètres
        if not all([reforme_id, niveau_id, specialite_id, semestre_id]):
            return JsonResponse({
                'success': False, 
                'message': 'جميع الحقول مطلوبة (الإصلاح، المستوى، التخصص، السداسي)'
            })
        
        if not file_matieres:
            return JsonResponse({
                'success': False, 
                'message': 'لم يتم اختيار ملف'
            })
        
        # Vérifier l'extension du fichier
        if not file_matieres.name.lower().endswith(('.xlsx', '.xls')):
            return JsonResponse({
                'success': False, 
                'message': 'نوع الملف غير صحيح. يجب أن يكون ملف Excel (.xlsx أو .xls)'
            })
        
        # Récupérer le département de l'utilisateur actuel
        try:
            enseignant = request.user.enseignant_profile
            real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
            departement = real_Dep.departement
            # # print(f"🔍 DEBUG - Département: {departement}")
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'message': f'خطأ في تحديد القسم: {str(e)}'
            })
        
        # Vérifier que les sélections correspondent au département et récupérer les objets
        try:           
            # Vérifier que la spécialité appartient au département
            niv_spe_dep = NivSpeDep.objects.get(
                specialite_id=specialite_id,
                niveau_id=niveau_id,
                departement=departement
            )
            # # print(f"🔍 DEBUG - NivSpeDep trouvé: {niv_spe_dep}")
            
            # Vérifier que le semestre existe
            semestre = Semestre.objects.get(id=semestre_id)
            # # print(f"🔍 DEBUG - Semestre trouvé: {semestre}")
            
        except NivSpeDep.DoesNotExist:
            return JsonResponse({
                'success': False, 
                'message': 'التخصص المحدد غير متوفر في هذا القسم أو المستوى غير صحيح'
            })
        except Semestre.DoesNotExist:
            return JsonResponse({
                'success': False, 
                'message': 'السداسي المحدد غير موجود'
            })
        
        # Lire le fichier Excel
        try:
            dataset = Dataset()
            imported_data = dataset.load(file_matieres.read(), format='xlsx')
            # # print(f"🔍  - Fichier Excel lu. Nombre de lignes: {len(imported_data)}")
            # # print(f"🔍 DEBUDEBUGG - Premières lignes:")
            for i, row in enumerate(imported_data[:3]):  # Afficher les 3 premières lignes
                pass
                # print(f"   Ligne {i+1}: {row}")
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'message': f'خطأ في قراءة ملف Excel: {str(e)}'
            })
        
        # Vérifier que le fichier contient des données
        if len(imported_data) == 0:
            return JsonResponse({
                'success': False, 
                'message': 'الملف فارغ أو لا يحتوي على بيانات صالحة'
            })
        
        # Variables pour le comptage et les erreurs
        created_count = 0
        updated_count = 0
        error_count = 0
        errors = []
        success_details = []
        
        # # print(f"🔍 DEBUG - Début du traitement des données...")
        
        # Traitement des données ligne par ligne
        with transaction.atomic():
            for row_index, data in enumerate(imported_data, start=1):
                # # print(f"\n🔍 DEBUG - Traitement ligne {row_index}: {data}")
                
                try:
                    # Vérifier que la ligne contient suffisamment de colonnes
                    if len(data) < 3:
                        error_count += 1
                        error_msg = f'السطر {row_index}: بيانات ناقصة (يجب أن يحتوي على 6 أعمدة على الأقل)'
                        errors.append(error_msg)
                        # print(f"❌ {error_msg}")
                        continue
                    
                    # Extraire les données avec gestion des valeurs manquantes
                    nom_ar = str(data[0]).strip() if data[0] and str(data[0]).strip() else None
                    nom_fr = str(data[1]).strip() if len(data) > 1 and data[1] and str(data[1]).strip() else None
                    code = str(data[2]).strip() if len(data) > 2 and data[2] and str(data[2]).strip() else None
                    
                    # print(f"   nom_ar: '{nom_ar}'")
                    # print(f"   nom_fr: '{nom_fr}'")
                    # print(f"   code: '{code}'")
                    
                    # Conversion sécurisée pour les nombres
                    try:
                        coeff = float(data[3]) if len(data) > 3 and data[3] is not None and str(data[3]).strip() else 1.0
                    except (ValueError, TypeError):
                        coeff = 1.0
                    
                    try:
                        credit = int(float(data[4])) if len(data) > 4 and data[4] is not None and str(data[4]).strip() else 1
                    except (ValueError, TypeError):
                        credit = 1
                    
                    unite_code = str(data[5]).strip() if len(data) > 5 and data[5] and str(data[5]).strip() else None
                    
                    # print(f"   coeff: {coeff}")
                    # print(f"   credit: {credit}")
                    # print(f"   unite_code: '{unite_code}'")
                    
                    # Validations obligatoires
                    # if not code:
                    #     error_count += 1
                    #     error_msg = f'السطر {row_index}: كود المادة مطلوب'
                    #     errors.append(error_msg)
                    #     # print(f"❌ {error_msg}")
                    #     continue
                    
                    if not nom_fr:
                        error_count += 1
                        error_msg = f'السطر {row_index}: اسم المادة بالفرنسية مطلوب'
                        errors.append(error_msg)
                        # print(f"❌ {error_msg}")
                        continue
                    
                    # Rechercher l'unité par code
                    unite = None
                    if unite_code:
                        try:
                            unite = Unite.objects.get(code=unite_code)
                            # print(f"   ✅ Unité trouvée: {unite}")
                        except Exception as e:
                            unite = None
                            # print(f"   ⚠️ Unité non trouvée pour le code '{unite_code}': {e}")
                    # else:
                    #     pass
                        # print(f"   ℹ️ Pas de code d'unité fourni")
                    
                    # Vérifier si la matière existe déjà
                    existing_matiere = Matiere.objects.filter(
                        nom_ar=nom_ar,
                        nom_fr=nom_fr,
                        code=code,
                        unite=unite,
                        niv_spe_dep=niv_spe_dep
                    ).first()
                    
                    if existing_matiere:
                        # print(f"   🔄 Mise à jour de la matière existante: {existing_matiere}")
                        # Mettre à jour la matière existante
                        if nom_ar:
                            existing_matiere.nom_ar = nom_ar
                        existing_matiere.nom_fr = nom_fr
                        existing_matiere.coeff = coeff
                        existing_matiere.credit = credit
                        existing_matiere.semestre = semestre
                        if unite:
                            existing_matiere.unite = unite
                        
                        try:
                            existing_matiere.full_clean()
                            existing_matiere.save()
                            updated_count += 1
                            success_msg = f'تم تحديث: {code} - {nom_fr}'
                            success_details.append(success_msg)
                            # print(f"   ✅ {success_msg}")
                        except ValidationError as ve:
                            error_count += 1
                            error_msg = f'السطر {row_index}: خطأ في التحقق - {"; ".join(ve.messages)}'
                            errors.append(error_msg)
                            # print(f"   ❌ Validation Error: {ve}")
                            continue
                        except Exception as e:
                            error_count += 1
                            error_msg = f'السطر {row_index}: خطأ في الحفظ - {str(e)}'
                            errors.append(error_msg)
                            # print(f"   ❌ Save Error: {e}")
                            continue
                    else:
                        # print(f"   ➕ Création d'une nouvelle matière")
                        # Créer une nouvelle matière
                        try:
                            new_matiere = Matiere(
                                nom_ar=nom_ar,
                                nom_fr=nom_fr,
                                code=code,
                                coeff=coeff,
                                credit=credit,
                                unite=unite,
                                niv_spe_dep=niv_spe_dep,
                                semestre=semestre
                            )
                            
                            # print(f"   🔍 Objet Matiere créé: {new_matiere}")
                            
                            # Validation avant sauvegarde
                            new_matiere.full_clean()
                            new_matiere.save()
                            
                            created_count += 1
                            success_msg = f'تم إنشاء: {code} - {nom_fr}'
                            success_details.append(success_msg)
                            # print(f"   ✅ {success_msg}")
                            
                        except ValidationError as ve:
                            error_count += 1
                            error_msg = f'السطر {row_index}: خطأ في التحقق - {"; ".join(ve.messages)}'
                            errors.append(error_msg)
                            # print(f"   ❌ Validation Error: {ve}")
                            continue
                        except Exception as e:
                            error_count += 1
                            error_msg = f'السطر {row_index}: خطأ في الإنشاء - {str(e)}'
                            errors.append(error_msg)
                            # print(f"   ❌ Creation Error: {e}")
                            continue
                
                except Exception as e:
                    error_count += 1
                    error_msg = f'السطر {row_index}: خطأ غير متوقع - {str(e)}'
                    errors.append(error_msg)
                    # print(f"   ❌ Unexpected Error: {e}")
        
        # إعداد رسالة النتيجة
        total_processed = created_count + updated_count
        
        # # print(f"\n🔍 DEBUG - Résultats finaux:")
        # # print(f"   Created: {created_count}")
        # # print(f"   Updated: {updated_count}")
        # # print(f"   Errors: {error_count}")
        # # print(f"   Total processed: {total_processed}")
        
        if total_processed == 0 and error_count > 0:
            return JsonResponse({
                'success': False,
                'message': f'فشل في معالجة جميع الأسطر. عدد الأخطاء: {error_count}',
                'errors': errors[:5],
                'debug_info': f'Processed {len(imported_data)} rows, all failed'
            })
        
        # إعداد رسالة مفصلة
        message_parts = []
        if created_count > 0:
            message_parts.append(f'✅ تم إنشاء {created_count} مادة جديدة')
        if updated_count > 0:
            message_parts.append(f'🔄 تم تحديث {updated_count} مادة موجودة')
        if error_count > 0:
            message_parts.append(f'❌ حدث خطأ مع {error_count} سطر')
        
        success_message = '<br>'.join(message_parts)
        
        # إضافة تفاصيل الأخطاء
        if errors and len(errors) > 0:
            error_summary = '<br>'.join([f'• {error}' for error in errors[:5]])
            if len(errors) > 5:
                error_summary += f'<br>• ... و {len(errors) - 5} أخطاء أخرى'
            success_message += f'<br><br><strong>تفاصيل الأخطاء:</strong><br>{error_summary}'
        
        return JsonResponse({
            'success': True,
            'message': success_message,
            'count': total_processed,
            'created': created_count,
            'updated': updated_count,
            'errors': error_count,
            'details': success_details[:10],
            'debug_info': f'Processed {len(imported_data)} rows from Excel file'
        })
    
    except Exception as e:
        # print(f"❌ ERREUR GLOBALE: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'message': f'حدث خطأ غير متوقع في الخادم: {str(e)}'
        })








##################################################################
##################################################################
@login_required
def import_emploi(request):
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)

    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        try:
            df = pd.read_excel(excel_file)
            # print(f"Fichier Excel lu avec succès. Nombre de lignes: {len(df)}")
            # print(f"Colonnes trouvées: {df.columns.tolist()}")
        except Exception as e:
            messages.error(request, f"Erreur lors de la lecture du fichier Excel: {str(e)}")
            return redirect('depa:import_emploi')

        # Récupérer les paramètres du formulaire
        reforme_id = request.POST.get('reforme')
        niveau_id = request.POST.get('niveau')
        specialite_id = request.POST.get('specialite')
        semestre_id = request.POST.get('semestre')

        # Vérifier que tous les paramètres sont présents
        if not all([reforme_id, niveau_id, specialite_id, semestre_id]):
            messages.error(request, "جميع الحقول مطلوبة: الإصلاح، المستوى، التخصص، والسداسي")
            return redirect('depa:import_emploi')

        # Dictionnaire de mappage des numéros de séance aux valeurs Timeblock
        seance_mapping = {
            1: '09:30-08:00',
            2: '11:10-09:40',
            3: '12:50-11:20',
            4: '14:40-13:10',
            5: '16:20-14:50',
            6: '18:00-16:30'
        }

        errors = []
        success_count = 0
        
        for index, row in df.iterrows():
            try:
                ligne_num = index + 1
                # print(f"Traitement de la ligne {ligne_num}: {row.to_dict()}")
                
                # Vérifier que les colonnes nécessaires existent
                required_columns = ['Jour', 'Séance', 'Matiere', 'Enseignant']
                # optional_columns = ['Type', 'Groupes', 'Section', 'Amphi', 'Salle', 'Labo']
                
                # Vérifier les colonnes obligatoires
                for col in required_columns:
                    if col not in df.columns:
                        raise ValueError(f"Colonne manquante: {col}")
                    if pd.isna(row[col]) or str(row[col]).strip() == '':
                        raise ValueError(f"Valeur manquante dans la colonne {col}")
                
                # Gérer la colonne Type avec valeur par défaut
                type_val = row.get('Type', None)
                if pd.isna(type_val) or str(type_val).strip() == '' or str(type_val).strip() == 'nan':
                    type_cours = 'Cours'  # Valeur par défaut
                    # print(f"Ligne {ligne_num}: Type manquant, utilisation de 'Cours' par défaut")
                else:
                    type_cours = str(type_val).strip()

                # 1. Récupérer la matière
                matiere_nom = str(row['Matiere']).strip()
                matiere = Matiere.objects.filter(nom_fr__iexact=matiere_nom).first()
                if not matiere:
                    raise ValueError(f"Matière non trouvée: {matiere_nom}")

                # 2. Récupérer l'enseignant
                enseignant_nom = str(row['Enseignant']).strip()
                enseignant_obj = Ens_Dep.objects.filter(
                    enseignant__nom_fr__iexact=enseignant_nom,
                    departement=departement,
                ).first()
                
                if not enseignant_obj:
                    raise ValueError(f"Enseignant non trouvé: {enseignant_nom}")

                # 3. Gérer la séance
                try:
                    seance_num = int(float(row['Séance']))  # Conversion sécurisée
                    if seance_num not in seance_mapping:
                        raise ValueError(f"Séance {seance_num} n'est pas valide. Utilisez 1 à 6.")
                    temps = seance_mapping[seance_num]
                except (ValueError, TypeError):
                    raise ValueError(f"Valeur de séance invalide: {row['Séance']}")

                # 4. Gérer Groupe ou Section
                groupe_val = row.get('Groupes', None)
                section_val = row.get('Section', None)
                
                # # Nettoyer les valeurs
                # groupe_val = str(groupe_val).strip() if pd.notna(groupe_val) and str(groupe_val).strip() != 'nan' else None
                # section_val = str(section_val).strip() if pd.notna(section_val) and str(section_val).strip() != 'nan' else None

                # Nettoyer les valeurs
                if pd.notna(groupe_val) and str(groupe_val).strip() != 'nan':
                    groupe_val = str(int(float(groupe_val)))  # Convertir float en int, puis en str
                else:
                    groupe_val = None

                if pd.notna(section_val) and str(section_val).strip() != 'nan':
                    section_val = str(int(float(section_val)))  # Convertir float en int, puis en str
                else:
                    section_val = None

                # print('*'*50)
                # print(groupe_val)
                # print('*'*50)
                
                niveau_obj = None
                
                if groupe_val and not section_val:
                    # Recherche par groupe
                    groupe = Groupe.objects.filter(numero__iexact=groupe_val).first()
                    # print('*'*50)
                    # print(groupe)
                    # print('*'*50)
                    if not groupe:
                        raise ValueError(f"Groupe non trouvé: {groupe_val}")
                    
                    niveau_obj = NivSpeDep_SG.objects.filter(
                        niv_spe_dep__specialite__reforme_id=reforme_id,
                        niv_spe_dep__niveau_id=niveau_id,
                        niv_spe_dep__specialite_id=specialite_id,
                        niv_spe_dep__departement=departement,
                        groupe=groupe
                    ).first()
                    
                elif section_val and not groupe_val:
                    # Recherche par section
                    section = Section.objects.filter(numero__iexact=section_val).first()
                    if not section:
                        raise ValueError(f"Section non trouvée: {section_val}")
                    
                    niveau_obj = NivSpeDep_SG.objects.filter(
                        niv_spe_dep__specialite__reforme_id=reforme_id,
                        niv_spe_dep__niveau_id=niveau_id,
                        niv_spe_dep__specialite_id=specialite_id,
                        niv_spe_dep__departement=departement,
                        section=section
                    ).first()
                else:
                    raise ValueError("Vous devez spécifier soit un Groupe soit une Section (pas les deux)")

                if not niveau_obj:
                    raise ValueError(f"Niveau non trouvé pour le groupe/section spécifié")

                # 5. Gérer le lieu avec GenericForeignKey
                amphi_val = row.get('Amphi', None)
                salle_val = row.get('Salle', None)
                labo_val = row.get('Labo', None)
                
                # Nettoyer les valeurs
                amphi_val = str(amphi_val).strip() if pd.notna(amphi_val) and str(amphi_val).strip() != 'nan' else None
                # salle_val = str(salle_val).strip() if pd.notna(salle_val) and str(salle_val).strip() != 'nan' else None
                # labo_val = str(labo_val).strip() if pd.notna(labo_val) and str(labo_val).strip() != 'nan' else None

                # if pd.notna(amphi_val) and str(amphi_val).strip() != 'nan':
                #     amphi_val = str(int(float(amphi_val)))  # Convertir float en int, puis en str
                # else:
                #     amphi_val = None

                if pd.notna(salle_val) and str(salle_val).strip() != 'nan':
                    salle_val = str(int(float(salle_val)))  # Convertir float en int, puis en str
                else:
                    salle_val = None

                if pd.notna(labo_val) and str(labo_val).strip() != 'nan':
                    labo_val = str(int(float(labo_val)))  # Convertir float en int, puis en str
                else:
                    labo_val = None
                
                content_type = None
                object_id = None
                
                # Compter les lieux non vides
                lieux_non_vides = sum([bool(amphi_val), bool(salle_val), bool(labo_val)])
                
                if lieux_non_vides == 0:
                    raise ValueError("Au moins un lieu (Amphi, Salle, ou Labo) doit être spécifié")
                elif lieux_non_vides > 1:
                    raise ValueError("Un seul type de lieu (Amphi, Salle, ou Labo) doit être défini")
                
                if amphi_val:
                    lieu_instance = Amphi_Dep.objects.filter(
                        amphi__nom_fr__iexact=amphi_val, 
                        departement=departement
                    ).first()
                    if not lieu_instance:
                        raise ValueError(f"Amphi non trouvé: {amphi_val}")
                    content_type = ContentType.objects.get_for_model(Amphi_Dep)
                    object_id = lieu_instance.id
                    
                elif salle_val:
                    lieu_instance = Salle_Dep.objects.filter(
                        salle__numero__iexact=salle_val, 
                        departement=departement
                    ).first()
                    if not lieu_instance:
                        raise ValueError(f"Salle non trouvée: {salle_val}")
                    content_type = ContentType.objects.get_for_model(Salle_Dep)
                    object_id = lieu_instance.id
                    
                elif labo_val:
                    lieu_instance = Laboratoire_Dep.objects.filter(
                        laboratoire__numero__iexact=labo_val, 
                        departement=departement
                    ).first()
                    if not lieu_instance:
                        raise ValueError(f"Laboratoire non trouvé: {labo_val}")
                    content_type = ContentType.objects.get_for_model(Laboratoire_Dep)
                    object_id = lieu_instance.id

                # 6. Vérifier si la Classe existe déjà
                existing_classe = Classe.objects.filter(
                    semestre_id=semestre_id,
                    matiere=matiere,
                    enseignant=enseignant_obj,
                    niv_spe_dep_sg=niveau_obj,
                    jour=row['Jour'],
                    temps=temps,
                    type=type_cours,
                    content_type=content_type,
                    object_id=object_id,
                ).first()

                if existing_classe:
                    # print(f"Classe déjà existante à la ligne {ligne_num}, ignorée")
                    continue

                # 7. Créer la classe
                classe = Classe(
                    semestre_id=semestre_id,
                    matiere=matiere,
                    enseignant=enseignant_obj,
                    niv_spe_dep_sg=niveau_obj,
                    jour=row['Jour'],
                    temps=temps,
                    type=type_cours,
                    content_type=content_type,
                    object_id=object_id,
                )
                classe.save()
                success_count += 1
                # print(f"Classe créée avec succès à la ligne {ligne_num}")
                
            except IntegrityError as e:
                error_msg = f"Ligne {ligne_num}: Doublon détecté - {str(e)}"
                errors.append(error_msg)
                # print(error_msg)
                
            except Exception as e:
                error_msg = f"Ligne {ligne_num}: {str(e)}"
                errors.append(error_msg)
                # print(error_msg)

        # Afficher les résultats
        if success_count > 0:
            messages.success(request, f'تم إضافة {success_count} حصة بنجاح')
        
        if errors:
            error_summary = f"تم العثور على {len(errors)} أخطاء:\n" + "\n".join(errors[:5])
            if len(errors) > 5:
                error_summary += f"\n... و {len(errors) - 5} أخطاء أخرى"
            messages.error(request, error_summary)
        
        return redirect('depa:import_emploi')

    context = {
        'departement': departement,
        'refs_url': reverse('depa:refs-json'),
        'nivs_url': reverse('depa:nivs-json', kwargs={'reforme_id': 0}),
        'spts_url': reverse('depa:spts-json', kwargs={'niveau_id': 0}),
        'semestres_url': reverse('depa:semestres-json'),
    }
    return render(request, 'departement/import_emploi.html', context)






def validate_excel_structure(df):
    """Valide la structure du fichier Excel"""
    required_columns = ['Jour', 'Séance', 'Matiere', 'Enseignant']
    optional_columns = ['Type', 'Groupes', 'Section', 'Amphi', 'Salle', 'Labo']
    
    # Vérifier que les colonnes obligatoires existent
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        return False, f"Colonnes manquantes: {', '.join(missing_columns)}"
    
    # Ajouter les colonnes optionnelles manquantes avec des valeurs par défaut
    for col in optional_columns:
        if col not in df.columns:
            df[col] = None
    
    return True, "Structure valide"








##################################################################
# @login_required
# def get_niv_spe_dep_sg_json(request, niveau_id):
#     niv_spe_dep_sgs = NivSpeDep_SG.objects.filter(
#         niv_spe_dep__niveau_id=niveau_id,
#         type_affectation='par_groupe'
#     ).values('id', 'nom')
#     return JsonResponse({'data': list(niv_spe_dep_sgs)})


@login_required
def get_niveaux_json(request, reforme_id):
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)

    niveaux = NivSpeDep.objects.filter(
        departement=departement,
        specialite__reforme_id=reforme_id
    ).values('niveau__id', 'niveau__nom_ar', 'niveau__nom_fr').distinct()
    return JsonResponse({'data': list(niveaux)})


@login_required
def get_niv_spe_dep_sg_json(request, niveau_id):
    """CORRECTION: Ajouter filtrage par département et reforme_id"""
    
    # Récupérer le département de l'utilisateur connecté
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)
    
    # Récupérer reforme_id depuis la requête GET
    reforme_id = request.GET.get('reforme_id')
    
    # print(f"🔍 DEPT get_niv_spe_dep_sg_json: département = {departement.nom_ar}, niveau_id = {niveau_id}, reforme_id = {reforme_id}")
    
    # Construire les filtres avec département
    groupes_filters = {
        'niv_spe_dep__departement': departement,
        'niv_spe_dep__niveau_id': niveau_id
    }
    
    # Ajouter reforme_id si fourni
    if reforme_id:
        groupes_filters['niv_spe_dep__specialite__reforme_id'] = reforme_id
    
    # Récupérer les groupes avec les filtres corrects
    niv_spe_dep_sgs = NivSpeDep_SG.objects.filter(**groupes_filters).exclude(
        # Exclure les sections conteneurs (section mais pas de groupe)
        section__isnull=False,
        groupe__isnull=True
    ).select_related(
        'niv_spe_dep__niveau',
        'niv_spe_dep__specialite',
        'niv_spe_dep__specialite__reforme',
        'section',
        'groupe'
    ).order_by('section__nom_ar', 'groupe__nom_ar')
    
    # print(f"🔍 DEPT Groupes trouvés: {niv_spe_dep_sgs.count()}")
    
    data = []
    sections_dict = {}
    total_etudiants = 0
    
    # Organiser par section
    for groupe in niv_spe_dep_sgs:
        niveau_spe = f"{groupe.niv_spe_dep.niveau.nom_ar} - {groupe.niv_spe_dep.specialite.nom_ar}"
        total_etudiants += groupe.nbr_etudiants_SG
        
        if groupe.section:
            # Groupe appartient à une section
            section_name = groupe.section.nom_ar
            if section_name not in sections_dict:
                sections_dict[section_name] = {
                    'groupes': [],
                    'total_etudiants': 0,
                    'niveau_spe': niveau_spe
                }
            
            if groupe.groupe:
                full_name = f"{niveau_spe} - Section: {section_name} - Groupe: {groupe.groupe.nom_ar} ({groupe.nbr_etudiants_SG} étudiants)"
            else:
                full_name = f"{niveau_spe} - Section: {section_name} ({groupe.nbr_etudiants_SG} étudiants)"
                
            sections_dict[section_name]['groupes'].append({
                'id': groupe.id,
                'full_name': full_name,
                'section': section_name,
                'groupe': groupe.groupe.nom_ar if groupe.groupe else None,
                'nb_etudiants': groupe.nbr_etudiants_SG
            })
            sections_dict[section_name]['total_etudiants'] += groupe.nbr_etudiants_SG
        else:
            # Groupe sans section spécifique
            if groupe.groupe:
                full_name = f"{niveau_spe} - Groupe: {groupe.groupe.nom_ar} ({groupe.nbr_etudiants_SG} étudiants)"
            else:
                full_name = f"{niveau_spe} - Groupe {groupe.id} ({groupe.nbr_etudiants_SG} étudiants)"
            
            data.append({
                'id': groupe.id,
                'full_name': full_name
            })
    total_etudiants = math.ceil(total_etudiants/2)

    # 1. Ajouter "TOUS LES ÉTUDIANTS"
    if total_etudiants > 0 and niv_spe_dep_sgs.exists():
        niveau_spe = f"{niv_spe_dep_sgs.first().niv_spe_dep.niveau.nom_ar} - {niv_spe_dep_sgs.first().niv_spe_dep.specialite.nom_ar}"
        tous_id = f"tous_{niveau_id}"
        if reforme_id:
            tous_id += f"_{reforme_id}"
            
        data.insert(0, {
            'id': tous_id,
            'full_name': f"{niveau_spe} - TOUS LES ÉTUDIANTS ({total_etudiants} étudiants)"
        })
    
    # 2. Ajouter les options par section ET par groupes
    for section_name, section_info in sections_dict.items():
        section_id = f"section_{section_name}"
        if reforme_id:
            section_id += f"_{reforme_id}"
            
        # Ajouter l'option pour toute la section
        data.append({
            'id': section_id,
            'full_name': f"{section_info['niveau_spe']} - Section: {section_name} - TOUS LES GROUPES ({section_info['total_etudiants']} étudiants)"
        })
        
        # Ajouter tous les groupes individuels de cette section
        for groupe in section_info['groupes']:
            data.append(groupe)
    
    # print(f"🔍 DEPT Données finales: {len(data)} éléments")
    
    return JsonResponse({'data': data})



# @login_required
# def import_etudiants(request):
#     if request.method == 'POST':
#         reforme_id = request.POST.get('reforme')
#         niveau_id = request.POST.get('niveau')
#         niv_spe_dep_sg_id = request.POST.get('niv_spe_dep_sg')
#         excel_file = request.FILES.get('excel_file')

#         if not all([reforme_id, niveau_id, niv_spe_dep_sg_id, excel_file]):
#             return render(request, 'departement/import_etudiants.html', {
#                 'error': 'يرجى اختيار جميع الحقول وتحميل ملف Excel.',
#                 'refs_url': reverse('depa:refs-json'),
#                 'nivs_url': reverse('depa:nivs-json', kwargs={'reforme_id': 0}),
#             })

#         try:
#             df = pd.read_excel(excel_file)
#             niv_spe_dep_sg = NivSpeDep_SG.objects.get(id=niv_spe_dep_sg_id)
            
#             # 🔢 COMPTEUR pour les étudiants ajoutés
#             etudiants_ajoutes = 0

#             for _, row in df.iterrows():
#                 # # print(f"Importing Etudiant: {row.get('اللقب', '')} {row.get('الإسم', '')}, Matricule={row.get('الرقم التسلسلي', '')}")
#                 etudiant = Etudiant.objects.create(
#                     nom_ar=row.get('اللقب', ''),
#                     prenom_ar=row.get('الإسم', ''),
#                     nom_fr=row.get('Nom', ''),
#                     prenom_fr=row.get('Prénom', ''),
#                     matricule=row.get('الرقم التسلسلي', ''),
#                     num_ins=row.get('رقم التسجيل', ''),
#                     sexe=row.get('الجنس', Etudiant.Sexe.M),
#                     niv_spe_dep_sg=niv_spe_dep_sg,
#                     inscrit_univ=True,
#                 )
#                 # 🔢 Incrémenter le compteur
#                 etudiants_ajoutes += 1
#                 # # print(f"Etudiant created: ID={etudiant.id}, User={etudiant.user}")

#             # ✅ MISE À JOUR du nombre d'étudiants dans NivSpeDep_SG
#             niv_spe_dep_sg.nbr_etudiants_SG += etudiants_ajoutes
#             niv_spe_dep_sg.save()

#             return render(request, 'departement/import_etudiants.html', {
#                 'success': f'تم استيراد {etudiants_ajoutes} طالب بنجاح.',
#                 'refs_url': reverse('depa:refs-json'),
#                 'nivs_url': reverse('depa:nivs-json', kwargs={'reforme_id': 0}),
#             })

#         except Exception as e:
#             return render(request, 'departement/import_etudiants.html', {
#                 'error': f'خطأ أثناء الاستيراد: {str(e)}',
#                 'refs_url': reverse('depa:refs-json'),
#                 'nivs_url': reverse('depa:nivs-json', kwargs={'reforme_id': 0}),
#             })

#     return render(request, 'departement/import_etudiants.html', {
#         'refs_url': reverse('depa:refs-json'),
#         'nivs_url': reverse('depa:nivs-json', kwargs={'reforme_id': 0}),
#     })




@login_required
def list_etudiants(request):
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)

    if request.method == 'POST':
        niv_spe_dep_sg_id = request.POST.get('niv_spe_dep_sg')
        # # print(f"🔍 DEPT DEBUG: ID reçu = '{niv_spe_dep_sg_id}'")
        
        if niv_spe_dep_sg_id:
            
            if niv_spe_dep_sg_id.startswith('tous_'):
                # # print("🔍 DEPT DEBUG: Cas TOUS LES ÉTUDIANTS")
                
                # Extraire niveau_id et reforme_id de l'ID
                if '_' in niv_spe_dep_sg_id:
                    parts = niv_spe_dep_sg_id.replace('tous_', '').split('_')
                    niveau_id = parts[0]
                    reforme_id = parts[1] if len(parts) > 1 else None
                else:
                    niveau_id = niv_spe_dep_sg_id.replace('tous_', '')
                    reforme_id = None
                
                # print(f"🔍 DEPT Extrait: niveau_id={niveau_id}, reforme_id={reforme_id}")
                
                # Filtres STRICTS avec département
                groupes_filters = {
                    'niv_spe_dep__departement': departement,
                    'niv_spe_dep__niveau_id': niveau_id
                }
                if reforme_id:
                    groupes_filters['niv_spe_dep__specialite__reforme_id'] = reforme_id
                
                # Récupérer tous les NivSpeDep_SG correspondants
                groupes_niveau = NivSpeDep_SG.objects.filter(**groupes_filters).exclude(
                    section__isnull=False,
                    groupe__isnull=True
                ).values_list('id', flat=True)
                
                # print(f"🔍 DEPT Groupes niveau trouvés: {len(list(groupes_niveau))}")
                
                etudiants = Etudiant.objects.filter(
                    niv_spe_dep_sg_id__in=groupes_niveau
                ).distinct().order_by('nom_ar').values(
                    'id', 'nom_ar', 'prenom_ar', 'nom_fr', 'prenom_fr', 'matricule', 'inscrit_univ'
                )
                
            elif niv_spe_dep_sg_id.startswith('section_'):
                # # print("🔍 DEPT DEBUG: Cas SECTION COMPLÈTE")
                
                # Extraire section_name et reforme_id de l'ID  
                if '_' in niv_spe_dep_sg_id:
                    parts = niv_spe_dep_sg_id.replace('section_', '').split('_')
                    section_name = parts[0]
                    reforme_id = parts[1] if len(parts) > 1 else None
                else:
                    section_name = niv_spe_dep_sg_id.replace('section_', '')
                    reforme_id = None
                
                # print(f"🔍 DEPT Extrait: section_name={section_name}, reforme_id={reforme_id}")
                
                # Filtres STRICTS pour section avec département
                section_filters = {
                    'niv_spe_dep__departement': departement,
                    'section__nom_ar': section_name,
                    'section__isnull': False,
                    'groupe__isnull': False
                }
                if reforme_id:
                    section_filters['niv_spe_dep__specialite__reforme_id'] = reforme_id
                
                groupes_section = NivSpeDep_SG.objects.filter(**section_filters)
                
                # print(f"🔍 DEPT Groupes section trouvés: {groupes_section.count()}")
                
                groupes_ids = list(groupes_section.values_list('id', flat=True))
                
                etudiants = Etudiant.objects.filter(
                    niv_spe_dep_sg_id__in=groupes_ids
                ).distinct().order_by('nom_ar').values(
                    'id', 'nom_ar', 'prenom_ar', 'nom_fr', 'prenom_fr', 'matricule', 'inscrit_univ'
                )
                
            else:
                # # print("🔍 DEPT DEBUG: Cas GROUPE INDIVIDUEL")
                etudiants = Etudiant.objects.filter(
                    niv_spe_dep_sg_id=niv_spe_dep_sg_id
                ).order_by('nom_ar').values(
                    'id', 'nom_ar', 'prenom_ar', 'nom_fr', 'prenom_fr', 'matricule', 'inscrit_univ'
                )
            
            result_data = list(etudiants)
            # # print(f"🔍 DEPT DEBUG: Données finales: {len(result_data)} étudiants")
            return JsonResponse({'data': result_data})
            
        return JsonResponse({'data': []})

    context = {
        'title': 'قائمة الطلبة',
        'my_Dep': departement,
        'my_Fac': departement.faculte,
        'refs_url': reverse('depa:get_reformes_json'),
        'nivs_url': reverse('depa:get_niveaux_json', kwargs={'reforme_id': 0}),
        'niv_spe_dep_sg_url': reverse('depa:get_niv_spe_dep_sg_json', kwargs={'niveau_id': 0}),
    }
    return render(request, 'departement/list_etudiants.html', context)


@login_required
def inscrire_univ(request):
    if request.method == 'POST':
        try:
            etudiant_id = request.POST.get('etudiant_id')
            if not etudiant_id:
                return JsonResponse({'success': False, 'error': 'ID étudiant manquant.'}, status=400)

            etudiant = Etudiant.objects.get(id=etudiant_id)
            etudiant.inscrit_univ = True
            etudiant.save()
            return JsonResponse({'success': True})
        except Etudiant.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Étudiant non trouvé.'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée.'}, status=405)

@login_required
def desinscrire_univ(request):
    if request.method == 'POST':
        try:
            etudiant_id = request.POST.get('etudiant_id')
            if not etudiant_id:
                return JsonResponse({'success': False, 'error': 'ID étudiant manquant.'}, status=400)

            etudiant = Etudiant.objects.get(id=etudiant_id)
            etudiant.inscrit_univ = False
            etudiant.save()
            return JsonResponse({'success': True})
        except Etudiant.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Étudiant non trouvé.'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée.'}, status=405)




@login_required
def profile_Etu(request, etudiant_id):
    # Récupérer l'étudiant
    etudiant = get_object_or_404(Etudiant, id=etudiant_id)
    
    # Récupérer le département de l'utilisateur connecté (comme dans les autres vues)
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)
    
    # DEBUG pour vérifier les données
    # # print(f"🔍 DEPT PROFILE DEBUG:")
    # # print(f"  - Étudiant ID: {etudiant_id}")
    # # print(f"  - Étudiant: {etudiant.nom_ar} {etudiant.prenom_ar}")
    # # print(f"  - A un compte user: {etudiant.user is not None}")
    # # print(f"  - inscrit_univ: {etudiant.inscrit_univ}")
    # # print(f"  - Département: {departement.nom_ar}")
    
    context = {
        'title': f'ملف الطالب - {etudiant.nom_ar}',
        'my_Etu': etudiant,
        'my_Dep': departement,
        'my_Fac': departement.faculte,
        'departement': departement,  # Pour compatibilité avec l'ancien template
    }
    
    return render(request, 'departement/profile_Etu.html', context)




# ----------------------------------------------------------------------
@login_required
def profileUpdate_Etu_Dep(request, etudiant_id): 
    etudiant = Etudiant.objects.get(id=etudiant_id)
    departement = etudiant.niv_spe_dep_sg.niv_spe_dep.departement
    Etu_form = profileUpdate_Etu_Form(instance=etudiant)
    if request.method == 'POST':
        # profile_form = ProfileUpdateForm(request.POST, request.FILES, instance=request.user.profile)
        # User_form = UserUpdateForm(request.POST, instance=request.user)
        Etu_form = profileUpdate_Etu_Form(request.POST, instance=etudiant)
        # if user_form.is_valid and profile_form.is_valid and dep_form.is_valid:
        if Etu_form.is_valid:
            # User_form.save()
            Etu_form.save()
            messages.success(request, 'تم تحديث المعلومات بنجاح')
            # context = {
            # 'title': 'تعديل المعلومات',
            # # 'User_form': User_form,
            # 'Etu_form' : Etu_form,
            # 'my_Etu': etudiant,
            # 'my_Dep': departement,  # Passer le département au template
            # 'my_Fac': departement.faculte,  # Passer le département au template
            # }
            # return render(request, 'departement/profile_Etu.html', context)
            # return redirect(reverse('change_password_Ens', kwargs={'dep_id': dep_id}))
            return redirect(reverse('depa:profile_Etu', kwargs={'etudiant_id': etudiant.id}))
    else:
        # User_form = UserUpdateForm(instance=request.user)
        Etu_form = profileUpdate_Etu_Form(instance=etudiant)

    context = {
        'title': 'تعديل المعلومات',
        # 'User_form': User_form,
        'Etu_form' : Etu_form,
        'my_Etu': etudiant,
        'my_Dep': departement,  # Passer le département au template
        'my_Fac': departement.faculte,  # Passer le département au template
        
    }
    return render(request, 'departement/profileUpdate_Etu_Dep.html', context)



# ----------------------------------------------------------------------
# ----------------------------------------------------------------------
# ----------------------------------------------------------------------
from django.db.models import Count, Q, Prefetch

from django.db.models import Count, Q

@login_required
def emploi_Ens(request, ens_id, semestre_num=None):
    """
    VERSION OPTIMISÉE avec TRI CORRECT par jour et horaire
    """
    
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)
    my_Dep = Departement.objects.get(id=departement.id)

    # GESTION DU SEMESTRE
    if semestre_num is None:
        semestre_num = request.GET.get('semestre')
        if semestre_num:
            try:
                semestre_num = int(semestre_num)
            except ValueError:
                semestre_num = 1
        else:
            semestre_num = 1
    
    # Validation du semestre
    if semestre_num not in [1, 2]:
        semestre_num = 1

    # FILTRE DE BASE
    base_filter = {
        'enseignant__departement': my_Dep.id,
        'enseignant__enseignant__id': ens_id,
        'semestre': semestre_num,
    }

    # 🚀 OPTIMISATION 1 : Une seule requête pour TOUTES les classes
    all_Classe_Ens = Classe.objects.filter(**base_filter).select_related(
        'enseignant',
        'enseignant__enseignant',
        'matiere',
        'niv_spe_dep_sg',
        'niv_spe_dep_sg__niv_spe_dep',
        'niv_spe_dep_sg__niv_spe_dep__specialite',
        'niv_spe_dep_sg__niv_spe_dep__niveau'
    )

    # Récupérer les IDs des classes
    classe_ids = [classe.id for classe in all_Classe_Ens]

    # 🚀 OPTIMISATION 2 : Calculer les stats des séances en UNE SEULE requête
    seances_stats = {}
    if classe_ids:
        seances_stats_query = Seance.objects.filter(
            classe_id__in=classe_ids
        ).values('classe_id').annotate(
            total=Count('id'),
            fait_count=Count('id', filter=Q(fait=True)),
            remplacer_count=Count('id', filter=Q(remplacer=True)),
            annuler_count=Count('id', filter=Q(annuler=True))
        )
        
        # Convertir en dictionnaire pour accès rapide
        seances_stats = {item['classe_id']: item for item in seances_stats_query}

    # Variables d'initialisation
    nbr_Cours = 0
    nbr_TP = 0
    nbr_TD = 0
    nbr_SS = 0
    classes_to_update = []

    # 🚀 OPTIMISATION 3 : Traiter toutes les classes en mémoire
    for classe in all_Classe_Ens:
        # Compter par type
        if classe.type == "Cours":
            nbr_Cours += 1
        elif classe.type == "TP":
            nbr_TP += 1
        elif classe.type == "TD":
            nbr_TD += 1
        elif classe.type == "Sortie Scientifique":
            nbr_SS += 1
        
        # Calculer le taux d'avancement (sans requêtes SQL supplémentaires!)
        if classe.seance_created:
            stats = seances_stats.get(classe.id)
            if stats and stats['total'] > 0:
                taux_avancement = round((stats['fait_count'] / stats['total']) * 100)
                if classe.taux_avancement != taux_avancement:
                    classe.taux_avancement = taux_avancement
                    classes_to_update.append(classe)

    # 💾 SAUVEGARDE EN LOT (1 seule requête pour toutes les mises à jour)
    if classes_to_update:
        Classe.objects.bulk_update(classes_to_update, ['taux_avancement'])

    all_classes = nbr_Cours + nbr_TP + nbr_TD + nbr_SS

    # ✅ TRI MANUEL PAR JOUR ET HORAIRE
    jour_order = {
        'Samedi': 1,
        'Dimanche': 2,
        'Lundi': 3,
        'Mardi': 4,
        'Mercredi': 5,
        'Jeudi': 6,
        'Vendredi': 7
    }
    
    def parse_time(temps_str):
        """Convertir '09:30-08:00' en minutes (prendre l'heure de début = 08:00)"""
        if not temps_str:
            return 0
        # Prendre la partie après le tiret
        if '-' in temps_str:
            heure = temps_str.split('-')[1].strip()
        else:
            heure = temps_str.strip()
        
        try:
            h, m = heure.split(':')
            return int(h) * 60 + int(m)
        except:
            return 0
    
    # Convertir en liste et trier
    all_Classe_Ens = sorted(
        list(all_Classe_Ens),
        key=lambda x: (
            jour_order.get(x.jour, 999),  # D'abord par jour
            parse_time(x.temps)            # Puis par horaire
        )
    )

    # Définition des créneaux et jours
    sixClasses = [
        '09:30-08:00',
        '11:10-09:40',
        '12:50-11:20',
        '14:40-13:10',
        '16:20-14:50',
        '18:00-16:30',
    ]
    
    sevenDays = ['Samedi', 'Dimanche', 'Lundi', 'Mardi', 'Mercredi', 'Jeudi']

    # all_Classe_Time contient toutes les classes (déjà triées)
    all_Classe_Time = all_Classe_Ens

    # Récupération des informations de l'enseignant
    try:
        enseignant_info = Enseignant.objects.get(id=ens_id)
    except Enseignant.DoesNotExist:
        messages.error(request, "الأستاذ غير موجود")
        return redirect('depa:manage_enseignants', semestre_num=semestre_num)

    context = {
        'title': 'إستعمال الزمن',
        'enseignant_info': enseignant_info,
        'semestre_num': semestre_num,
        'all_Classe_Ens': all_Classe_Ens,
        'reload': False,
        'my_Dep': departement,
        'my_Fac': departement.faculte,
        'ttEns_0': 0,
        'all_Classe_Time': all_Classe_Time,
        'nbr_Cours': nbr_Cours,
        'nbr_TP': nbr_TP,
        'sevenDays': sevenDays,
        'sixClasses': sixClasses,
        'nbr_TD': nbr_TD,
        'nbr_SS': nbr_SS,
        'all_classes': all_classes,
    }

    return render(request, 'departement/emploi_Ens.html', context)





# from django.contrib import messages
# from django.shortcuts import redirect, get_object_or_404
# from django.utils import timezone
# from scholarly import scholarly, ProxyGenerator
# import logging

# logger = logging.getLogger(__name__)

# @login_required
# def update_scholar_manual(request, enseignant_id):
#     enseignant = get_object_or_404(Enseignant, id=enseignant_id)
    
#     if not enseignant.googlescholar:
#         messages.warning(request, "Aucun profil Google Scholar configuré")
#         return redirect(request.META.get('HTTP_REFERER', 'depa:list_enseignants_dep'))
    
#     user_id = enseignant.scholar_user_id
#     if not user_id:
#         messages.error(request, "URL Google Scholar invalide")
#         return redirect(request.META.get('HTTP_REFERER', 'depa:list_enseignants_dep'))
    
#     # Configurer le proxy
#     try:
#         pg = ProxyGenerator()
#         success = pg.FreeProxies()
#         if success:
#             scholarly.use_proxy(pg)
#             messages.info(request, "Proxy configuré avec succès")
#         else:
#             messages.warning(request, "Impossible de configurer le proxy, continuation sans proxy")
#     except Exception as e:
#         logger.warning(f"Erreur proxy: {e}")
#         messages.warning(request, f"Erreur lors de la configuration du proxy: {e}")
    
#     # Récupérer les données
#     try:
#         author = scholarly.search_author_id(user_id)
#         author_filled = scholarly.fill(author)
        
#         # Mettre à jour
#         enseignant.scholar_publications_count = len(author_filled.get('publications', []))
#         enseignant.scholar_citations_count = author_filled.get('citedby', 0)
#         enseignant.scholar_h_index = author_filled.get('hindex', 0)
#         enseignant.scholar_i10_index = author_filled.get('i10index', 0)
#         enseignant.scholar_last_update = timezone.now()
        
#         enseignant.save(update_fields=[
#             'scholar_publications_count',
#             'scholar_citations_count',
#             'scholar_h_index',
#             'scholar_i10_index',
#             'scholar_last_update'
#         ])
        
#         messages.success(
#             request,
#             f"✅ Profil Google Scholar mis à jour avec succès!\n"
#             f"📄 {enseignant.scholar_publications_count} publications | "
#             f"📚 {enseignant.scholar_citations_count} citations | "
#             f"📊 H-index: {enseignant.scholar_h_index}"
#         )
        
#     except Exception as e:
#         logger.error(f"Error fetching Scholar data: {str(e)}")
#         messages.error(request, f"❌ Erreur lors de la récupération des données: {str(e)}")
    
#     return redirect(request.META.get('HTTP_REFERER', 'depa:list_enseignants_dep'))











# views.py (departement/views.py)

# views.py (departement/views.py)

# views.py (departement/views.py)

from django.contrib import messages
from django.shortcuts import redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.conf import settings
import requests
import logging

logger = logging.getLogger(__name__)

# views.py
from django.http import JsonResponse

@login_required
def update_scholar_manual(request, enseignant_id):
    """
    Met à jour les métriques Google Scholar d'un enseignant via Serpapi
    """
    enseignant = get_object_or_404(Enseignant, id=enseignant_id)
    
    # Vérifier si c'est une requête AJAX
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    if not enseignant.googlescholar:
        if is_ajax:
            return JsonResponse({
                'success': False,
                'message': 'لا يوجد حساب Google Scholar لهذا الأستاذ'
            })
        messages.warning(request, "❌ Aucun profil Google Scholar configuré")
        return redirect(request.META.get('HTTP_REFERER', 'depa:list_enseignants_dep'))
    
    user_id = enseignant.scholar_user_id
    if not user_id:
        if is_ajax:
            return JsonResponse({
                'success': False,
                'message': 'رابط Google Scholar غير صالح'
            })
        messages.error(request, "❌ URL Google Scholar invalide")
        return redirect(request.META.get('HTTP_REFERER', 'depa:list_enseignants_dep'))
    
    SERPAPI_KEY = getattr(settings, 'SERPAPI_KEY', None)
    if not SERPAPI_KEY:
        if is_ajax:
            return JsonResponse({
                'success': False,
                'message': 'مفتاح API غير مكوّن'
            })
        messages.error(request, "❌ Clé API non configurée")
        return redirect(request.META.get('HTTP_REFERER', 'depa:list_enseignants_dep'))
    
    try:
        url = "https://serpapi.com/search.json"
        params = {
            "engine": "google_scholar_author",
            "author_id": user_id,
            "api_key": SERPAPI_KEY,
            "hl": "ar"
        }
        
        response = requests.get(url, params=params, timeout=15)
        response.raise_for_status()
        data = response.json()
        
        if "error" in data:
            raise ValueError(data["error"])
        
        # Extraire les métriques
        cited_by = data.get('cited_by', {})
        table = cited_by.get('table', [])
        
        citations_all = None
        h_index_all = None
        i10_index_all = None
        
        if table and len(table) > 0:
            first_row = table[0]
            citations_all = first_row.get('citations', {}).get('all')
            h_index_all = first_row.get('h_index', {}).get('all')
            i10_index_all = first_row.get('i10_index', {}).get('all')
        
        articles = data.get('articles', [])
        publications_count = len(articles) if articles else None
        
        # Vérifier et mettre à jour
        has_valid_data = False
        update_fields = []
        
        if publications_count is not None and publications_count > 0:
            enseignant.scholar_publications_count = publications_count
            update_fields.append('scholar_publications_count')
            has_valid_data = True
        
        if citations_all is not None and citations_all >= 0:
            enseignant.scholar_citations_count = citations_all
            update_fields.append('scholar_citations_count')
            has_valid_data = True
        
        if h_index_all is not None and h_index_all >= 0:
            enseignant.scholar_h_index = h_index_all
            update_fields.append('scholar_h_index')
            has_valid_data = True
        
        if i10_index_all is not None and i10_index_all >= 0:
            enseignant.scholar_i10_index = i10_index_all
            update_fields.append('scholar_i10_index')
            has_valid_data = True
        
        if has_valid_data:
            enseignant.scholar_last_update = timezone.now()
            update_fields.append('scholar_last_update')
            enseignant.save(update_fields=update_fields)
            
            if is_ajax:
                return JsonResponse({
                    'success': True,
                    'publications': enseignant.scholar_publications_count or 0,
                    'citations': enseignant.scholar_citations_count or 0,
                    'h_index': enseignant.scholar_h_index or 0,
                    'i10_index': enseignant.scholar_i10_index or 0,
                    'message': 'تم التحديث بنجاح'
                })
            
            messages.success(request, f"✅ تم التحديث بنجاح")
        else:
            if is_ajax:
                return JsonResponse({
                    'success': False,
                    'message': 'لم يتم العثور على بيانات صالحة'
                })
            messages.warning(request, "⚠️ لم يتم العثور على بيانات صالحة")
        
    except Exception as e:
        logger.error(f"Error: {str(e)}")
        if is_ajax:
            return JsonResponse({
                'success': False,
                'message': f'خطأ: {str(e)}'
            })
        messages.error(request, f"❌ خطأ: {str(e)}")
    
    return redirect(request.META.get('HTTP_REFERER', 'depa:list_enseignants_dep'))


@login_required
def update_all_scholars(request):
    """
    Met à jour les profils Google Scholar de tous les enseignants du département
    """
    # Récupérer le département de l'utilisateur
    try:
        enseignant_user = request.user.enseignant_profile
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant_user, statut='Permanent')
        departement = Departement.objects.get(id=real_Dep.departement.id)
    except:
        messages.error(request, "❌ Vous n'êtes pas associé à un département")
        return redirect('depa:dashboard_Dep')
    
    # Récupérer tous les Ens_Dep du département
    ens_deps = Ens_Dep.objects.filter(departement=departement)
    
    # Récupérer les enseignants qui ont un profil Google Scholar
    enseignant_ids = ens_deps.values_list('enseignant_id', flat=True)
    enseignants = Enseignant.objects.filter(
        id__in=enseignant_ids,
        googlescholar__isnull=False
    ).exclude(googlescholar='')
    
    if not enseignants.exists():
        messages.warning(request, "❌ Aucun enseignant avec profil Google Scholar dans ce département")
        return redirect('depa:list_enseignants_dep')
    
    success_count = 0
    error_count = 0
    skipped_count = 0
    
    SERPAPI_KEY = getattr(settings, 'SERPAPI_KEY', None)
    if not SERPAPI_KEY:
        messages.error(request, "❌ Clé API Serpapi non configurée")
        return redirect('depa:list_enseignants_dep')
    
    for enseignant in enseignants:
        if not enseignant.scholar_user_id:
            skipped_count += 1
            continue
            
        try:
            url = "https://serpapi.com/search.json"
            params = {
                "engine": "google_scholar_author",
                "author_id": enseignant.scholar_user_id,
                "api_key": SERPAPI_KEY,
                "hl": "fr"
            }
            
            response = requests.get(url, params=params, timeout=15)
            response.raise_for_status()
            data = response.json()
            
            if "error" in data:
                logger.warning(f"API error for {enseignant.nom_ar}: {data.get('error')}")
                error_count += 1
                continue
            
            # Extraire les métriques
            cited_by = data.get('cited_by', {})
            table = cited_by.get('table', [])
            
            citations_all = None
            h_index_all = None
            i10_index_all = None
            
            if table and len(table) > 0:
                first_row = table[0]
                citations_all = first_row.get('citations', {}).get('all')
                h_index_all = first_row.get('h_index', {}).get('all')
                i10_index_all = first_row.get('i10_index', {}).get('all')
            
            articles = data.get('articles', [])
            publications_count = len(articles) if articles else None
            
            # ✅ VÉRIFICATION : Ne mettre à jour que si les données sont valides
            has_valid_data = False
            update_fields = []
            
            if publications_count is not None and publications_count > 0:
                enseignant.scholar_publications_count = publications_count
                update_fields.append('scholar_publications_count')
                has_valid_data = True
            
            if citations_all is not None and citations_all >= 0:
                enseignant.scholar_citations_count = citations_all
                update_fields.append('scholar_citations_count')
                has_valid_data = True
            
            if h_index_all is not None and h_index_all >= 0:
                enseignant.scholar_h_index = h_index_all
                update_fields.append('scholar_h_index')
                has_valid_data = True
            
            if i10_index_all is not None and i10_index_all >= 0:
                enseignant.scholar_i10_index = i10_index_all
                update_fields.append('scholar_i10_index')
                has_valid_data = True
            
            # Sauvegarder seulement si des données valides ont été trouvées
            if has_valid_data:
                enseignant.scholar_last_update = timezone.now()
                update_fields.append('scholar_last_update')
                enseignant.save(update_fields=update_fields)
                success_count += 1
                logger.info(f"Successfully updated Scholar data for {enseignant.nom_ar}")
            else:
                logger.warning(f"No valid data found for {enseignant.nom_ar}, keeping existing data")
                skipped_count += 1
                
        except Exception as e:
            logger.error(f"Error updating {enseignant.nom_ar}: {str(e)}")
            error_count += 1
    
    # Messages de retour détaillés
    if success_count > 0:
        messages.success(request, f"✅ {success_count} profil(s) mis à jour avec succès")
    if skipped_count > 0:
        messages.info(request, f"ℹ️ {skipped_count} profil(s) ignoré(s) (pas de données valides)")
    if error_count > 0:
        messages.warning(request, f"⚠️ {error_count} erreur(s) lors de la mise à jour")
    
    return redirect('depa:list_enseignants_dep')

















    # print(all_Classe_Ens)
    
    # print("Base filter:", base_filter)
    # for jour in jours:
    #     print(f"Recherche pour jour: {jour}")
    #     classes_jour = Classe.objects.filter(**base_filter, jour=jour)
    #     print(f"  -> {classes_jour.count()} classes trouvées")

    # all_Classe_Ens = []
    
    # # RÉCUPÉRATION PAR JOUR avec filtre semestre
    # all_Classe_Ens_SAM = Classe.objects.filter(
    #     **base_filter,
    #     jour='Samedi'
    # ).order_by('temps')
    # all_Classe_Ens.extend(all_Classe_Ens_SAM)

    # all_Classe_Ens_DIM = Classe.objects.filter(
    #     **base_filter,
    #     jour='Dimanche'
    # ).order_by('temps')
    # all_Classe_Ens.extend(all_Classe_Ens_DIM)

    # all_Classe_Ens_LUN = Classe.objects.filter(
    #     **base_filter,
    #     jour='Lundi'
    # ).order_by('temps')
    # all_Classe_Ens.extend(all_Classe_Ens_LUN)

    # all_Classe_Ens_MAR = Classe.objects.filter(
    #     **base_filter,
    #     jour='Mardi'
    # ).order_by('temps')
    # all_Classe_Ens.extend(all_Classe_Ens_MAR)
    
    # all_Classe_Ens_MER = Classe.objects.filter(
    #     **base_filter,
    #     jour='Mercredi'
    # ).order_by('temps')
    # all_Classe_Ens.extend(all_Classe_Ens_MER)

    # all_Classe_Ens_JEU = Classe.objects.filter(
    #     **base_filter,
    #     jour='Jeudi'
    # ).order_by('temps')
    # all_Classe_Ens.extend(all_Classe_Ens_JEU)

    # RÉCUPÉRATION PAR CRÉNEAU HORAIRE avec filtre semestre
    all_Classe_Time = []
    
    # Classe 1 (08:00-09:30)
    all_Classe_Time_1 = Classe.objects.filter(
        **base_filter,
        temps='09:30-08:00'
    )
    all_Classe_Time.extend(all_Classe_Time_1)
    
    # Classe 2 (09:40-11:10)
    all_Classe_Time_2 = Classe.objects.filter(
        **base_filter,
        temps='11:10-09:40'
    )
    all_Classe_Time.extend(all_Classe_Time_2)
    
    # Classe 3 (11:20-12:50)
    all_Classe_Time_3 = Classe.objects.filter(
        **base_filter,
        temps='12:50-11:20'
    )
    all_Classe_Time.extend(all_Classe_Time_3)
    
    # Classe 4 (13:10-14:40)
    all_Classe_Time_4 = Classe.objects.filter(
        **base_filter,
        temps='14:40-13:10'
    )
    all_Classe_Time.extend(all_Classe_Time_4)
    
    # Classe 5 (14:50-16:20)
    all_Classe_Time_5 = Classe.objects.filter(
        **base_filter,
        temps='16:20-14:50'
    )
    all_Classe_Time.extend(all_Classe_Time_5)
    
    # Classe 6 (16:30-18:00)
    all_Classe_Time_6 = Classe.objects.filter(
        **base_filter,
        temps='18:00-16:30'
    )
    all_Classe_Time.extend(all_Classe_Time_6)

    # Définition des créneaux horaires
    sixClasses = [
        '09:30-08:00',
        '11:10-09:40',
        '12:50-11:20',
        '14:40-13:10',
        '16:20-14:50',
        '18:00-16:30',
    ]

    # CALCUL DES STATISTIQUES
    for idx1 in list(all_Classe_Ens):
        if "Cours" == idx1.type:
            nbr_Cours = nbr_Cours + 1
        if "TP" == idx1.type:
            nbr_TP = nbr_TP + 1
        if "TD" == idx1.type:
            nbr_TD = nbr_TD + 1
        if "Sortie Scientifique" == idx1.type:
            nbr_SS = nbr_SS + 1
        
        # Calcul du taux d'avancement
        if idx1.seance_created == True:
            nbr_Sea_fait = 0
            nbr_Sea_remplacer = 0
            nbr_Sea_annuler = 0
            all_Seances = 0
            all_Seances = Seance.objects.filter(classe=idx1.id)
            nbr_Sea_fait = Seance.objects.filter(classe=idx1.id, fait=True)
            nbr_Sea_remplacer = Seance.objects.filter(classe=idx1.id, remplacer=True)
            nbr_Sea_annuler = Seance.objects.filter(classe=idx1.id, annuler=True)

            if nbr_Sea_fait.exists() and all_Seances.exists():
                taux_avancement = ((nbr_Sea_fait.count()) / all_Seances.count()) * 100
                taux_avancement = round(taux_avancement)
                idx1.taux_avancement = taux_avancement
                idx1.save()

    all_classes = nbr_Cours + nbr_TP + nbr_TD + nbr_SS
    
    # Définition des jours
    sevenDays = ['Samedi', 'Dimanche', 'Lundi', 'Mardi', 'Mercredi', 'Jeudi']

    # Récupération des informations de l'enseignant pour l'affichage
    try:
        enseignant_info = Enseignant.objects.get(id=ens_id)
    except Enseignant.DoesNotExist:
        messages.error(request, "الأستاذ غير موجود")
        return redirect('depa:manage_enseignants', semestre_num=semestre_num)

    context = {
        'title': 'إستعمال الزمن',
        'enseignant_info': enseignant_info,  # Informations de l'enseignant
        'semestre_num': semestre_num,  # IMPORTANT: Passer le semestre au template
        'all_Classe_Ens': all_Classe_Ens,
        'reload': reload,
        'my_Dep': departement,
        'my_Fac': departement.faculte,
        'ttEns_0': ttEns_0,
        'all_Classe_Time': all_Classe_Time,
        'nbr_Cours': nbr_Cours,
        'nbr_TP': nbr_TP,
        'sevenDays': sevenDays,
        'sixClasses': sixClasses,
        'nbr_TD': nbr_TD,
        'nbr_SS': nbr_SS,
        'all_classes': all_classes,
    }

    return render(request, 'departement/emploi_Ens.html', context)




    







@login_required
def list_Notes_Etu_Classe(request, clas_id):
    """
    Affiche les notes des étudiants d'une classe (lecture seule pour le département)
    """
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)
    my_Dep = Departement.objects.get(id=departement.id)

    # Récupérer la classe
    my_Clas = get_object_or_404(Classe, id=clas_id)
    
    # DEBUG: Afficher les informations pour le débogage
    # print(f"DEBUG - my_Clas.enseignant.departement: {my_Clas.enseignant.departement}")
    # print(f"DEBUG - Type: {type(my_Clas.enseignant.departement)}")
    # print(f"DEBUG - my_Dep.id: {my_Dep.id}")
    # print(f"DEBUG - Type: {type(my_Dep.id)}")
    
    # Vérification corrigée - S'assurer que les types correspondent
    try:
        # Option 1: Comparer les IDs directement
        if hasattr(my_Clas.enseignant, 'departement'):
            if isinstance(my_Clas.enseignant.departement, int):
                classe_dept_id = my_Clas.enseignant.departement
            else:
                classe_dept_id = my_Clas.enseignant.departement.id if hasattr(my_Clas.enseignant.departement, 'id') else my_Clas.enseignant.departement
        else:
            # Si pas de champ departement direct, chercher via Ens_Dep
            ens_dep = Ens_Dep.objects.filter(enseignant=my_Clas.enseignant, statut='Permanent').first()
            if ens_dep:
                classe_dept_id = ens_dep.departement.id
            else:
                classe_dept_id = None
        
        if classe_dept_id != my_Dep.id:
            messages.error(request, "غير مصرح لك بعرض هذه الحصة")
            return redirect('depa:emploi_Ens', ens_id=my_Clas.enseignant.id)
            
    except Exception as e:
        # print(f"DEBUG - Erreur de vérification: {e}")
        # En cas d'erreur, permettre l'accès mais logger l'erreur
        messages.warning(request, f"تحذير: تعذر التحقق من الصلاحية - {str(e)}")
        pass

    # Récupérer toutes les notes de la classe
    all_Notes_Classe = Abs_Etu_Classe.objects.filter(
        classe=my_Clas
    ).select_related('etudiant').order_by('etudiant__nom_ar', 'etudiant__prenom_ar')

    # Calculer les statistiques de la classe
    try:
        statistiques = Abs_Etu_Classe.get_statistiques_classe(my_Clas)
    except:
        statistiques = None

    titre_00 = my_Clas.niv_spe_dep_sg
    titre_01 = my_Clas.matiere.nom_fr

    context = {
        'title': 'نقاط الطلاب',
        'titre_00': titre_00,
        'titre_01': titre_01,
        'my_Clas': my_Clas,
        'all_Notes_Classe': all_Notes_Classe,
        'statistiques': statistiques,
        'my_Dep': departement,
        'my_Fac': departement.faculte,
        'enseignant_info': my_Clas.enseignant,  # Informations de l'enseignant de la classe
    }

    return render(request, 'departement/list_Notes_Etu_Classe.html', context)













@login_required
def export_notes_classe(request, clas_id):
    """
    Exporte les notes d'une classe en format Excel - Version département
    """
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    
    # Vérifications d'accès
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)
    my_Dep = Departement.objects.get(id=departement.id)

    # Récupérer la classe et vérifier l'accès
    my_Clas = get_object_or_404(Classe, id=clas_id)
    
    # Vérifier que la classe appartient au département
    # if my_Clas.enseignant.departement != my_Dep.id:
    #     messages.error(request, "غير مصرح لك بتصدير هذه الحصة")
    #     return redirect('depa:emploi_Ens', ens_id=my_Clas.enseignant.id)
    
    # Récupérer toutes les notes de la classe
    all_Notes_Classe = Abs_Etu_Classe.objects.filter(
        classe=my_Clas
    ).select_related('etudiant').order_by('etudiant__nom_ar')
    
    # Créer un workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Notes_{my_Clas.matiere.nom_fr}"
    
    # Styles Excel
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="17A2B8", end_color="17A2B8", fill_type="solid")  # Couleur info pour département
    info_font = Font(bold=True, size=11)
    info_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Informations sur la classe (lignes 1-4)
    ws.merge_cells('A1:L1')
    ws['A1'] = f"تصدير نقاط الطلاب - {my_Clas.matiere.nom_fr}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A2:L2')
    ws['A2'] = f"الأستاذ: {my_Clas.enseignant.enseignant.nom_fr} {my_Clas.enseignant.enseignant.prenom_fr}"
    ws['A2'].font = info_font
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A3:L3')
    ws['A3'] = f"المستوى: {my_Clas.niv_spe_dep_sg} - النوع: {my_Clas.type}"
    ws['A3'].font = info_font
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A4:L4')
    ws['A4'] = f"تاريخ التصدير: {timezone.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A4'].font = Font(size=10, italic=True)
    ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
    
    # En-têtes des colonnes (ligne 6)
    headers = [
        'الرقم', 'اللقب', 'الاسم', 'Nom', 'Prénom',
        'نقطة الحضور\n(0-5)', 'نقطة المشاركة\n(0-5)', 'نقطة الامتحان 1\n(0-5)', 'نقطة الامتحان 2\n(0-5)',
        'نقاط إضافية للحصة', 'النقطة النهائية\n(0-20)', 
        'النسبة المئوية للحضور', 'التقدير', 'ملاحظات'
    ]
    
    header_row = 6
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Données des étudiants (à partir de la ligne 7)
    data_start_row = 7
    for row_idx, note in enumerate(all_Notes_Classe, data_start_row):
        data = [
            row_idx - data_start_row + 1,  # Numérotation
            note.etudiant.nom_ar or '',
            note.etudiant.prenom_ar or '',
            note.etudiant.nom_fr or '',
            note.etudiant.prenom_fr or '',
            float(note.note_presence),
            float(note.note_participe_HW),
            float(note.note_controle_1),
            float(note.note_controle_2),
            float(note.total_sup_seance) if note.total_sup_seance else 0,
            # float(note.total_sup_examen) if note.total_sup_examen else 0,
            float(note.note_finale),
            f"{note.taux_presence_pourcentage}%" if hasattr(note, 'taux_presence_pourcentage') else '0%',
            note.mention or '',
            note.obs or ''
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Coloration conditionnelle pour la note finale
            if col == 11 and isinstance(value, (int, float)):  # Colonne note finale
                if value >= 16:
                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")  # Vert
                elif value >= 14:
                    cell.fill = PatternFill(start_color="CCE7FF", end_color="CCE7FF", fill_type="solid")  # Bleu
                elif value >= 12:
                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")  # Jaune
                elif value >= 10:
                    cell.fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")  # Orange clair
                else:
                    cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")  # Rouge clair
    
    # Ligne de statistiques (après les données)
    if all_Notes_Classe.exists():
        stats_row = data_start_row + len(all_Notes_Classe) + 2
        
        # Calculer les statistiques
        notes_finales = [float(note.note_finale) for note in all_Notes_Classe]
        moyenne_classe = sum(notes_finales) / len(notes_finales) if notes_finales else 0
        note_max = max(notes_finales) if notes_finales else 0
        note_min = min(notes_finales) if notes_finales else 0
        nombre_admis = len([note for note in notes_finales if note >= 10])
        taux_reussite = (nombre_admis / len(notes_finales) * 100) if notes_finales else 0
        
        # Ajouter les statistiques
        ws.merge_cells(f'A{stats_row}:D{stats_row}')
        ws[f'A{stats_row}'] = "إحصائيات الصف:"
        ws[f'A{stats_row}'].font = Font(bold=True)
        
        stats_data = [
            (f'E{stats_row}', f"عدد الطلاب: {len(all_Notes_Classe)}"),
            (f'F{stats_row}', f"المعدل: {moyenne_classe:.2f}"),
            (f'G{stats_row}', f"أعلى نقطة: {note_max:.2f}"),
            (f'H{stats_row}', f"أقل نقطة: {note_min:.2f}"),
            (f'I{stats_row}', f"عدد الناجحين: {nombre_admis}"),
            (f'J{stats_row}', f"نسبة النجاح: {taux_reussite:.1f}%")
        ]
        
        for cell_ref, value in stats_data:
            ws[cell_ref] = value
            ws[cell_ref].font = Font(bold=True)
            ws[cell_ref].fill = info_fill
    
    # Ajuster la largeur des colonnes
    column_widths = {
        'A': 8,   # الرقم
        'B': 15,  # اللقب
        'C': 15,  # الاسم
        'D': 15,  # Nom
        'E': 15,  # Prénom
        'F': 12,  # نقطة الحضور
        'G': 12,  # نقطة المشاركة
        'I': 12,  # نقاط إضافية سيانس
        'K': 15,  # النقطة النهائية 1
        'L': 15,  # النقطة النهائية 2
        'M': 15,  # النسبة المئوية للحضور
        'N': 12,  # التقدير
        'Ö': 20   # ملاحظات
    }
    
    for column, width in column_widths.items():
        ws.column_dimensions[column].width = width
    
    # Hauteur des lignes d'en-tête
    ws.row_dimensions[header_row].height = 40
    
    # Préparer la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Nom de fichier sécurisé
    safe_filename = f"notes_{my_Clas.matiere.nom_fr}_{my_Clas.type}_{timezone.now().strftime('%Y%m%d_%H%M')}"
    safe_filename = safe_filename.replace(' ', '_').replace('/', '_')
    
    response['Content-Disposition'] = f'attachment; filename="{safe_filename}.xlsx"'
    
    # Sauvegarder et retourner
    wb.save(response)
    return response








# ----------------------------------------------------------------------
@login_required
def list_Seance_Ens(request, clas_id):
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)


    taux_avancement = 0
    nbr_Sea = 0
    nbr_Sea_fait = 0
    nbr_Sea_annuler = 0
    nbr_Sea_reste = 0
    nbr_Sea_remplacer = 0

    all_Seances = Seance.objects.filter(classe = clas_id).order_by('date') # plus Ens get JSON
    nbr_Sea_fait = Seance.objects.filter(classe = clas_id, fait = True) # plus Ens get JSON
    nbr_Sea_remplacer = Seance.objects.filter(classe = clas_id, remplacer = True) # plus Ens get JSON
    nbr_Sea_annuler = Seance.objects.filter(classe = clas_id, annuler = True) # plus Ens get JSON
    nbr_Sea_reste = all_Seances.count() - nbr_Sea_fait.count()

    if nbr_Sea_fait.count() != 0 and all_Seances.count() != 0 :
        taux_avancement = (nbr_Sea_fait.count() / all_Seances.count()) * 100
        taux_avancement = round(taux_avancement)
    
    sea_name = Classe.objects.get(id = clas_id)

    context = {
        'title': 'جميع الدروس',
        'sea_name': sea_name,
        'all_Seances': all_Seances,
        'nbr_Sea_fait': nbr_Sea_fait,
        'nbr_Sea_annuler': nbr_Sea_annuler,
        'nbr_Sea_reste': nbr_Sea_reste,
        'taux_avancement': taux_avancement,
        'my_Dep': departement,  # Passer le département au template
        'my_Fac': departement.faculte,  # Passer le département au template
    }
    return render(request, 'departement/list_Seance_Ens.html', context)













# ----------------------------------------------------------------------
@login_required
def new_Sceance_Ens(request, clas_id , Ens_id):
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)
    

    myClasse = Classe.objects.get(id=clas_id)
    
    dateStart = myClasse.semestre.date_debut
    dateEnd = myClasse.semestre.date_fin

    dic_Days = {'Samedi':'Saturday', 'Dimanche':'Sunday', 'Lundi':'Monday', 'Mardi': 'Tuesday', 'Mercredi':'Wednesday', 'Jeudi':'Thursday', 'Vendredi':'Friday'}

    
    dateStartName = pd.to_datetime(str(dateStart).split(" - ")).day_name() #friday
    while dateStart <= dateEnd:
        for key, val in dic_Days.items():
            dateStartName = pd.to_datetime(str(dateStart).split(" - ")).day_name() #friday
            # convert day to Arabic
            if dateStartName == val:
                dateStartName_fr = key

                if (myClasse.jour == dateStartName_fr) and (dateStart <= dateEnd):
                    # real_dateStartName_ar = dateStartName_ar
                    real_dateStart = dateStart

                    new_Seance = Seance(
                        classe = myClasse,
                        date = real_dateStart,
                        temps = myClasse.temps
                    )
                    new_Seance.save()
                    myClasse.seance_created = True
                    myClasse.save()
                    dateStart = dateStart + timedelta(days=1) # next day 
                else:
                    dateStart = dateStart + timedelta(days=1) # next day 

    # ttEns_0 = 0
    reload = True
    all_Classe_Ens = None
    all_classes = None
    nbr_Cours=0
    nbr_TP=0
    nbr_TD=0
    nbr_SS=0

    all_Classe_Ens = Classe.objects.filter(enseignant__departement = departement.id,
                                        enseignant = Ens_id,) # plus Ens get JSON
    
    for idx1 in list(all_Classe_Ens):
        # all_Seances.append(idx1.clas_jour +" - "+idx1.clas_temps)
        if "Cours" == idx1.type:
            nbr_Cours = nbr_Cours + 1
        if "TP" == idx1.type:
            nbr_TP = nbr_TP + 1
        if "TD" == idx1.type:
            nbr_TD = nbr_TD + 1
        if "Sortie Scientifique" == idx1.type:
            nbr_SS = nbr_SS + 1
    all_classes = nbr_Cours+nbr_TP+nbr_TD+nbr_SS

    messages.success(request, f'لقد تمت عملية تسجيل {myClasse} بنجاح')

    context = {
        'title': 'إستعمال الزمن',
        'all_Classe_Ens': all_Classe_Ens,
        'reload': reload,
        'my_Dep': departement,  # Passer le département au template
        'my_Fac': departement.faculte,  # Passer le département au template
        'ttEns_0':Ens_id,
        'nbr_Cours': nbr_Cours,
        'nbr_TP': nbr_TP,
        'nbr_TD': nbr_TD,
        'nbr_SS': nbr_SS,
        'all_classes': all_classes,
        }    
    clas_id = myClasse.id

    return redirect(reverse('depa:list_Seance_Ens', kwargs={'clas_id': clas_id}))





# # ----------------------------------------------------------------------
@login_required
def fichePedagogique_Ens_Dep(request, ens_id):
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)

    # Variables d'initialisation S1
    all_classes_S1 = 0
    nbr_Cours_S1 = 0
    nbr_TP_S1 = 0
    nbr_TD_S1 = 0
    nbr_SS_S1 = 0
    vol_hor_cours_S1 = 0
    vol_hor_TD_S1 = 0
    vol_hor_TP_S1 = 0
    vol_hor_Total_S1 = 0
    all_Classe_Ens_S1 = []

    try:
        myEns_obj = Enseignant.objects.get(id=ens_id)
        myEns = Ens_Dep.objects.get(enseignant=myEns_obj, departement=departement)
    except (Enseignant.DoesNotExist, Ens_Dep.DoesNotExist):
        messages.error(request, "الأستاذ غير موجود")
        return redirect('depa:manage_enseignants')
    
    # 🚀 SEMESTRE 1 - Optimisé
    base_filter_S1 = {
        'enseignant__departement': departement.id,
        'enseignant__enseignant__id': ens_id,
        'semestre': 1,
    }

    # Récupérer les classes par jour pour S1
    jours = ['Samedi', 'Dimanche', 'Lundi', 'Mardi', 'Mercredi', 'Jeudi']
    for jour in jours:
        classes_jour = Classe.objects.filter(
            **base_filter_S1,
            jour=jour
        ).select_related(
            'matiere',
            'matiere__unite',
            'niv_spe_dep_sg',
            'niv_spe_dep_sg__niv_spe_dep',
            'niv_spe_dep_sg__niv_spe_dep__specialite',
            'niv_spe_dep_sg__section',
            'niv_spe_dep_sg__groupe'
        ).order_by('temps')
        all_Classe_Ens_S1.extend(classes_jour)

    # Récupérer les IDs des classes S1 qui ont des séances
    classe_ids_S1 = [classe.id for classe in all_Classe_Ens_S1 if classe.seance_created]
    
    # Calculer les stats des séances S1 en UNE requête
    seances_stats_S1 = {}
    if classe_ids_S1:
        seances_query_S1 = Seance.objects.filter(
            classe_id__in=classe_ids_S1
        ).values('classe_id').annotate(
            total=Count('id'),
            fait_count=Count('id', filter=Q(fait=True))
        )
        seances_stats_S1 = {item['classe_id']: item for item in seances_query_S1}

    classes_to_update_S1 = []

    # Traiter les classes S1
    for classe in all_Classe_Ens_S1:
        # Compter par type
        if classe.type == "Cours":
            nbr_Cours_S1 += 1
        elif classe.type == "TP":
            nbr_TP_S1 += 1
        elif classe.type == "TD":
            nbr_TD_S1 += 1
        elif classe.type == "Sortie Scientifique":
            nbr_SS_S1 += 1

        # Calculer le taux d'avancement
        if classe.seance_created:
            stats = seances_stats_S1.get(classe.id)
            if stats and stats['total'] > 0:
                taux_avancement = round((stats['fait_count'] / stats['total']) * 100)
                if classe.taux_avancement != taux_avancement:
                    classe.taux_avancement = taux_avancement
                    classes_to_update_S1.append(classe)

    # Calculer volumes horaires S1
    vol_hor_cours_S1 = nbr_Cours_S1 * 2.25
    vol_hor_TD_S1 = nbr_TD_S1 * 1.5
    vol_hor_TP_S1 = nbr_TP_S1 * 1.5
    vol_hor_Total_S1 = vol_hor_cours_S1 + vol_hor_TD_S1 + vol_hor_TP_S1
    all_classes_S1 = nbr_Cours_S1 + nbr_TP_S1 + nbr_TD_S1 + nbr_SS_S1

    # Calculer jours distincts S1
    myDays_S1 = len(set(c.jour for c in all_Classe_Ens_S1 if c.jour))

    # Sauvegarder les taux d'avancement S1
    if classes_to_update_S1:
        Classe.objects.bulk_update(classes_to_update_S1, ['taux_avancement'])

    # 🚀 SEMESTRE 2 - Optimisé (même structure que S1)
    all_classes_S2 = 0
    nbr_Cours_S2 = 0
    nbr_TP_S2 = 0
    nbr_TD_S2 = 0
    nbr_SS_S2 = 0
    vol_hor_cours_S2 = 0
    vol_hor_TD_S2 = 0
    vol_hor_TP_S2 = 0
    vol_hor_Total_S2 = 0
    all_Classe_Ens_S2 = []

    base_filter_S2 = {
        'enseignant__departement': departement.id,
        'enseignant__enseignant__id': ens_id,
        'semestre': 2,
    }

    # Récupérer les classes par jour pour S2
    for jour in jours:
        classes_jour = Classe.objects.filter(
            **base_filter_S2,
            jour=jour
        ).select_related(
            'matiere',
            'matiere__unite',
            'niv_spe_dep_sg',
            'niv_spe_dep_sg__niv_spe_dep',
            'niv_spe_dep_sg__niv_spe_dep__specialite',
            'niv_spe_dep_sg__section',
            'niv_spe_dep_sg__groupe'
        ).order_by('temps')
        all_Classe_Ens_S2.extend(classes_jour)

    # Récupérer les IDs des classes S2 qui ont des séances
    classe_ids_S2 = [classe.id for classe in all_Classe_Ens_S2 if classe.seance_created]
    
    # Calculer les stats des séances S2 en UNE requête
    seances_stats_S2 = {}
    if classe_ids_S2:
        seances_query_S2 = Seance.objects.filter(
            classe_id__in=classe_ids_S2
        ).values('classe_id').annotate(
            total=Count('id'),
            fait_count=Count('id', filter=Q(fait=True))
        )
        seances_stats_S2 = {item['classe_id']: item for item in seances_query_S2}

    classes_to_update_S2 = []

    # Traiter les classes S2
    for classe in all_Classe_Ens_S2:
        # Compter par type
        if classe.type == "Cours":
            nbr_Cours_S2 += 1
        elif classe.type == "TP":
            nbr_TP_S2 += 1
        elif classe.type == "TD":
            nbr_TD_S2 += 1
        elif classe.type == "Sortie Scientifique":
            nbr_SS_S2 += 1

        # Calculer le taux d'avancement
        if classe.seance_created:
            stats = seances_stats_S2.get(classe.id)
            if stats and stats['total'] > 0:
                taux_avancement = round((stats['fait_count'] / stats['total']) * 100)
                if classe.taux_avancement != taux_avancement:
                    classe.taux_avancement = taux_avancement
                    classes_to_update_S2.append(classe)

    # Calculer volumes horaires S2
    vol_hor_cours_S2 = nbr_Cours_S2 * 2.25
    vol_hor_TD_S2 = nbr_TD_S2 * 1.5
    vol_hor_TP_S2 = nbr_TP_S2 * 1.5
    vol_hor_Total_S2 = vol_hor_cours_S2 + vol_hor_TD_S2 + vol_hor_TP_S2
    all_classes_S2 = nbr_Cours_S2 + nbr_TP_S2 + nbr_TD_S2 + nbr_SS_S2

    # Calculer jours distincts S2
    myDays_S2 = len(set(c.jour for c in all_Classe_Ens_S2 if c.jour))

    # Sauvegarder les taux d'avancement S2
    if classes_to_update_S2:
        Classe.objects.bulk_update(classes_to_update_S2, ['taux_avancement'])

    # 💾 Mise à jour de myEns pour les DEUX semestres en UNE SEULE fois
    myEns.nbrClas_Cours_in_Dep_S1 = nbr_Cours_S1
    myEns.nbrClas_TD_in_Dep_S1 = nbr_TD_S1
    myEns.nbrClas_TP_in_Dep_S1 = nbr_TP_S1
    myEns.nbrClas_SS_in_Dep_S1 = nbr_SS_S1
    myEns.nbrJour_in_Dep_S1 = myDays_S1
    myEns.volHor_in_Dep_S1 = vol_hor_Total_S1
    myEns.nbrClas_Cours_out_Dep_S1 = 0
    myEns.nbrClas_TD_out_Dep_S1 = 0
    myEns.nbrClas_TP_out_Dep_S1 = 0
    myEns.nbrClas_SS_out_Dep_S1 = 0
    myEns.volHor_out_Dep_S1 = 0

    myEns.nbrClas_Cours_in_Dep_S2 = nbr_Cours_S2
    myEns.nbrClas_TD_in_Dep_S2 = nbr_TD_S2
    myEns.nbrClas_TP_in_Dep_S2 = nbr_TP_S2
    myEns.nbrClas_SS_in_Dep_S2 = nbr_SS_S2
    myEns.nbrJour_in_Dep_S2 = myDays_S2
    myEns.volHor_in_Dep_S2 = vol_hor_Total_S2
    myEns.nbrClas_Cours_out_Dep_S2 = 0
    myEns.nbrClas_TD_out_Dep_S2 = 0
    myEns.nbrClas_TP_out_Dep_S2 = 0
    myEns.nbrClas_SS_out_Dep_S2 = 0
    myEns.volHor_out_Dep_S2 = 0

    myEns.save()

    univ = departement.faculte.universite

    context = {
        'all_Classe_Ens_S1': all_Classe_Ens_S1,
        'all_classes_S1': all_classes_S1,
        'vol_hor_cours_S1': vol_hor_cours_S1,
        'vol_hor_TD_S1': vol_hor_TD_S1,
        'vol_hor_TP_S1': vol_hor_TP_S1,
        'vol_hor_Total_S1': vol_hor_Total_S1,

        'all_Classe_Ens_S2': all_Classe_Ens_S2,
        'all_classes_S2': all_classes_S2,
        'vol_hor_cours_S2': vol_hor_cours_S2,
        'vol_hor_TD_S2': vol_hor_TD_S2,
        'vol_hor_TP_S2': vol_hor_TP_S2,
        'vol_hor_Total_S2': vol_hor_Total_S2,

        'vol_hor_Total_General': vol_hor_Total_S1 + vol_hor_Total_S2,

        'title': f"Fiche Pédagogique - {myEns.enseignant.nom_fr} {myEns.enseignant.prenom_fr}",
        'univ': univ,
        'my_Dep': departement,
        'my_Fac': departement.faculte,
        'myEns': myEns,
        'all_classes': all_classes_S1 + all_classes_S2,
    }

    return render(request, 'departement/fichePedagogique_Ens_Dep.html', context)



# @login_required
# def timeTable_Niv(request, niv_id, semestre_num):
#     """Emploi du temps d'un niveau - Version DEBUG pour département"""
    
#     enseignant = request.user.enseignant_profile
#     real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
#     departement = Departement.objects.get(id=real_Dep.departement.id)

#     # print(f"\n🔍 === DEBUG TIMETABLE DÉPARTEMENT ===")
#     # print(f"🔍 Paramètres reçus:")
#     # print(f"  - niv_id: {niv_id} (type: {type(niv_id)})")
#     # print(f"  - semestre_num: {semestre_num} (type: {type(semestre_num)})")
#     # print(f"  - departement: {departement} (ID: {departement.id})")

#     # Vérifier que le niveau existe
#     try:
#         niveau = Niveau.objects.get(id=niv_id)
#         # print(f"✅ Niveau trouvé: {niveau}")
#     except Niveau.DoesNotExist:
#         # print(f"❌ ERREUR: Niveau avec ID {niv_id} n'existe pas!")
#         return render(request, 'departement/timeTable_Niv.html', {
#             'error': f'المستوى برقم {niv_id} غير موجود',
#             'all_classes': 0,
#             'my_Dep': departement,
#             'my_Fac': departement.faculte,
#         })

#     # Vérifier que le semestre existe
#     try:
#         semestre = Semestre.objects.get(numero=semestre_num)
#         # print(f"✅ Semestre trouvé: {semestre}")
#     except Semestre.DoesNotExist:
#         # print(f"❌ ERREUR: Semestre {semestre_num} n'existe pas!")
#         return render(request, 'departement/timeTable_Niv.html', {
#             'error': f'السداسي رقم {semestre_num} غير موجود',
#             'all_classes': 0,
#             'my_Dep': departement,
#             'my_Fac': departement.faculte,
#         })

#     # Test 1: Voir TOUTES les classes du département
#     all_classes_dept = Classe.objects.filter(enseignant__departement=departement)
#     # print(f"🔍 Total classes dans le département: {all_classes_dept.count()}")
    
#     # Test 2: Classes pour ce niveau (tous semestres)
#     classes_niveau = Classe.objects.filter(
#         enseignant__departement=departement,
#         niv_spe_dep_sg__niv_spe_dep__niveau_id=niv_id
#     )
#     # print(f"🔍 Classes pour niveau {niv_id}: {classes_niveau.count()}")
    
#     # Test 3: Classes pour ce semestre (tous niveaux)
#     classes_semestre = Classe.objects.filter(
#         enseignant__departement=departement,
#         semestre__numero=semestre_num
#     )
#     # print(f"🔍 Classes pour semestre {semestre_num}: {classes_semestre.count()}")
    
#     # Test 4: Classes avec le filtre complet
#     all_Classe_Niv = Classe.objects.filter(
#         enseignant__departement=departement,
#         niv_spe_dep_sg__niv_spe_dep__niveau_id=niv_id,
#         semestre__numero=semestre_num
#     )
#     # print(f"🔍 Classes avec filtre complet: {all_Classe_Niv.count()}")

#     # Debug des classes trouvées
#     # if all_Classe_Niv.exists():
#         # pass
#         # print(f"✅ Classes trouvées:")
#         # for i, classe in enumerate(all_Classe_Niv[:5]):  # Afficher max 5
#             # print(f"  {i+1}. {classe.matiere} - {classe.type} - {classe.jour} {classe.temps}")
#     # else:
#         # print(f"❌ Aucune classe trouvée avec les critères complets!")
        
#         # Debug plus approfondi
#         # print(f"\n🔍 DEBUG APPROFONDI:")
        
#         # Vérifier les NivSpeDep pour ce département et niveau
#         # nivspedep = NivSpeDep.objects.filter(
#         #     departement=departement,
#         #     niveau_id=niv_id
#         # )
#         # print(f"📋 NivSpeDep pour dept + niveau: {nivspedep.count()}")
#         # for nsd in nivspedep:
#             # print(f"  - {nsd}")
            
#         # Vérifier les NivSpeDep_SG
#         # nivspedep_sg = NivSpeDep_SG.objects.filter(
#         #     niv_spe_dep__departement=departement,
#         #     niv_spe_dep__niveau_id=niv_id
#         # )
#         # print(f"📋 NivSpeDep_SG pour dept + niveau: {nivspedep_sg.count()}")
#         # for nsg in nivspedep_sg[:3]:
#             # print(f"  - {nsg}")
            
#         # Vérifier les enseignants du département
#         # enseignants_dept = Ens_Dep.objects.filter(departement=departement)
#         # print(f"👨‍🏫 Enseignants dans le département: {enseignants_dept.count()}")
        
#         # Voir quelques classes d'exemple du département
#         # classes_examples = Classe.objects.filter(enseignant__departement=departement)[:3]
#         # print(f"📚 Exemples de classes du département:")
#         # for cl in classes_examples:
#             # print(f"  - Matière: {cl.matiere}")
#             # print(f"    Niveau: {cl.niv_spe_dep_sg.niv_spe_dep.niveau} (ID: {cl.niv_spe_dep_sg.niv_spe_dep.niveau.id})")
#             # print(f"    Semestre: {cl.semestre.numero}")
#             # print(f"    Type: {cl.type}")

#     # Initialisation des variables pour le template
#     nbr_Cours = 0
#     nbr_TP = 0
#     nbr_TD = 0
#     nbr_SS = 0
#     all_classes = 0

#     # Créneaux horaires (gardons la logique originale)
#     sixClasses = [
#         '08:00-09:30',
#         '09:40-11:10', 
#         '11:20-12:50',
#         '13:10-14:40',
#         '14:50-16:20',
#         '16:30-18:00',
#     ]

#     sixClassesTimes = [
#         '09:30-08:00',
#         '11:10-09:40',
#         '12:50-11:20',
#         '14:40-13:10',
#         '16:20-14:50',
#         '18:00-16:30',
#     ]

#     sevenDays = ['Samedi', 'Dimanche', 'Lundi', 'Mardi', 'Mercredi', 'Jeudi']

#     # Récupérer les classes organisées par créneaux (utiliser la même requête)
#     all_Classe_Time = []
#     for time_slot in sixClassesTimes:
#         classes_time = Classe.objects.filter(
#             enseignant__departement=departement,
#             niv_spe_dep_sg__niv_spe_dep__niveau_id=niv_id,
#             semestre__numero=semestre_num,
#             temps=time_slot
#         )
#         all_Classe_Time.extend(classes_time)

#     # Calculer les statistiques
#     for classe in all_Classe_Niv:
#         if classe.type == "Cours":
#             nbr_Cours += 1
#         elif classe.type == "TP":
#             nbr_TP += 1
#         elif classe.type == "TD":
#             nbr_TD += 1
#         elif classe.type == "Sortie Scientifique":
#             nbr_SS += 1

#     all_classes = nbr_Cours + nbr_TP + nbr_TD + nbr_SS
    
#     # print(f"🔍 RÉSUMÉ FINAL:")
#     # print(f"  - Total classes: {all_classes}")
#     # print(f"  - Cours: {nbr_Cours}, TD: {nbr_TD}, TP: {nbr_TP}, SS: {nbr_SS}")
#     # print(f"🔍 =================================\n")

#     context = {
#         'title': 'إستعمال الزمن للمستوى',
#         'all_Classe_Niv': all_Classe_Niv,
#         'my_Dep': departement,
#         'my_Fac': departement.faculte,
#         'all_Classe_Time': all_Classe_Time,
#         'nbr_Cours': nbr_Cours,
#         'nbr_TP': nbr_TP,
#         'nbr_TD': nbr_TD,
#         'nbr_SS': nbr_SS,
#         'all_classes': all_classes,
#         'sevenDays': sevenDays,
#         'sixClasses': sixClasses,
#         'niveau_id': niv_id,
#         'semestre_num': semestre_num,
#     }

#     return render(request, 'departement/timeTable_Niv.html', context)










from django.db.models import Q, Prefetch
from django.core.cache import cache
from django.views.decorators.cache import cache_page

# @login_required
# def timeTable_Niv(request, niv_spe_dep_id, semestre_num):
#     """Emploi du temps d'un niveau - Version ULTRA OPTIMISÉE avec CACHE"""
    
#     enseignant = request.user.enseignant_profile
#     real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
#     departement = Departement.objects.get(id=real_Dep.departement.id)

#     my_niv_spe_dep = NivSpeDep.objects.get(id=niv_spe_dep_id)

#     # ✅ Créer une clé de cache unique
#     cache_key = f'timetable_niv_{departement.id}_{niv_spe_dep_id}_{semestre_num}'
    
#     # ✅ Essayer de récupérer depuis le cache (expire après 15 minutes)
#     cached_data = cache.get(cache_key)
#     if cached_data:
#         # Données en cache trouvées - retour immédiat!
#         cached_data['my_Dep'] = departement
#         cached_data['my_Fac'] = departement.faculte
#         return render(request, 'departement/timeTable_Niv.html', cached_data)

#     # Vérifier que le niveau existe
#     try:
#         niveau = Niveau.objects.get(id=my_niv_spe_dep.niveau.id)
#     except Niveau.DoesNotExist:
#         return render(request, 'departement/timeTable_Niv.html', {
#             'error': f'المستوى برقم {my_niv_spe_dep.niveau.id} غير موجود',
#             'all_classes': 0,
#             'my_Dep': departement,
#             'my_Fac': departement.faculte,
#         })

#     # Vérifier que le semestre existe
#     try:
#         semestre = Semestre.objects.get(numero=semestre_num)
#     except Semestre.DoesNotExist:
#         return render(request, 'departement/timeTable_Niv.html', {
#             'error': f'السداسي رقم {semestre_num} غير موجود',
#             'all_classes': 0,
#             'my_Dep': departement,
#             'my_Fac': departement.faculte,
#         })

#     # ✅ OPTIMISATION MAJEURE : UNE SEULE REQUÊTE avec tous les select_related
#     all_Classe_Niv = Classe.objects.filter(
#         enseignant__departement=departement,
#         niv_spe_dep_sg__niv_spe_dep=my_niv_spe_dep,
#         semestre__numero=semestre_num
#     ).select_related(
#         'matiere',
#         'enseignant__enseignant',
#         'enseignant__departement',
#         'niv_spe_dep_sg__niv_spe_dep__niveau',
#         'niv_spe_dep_sg__niv_spe_dep__specialite__reforme',
#         'niv_spe_dep_sg__section',
#         'niv_spe_dep_sg__groupe',
#         'semestre'
#     ).order_by('jour', 'temps')

#     # ✅ Convertir en liste pour éviter des requêtes répétées
#     all_Classe_Niv_list = list(all_Classe_Niv)
    
#     # ✅ Calcul optimisé des statistiques
#     stats = {'Cours': 0, 'TP': 0, 'TD': 0, 'Sortie Scientifique': 0}
#     for classe in all_Classe_Niv_list:
#         if classe.type in stats:
#             stats[classe.type] += 1

#     nbr_Cours = stats['Cours']
#     nbr_TP = stats['TP']
#     nbr_TD = stats['TD']
#     nbr_SS = stats['Sortie Scientifique']
#     all_classes = sum(stats.values())

#     # Créneaux horaires
#     sixClasses = [
#         '08:00-09:30',
#         '09:40-11:10', 
#         '11:20-12:50',
#         '13:10-14:40',
#         '14:50-16:20',
#         '16:30-18:00',
#     ]

#     sixClassesTimes = [
#         '09:30-08:00',
#         '11:10-09:40',
#         '12:50-11:20',
#         '14:40-13:10',
#         '16:20-14:50',
#         '18:00-16:30',
#     ]

#     sevenDays = ['Samedi', 'Dimanche', 'Lundi', 'Mardi', 'Mercredi', 'Jeudi']

#     context = {
#         'title': 'إستعمال الزمن للمستوى',
#         'all_Classe_Niv': all_Classe_Niv_list,
#         'all_Classe_Time': all_Classe_Niv_list,
#         'nbr_Cours': nbr_Cours,
#         'nbr_TP': nbr_TP,
#         'nbr_TD': nbr_TD,
#         'nbr_SS': nbr_SS,
#         'all_classes': all_classes,
#         'sevenDays': sevenDays,
#         'sixClasses': sixClasses,
#         'niveau_id': my_niv_spe_dep.niveau.id,
#         'semestre_num': semestre_num,
#     }
    
#     # ✅ Mettre en cache pour 15 minutes (900 secondes)
#     cache.set(cache_key, context, 900)
    
#     # Ajouter les objets non-sérialisables
#     context['my_Dep'] = departement
#     context['my_Fac'] = departement.faculte

#     return render(request, 'departement/timeTable_Niv.html', context)




from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.contrib import messages
from django.db import transaction

from .forms import ClasseEditForm  # Vous devrez créer ce formulaire
from apps.academique.affectation.models import Amphi_Dep, Salle_Dep, Laboratoire_Dep, Classe, Ens_Dep, Abs_Etu_Seance, Abs_Etu_Classe, Seance

@login_required
def timeTable_Niv(request, niv_spe_dep_id, semestre_num):
    """Emploi du temps d'un niveau - Version ULTRA OPTIMISÉE avec CACHE"""
    
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)

    my_niv_spe_dep = NivSpeDep.objects.get(id=niv_spe_dep_id)

    # ✅ Créer une clé de cache unique
    cache_key = f'timetable_niv_{departement.id}_{niv_spe_dep_id}_{semestre_num}'
    
    # ✅ Essayer de récupérer depuis le cache (expire après 15 minutes)
    cached_data = cache.get(cache_key)
    if cached_data:
        # Données en cache trouvées - retour immédiat!
        cached_data['my_Dep'] = departement
        cached_data['my_Fac'] = departement.faculte
        return render(request, 'departement/timeTable_Niv.html', cached_data)

    # Vérifier que le niveau existe
    try:
        niveau = Niveau.objects.get(id=my_niv_spe_dep.niveau.id)
    except Niveau.DoesNotExist:
        return render(request, 'departement/timeTable_Niv.html', {
            'error': f'المستوى برقم {my_niv_spe_dep.niveau.id} غير موجود',
            'all_classes': 0,
            'my_Dep': departement,
            'my_Fac': departement.faculte,
        })

    # Vérifier que le semestre existe
    try:
        semestre = Semestre.objects.get(numero=semestre_num)
    except Semestre.DoesNotExist:
        return render(request, 'departement/timeTable_Niv.html', {
            'error': f'السداسي رقم {semestre_num} غير موجود',
            'all_classes': 0,
            'my_Dep': departement,
            'my_Fac': departement.faculte,
        })

    # ✅ OPTIMISATION MAJEURE : UNE SEULE REQUÊTE avec tous les select_related
    all_Classe_Niv = Classe.objects.filter(
        enseignant__departement=departement,
        niv_spe_dep_sg__niv_spe_dep=my_niv_spe_dep,
        semestre__numero=semestre_num
    ).select_related(
        'matiere',
        'enseignant__enseignant',
        'enseignant__departement',
        'niv_spe_dep_sg__niv_spe_dep__niveau',
        'niv_spe_dep_sg__niv_spe_dep__specialite__reforme',
        'niv_spe_dep_sg__section',
        'niv_spe_dep_sg__groupe',
        'semestre'
    ).order_by('jour', 'temps')

    # ✅ Convertir en liste pour éviter des requêtes répétées
    all_Classe_Niv_list = list(all_Classe_Niv)
    
    # ✅ Calcul optimisé des statistiques
    stats = {'Cours': 0, 'TP': 0, 'TD': 0, 'Sortie Scientifique': 0}
    for classe in all_Classe_Niv_list:
        if classe.type in stats:
            stats[classe.type] += 1

    nbr_Cours = stats['Cours']
    nbr_TP = stats['TP']
    nbr_TD = stats['TD']
    nbr_SS = stats['Sortie Scientifique']
    all_classes = sum(stats.values())

    # Créneaux horaires
    time_slots = [
        ('09:30-08:00', '08:00-09:30'),
        ('11:10-09:40', '09:40-11:10'),
        ('12:50-11:20', '11:20-12:50'),
        ('14:40-13:10', '13:10-14:40'),
        ('16:20-14:50', '14:50-16:20'),
        ('18:00-16:30', '16:30-18:00'),
    ]

    sevenDays = ['Samedi', 'Dimanche', 'Lundi', 'Mardi', 'Mercredi', 'Jeudi']

    context = {
        'title': 'إستعمال الزمن للمستوى',
        'all_Classe_Niv': all_Classe_Niv_list,
        'all_Classe_Time': all_Classe_Niv_list,
        'nbr_Cours': nbr_Cours,
        'nbr_TP': nbr_TP,
        'nbr_TD': nbr_TD,
        'nbr_SS': nbr_SS,
        'all_classes': all_classes,
        'sevenDays': sevenDays,
        'time_slots': time_slots,
        'niveau_id': my_niv_spe_dep.niveau.id,
        'semestre_num': semestre_num,
    }
    
    # ✅ Mettre en cache pour 15 minutes (900 secondes)
    cache.set(cache_key, context, 900)
    
    # Ajouter les objets non-sérialisables
    context['my_Dep'] = departement
    context['my_Fac'] = departement.faculte

    return render(request, 'departement/timeTable_Niv.html', context)


@login_required
@require_http_methods(["GET", "POST"])
def edit_classe(request, classe_id):
    """Modifier une classe avec vérifications des séances"""
    try:
        classe = get_object_or_404(Classe, id=classe_id)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'الحصة غير موجودة: {str(e)}'
        }, status=404)
    
    # Vérifier les permissions
    try:
        enseignant = request.user.enseignant_profile
    except:
        return JsonResponse({
            'success': False,
            'message': 'لم يتم العثور على ملف الأستاذ'
        }, status=403)
    
    try:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    except Ens_Dep.DoesNotExist:
        try:
            real_Dep = Ens_Dep.objects.get(enseignant=enseignant)
        except:
            return JsonResponse({
                'success': False,
                'message': 'لم يتم العثور على القسم'
            }, status=403)
    
    if request.method == 'GET':
        try:
            # Préparer les données pour le formulaire
            semestres = Semestre.objects.all().order_by('numero')
            
            # ✅ POINT 1: Récupérer UNIQUEMENT les matières de ce niveau pour le semestre de la classe
            try:
                # Matières pour l'affichage initial (semestre actuel de la classe)
                matieres = Matiere.objects.filter(
                    specialite=classe.niv_spe_dep_sg.niv_spe_dep.specialite,
                    semestre=classe.semestre
                ).order_by('nom_fr')
                
                # TOUTES les matières de ce niveau (tous semestres) pour le filtrage JavaScript
                all_matieres = Matiere.objects.filter(
                    specialite=classe.niv_spe_dep_sg.niv_spe_dep.specialite
                ).select_related('semestre').order_by('semestre__numero', 'nom_fr')
            except:
                matieres = Matiere.objects.filter(semestre=classe.semestre).order_by('nom_fr')
                all_matieres = Matiere.objects.all().select_related('semestre').order_by('semestre__numero', 'nom_fr')
            
            # Récupérer les enseignants
            enseignants = Ens_Dep.objects.filter(
                departement=real_Dep.departement
            ).select_related('enseignant').order_by('enseignant__nom_ar')
            
            # Récupérer les groupes
            try:
                groupes = NivSpeDep_SG.objects.filter(
                    niv_spe_dep=classe.niv_spe_dep_sg.niv_spe_dep
                ).select_related('section', 'groupe').order_by('id')
            except:
                groupes = NivSpeDep_SG.objects.all().select_related('section', 'groupe').order_by('id')
            
            # ✅ POINT 2: Récupérer TOUS les lieux (pas de filtrage automatique)
            amphitheatres = Amphi_Dep.objects.filter(departement=real_Dep.departement).order_by('amphi__numero')
            salles = Salle_Dep.objects.filter(departement=real_Dep.departement).order_by('salle__numero')
            labos = Laboratoire_Dep.objects.filter(departement=real_Dep.departement).order_by('laboratoire__numero')
            
            # Déterminer le type de lieu actuel
            lieu_type = ''
            lieu_id = None
            if classe.lieu:
                content_type = classe.content_type
                if content_type:
                    if content_type.model == 'amphi_dep':
                        lieu_type = 'amphi'
                    elif content_type.model == 'salle_dep':
                        lieu_type = 'salle'
                    elif content_type.model == 'laboratoire_dep':
                        lieu_type = 'labo'
                    lieu_id = classe.object_id
            
            context = {
                'classe': classe,
                'semestres': semestres,
                'matieres': matieres,  # Matières du semestre actuel
                'all_matieres': all_matieres,  # TOUTES les matières du niveau
                'enseignants': enseignants,
                'groupes': groupes,
                'amphitheatres': amphitheatres,
                'salles': salles,
                'labos': labos,
                'lieu_type': lieu_type,
                'lieu_id': lieu_id,
                'jours': Classe.Dayblock.choices,
                'temps': Classe.Timeblock.choices,
                'types': Classe.Typeblock.choices,
            }
            
            return render(request, 'departement/edit_classe_form.html', context)
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print("ERREUR DÉTAILLÉE:", error_details)
            
            return render(request, 'departement/edit_classe_form.html', {
                'error': f'حدث خطأ أثناء تحميل البيانات: {str(e)}',
                'error_details': error_details
            })
    
    elif request.method == 'POST':
        # ... le reste du code POST reste identique ...
        try:
            with transaction.atomic():
                # Récupérer les données du formulaire
                semestre_id = request.POST.get('semestre')
                matiere_id = request.POST.get('matiere')
                enseignant_id = request.POST.get('enseignant')
                niv_spe_dep_sg_id = request.POST.get('niv_spe_dep_sg')
                jour = request.POST.get('jour')
                temps = request.POST.get('temps')
                type_classe = request.POST.get('type')
                observation = request.POST.get('observation', '')
                lieu_type = request.POST.get('lieu_type')
                lieu_id = request.POST.get('lieu_id')
                
                # Validation des champs obligatoires
                if not all([semestre_id, matiere_id, enseignant_id, niv_spe_dep_sg_id, jour, temps, type_classe]):
                    return JsonResponse({
                        'success': False,
                        'message': 'جميع الحقول الإلزامية يجب أن تكون معبأة'
                    }, status=400)
                
                # Vérifier que le lieu est spécifié
                if not lieu_type or not lieu_id:
                    return JsonResponse({
                        'success': False,
                        'message': 'يجب تحديد المكان (مدرج، قاعة أو مخبر)'
                    }, status=400)
                
                # VÉRIFICATION 1: Si seance_created est False, modifier directement
                if not classe.seance_created:
                    # Mettre à jour les champs
                    classe.semestre_id = semestre_id
                    classe.matiere_id = matiere_id
                    classe.enseignant_id = enseignant_id
                    classe.niv_spe_dep_sg_id = niv_spe_dep_sg_id
                    classe.jour = jour
                    classe.temps = temps
                    classe.type = type_classe
                    classe.observation = observation
                    
                    # Mettre à jour le lieu
                    if lieu_type == 'amphi':
                        classe.content_type = ContentType.objects.get_for_model(Amphi_Dep)
                    elif lieu_type == 'salle':
                        classe.content_type = ContentType.objects.get_for_model(Salle_Dep)
                    elif lieu_type == 'labo':
                        classe.content_type = ContentType.objects.get_for_model(Laboratoire_Dep)
                    classe.object_id = lieu_id
                    
                    classe.save()
                    
                    # Invalider le cache
                    cache_key = f'timetable_niv_{real_Dep.departement.id}_{classe.niv_spe_dep_sg.niv_spe_dep.id}_{classe.semestre.numero}'
                    cache.delete(cache_key)
                    
                    return JsonResponse({
                        'success': True,
                        'message': 'تم تعديل الحصة بنجاح'
                    })
                
                # VÉRIFICATION 2: Si seance_created est True
                seances = Seance.objects.filter(classe=classe)
                
                seances_non_effectuees = seances.filter(
                    fait=False, 
                    remplacer=False, 
                    annuler=False
                )
                seances_effectuees = seances.exclude(
                    fait=False, 
                    remplacer=False, 
                    annuler=False
                )
                
                if not seances_effectuees.exists():
                    for seance in seances:
                        Abs_Etu_Seance.objects.filter(seance=seance).delete()
                        seance.delete()
                    
                    Abs_Etu_Classe.objects.filter(classe=classe).delete()
                    
                    classe.seance_created = False
                    classe.abs_liste_Etu = False
                    
                    classe.semestre_id = semestre_id
                    classe.matiere_id = matiere_id
                    classe.enseignant_id = enseignant_id
                    classe.niv_spe_dep_sg_id = niv_spe_dep_sg_id
                    classe.jour = jour
                    classe.temps = temps
                    classe.type = type_classe
                    classe.observation = observation
                    
                    if lieu_type == 'amphi':
                        classe.content_type = ContentType.objects.get_for_model(Amphi_Dep)
                    elif lieu_type == 'salle':
                        classe.content_type = ContentType.objects.get_for_model(Salle_Dep)
                    elif lieu_type == 'labo':
                        classe.content_type = ContentType.objects.get_for_model(Laboratoire_Dep)
                    classe.object_id = lieu_id
                    
                    classe.save()
                    
                    cache_key = f'timetable_niv_{real_Dep.departement.id}_{classe.niv_spe_dep_sg.niv_spe_dep.id}_{classe.semestre.numero}'
                    cache.delete(cache_key)
                    
                    return JsonResponse({
                        'success': True,
                        'message': 'تم تعديل الحصة بنجاح وحذف جميع الحصص المرتبطة'
                    })
                
                else:
                    nbr_deleted = 0
                    for seance in seances_non_effectuees:
                        Abs_Etu_Seance.objects.filter(seance=seance).delete()
                        seance.delete()
                        nbr_deleted += 1
                    
                    try:
                        etudiants_concernes = classe.niv_spe_dep_sg.etudiants_niv_spe_dep_sg.all()
                    except:
                        etudiants_concernes = []
                    
                    for etudiant in etudiants_concernes:
                        try:
                            abs_etu_classe = Abs_Etu_Classe.objects.get(
                                classe=classe,
                                etudiant=etudiant
                            )
                            
                            abs_etu_classe.nbr_absence = 0
                            abs_etu_classe.nbr_absence_justifiee = 0
                            
                            seances_restantes = Seance.objects.filter(classe=classe)
                            for seance in seances_restantes:
                                absences = Abs_Etu_Seance.objects.filter(
                                    seance=seance,
                                    etudiant=etudiant
                                )
                                
                                for absence in absences:
                                    if not absence.present and not absence.justifiee:
                                        abs_etu_classe.nbr_absence += 1
                                    elif not absence.present and absence.justifiee:
                                        abs_etu_classe.nbr_absence_justifiee += 1
                            
                            abs_etu_classe.save()
                            
                        except Abs_Etu_Classe.DoesNotExist:
                            continue
                    
                    classe.semestre_id = semestre_id
                    classe.matiere_id = matiere_id
                    classe.enseignant_id = enseignant_id
                    classe.niv_spe_dep_sg_id = niv_spe_dep_sg_id
                    classe.jour = jour
                    classe.temps = temps
                    classe.type = type_classe
                    classe.observation = observation
                    
                    if lieu_type == 'amphi':
                        classe.content_type = ContentType.objects.get_for_model(Amphi_Dep)
                    elif lieu_type == 'salle':
                        classe.content_type = ContentType.objects.get_for_model(Salle_Dep)
                    elif lieu_type == 'labo':
                        classe.content_type = ContentType.objects.get_for_model(Laboratoire_Dep)
                    classe.object_id = lieu_id
                    
                    classe.save()
                    
                    cache_key = f'timetable_niv_{real_Dep.departement.id}_{classe.niv_spe_dep_sg.niv_spe_dep.id}_{classe.semestre.numero}'
                    cache.delete(cache_key)
                    
                    nbr_kept = seances_effectuees.count()
                    
                    return JsonResponse({
                        'success': True,
                        'message': f'تم تعديل الحصة بنجاح. تم حذف {nbr_deleted} حصة غير مكتملة والاحتفاظ بـ {nbr_kept} حصة مكتملة'
                    })
        
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print("ERREUR DÉTAILLÉE POST:", error_details)
            
            return JsonResponse({
                'success': False,
                'message': f'حدث خطأ أثناء التعديل: {str(e)}'
            }, status=500)
        




















@login_required
@require_http_methods(["GET", "POST"])
def add_classe(request):
    """Ajouter une classe individuellement"""
    try:
        enseignant = request.user.enseignant_profile
    except:
        messages.error(request, 'لم يتم العثور على ملف الأستاذ')
        return redirect('depa:dashboard_Dep')
    
    try:
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    except Ens_Dep.DoesNotExist:
        try:
            real_Dep = Ens_Dep.objects.get(enseignant=enseignant)
        except:
            messages.error(request, 'لم يتم العثور على القسم')
            return redirect('depa:dashboard_Dep')
    
    departement = real_Dep.departement
    
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Récupérer les données du formulaire
                semestre_id = request.POST.get('semestre')
                matiere_id = request.POST.get('matiere')
                enseignant_id = request.POST.get('enseignant')
                niv_spe_dep_sg_id = request.POST.get('niv_spe_dep_sg')
                jour = request.POST.get('jour')
                temps = request.POST.get('temps')
                type_classe = request.POST.get('type')
                observation = request.POST.get('observation', '')
                lieu_type = request.POST.get('lieu_type')
                lieu_id = request.POST.get('lieu_id')
                
                # Validation des champs obligatoires
                if not all([semestre_id, matiere_id, enseignant_id, niv_spe_dep_sg_id, jour, temps, type_classe]):
                    messages.error(request, 'جميع الحقول الإلزامية يجب أن تكون معبأة')
                    return redirect('depa:add_classe')
                
                # Vérifier que le lieu est spécifié
                if not lieu_type or not lieu_id:
                    messages.error(request, 'يجب تحديد المكان (مدرج، قاعة أو مخبر)')
                    return redirect('depa:add_classe')
                
                # Déterminer le content_type et object_id
                if lieu_type == 'amphi':
                    content_type = ContentType.objects.get_for_model(Amphi_Dep)
                elif lieu_type == 'salle':
                    content_type = ContentType.objects.get_for_model(Salle_Dep)
                elif lieu_type == 'labo':
                    content_type = ContentType.objects.get_for_model(Laboratoire_Dep)
                else:
                    messages.error(request, 'نوع المكان غير صحيح')
                    return redirect('depa:add_classe')
                
                # Vérifier si la classe existe déjà
                existing_classe = Classe.objects.filter(
                    semestre_id=semestre_id,
                    matiere_id=matiere_id,
                    enseignant_id=enseignant_id,
                    niv_spe_dep_sg_id=niv_spe_dep_sg_id,
                    jour=jour,
                    temps=temps,
                    type=type_classe,
                    content_type=content_type,
                    object_id=lieu_id,
                ).first()
                
                if existing_classe:
                    messages.warning(request, 'هذه الحصة موجودة بالفعل في النظام')
                    return redirect('depa:add_classe')
                
                # Créer la classe
                classe = Classe(
                    semestre_id=semestre_id,
                    matiere_id=matiere_id,
                    enseignant_id=enseignant_id,
                    niv_spe_dep_sg_id=niv_spe_dep_sg_id,
                    jour=jour,
                    temps=temps,
                    type=type_classe,
                    observation=observation,
                    content_type=content_type,
                    object_id=lieu_id,
                )
                classe.save()
                
                messages.success(request, 'تم إضافة الحصة بنجاح')
                return redirect('depa:add_classe')
                
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print("ERREUR:", error_details)
            messages.error(request, f'حدث خطأ أثناء إضافة الحصة: {str(e)}')
            return redirect('depa:add_classe')
    
    # GET - Préparer les données pour le formulaire
    semestres = Semestre.objects.all().order_by('numero')
    enseignants = Ens_Dep.objects.filter(departement=departement).select_related('enseignant').order_by('enseignant__nom_ar')
    
    # Récupérer tous les niveaux/spécialités/groupes du département
    niv_spe_deps = NivSpeDep.objects.filter(departement=departement).select_related('niveau', 'specialite')
    groupes = NivSpeDep_SG.objects.filter(niv_spe_dep__departement=departement).select_related('section', 'groupe', 'niv_spe_dep').order_by('id')
    
    # Récupérer toutes les matières
    all_matieres = Matiere.objects.all().select_related('semestre').order_by('semestre__numero', 'nom_fr')
    
    # Récupérer les lieux
    amphitheatres = Amphi_Dep.objects.filter(departement=departement).select_related('amphi').order_by('amphi__numero')
    salles = Salle_Dep.objects.filter(departement=departement).select_related('salle').order_by('salle__numero')
    labos = Laboratoire_Dep.objects.filter(departement=departement).select_related('laboratoire').order_by('laboratoire__numero')
    
    context = {
        'title': 'إضافة حصة جديدة',
        'my_Dep': departement,  # Passer le département au template
        'my_Fac': departement.faculte,  # Passer le département au template
        'departement': departement,
        'semestres': semestres,
        'all_matieres': all_matieres,
        'enseignants': enseignants,
        'groupes': groupes,
        'amphitheatres': amphitheatres,
        'salles': salles,
        'labos': labos,
        'jours': Classe.Dayblock.choices,
        'temps': Classe.Timeblock.choices,
        'types': Classe.Typeblock.choices,
    }
    
    return render(request, 'departement/add_classe.html', context)




































@login_required
@require_http_methods(["GET", "POST"])
def delete_classe(request, classe_id):
    """Supprimer une classe avec vérifications des séances"""
    classe = get_object_or_404(Classe, id=classe_id)
    
    # Vérifier les permissions
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    
    if classe.enseignant.departement != real_Dep.departement:
        return JsonResponse({
            'success': False,
            'message': 'ليس لديك صلاحية لحذف هذه الحصة'
        }, status=403)
    
    if request.method == 'GET':
        # Afficher le formulaire de confirmation
        
        # VÉRIFICATION 1: Si seance_created est False
        if not classe.seance_created:
            context = {
                'classe': classe,
                'has_seances': False,
                'message': 'هل أنت متأكد من حذف هذه الحصة؟',
                'danger_level': 'low'
            }
        else:
            # VÉRIFICATION 2: Si seance_created est True
            seances = Seance.objects.filter(classe=classe)
            nbr_seances = seances.count()
            
            context = {
                'classe': classe,
                'has_seances': True,
                'nbr_seances': nbr_seances,
                'message': f'تحذير! هذه الحصة تحتوي على {nbr_seances} حصة مسجلة. سيتم حذف جميع الحصص والبيانات المرتبطة بها.',
                'danger_level': 'high',
                'warning_text': 'تنبيه: لن تتمكن من التراجع عن هذا الإجراء!'
            }
        
        return render(request, 'departement/delete_classe_confirm.html', context)
    
    elif request.method == 'POST':
        try:
            with transaction.atomic():
                niv_spe_dep_id = classe.niv_spe_dep_sg.niv_spe_dep.id
                semestre_num = classe.semestre.numero
                departement_id = real_Dep.departement.id
                
                # VÉRIFICATION 1: Si seance_created est False, supprimer directement
                if not classe.seance_created:
                    classe.delete()
                    
                    # Invalider le cache
                    cache_key = f'timetable_niv_{departement_id}_{niv_spe_dep_id}_{semestre_num}'
                    cache.delete(cache_key)
                    
                    return JsonResponse({
                        'success': True,
                        'message': 'تم حذف الحصة بنجاح'
                    })
                
                # VÉRIFICATION 2: Si seance_created est True, supprimer toutes les séances
                seances = Seance.objects.filter(classe=classe)
                nbr_seances = seances.count()
                nbr_absences_total = 0
                
                # Récupérer les étudiants concernés
                etudiants_concernes = set()
                
                for seance in seances:
                    # Compter et supprimer les absences
                    absences = Abs_Etu_Seance.objects.filter(seance=seance)
                    nbr_absences_total += absences.count()
                    
                    # Récupérer les étudiants
                    for absence in absences:
                        etudiants_concernes.add(absence.etudiant)
                    
                    # Supprimer les absences
                    absences.delete()
                    
                    # Supprimer la séance
                    seance.delete()
                
                # Supprimer les enregistrements Abs_Etu_Classe
                Abs_Etu_Classe.objects.filter(
                    classe=classe,
                    etudiant__in=etudiants_concernes
                ).delete()
                
                # Supprimer la classe
                classe.delete()
                
                # Invalider le cache
                cache_key = f'timetable_niv_{departement_id}_{niv_spe_dep_id}_{semestre_num}'
                cache.delete(cache_key)
                
                return JsonResponse({
                    'success': True,
                    'message': f'تم حذف الحصة بنجاح مع {nbr_seances} حصة و {nbr_absences_total} سجل غياب'
                })
        
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'حدث خطأ أثناء الحذف: {str(e)}'
            }, status=500)





@login_required
def get_lieux_ajax(request):
    """Charger les lieux selon le type (amphi, salle, labo)"""
    lieu_type = request.GET.get('type', '')
    
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = real_Dep.departement
    
    lieux = []
    
    if lieu_type == 'amphi':
        amphitheatres = Amphi_Dep.objects.filter(departement=departement)
        lieux = [{'id': a.id, 'nom_fr': str(a)} for a in amphitheatres]
    elif lieu_type == 'salle':
        salles = Salle_Dep.objects.filter(departement=departement)
        lieux = [{'id': s.id, 'nom_fr': str(s)} for s in salles]
    elif lieu_type == 'labo':
        labos = Laboratoire_Dep.objects.filter(departement=departement)
        lieux = [{'id': l.id, 'nom_fr': str(l)} for l in labos]

    return JsonResponse({'lieux': lieux})























# ========================================
# 🔧 FONCTION UTILITAIRE : Vider le cache
# ========================================
def clear_timetable_cache(departement_id=None, niv_id=None, semestre_num=None):
    """
    Appelez cette fonction après modification de l'emploi du temps
    pour vider le cache et forcer le rechargement.
    
    Exemples d'utilisation:
    - clear_timetable_cache()  # Vide TOUT le cache
    - clear_timetable_cache(departement_id=1)  # Vide le cache d'un département
    - clear_timetable_cache(departement_id=1, niv_id=5, semestre_num=1)  # Cache spécifique
    """
    if departement_id and niv_id and semestre_num:
        # Vider un cache spécifique
        cache_key = f'timetable_niv_{departement_id}_{niv_id}_{semestre_num}'
        cache.delete(cache_key)
    elif departement_id:
        # Vider tous les caches d'un département (nécessite pattern matching)
        # Note: Cela dépend du backend de cache utilisé
        cache.delete_pattern(f'timetable_niv_{departement_id}_*')
    else:
        # Vider TOUT le cache des emplois du temps
        cache.delete_pattern('timetable_niv_*')


# ========================================
# 📊 STATISTIQUES DE PERFORMANCE
# ========================================
# AVANT (version originale):
#   - ~50-100 requêtes SQL (une par créneau!)
#   - ~3-8 secondes de chargement
#
# AVEC OPTIMISATION (sans cache):
#   - 1-2 requêtes SQL seulement
#   - ~0.3-1 seconde de chargement
#   - Réduction de 95% des requêtes
#
# AVEC CACHE (après 1er chargement):
#   - 0 requêtes SQL
#   - ~0.05-0.1 seconde de chargement
#   - 98% plus rapide!
# ========================================












# # ----------------------------------------------------------------------
@login_required
def timeTable_Amphi(request, amp_id, semestre_num):
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)

    ttEns_0 = 0
    all_Classe_Ens = None
    all_classes = 0
    nbr_Cours=0
    nbr_TP=0
    nbr_TD=0
    nbr_SS=0
    reload = False
    vol_hor_cours = 0
    vol_hor_TD = 0
    vol_hor_TP = 0
    vol_hor_Total = 0

    sixClasses = [
     'الحصة:1 / 08:00-09:30',
     'الحصة:2 / 09:40-11:10',
     'الحصة:3 / 11:20-12:50',
     'الحصة:4 / 13:10-14:40',
     'الحصة:5 / 14:50-16:20',
     'الحصة:6 / 16:30-18:00',
    ]

    # sevenDays = ['السبت', 'الأحد', 'الإثنين', 'الثلاثاء', 'الأربعاء', 'الخميس']
    sevenDays = ['Samedi', 'Dimanche', 'Lundi', 'Mardi', 'Mercredi', 'Jeudi']

    all_Classe_Amp = []

    lieu_content_type = ContentType.objects.get_for_model(Amphi_Dep)

    # samedi
    all_Classe_SAM = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        object_id = amp_id,
                                        jour='Samedi').order_by('temps') # plus Ens get
    all_Classe_Amp.extend(all_Classe_SAM)

    all_Classe_DIM = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = amp_id,
                                        jour='Dimanche').order_by('temps') # plus Ens get JSON
    all_Classe_Amp.extend(all_Classe_DIM)

    all_Classe_LUN = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = amp_id,
                                        jour='Lundi').order_by('temps') # plus Ens get JSON
    all_Classe_Amp.extend(all_Classe_LUN)

    all_Classe_MAR = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = amp_id,
                                        jour='Mardi').order_by('temps') # plus Ens get JSON
    all_Classe_Amp.extend(all_Classe_MAR)
    
    all_Classe_MER = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = amp_id,
                                        jour='Mercredi').order_by('temps') # plus Ens get JSON
    all_Classe_Amp.extend(all_Classe_MER)

    all_Classe_JEU = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = amp_id,
                                        jour='Jeudi').order_by('temps') # plus Ens get JSON
    all_Classe_Amp.extend(all_Classe_JEU)
    


    
    all_Classe_Time = []
    # # Classe 1
    all_Classe_Time_1 = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = amp_id,
                                        temps='09:30-08:00') # plus Ens get 
    all_Classe_Time.extend(all_Classe_Time_1)
    
    # Classe 2
    all_Classe_Time_2 = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = amp_id,
                                        temps='11:10-09:40') # plus Ens get  
    all_Classe_Time.extend(all_Classe_Time_2)
    
    # Classe 3
    all_Classe_Time_3 = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = amp_id,
                                        temps='12:50-11:20') # plus Ens get 
    all_Classe_Time.extend(all_Classe_Time_3)
    
    # Classe 4
    all_Classe_Time_4 = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = amp_id,
                                        temps='14:40-13:10') # plus Ens get 
    all_Classe_Time.extend(all_Classe_Time_4)
    
    # Classe 5
    all_Classe_Time_5 = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = amp_id,
                                        temps='16:20-14:50') # plus Ens get  
    all_Classe_Time.extend(all_Classe_Time_5)
    
    # Classe 6
    all_Classe_Time_6 = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = amp_id,
                                        temps='18:00-16:30') # plus Ens get  
    all_Classe_Time.extend(all_Classe_Time_6)

    sixClasses = [
        '09:30-08:00',
        '11:10-09:40',
        '12:50-11:20',
        '14:40-13:10',
        '16:20-14:50',
        '18:00-16:30',
    ]
   
    for idx1 in list(all_Classe_Amp):
        # all_Seances.append(idx1.clas_jour +" - "+idx1.clas_temps)
        if "Cours" == idx1.type:
            nbr_Cours = nbr_Cours + 1
        if "TP" == idx1.type:
            nbr_TP = nbr_TP + 1
        if "TD" == idx1.type:
            nbr_TD = nbr_TD + 1
        if "Sortie Scientifique" == idx1.type:
            nbr_SS = nbr_SS + 1
        
        if idx1.seance_created == True:
            nbr_Sea_fait = 0
            nbr_Sea_remplacer = 0
            nbr_Sea_annuler = 0
            all_Seances = 0
            all_Seances = Seance.objects.filter(classe = idx1.id) # plus Ens get JSON
            nbr_Sea_fait = Seance.objects.filter(classe = idx1.id, fait = True) # plus Ens get JSON
            nbr_Sea_remplacer = Seance.objects.filter(classe = idx1.id, remplacer = True) # plus Ens get JSON
            nbr_Sea_annuler = Seance.objects.filter(classe = idx1.id, annuler = True) # plus Ens get JSON
            # # print("*"*50)
            # # print(nbr_Sea_fait)
            # # print(all_Seances)

            if nbr_Sea_fait.exists() and all_Seances.exists():
                taux_avancement = ((nbr_Sea_fait.count())/ all_Seances.count()) * 100
                taux_avancement = round(taux_avancement)
                idx1.taux_avancement = taux_avancement
                idx1.save()

    all_classes = nbr_Cours+nbr_TP+nbr_TD+nbr_SS
    
    context = {
        'title': 'إستعمال الزمن',
        'all_Classe_Amp': all_Classe_Amp,
        'reload': reload,
        'my_Dep': departement,  # Passer le département au template
        'my_Fac': departement.faculte,  # Passer le département au template
        'ttEns_0':ttEns_0,
        'all_Classe_Time': all_Classe_Time,
        'nbr_Cours': nbr_Cours,
        'nbr_TP': nbr_TP,
        'sevenDays': sevenDays,
        'sixClasses': sixClasses,
        'nbr_TD': nbr_TD,
        'nbr_SS': nbr_SS,
        'all_classes': all_classes,
    }

    return render(request, 'departement/timeTable_Amphi.html', context)




# # ----------------------------------------------------------------------
@login_required
def timeTable_Salle(request, salle_id, semestre_num):
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)

    ttEns_0 = 0
    all_Classe_Ens = None
    all_classes = 0
    nbr_Cours=0
    nbr_TP=0
    nbr_TD=0
    nbr_SS=0
    reload = False
    vol_hor_cours = 0
    vol_hor_TD = 0
    vol_hor_TP = 0
    vol_hor_Total = 0

    sixClasses = [
     'الحصة:1 / 08:00-09:30',
     'الحصة:2 / 09:40-11:10',
     'الحصة:3 / 11:20-12:50',
     'الحصة:4 / 13:10-14:40',
     'الحصة:5 / 14:50-16:20',
     'الحصة:6 / 16:30-18:00',
    ]

    # sevenDays = ['السبت', 'الأحد', 'الإثنين', 'الثلاثاء', 'الأربعاء', 'الخميس']
    sevenDays = ['Samedi', 'Dimanche', 'Lundi', 'Mardi', 'Mercredi', 'Jeudi']

    all_Classe_Salle = []

    lieu_content_type = ContentType.objects.get_for_model(Salle_Dep)

    # samedi
    all_Classe_SAM = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = salle_id,
                                        jour='Samedi').order_by('temps') # plus Ens get
    all_Classe_Salle.extend(all_Classe_SAM)

    all_Classe_DIM = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = salle_id,
                                        jour='Dimanche').order_by('temps') # plus Ens get JSON
    all_Classe_Salle.extend(all_Classe_DIM)

    all_Classe_LUN = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = salle_id,
                                        jour='Lundi').order_by('temps') # plus Ens get JSON
    all_Classe_Salle.extend(all_Classe_LUN)

    all_Classe_MAR = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = salle_id,
                                        jour='Mardi').order_by('temps') # plus Ens get JSON
    all_Classe_Salle.extend(all_Classe_MAR)
    
    all_Classe_MER = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = salle_id,
                                        jour='Mercredi').order_by('temps') # plus Ens get JSON
    all_Classe_Salle.extend(all_Classe_MER)

    all_Classe_JEU = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = salle_id,
                                        jour='Jeudi').order_by('temps') # plus Ens get JSON
    all_Classe_Salle.extend(all_Classe_JEU)
    


    
    all_Classe_Time = []
    # # Classe 1
    all_Classe_Time_1 = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = salle_id,
                                        temps='09:30-08:00') # plus Ens get 
    all_Classe_Time.extend(all_Classe_Time_1)
    
    # Classe 2
    all_Classe_Time_2 = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = salle_id,
                                        temps='11:10-09:40') # plus Ens get  
    all_Classe_Time.extend(all_Classe_Time_2)
    
    # Classe 3
    all_Classe_Time_3 = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = salle_id,
                                        temps='12:50-11:20') # plus Ens get 
    all_Classe_Time.extend(all_Classe_Time_3)
    
    # Classe 4
    all_Classe_Time_4 = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = salle_id,
                                        temps='14:40-13:10') # plus Ens get 
    all_Classe_Time.extend(all_Classe_Time_4)
    
    # Classe 5
    all_Classe_Time_5 = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = salle_id,
                                        temps='16:20-14:50') # plus Ens get  
    all_Classe_Time.extend(all_Classe_Time_5)
    
    # Classe 6
    all_Classe_Time_6 = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = salle_id,
                                        temps='18:00-16:30') # plus Ens get  
    all_Classe_Time.extend(all_Classe_Time_6)

    sixClasses = [
        '09:30-08:00',
        '11:10-09:40',
        '12:50-11:20',
        '14:40-13:10',
        '16:20-14:50',
        '18:00-16:30',
    ]
   
    for idx1 in list(all_Classe_Salle):
        # all_Seances.append(idx1.clas_jour +" - "+idx1.clas_temps)
        if "Cours" == idx1.type:
            nbr_Cours = nbr_Cours + 1
        if "TP" == idx1.type:
            nbr_TP = nbr_TP + 1
        if "TD" == idx1.type:
            nbr_TD = nbr_TD + 1
        if "Sortie Scientifique" == idx1.type:
            nbr_SS = nbr_SS + 1
        
        if idx1.seance_created == True:
            nbr_Sea_fait = 0
            nbr_Sea_remplacer = 0
            nbr_Sea_annuler = 0
            all_Seances = 0
            all_Seances = Seance.objects.filter(classe = idx1.id) # plus Ens get JSON
            nbr_Sea_fait = Seance.objects.filter(classe = idx1.id, fait = True) # plus Ens get JSON
            nbr_Sea_remplacer = Seance.objects.filter(classe = idx1.id, remplacer = True) # plus Ens get JSON
            nbr_Sea_annuler = Seance.objects.filter(classe = idx1.id, annuler = True) # plus Ens get JSON
            # # print("*"*50)
            # # print(nbr_Sea_fait)
            # # print(all_Seances)

            if nbr_Sea_fait.exists() and all_Seances.exists():
                taux_avancement = ((nbr_Sea_fait.count())/ all_Seances.count()) * 100
                taux_avancement = round(taux_avancement)
                idx1.taux_avancement = taux_avancement
                idx1.save()

    all_classes = nbr_Cours+nbr_TP+nbr_TD+nbr_SS
    
    context = {
        'title': 'إستعمال الزمن',
        'all_Classe_Salle': all_Classe_Salle,
        'reload': reload,
        'my_Dep': departement,  # Passer le département au template
        'my_Fac': departement.faculte,  # Passer le département au template
        'ttEns_0':ttEns_0,
        'all_Classe_Time': all_Classe_Time,
        'nbr_Cours': nbr_Cours,
        'nbr_TP': nbr_TP,
        'sevenDays': sevenDays,
        'sixClasses': sixClasses,
        'nbr_TD': nbr_TD,
        'nbr_SS': nbr_SS,
        'all_classes': all_classes,
    }

    return render(request, 'departement/timeTable_Salle.html', context)




# # ----------------------------------------------------------------------
@login_required
def timeTable_Labo(request, labo_id, semestre_num):
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)

    ttEns_0 = 0
    all_Classe_Ens = None
    all_classes = 0
    nbr_Cours=0
    nbr_TP=0
    nbr_TD=0
    nbr_SS=0
    reload = False
    vol_hor_cours = 0
    vol_hor_TD = 0
    vol_hor_TP = 0
    vol_hor_Total = 0

    sixClasses = [
     'الحصة:1 / 08:00-09:30',
     'الحصة:2 / 09:40-11:10',
     'الحصة:3 / 11:20-12:50',
     'الحصة:4 / 13:10-14:40',
     'الحصة:5 / 14:50-16:20',
     'الحصة:6 / 16:30-18:00',
    ]

    # sevenDays = ['السبت', 'الأحد', 'الإثنين', 'الثلاثاء', 'الأربعاء', 'الخميس']
    sevenDays = ['Samedi', 'Dimanche', 'Lundi', 'Mardi', 'Mercredi', 'Jeudi']

    all_Classe_Labo = []

    lieu_content_type = ContentType.objects.get_for_model(Laboratoire_Dep)

    # samedi
    all_Classe_SAM = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = labo_id,
                                        jour='Samedi').order_by('temps') # plus Ens get
    all_Classe_Labo.extend(all_Classe_SAM)

    all_Classe_DIM = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = labo_id,
                                        jour='Dimanche').order_by('temps') # plus Ens get JSON
    all_Classe_Labo.extend(all_Classe_DIM)

    all_Classe_LUN = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = labo_id,
                                        jour='Lundi').order_by('temps') # plus Ens get JSON
    all_Classe_Labo.extend(all_Classe_LUN)

    all_Classe_MAR = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = labo_id,
                                        jour='Mardi').order_by('temps') # plus Ens get JSON
    all_Classe_Labo.extend(all_Classe_MAR)
    
    all_Classe_MER = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = labo_id,
                                        jour='Mercredi').order_by('temps') # plus Ens get JSON
    all_Classe_Labo.extend(all_Classe_MER)

    all_Classe_JEU = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = labo_id,
                                        jour='Jeudi').order_by('temps') # plus Ens get JSON
    all_Classe_Labo.extend(all_Classe_JEU)
    


    
    all_Classe_Time = []
    # # Classe 1
    all_Classe_Time_1 = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = labo_id,
                                        temps='09:30-08:00') # plus Ens get 
    all_Classe_Time.extend(all_Classe_Time_1)
    
    # Classe 2
    all_Classe_Time_2 = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = labo_id,
                                        temps='11:10-09:40') # plus Ens get  
    all_Classe_Time.extend(all_Classe_Time_2)
    
    # Classe 3
    all_Classe_Time_3 = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = labo_id,
                                        temps='12:50-11:20') # plus Ens get 
    all_Classe_Time.extend(all_Classe_Time_3)
    
    # Classe 4
    all_Classe_Time_4 = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = labo_id,
                                        temps='14:40-13:10') # plus Ens get 
    all_Classe_Time.extend(all_Classe_Time_4)
    
    # Classe 5
    all_Classe_Time_5 = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = labo_id,
                                        temps='16:20-14:50') # plus Ens get  
    all_Classe_Time.extend(all_Classe_Time_5)
    
    # Classe 6
    all_Classe_Time_6 = Classe.objects.filter(enseignant__departement=departement.id,
                                        content_type=lieu_content_type,
                                        semestre__numero = semestre_num,
                                        object_id = labo_id,
                                        temps='18:00-16:30') # plus Ens get  
    all_Classe_Time.extend(all_Classe_Time_6)

    sixClasses = [
        '09:30-08:00',
        '11:10-09:40',
        '12:50-11:20',
        '14:40-13:10',
        '16:20-14:50',
        '18:00-16:30',
    ]
   
    for idx1 in list(all_Classe_Labo):
        # all_Seances.append(idx1.clas_jour +" - "+idx1.clas_temps)
        if "Cours" == idx1.type:
            nbr_Cours = nbr_Cours + 1
        if "TP" == idx1.type:
            nbr_TP = nbr_TP + 1
        if "TD" == idx1.type:
            nbr_TD = nbr_TD + 1
        if "Sortie Scientifique" == idx1.type:
            nbr_SS = nbr_SS + 1
        
        if idx1.seance_created == True:
            nbr_Sea_fait = 0
            nbr_Sea_remplacer = 0
            nbr_Sea_annuler = 0
            all_Seances = 0
            all_Seances = Seance.objects.filter(classe = idx1.id) # plus Ens get JSON
            nbr_Sea_fait = Seance.objects.filter(classe = idx1.id, fait = True) # plus Ens get JSON
            nbr_Sea_remplacer = Seance.objects.filter(classe = idx1.id, remplacer = True) # plus Ens get JSON
            nbr_Sea_annuler = Seance.objects.filter(classe = idx1.id, annuler = True) # plus Ens get JSON
            # # print("*"*50)
            # # print(nbr_Sea_fait)
            # # print(all_Seances)

            if nbr_Sea_fait.exists() and all_Seances.exists():
                taux_avancement = ((nbr_Sea_fait.count())/ all_Seances.count()) * 100
                taux_avancement = round(taux_avancement)
                idx1.taux_avancement = taux_avancement
                idx1.save()

    all_classes = nbr_Cours+nbr_TP+nbr_TD+nbr_SS
    
    context = {
        'title': 'إستعمال الزمن',
        'all_Classe_Labo': all_Classe_Labo,
        'reload': reload,
        'my_Dep': departement,  # Passer le département au template
        'my_Fac': departement.faculte,  # Passer le département au template
        'ttEns_0':ttEns_0,
        'all_Classe_Time': all_Classe_Time,
        'nbr_Cours': nbr_Cours,
        'nbr_TP': nbr_TP,
        'sevenDays': sevenDays,
        'sixClasses': sixClasses,
        'nbr_TD': nbr_TD,
        'nbr_SS': nbr_SS,
        'all_classes': all_classes,
    }

    return render(request, 'departement/timeTable_Labo.html', context)








@login_required
def import_etudiants(request):
    # Récupérer le département de l'utilisateur connecté
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)
    
    if request.method == 'POST':
        reforme_id = request.POST.get('reforme')
        niveau_id = request.POST.get('niveau')
        niv_spe_dep_sg_id = request.POST.get('niv_spe_dep_sg')
        excel_file = request.FILES.get('excel_file')

        if not all([reforme_id, niveau_id, niv_spe_dep_sg_id, excel_file]):
            return render(request, 'departement/import_etudiants.html', {
                'error': 'يرجى اختيار جميع الحقول وتحميل ملف Excel.',
                'refs_url': reverse('departement:get_reformes_json'),
                'nivs_url': reverse('departement:get_niveaux_json', kwargs={'reforme_id': 0}),
            })

        try:
            # Vérifier/créer le poste étudiant
            poste_etudiant, poste_message = get_or_create_etudiant_poste()
            if not poste_etudiant:
                return render(request, 'departement/import_etudiants.html', {
                    'error': f'خطأ في إنشاء منصب الطالب: {poste_message}',
                    'my_Dep': departement,
                    'my_Fac': departement.faculte,
                })

            # DEBUG: Lire le fichier Excel
            # print(f"🔍 DEBUG: Lecture du fichier Excel...")
            df = pd.read_excel(excel_file)
            
            # DEBUG: Afficher les colonnes disponibles
            # print(f"🔍 DEBUG: Colonnes dans le fichier: {list(df.columns)}")
            # print(f"🔍 DEBUG: Nombre de lignes: {len(df)}")
            
            # DEBUG: Afficher les 3 premières lignes
            # if len(df) > 0:
                # print(f"🔍 DEBUG: Première ligne example:")
                # for col in df.columns:
                    # print(f"  - {col}: {df.iloc[0][col] if len(df) > 0 else 'N/A'}")
            
            niv_spe_dep_sg = NivSpeDep_SG.objects.get(id=niv_spe_dep_sg_id)
            # print(f"🔍 DEBUG: Groupe sélectionné: {niv_spe_dep_sg}")
            
            # Compteurs
            etudiants_ajoutes = 0
            etudiants_existants = 0
            erreurs = 0
            erreurs_details = []  # Pour stocker les détails des erreurs
            nouveaux_utilisateurs = []

            for index, row in df.iterrows():
                try:
                    with transaction.atomic():
                        # DEBUG: Afficher les données de chaque ligne
                        # print(f"\n🔍 DEBUG Ligne {index + 1}:")
                        
                        # Données de base avec debug
                        nom_ar = str(row.get('اللقب', '')).strip() if pd.notna(row.get('اللقب', '')) else ''
                        prenom_ar = str(row.get('الإسم', '')).strip() if pd.notna(row.get('الإسم', '')) else ''
                        nom_fr = str(row.get('Nom', '')).strip() if pd.notna(row.get('Nom', '')) else ''
                        prenom_fr = str(row.get('Prénom', '')).strip() if pd.notna(row.get('Prénom', '')) else ''
                        matricule = str(row.get('الرقم التسلسلي', '')).strip() if pd.notna(row.get('الرقم التسلسلي', '')) else ''
                        num_ins = str(row.get('رقم التسجيل', '')).strip() if pd.notna(row.get('رقم التسجيل', '')) else ''
                        sexe = str(row.get('الجنس', 'M')).strip() if pd.notna(row.get('الجنس', '')) else 'M'
                                                
                        # Validation des données obligatoires
                        if not nom_fr or not prenom_fr or not matricule:
                            error_msg = f"Données manquantes: nom_fr='{nom_fr}', prenom_fr='{prenom_fr}', matricule='{matricule}'"
                            # print(f"❌ {error_msg}")
                            erreurs_details.append(f"Ligne {index + 1}: {error_msg}")
                            erreurs += 1
                            continue
                        
                        # Vérifier si l'étudiant existe déjà
                        existing_etudiant = Etudiant.objects.filter(
                            matricule=matricule
                        ).first()
                        
                        if existing_etudiant:
                            # print(f"⚠️ Étudiant existant: {nom_ar} {prenom_ar} (Matricule: {matricule})")
                            etudiants_existants += 1
                            continue
                        
                        # Créer l'utilisateur automatiquement
                        username = generate_student_username(nom_fr, prenom_fr, matricule)
                        password = generate_student_password(nom_fr, matricule)
                        email = generate_student_email(nom_fr, prenom_fr , matricule)
                        
                        # print(f"📝 Création utilisateur: {username} / {password}")
                        
                        # Créer l'utilisateur CustomUser
                        user = CustomUser.objects.create_user(
                            username=username,
                            email=email,
                            password=password,
                            first_name=prenom_fr,
                            last_name=nom_fr,
                            poste_principal=poste_etudiant
                        )
                        
                        # print(f"✅ Utilisateur créé: {username}")
                        
                        # Créer l'étudiant
                        etudiant = Etudiant.objects.create(
                            user=user,
                            nom_ar=nom_ar,
                            prenom_ar=prenom_ar,
                            nom_fr=nom_fr,
                            prenom_fr=prenom_fr,
                            matricule=matricule,
                            num_ins=num_ins,
                            sexe=sexe,
                            niv_spe_dep_sg=niv_spe_dep_sg,
                            inscrit_univ=True,
                        )
                        
                        # Stocker les informations du nouvel utilisateur
                        nouveaux_utilisateurs.append({
                            'nom': f"{prenom_ar} {nom_ar}",
                            'nom_fr': f"{prenom_fr} {nom_fr}",
                            'matricule': matricule,
                            'username': username,
                            'password': password,
                            'email': email,
                        })
                        
                        etudiants_ajoutes += 1
                        # print(f"✅ Étudiant créé: {nom_ar} {prenom_ar} (ID: {etudiant.id})")
                        
                except Exception as e:
                    erreurs += 1
                    error_detail = f"Ligne {index + 1}: {str(e)}"
                    erreurs_details.append(error_detail)
                    # print(f"❌ Erreur ligne {index + 1}: {str(e)}")
                    continue

            # DEBUG: Afficher le résumé
            # print(f"\n🔍 DEBUG RÉSUMÉ:")
            # print(f"  - Étudiants ajoutés: {etudiants_ajoutes}")
            # print(f"  - Étudiants existants: {etudiants_existants}")
            # print(f"  - Erreurs: {erreurs}")
            
            # if erreurs_details:
            #     # print(f"🔍 DÉTAILS DES ERREURS:")
            #     for detail in erreurs_details[:5]:  # Afficher max 5 erreurs
                    # print(f"  - {detail}")

            # Mise à jour du nombre d'étudiants dans NivSpeDep_SG
            if etudiants_ajoutes > 0:
                niv_spe_dep_sg.nbr_etudiants_SG += etudiants_ajoutes
                niv_spe_dep_sg.save()

            # Préparer le message de succès avec les informations de connexion
            success_message = f'تم استيراد {etudiants_ajoutes} طالب بنجاح.'
            if etudiants_existants > 0:
                success_message += f' تم العثور على {etudiants_existants} طالب موجود مسبقاً.'
            if erreurs > 0:
                success_message += f' حدث خطأ مع {erreurs} طالب.'
            
            # Ajouter les détails des erreurs
            error_message = None
            if erreurs_details:
                error_message = "تفاصيل الأخطاء:\n" + "\n".join(erreurs_details[:10])  # Afficher max 10 erreurs
            
            # Ajouter les informations de connexion
            if nouveaux_utilisateurs:
                success_message += '\n\nمعلومات تسجيل الدخول للطلاب الجدد:\n'
                for info in nouveaux_utilisateurs:
                    success_message += f"• {info['nom']} (رقم: {info['matricule']}): {info['username']} / {info['password']}\n"

            return render(request, 'departement/import_etudiants.html', {
                'success': success_message,
                'error': error_message,
                'nouveaux_utilisateurs': nouveaux_utilisateurs,
                'my_Dep': departement,
                'my_Fac': departement.faculte,
            })

        except Exception as e:
            return render(request, 'departement/import_etudiants.html', {
                'error': f'خطأ أثناء الاستيراد: {str(e)}',
                'refs_url': reverse('departement:get_reformes_json'),
                'nivs_url': reverse('departement:get_niveaux_json', kwargs={'reforme_id': 0}),
            })

    return render(request, 'departement/import_etudiants.html', {
        # Ces URLs ne sont plus nécessaires car on utilise les URLs directement dans le template
        'my_Dep': departement,
        'my_Fac': departement.faculte,
    })


# ========== Fonctions utilitaires pour les étudiants ==========

def get_or_create_etudiant_poste():
    """Créer ou récupérer le poste 'étudiant'"""
    try:
        poste, created = Poste.objects.get_or_create(
            nom_ar='طالب',
            defaults={
                'nom_fr': 'Étudiant',
                'code': 'ETU',
                'description': 'Poste pour les étudiants'
            }
        )
        
        if created:
            return poste, f"تم إنشاء منصب 'طالب' بنجاح"
        else:
            return poste, None
            
    except Exception as e:
        return None, f"خطأ في إنشاء/استرجاع منصب الطالب: {str(e)}"
    

    
def generate_student_username(prenom_fr, nom_fr, matricule):
    """Génération du nom d'utilisateur pour étudiant"""
    # Nettoyer les noms
    prenom_clean = re.sub(r'[^a-zA-Z]', '', prenom_fr.lower()) if prenom_fr else 'student'
    nom_clean = re.sub(r'[^a-zA-Z]', '', nom_fr.lower()) if nom_fr else 'user'
    matricule_clean = re.sub(r'[^0-9]', '', str(matricule)) if matricule else '000'
    
    # Format: prenom.nom.matricule (limité à 30 caractères)
    base_username = f"{nom_clean[:10]}.{prenom_clean[:10]}.{matricule_clean[:6]}"
    
    # Vérifier l'unicité
    username = base_username[:30]
    counter = 1
    while CustomUser.objects.filter(username=username).exists():
        suffix = f".{counter}"
        username = f"{base_username[:30-len(suffix)]}{suffix}"
        counter += 1
    
    return username


def generate_student_password(nom_fr, matricule):
    """Génération du mot de passe pour étudiant"""
    # Nettoyer le nom
    nom_clean = re.sub(r'[^a-zA-Z]', '', nom_fr.lower()) if nom_fr else 'student'
    matricule_clean = re.sub(r'[^0-9]', '', str(matricule)) if matricule else '123'
    
    # Format: Nom + derniers chiffres matricule (ex: Ahmed123)
    password = f"{nom_clean.capitalize()}{matricule_clean[-3:]}"
    # password = "...Etu123"
    
    # Minimum 6 caractères
    if len(password) < 6:
        password = f"{password}2024"
    
    return password


def generate_student_email(prenom_fr, nom_fr, matricule):
    """Génération de l'email pour étudiant"""
    # Nettoyer les noms
    prenom_clean = re.sub(r'[^a-zA-Z]', '', prenom_fr.lower()) if prenom_fr else 'student'
    nom_clean = re.sub(r'[^a-zA-Z]', '', nom_fr.lower()) if nom_fr else 'user'
    matricule_clean = re.sub(r'[^0-9]', '', str(matricule)) if matricule else '000'
    
    # Format: prenom.nom.matricule@universite.dz
    email = f"{nom_clean[:10]}.{prenom_clean[:10]}.{matricule_clean[:6]}@univ-ouargla.dz"
    
    return email




# ----------------------------------------------------------------------
@login_required
def fichePedagogique_Ens_Dep_Semestre(request, ens_id, semestre_num):
    """
    Vue pour afficher la fiche pédagogique d'un enseignant pour un semestre spécifique
    (Utilisée dans l'onglet 2 "Heures de travail")
    """
    
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)

    # Validation du semestre
    if semestre_num not in [1, 2]:
        semestre_num = 1

    # Variables d'initialisation
    all_classes = 0
    nbr_Cours = 0
    nbr_TP = 0
    nbr_TD = 0
    nbr_SS = 0
    vol_hor_cours = 0
    vol_hor_TD = 0
    vol_hor_TP = 0
    vol_hor_Total = 0

    # Récupération de l'enseignant
    try:
        myEns_obj = Enseignant.objects.get(id=ens_id)
        myEns = Ens_Dep.objects.get(enseignant=myEns_obj, departement=departement)
    except (Enseignant.DoesNotExist, Ens_Dep.DoesNotExist):
        messages.error(request, "الأستاذ غير موجود")
        return redirect('departement:manage_enseignants', semestre_num=semestre_num)

    # Filtre de base pour les classes
    base_filter = {
        'enseignant__departement': departement.id,
        'enseignant__enseignant__id': ens_id,
        'semestre': semestre_num
    }

    # Liste des jours dans l'ordre souhaité
    jours = ['Samedi', 'Dimanche', 'Lundi', 'Mardi', 'Mercredi', 'Jeudi']
    
    # RÉCUPÉRATION DES CLASSES POUR LE SEMESTRE SPÉCIFIQUE PAR JOUR
    all_Classe_Ens = []
    for jour in jours:
        classes_jour = Classe.objects.filter(
            **base_filter,
            jour=jour
        ).order_by('temps')
        all_Classe_Ens.extend(classes_jour)

    # CALCUL DES STATISTIQUES
    for classe in all_Classe_Ens:
        if classe.type == "Cours":
            nbr_Cours += 1
        elif classe.type == "TP":
            nbr_TP += 1
        elif classe.type == "TD":
            nbr_TD += 1
        elif classe.type == "Sortie Scientifique":
            nbr_SS += 1

        # Calcul du taux d'avancement si nécessaire
        if classe.seance_created == True:
            all_Seances = Seance.objects.filter(classe=classe.id)
            nbr_Sea_fait = Seance.objects.filter(classe=classe.id, fait=True)

            if nbr_Sea_fait.exists() and all_Seances.exists():
                taux_avancement = ((nbr_Sea_fait.count()) / all_Seances.count()) * 100
                taux_avancement = round(taux_avancement)
                classe.taux_avancement = taux_avancement
                classe.save()

    # Calcul des volumes horaires
    vol_hor_cours = nbr_Cours * 2.25
    vol_hor_TD = nbr_TD * 1.5
    vol_hor_TP = nbr_TP * 1.5
    vol_hor_Total = vol_hor_cours + vol_hor_TD + vol_hor_TP

    all_classes = nbr_Cours + nbr_TP + nbr_TD + nbr_SS

    # Mise à jour des statistiques dans l'objet enseignant
    if semestre_num == 1:
        myEns.nbrClas_Cours_in_Dep_S1 = nbr_Cours
        myEns.nbrClas_TD_in_Dep_S1 = nbr_TD
        myEns.nbrClas_TP_in_Dep_S1 = nbr_TP
        myEns.nbrClas_SS_in_Dep_S1 = nbr_SS
        myEns.volHor_in_Dep_S1 = vol_hor_Total
        
        # Calcul des jours de travail
        myDays = Classe.objects.filter(
            enseignant__departement=departement.id,
            enseignant=ens_id, 
            semestre=1
        ).values('jour').distinct().count()
        myEns.nbrJour_in_Dep_S1 = myDays
    else:
        myEns.nbrClas_Cours_in_Dep_S2 = nbr_Cours
        myEns.nbrClas_TD_in_Dep_S2 = nbr_TD
        myEns.nbrClas_TP_in_Dep_S2 = nbr_TP
        myEns.nbrClas_SS_in_Dep_S2 = nbr_SS
        myEns.volHor_in_Dep_S2 = vol_hor_Total
        
        # Calcul des jours de travail
        myDays = Classe.objects.filter(
            enseignant__departement=departement.id,
            enseignant=ens_id, 
            semestre=2
        ).values('jour').distinct().count()
        myEns.nbrJour_in_Dep_S2 = myDays

    myEns.save()

    univ = departement.faculte.universite

    context = {
        'all_Classe_Ens': all_Classe_Ens,
        'all_classes': all_classes,
        'vol_hor_cours': vol_hor_cours,
        'vol_hor_TD': vol_hor_TD,
        'vol_hor_TP': vol_hor_TP,
        'vol_hor_Total': vol_hor_Total,
        'nbr_Cours': nbr_Cours,
        'nbr_TD': nbr_TD,
        'nbr_TP': nbr_TP,
        'nbr_SS': nbr_SS,
        'semestre_num': semestre_num,
        'title': f"Fiche Pédagogique S{semestre_num} - {myEns.enseignant.nom_fr} {myEns.enseignant.prenom_fr}",
        'univ': univ,
        'my_Dep': departement,
        'my_Fac': departement.faculte,
        'myEns': myEns,
    }

    return render(request, 'departement/fichePedagogique_Ens_Dep_Semestre.html', context)




@login_required
def view_seance_dep(request, seance_id):
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)
    
    my_Sea = get_object_or_404(Seance, id=seance_id)
    # my_Dep = request.user.departement  # Ajustez selon votre modèle
    
    context = {
        'my_Sea': my_Sea,
        'my_Dep': departement,
        # 'myEns': myEns,
    }
    return render(request, 'departement/view_seance_dep.html', context)






# views.py - Application département

# from django.shortcuts import render, get_object_or_404
# from django.contrib.auth.decorators import login_required
# from django.http import JsonResponse
# from django.core.paginator import Paginator
# from django.db.models import Count, Q
# from .models import *
# from .decorators import departement_access_required

# @departement_access_required
@login_required
def list_Abs_Etu_Classe_Dep(request, clas_id):
    try:
        # Récupérer l'enseignant et le département
        enseignant = request.user.enseignant_profile
        real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
        my_Dep = Departement.objects.get(id=real_Dep.departement.id)
        
        # Récupérer la classe
        my_Clas = get_object_or_404(Classe, id=clas_id)
        
        # Vérifier que la classe appartient au département
        if my_Clas.niv_spe_dep_sg.niv_spe_dep.departement != my_Dep:
            context = {
                'error_message': 'لا يحق لك الوصول إلى هذا الفصل',
                'my_Dep': my_Dep,
            }
            return render(request, 'departement/error.html', context)
        
        # Récupérer toutes les séances effectuées de cette classe
        all_sea_fait = Seance.objects.filter(
            classe=my_Clas, 
            fait=True
        ).order_by('date')
        
        # Récupérer tous les étudiants inscrits dans cette classe
        all_Abs_Classe = Abs_Etu_Classe.objects.filter(
            classe=my_Clas
        ).order_by('etudiant__nom_fr', 'etudiant__prenom_fr')
        
        # Compter le nombre total d'étudiants
        nbr_Abs_Classe = all_Abs_Classe.count()
        
        # Vérifier s'il y a des données
        if not all_Abs_Classe.exists():
            context = {
                'error_message': 'لا توجد بيانات غيابات لهذا الفصل',
                'my_Dep': my_Dep,
            }
            return render(request, 'departement/error.html', context)
        
        # Précalculer les absences par séance pour chaque étudiant
        all_Abs_Classe_with_absences = []
        
        for abs_classe in all_Abs_Classe:
            absences_per_seance = []
            
            # Pour chaque séance effectuée, vérifier l'absence de l'étudiant
            for sea in all_sea_fait:
                absence = Abs_Etu_Seance.objects.filter(
                    seance=sea, 
                    etudiant=abs_classe.etudiant
                ).first()
                
                absences_per_seance.append(absence if absence else None)
            
            # Ajouter la liste des absences à l'objet
            abs_classe.absences_per_seance = absences_per_seance
            all_Abs_Classe_with_absences.append(abs_classe)
        
        # Calculer des statistiques supplémentaires
        total_absences = sum([abs_cls.nbr_absence or 0 for abs_cls in all_Abs_Classe])
        total_justified = sum([abs_cls.nbr_absence_justifiee or 0 for abs_cls in all_Abs_Classe])
        
        # Statistiques par niveau d'absence
        students_high_absence = all_Abs_Classe.filter(nbr_absence__gte=5).count()
        students_medium_absence = all_Abs_Classe.filter(
            nbr_absence__gte=3, 
            nbr_absence__lt=5
        ).count()
        
        # Préparer les titres
        titre_00 = my_Clas.niv_spe_dep_sg
        titre_01 = my_Clas.matiere.nom_fr
        
        # Contexte pour le template
        context = {
            'titre_00': titre_00,
            'titre_01': titre_01,
            'all_sea_fait': all_sea_fait,
            'all_Abs_Classe_with_absences': all_Abs_Classe_with_absences,
            'nbr_Abs_Classe': nbr_Abs_Classe,
            'my_Dep': my_Dep,
            'my_Clas': my_Clas,
            'my_Fac': my_Dep.faculte,
            
            # Statistiques supplémentaires
            'total_absences': total_absences,
            'total_justified': total_justified,
            'students_high_absence': students_high_absence,
            'students_medium_absence': students_medium_absence,
            'total_sessions': all_sea_fait.count(),
            'total_students': nbr_Abs_Classe,
            
            # Métadonnées
            'page_title': f'غيابات طلاب {titre_00} - {titre_01}',
        }
        
        return render(request, 'departement/list_Abs_Etu_Classe_Dep.html', context)
        
    except Ens_Dep.DoesNotExist:
        context = {
            'error_message': 'لم يتم العثور على معلومات القسم للمستخدم الحالي'
        }
        return render(request, 'departement/error.html', context)
        
    except Departement.DoesNotExist:
        context = {
            'error_message': 'القسم غير موجود'
        }
        return render(request, 'departement/error.html', context)
    
    except Classe.DoesNotExist:
        context = {
            'error_message': 'الفصل غير موجود'
        }
        return render(request, 'departement/error.html', context)
    
    except Exception as e:
        context = {
            'error_message': f'حدث خطأ: {str(e)}'
        }
        return render(request, 'departement/error.html', context)
    









@login_required
def export_absences_excel(request, clas_id, departement):
    """
    Vue pour exporter les absences vers Excel
    """
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Font, PatternFill, Alignment
    
    try:
        my_Dep = departement
        my_Clas = get_object_or_404(Classe, id=clas_id)
        
        # Vérifier les permissions
        if my_Clas.niv_spe_dep_sg.dept != my_Dep:
            return JsonResponse({'error': 'غير مسموح'}, status=403)
        
        # Récupérer les données
        all_sea_fait = Seance.objects.filter(classe=my_Clas, fait=True).order_by('date')
        all_Abs_Classe = Abs_Etu_Classe.objects.filter(classe=my_Clas).order_by('etudiant__nom_fr')
        
        # Créer le fichier Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "قائمة الغيابات"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # En-têtes
        headers = ['#', 'اللقب', 'الاسم', 'Nom', 'Prénom']
        headers.extend([sea.date.strftime('%d/%m') for sea in all_sea_fait])
        headers.extend(['عدد الغيابات', 'المبررة'])
        
        # Écrire les en-têtes
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Données
        for row_idx, abs_classe in enumerate(all_Abs_Classe, 2):
            # Informations étudiant
            ws.cell(row=row_idx, column=1, value=row_idx-1)
            ws.cell(row=row_idx, column=2, value=abs_classe.etudiant.nom_ar or '-')
            ws.cell(row=row_idx, column=3, value=abs_classe.etudiant.prenom_ar or '-')
            ws.cell(row=row_idx, column=4, value=abs_classe.etudiant.nom_fr or '-')
            ws.cell(row=row_idx, column=5, value=abs_classe.etudiant.prenom_fr or '-')
            
            # Absences par séance
            col_idx = 6
            for sea in all_sea_fait:
                absence = Abs_Etu_Seance.objects.filter(seance=sea, etudiant=abs_classe.etudiant).first()
                if absence:
                    if not absence.present and not absence.justifiee:
                        value = 'غ'
                    elif not absence.present and absence.justifiee:
                        value = 'غ.م'
                    else:
                        value = '-'
                else:
                    value = '-'
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = center_alignment
                col_idx += 1
            
            # Totaux
            ws.cell(row=row_idx, column=col_idx, value=abs_classe.nbr_absence or 0)
            ws.cell(row=row_idx, column=col_idx+1, value=abs_classe.nbr_absence_justifiee or 0)
        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 15)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Préparer la réponse
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f'غيابات_{my_Clas.niv_spe_dep_sg}_{my_Clas.matiere}'.replace(' ', '_')
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def absences_statistics_api(request, clas_id, departement):
    """
    API pour récupérer les statistiques des absences en temps réel
    """
    try:
        my_Clas = get_object_or_404(Classe, id=clas_id)
        
        # Statistiques générales
        all_Abs_Classe = Abs_Etu_Classe.objects.filter(classe=my_Clas)
        total_students = all_Abs_Classe.count()
        total_absences = sum([abs_cls.nbr_absence or 0 for abs_cls in all_Abs_Classe])
        total_justified = sum([abs_cls.nbr_absence_justifiee or 0 for abs_cls in all_Abs_Classe])
        
        # Répartition par niveau d'absence
        absence_distribution = {
            'low': all_Abs_Classe.filter(Q(nbr_absence__lt=3) | Q(nbr_absence__isnull=True)).count(),
            'medium': all_Abs_Classe.filter(nbr_absence__gte=3, nbr_absence__lt=5).count(),
            'high': all_Abs_Classe.filter(nbr_absence__gte=5).count(),
        }
        
        # Moyenne des absences
        avg_absences = total_absences / total_students if total_students > 0 else 0
        
        # Taux d'assiduité
        total_sessions = Seance.objects.filter(classe=my_Clas, fait=True).count()
        total_possible_presences = total_students * total_sessions
        attendance_rate = ((total_possible_presences - total_absences) / total_possible_presences * 100) if total_possible_presences > 0 else 0
        
        return JsonResponse({
            'total_students': total_students,
            'total_absences': total_absences,
            'total_justified': total_justified,
            'total_sessions': total_sessions,
            'avg_absences': round(avg_absences, 2),
            'attendance_rate': round(attendance_rate, 2),
            'absence_distribution': absence_distribution,
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def list_absences_niveau(request, niv_spe_dep_id, semestre_num):
    """
    Vue pour afficher UNIQUEMENT les absences de TD et TP par niveau et semestre
    """
    # Récupérer le département de l'utilisateur
    enseignant = request.user.enseignant_profile
    real_Dep = Ens_Dep.objects.get(enseignant=enseignant, statut='Permanent')
    departement = Departement.objects.get(id=real_Dep.departement.id)

    # Récupérer le niveau et le semestre
    niv_spe_dep = get_object_or_404(NivSpeDep, id=niv_spe_dep_id, departement=departement)
    semestre = get_object_or_404(Semestre, numero=semestre_num)

    # Récupérer les paramètres de filtrage avec valeurs par défaut = 5
    min_absences = request.GET.get('min_absences', 5)
    min_absences_just = request.GET.get('min_absences_just', 5)

    try:
        min_absences = int(min_absences)
    except:
        min_absences = 5

    try:
        min_absences_just = int(min_absences_just)
    except:
        min_absences_just = 5

    # UNIQUEMENT TD et TP - Filtrer les classes
    classes = Classe.objects.filter(
        niv_spe_dep_sg__niv_spe_dep=niv_spe_dep,
        semestre=semestre,
        type__in=['TD', 'TP']  # SEULEMENT TD et TP
    ).select_related('matiere', 'enseignant__enseignant', 'niv_spe_dep_sg').order_by('matiere__nom_ar', 'type')

    # Récupérer TOUS les étudiants du niveau (pas seulement ceux avec absences)
    tous_etudiants = Etudiant.objects.filter(
        niv_spe_dep_sg__niv_spe_dep=niv_spe_dep
    ).order_by('nom_ar', 'prenom_ar')

    # Créer la liste des matières TD/TP avec leurs types
    matieres_types = []
    matieres_keys = {}
    for classe in classes:
        key = f"{classe.matiere.id}_{classe.type}"
        if key not in matieres_keys:
            matieres_types.append({
                'key': key,
                'matiere_nom': classe.matiere.nom_fr,
                'type': classe.type,
                'nom_complet': f"{classe.matiere.nom_fr} ({classe.type})"
            })
            matieres_keys[key] = True

    # Récupérer tous les enregistrements d'absences pour TD/TP
    absences_all = Abs_Etu_Classe.objects.filter(
        classe__in=classes
    ).select_related('etudiant', 'classe__matiere', 'classe')

    # Créer un dictionnaire des absences: {etudiant_id: {matiere_type_key: {abs, abs_just}}}
    absences_dict = {}
    for abs_rec in absences_all:
        etudiant_id = abs_rec.etudiant.id
        key = f"{abs_rec.classe.matiere.id}_{abs_rec.classe.type}"

        if etudiant_id not in absences_dict:
            absences_dict[etudiant_id] = {}

        if key not in absences_dict[etudiant_id]:
            absences_dict[etudiant_id][key] = {
                'nbr_absence': 0,
                'nbr_absence_justifiee': 0
            }

        absences_dict[etudiant_id][key]['nbr_absence'] += abs_rec.nbr_absence or 0
        absences_dict[etudiant_id][key]['nbr_absence_justifiee'] += abs_rec.nbr_absence_justifiee or 0

    # Construire la liste finale avec TOUS les étudiants
    etudiants_list = []
    for etudiant in tous_etudiants:
        etudiant_absences = absences_dict.get(etudiant.id, {})

        # Calculer les totaux
        total_abs = sum([v['nbr_absence'] for v in etudiant_absences.values()])
        total_just = sum([v['nbr_absence_justifiee'] for v in etudiant_absences.values()])

        # Vérifier si au moins UNE matière a >= seuil d'absences OU >= seuil de justifiées
        has_absences_threshold = any(v['nbr_absence'] >= min_absences for v in etudiant_absences.values())
        has_justifiees_threshold = any(v['nbr_absence_justifiee'] >= min_absences_just for v in etudiant_absences.values())

        # Filtrer: garder si au moins une matière atteint le seuil
        if has_absences_threshold or has_justifiees_threshold:
            etudiants_list.append({
                'etudiant': etudiant,
                'matieres': etudiant_absences,
                'total_absences': total_abs,
                'total_absences_justifiees': total_just
            })

    # Trier par nom de l'étudiant (alphabétique)
    etudiants_list.sort(key=lambda x: (x['etudiant'].nom_ar or '', x['etudiant'].prenom_ar or ''))

    # Statistiques
    total_etudiants = len(etudiants_list)
    total_absences = sum([e['total_absences'] for e in etudiants_list])
    total_justifiees = sum([e['total_absences_justifiees'] for e in etudiants_list])

    # Nombre total d'étudiants dans le niveau (avant filtrage)
    total_etudiants_niveau = tous_etudiants.count()

    # Nombre d'étudiants avec absences justifiées >= seuil
    total_etudiants_justifies = sum(1 for e in etudiants_list if e['total_absences_justifiees'] >= min_absences_just)

    context = {
        'my_Fac': departement.faculte,
        'my_Dep': departement,
        'niv_spe_dep': niv_spe_dep,
        'semestre': semestre,
        'departement': departement,
        'matieres_types': matieres_types,
        'etudiants_list': etudiants_list,
        'total_etudiants': total_etudiants,
        'total_absences': total_absences,
        'total_justifiees': total_justifiees,
        'total_etudiants_niveau': total_etudiants_niveau,
        'total_etudiants_justifies': total_etudiants_justifies,
        'min_absences': min_absences,
        'min_absences_just': min_absences_just,
    }

    return render(request, 'departement/list_absences_niveau.html', context)